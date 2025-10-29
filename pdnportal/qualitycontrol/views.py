from django.shortcuts import render, redirect, get_object_or_404
from django.core.paginator import Paginator
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.views.decorators.http import require_http_methods
from .models import TrialRunRequest, LotOutInspection, CutAway
from .forms import TrialRunRequestForm, LotOutInspectionForm, CutAwayForm
from settings.models import Line
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
from django.utils import timezone

def quality_control_home(request):
    if request.method == 'POST':
        if 'trial_run_id' in request.POST:  # Edit operation
            trial_run = get_object_or_404(TrialRunRequest, id=request.POST['trial_run_id'], requested_by=request.user)
            form = TrialRunRequestForm(request.POST, instance=trial_run)
            if form.is_valid():
                form.save()
                messages.success(request, 'Trial run request updated successfully!')
                return redirect(reverse('quality_control_home') + '?tab=TrialRun')
            else:
                messages.error(request, 'Please correct the errors below.')
        elif 'lot_out_id' in request.POST:  # Edit lot out operation
            lot_out = get_object_or_404(LotOutInspection, id=request.POST['lot_out_id'], requested_by=request.user)
            form = LotOutInspectionForm(request.POST, instance=lot_out)
            if form.is_valid():
                form.save()
                messages.success(request, 'Lot out inspection request updated successfully!')
                return redirect(reverse('quality_control_home') + '?tab=LotOut')
            else:
                messages.error(request, 'Please correct the errors below.')
        elif 'cut_away_id' in request.POST:  # Edit cut-away operation
            cut_away = get_object_or_404(CutAway, id=request.POST['cut_away_id'], requested_by=request.user)
            form = CutAwayForm(request.POST, instance=cut_away)
            if form.is_valid():
                form.save()
                messages.success(request, 'Cut-away request updated successfully!')
                return redirect(reverse('quality_control_home') + '?tab=CutAway')
            else:
                messages.error(request, 'Please correct the errors below.')
        elif request.POST.get('form_type') == 'lot_out_inspection':  # Lot out create operation
            form = LotOutInspectionForm(request.POST)
            if form.is_valid():
                lot_out_request = form.save(commit=False)
                lot_out_request.requested_by = request.user
                lot_out_request.save()
                messages.success(request, 'Lot out inspection request submitted successfully!')
                return redirect(reverse('quality_control_home') + '?tab=LotOut')
            else:
                messages.error(request, 'Please correct the errors below.')
        elif request.POST.get('form_type') == 'cut_away_request':  # Cut-Away create operation
            form = CutAwayForm(request.POST)
            if form.is_valid():
                cut_away_request = form.save(commit=False)
                cut_away_request.requested_by = request.user
                # Set default values for date_prepared
                cut_away_request.date_prepared = timezone.now().date()
                cut_away_request.save()
                messages.success(request, 'Cut-away request submitted successfully!')
                return redirect(reverse('quality_control_home') + '?tab=CutAway')
            else:
                messages.error(request, 'Please correct the errors below.')
        else:  # Trial run create operation
            form = TrialRunRequestForm(request.POST)
            if form.is_valid():
                trial_request = form.save(commit=False)
                trial_request.requested_by = request.user
                trial_request.save()
                messages.success(request, 'Trial run request submitted successfully!')
                return redirect(reverse('quality_control_home') + '?tab=TrialRun')
            else:
                messages.error(request, 'Please correct the errors below.')
    else:
        form = TrialRunRequestForm()
    
    trial_run_requests = TrialRunRequest.objects.all()
    paginator = Paginator(trial_run_requests, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    lot_out_inspections = LotOutInspection.objects.all()
    lot_out_paginator = Paginator(lot_out_inspections, 10)
    lot_out_page_number = request.GET.get('lot_out_page')
    lot_out_page_obj = lot_out_paginator.get_page(lot_out_page_number)
    
    cut_aways = CutAway.objects.all()
    cut_away_paginator = Paginator(cut_aways, 10)
    cut_away_page_number = request.GET.get('cut_away_page')
    cut_away_page_obj = cut_away_paginator.get_page(cut_away_page_number)
    
    context = {
        'trial_run_requests': page_obj,
        'lot_out_inspections': lot_out_page_obj,
        'cut_aways': cut_away_page_obj,
        'form': form,
        'available_lines': Line.objects.all(),
    }
    return render(request, 'qualitycontrol/qualitycontrol.html', context)

def cancel_trial_run(request, trial_run_id):
    if request.method == 'POST':
        trial_run = get_object_or_404(TrialRunRequest, id=trial_run_id, requested_by=request.user)
        trial_run.status = 'cancelled'
        trial_run.save()
        return JsonResponse({'success': True, 'message': 'Trial run request cancelled successfully!'})
    return JsonResponse({'success': False, 'message': 'Invalid request method.'})

def cancel_lot_out(request, lot_out_id):
    if request.method == 'POST':
        lot_out = get_object_or_404(LotOutInspection, id=lot_out_id, requested_by=request.user)
        lot_out.status = 'cancelled'
        lot_out.save()
        return JsonResponse({'success': True, 'message': 'Lot out inspection request cancelled successfully!'})
    return JsonResponse({'success': False, 'message': 'Invalid request method.'})

def cancel_cut_away(request, cut_away_id):
    if request.method == 'POST':
        cut_away = get_object_or_404(CutAway, id=cut_away_id, requested_by=request.user)
        cut_away.status_of_cut_away = 'cancelled'
        cut_away.save()
        return JsonResponse({'success': True, 'message': 'Cut-away request cancelled successfully!'})
    return JsonResponse({'success': False, 'message': 'Invalid request method.'})

def qa_cut_away(request, cut_away_id):
    if request.method == 'POST':
        cut_away = get_object_or_404(CutAway, id=cut_away_id)
        
        # Update the cut-away with QA information
        cut_away.schedule_of_cut_away = request.POST.get('schedule_of_cut_away')
        cut_away.qa_remarks = request.POST.get('qa_remarks', '')
        cut_away.received_date_by_qa = timezone.now().date()
        cut_away.qa_inspector = request.user
        cut_away.status_of_cut_away = 'scheduled'
        cut_away.save()
        
        return JsonResponse({'success': True, 'message': 'Cut-away request processed successfully!'})
    return JsonResponse({'success': False, 'message': 'Invalid request method.'})

def search_trial_runs(request):
    search_query = request.GET.get('q', '').strip()
    
    if not search_query:
        # Return all trial runs if no search query
        trial_runs = TrialRunRequest.objects.all()
    else:
        # Search across multiple fields
        from django.db.models import Q
        trial_runs = TrialRunRequest.objects.filter(
            Q(trr_number__icontains=search_query) |
            Q(product_name__icontains=search_query) |
            Q(product_number__icontains=search_query) |
            Q(line__line_name__icontains=search_query) |
            Q(process_name__icontains=search_query) |
            Q(trial_operator__icontains=search_query) |
            Q(requested_by__name__icontains=search_query)
        )
    
    # Prepare the data to return
    results = []
    for trial_run in trial_runs:
        results.append({
            'id': trial_run.id,
            'trr_number': trial_run.trr_number,
            'trial_date': trial_run.trial_date.strftime('%b %d, %Y'),
            'product_name': trial_run.product_name,
            'product_number': trial_run.product_number,
            'line': trial_run.line.line_name,
            'process_name': trial_run.process_name,
            'trial_operator': trial_run.trial_operator,
            'requested_by': trial_run.requested_by.name,
            'status': trial_run.status,
            'created_at': trial_run.created_at.strftime('%b %d, %Y'),
            'updated_at': trial_run.updated_at.strftime('%b %d, %Y'),
            'is_updated': trial_run.updated_at != trial_run.created_at,
            'change_from': trial_run.change_from,
            'change_to': trial_run.change_to,
            'is_owner': trial_run.requested_by == request.user,
        })
    
    return JsonResponse({
        'success': True,
        'results': results,
        'count': len(results)
    })

def search_lot_out(request):
    search_query = request.GET.get('q', '').strip()
    
    if not search_query:
        # Return all lot out inspections if no search query
        lot_out_inspections = LotOutInspection.objects.all()
    else:
        # Search across multiple fields
        from django.db.models import Q
        lot_out_inspections = LotOutInspection.objects.filter(
            Q(rli_number__icontains=search_query) |
            Q(product_name__icontains=search_query) |
            Q(product_number__icontains=search_query) |
            Q(line_affected__line_name__icontains=search_query) |
            Q(reason_for_lot_out__icontains=search_query) |
            Q(requested_by__name__icontains=search_query)
        )
    
    # Prepare the data to return
    results = []
    for lot_out in lot_out_inspections:
        results.append({
            'id': lot_out.id,
            'rli_number': lot_out.rli_number,
            'request_date': lot_out.request_date.strftime('%b %d, %Y'),
            'product_name': lot_out.product_name,
            'product_number': lot_out.product_number,
            'line_affected': lot_out.line_affected.line_name,
            'line_affected_id': lot_out.line_affected.id,
            'reason_for_lot_out': lot_out.reason_for_lot_out,
            'requested_by': lot_out.requested_by.name,
            'status': lot_out.status,
            'created_at': lot_out.created_at.strftime('%b %d, %Y'),
            'updated_at': lot_out.updated_at.strftime('%b %d, %Y'),
            'is_updated': lot_out.updated_at != lot_out.created_at,
            'is_owner': lot_out.requested_by == request.user,
        })
    
    return JsonResponse({
        'success': True,
        'results': results,
        'count': len(results)
    })

def export_trial_run_report(request):
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    status = request.GET.get('status')
    
    if not date_from or not date_to or not status:
        messages.error(request, 'Please provide all required fields.')
        return redirect('quality_control_home')
    
    # Convert string dates to date objects
    try:
        date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
        date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, 'Invalid date format.')
        return redirect('quality_control_home')
    
    # Filter trial runs by date range
    trial_runs = TrialRunRequest.objects.filter(
        trial_date__gte=date_from_obj,
        trial_date__lte=date_to_obj
    )
    
    # Filter by status if not 'all'
    if status != 'all':
        trial_runs = trial_runs.filter(status=status)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Trial Run Report"
    
    # Format dates for display
    date_from_formatted = date_from_obj.strftime('%B %d, %Y')
    date_to_formatted = date_to_obj.strftime('%B %d, %Y')
    
    # Add header
    ws.merge_cells('A1:K1')
    header_cell = ws['A1']
    header_cell.value = "Ryonan Electric Philippines Corporation"
    header_cell.font = Font(bold=True, size=14)
    header_cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Add subtitle
    ws.merge_cells('A2:K2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = f"Trial Run Summary Report for {date_from_formatted} to {date_to_formatted}"
    subtitle_cell.font = Font(italic=True, size=11)
    subtitle_cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Add empty row
    ws.append([])
    
    # Define table headers
    headers = [
        'TRR Number',
        'Date of Trial',
        'Product Name',
        'Product Number',
        'Line',
        'Process Name',
        'Trial Operator',
        'Change From',
        'Change To',
        'Requested By',
        'Status'
    ]
    
    # Add headers with formatting
    ws.append(headers)
    header_row = ws[4]
    
    # Style for headers
    header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    header_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in header_row:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data rows
    for trial_run in trial_runs:
        row_data = [
            trial_run.trr_number,
            trial_run.trial_date.strftime('%B %d, %Y'),
            trial_run.product_name,
            trial_run.product_number,
            trial_run.line.line_name,
            trial_run.process_name,
            trial_run.trial_operator,
            trial_run.change_from,
            trial_run.change_to,
            trial_run.requested_by.name,
            'Submitted' if trial_run.status == 'submitted' else 'Cancelled'
        ]
        ws.append(row_data)
        
        # Get the current row
        current_row = ws.max_row
        
        # Apply borders to all cells in the row
        for col_num in range(1, 12):
            cell = ws.cell(row=current_row, column=col_num)
            cell.border = border
            
            # Apply status color
            if col_num == 11:  # Status column
                if trial_run.status == 'submitted':
                    cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
    
    # Adjust column widths
    column_widths = {
        'A': 15,  # TRR Number
        'B': 18,  # Date of Trial
        'C': 25,  # Product Name
        'D': 18,  # Product Number
        'E': 15,  # Line
        'F': 20,  # Process Name
        'G': 20,  # Trial Operator
        'H': 30,  # Change From
        'I': 30,  # Change To
        'J': 20,  # Requested By
        'K': 12   # Status
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'Trial_Run_Report_{date_from}_to_{date_to}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

def export_lot_out_report(request):
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    status = request.GET.get('status')
    
    if not date_from or not date_to or not status:
        messages.error(request, 'Please provide all required fields.')
        return redirect('quality_control_home')
    
    # Convert string dates to date objects
    try:
        date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
        date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, 'Invalid date format.')
        return redirect('quality_control_home')
    
    # Filter lot out inspections by date range
    lot_out_inspections = LotOutInspection.objects.filter(
        request_date__gte=date_from_obj,
        request_date__lte=date_to_obj
    )
    
    # Filter by status if not 'all'
    if status != 'all':
        lot_out_inspections = lot_out_inspections.filter(status=status)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Lot Out Inspection Report"
    
    # Format dates for display
    date_from_formatted = date_from_obj.strftime('%B %d, %Y')
    date_to_formatted = date_to_obj.strftime('%B %d, %Y')
    
    # Add header
    ws.merge_cells('A1:H1')
    header_cell = ws['A1']
    header_cell.value = "Ryonan Electric Philippines Corporation"
    header_cell.font = Font(bold=True, size=14)
    header_cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Add subtitle
    ws.merge_cells('A2:H2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = f"Lot Out Inspection Summary Report for {date_from_formatted} to {date_to_formatted}"
    subtitle_cell.font = Font(italic=True, size=11)
    subtitle_cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Add empty row
    ws.append([])
    
    # Define table headers
    headers = [
        'RLI Number',
        'Request Date',
        'Product Name',
        'Product Number',
        'Line Affected',
        'Reason for Lot Out',
        'Requested By',
        'Status'
    ]
    
    # Add headers with formatting
    ws.append(headers)
    header_row = ws[4]
    
    # Style for headers
    header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    header_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in header_row:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data rows
    for lot_out in lot_out_inspections:
        row_data = [
            lot_out.rli_number,
            lot_out.request_date.strftime('%B %d, %Y'),
            lot_out.product_name,
            lot_out.product_number,
            lot_out.line_affected.line_name,
            lot_out.reason_for_lot_out,
            lot_out.requested_by.name,
            'Submitted' if lot_out.status == 'submitted' else 'Cancelled'
        ]
        ws.append(row_data)
        
        # Get the current row
        current_row = ws.max_row
        
        # Apply borders to all cells in the row
        for col_num in range(1, 9):
            cell = ws.cell(row=current_row, column=col_num)
            cell.border = border
            
            # Apply status color
            if col_num == 8:  # Status column
                if lot_out.status == 'submitted':
                    cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
    
    # Adjust column widths
    column_widths = {
        'A': 15,  # RLI Number
        'B': 18,  # Request Date
        'C': 25,  # Product Name
        'D': 18,  # Product Number
        'E': 15,  # Line Affected
        'F': 40,  # Reason for Lot Out
        'G': 20,  # Requested By
        'H': 12   # Status
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'Lot_Out_Inspection_Report_{date_from}_to_{date_to}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@require_http_methods(["GET"])
def export_cut_away_report(request):
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    status = request.GET.get('status')
    
    if not date_from or not date_to or not status:
        messages.error(request, 'Please provide all required fields.')
        return redirect('quality_control_home')
    
    # Convert string dates to date objects
    try:
        date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
        date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, 'Invalid date format.')
        return redirect('quality_control_home')
    
    # Filter cut-away requests by date range
    cut_aways = CutAway.objects.filter(
        date_prepared__gte=date_from_obj,
        date_prepared__lte=date_to_obj
    )
    
    # Filter by status if not 'all'
    if status != 'all':
        cut_aways = cut_aways.filter(status_of_cut_away=status)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Cut-Away Report"
    
    # Format dates for display
    date_from_formatted = date_from_obj.strftime('%B %d, %Y')
    date_to_formatted = date_to_obj.strftime('%B %d, %Y')
    
    # Add header
    ws.merge_cells('A1:P1')
    header_cell = ws['A1']
    header_cell.value = "Ryonan Electric Philippines Corporation"
    header_cell.font = Font(bold=True, size=14)
    header_cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Add subtitle
    ws.merge_cells('A2:P2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = f"Cut-Away Summary Report for {date_from_formatted} to {date_to_formatted}"
    subtitle_cell.font = Font(italic=True, size=11)
    subtitle_cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Add empty row
    ws.append([])
    
    # Define table headers
    headers = [
        'Control Number',
        'Date Prepared',
        'Assembly Line',
        'Product Number',
        'Schedule Date',
        'Date Received',
        'Machine Number',
        'Applicator Code',
        'Crimped Quantity',
        'Terminal Part Number',
        'Purpose',
        'Reason for Changing Parts',
        'Leadwire Partnumber',
        'Rubber Seal Partnumber',
        'Requested By',
        'Status'
    ]
    
    # Add headers with formatting
    ws.append(headers)
    header_row = ws[4]
    
    # Style for headers
    header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    header_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in header_row:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data rows
    for cut_away in cut_aways:
        status_text = {
            'pending': 'Pending',
            'scheduled': 'Scheduled',
            'in_progress': 'In Progress',
            'completed': 'Completed',
            'cancelled': 'Cancelled'
        }.get(cut_away.status_of_cut_away, cut_away.status_of_cut_away)
        
        # Map purpose codes to readable labels
        purpose_labels = {
            'new_product': 'New Product',
            'reached_50000': 'Reached 50,000',
            'reached_100000': 'Reached 100,000',
            'new_applicator': 'New Applicator',
            'change_wire_crimper': 'Change Wire Crimper',
            'change_wire_anvil': 'Change Wire Anvil',
            'new_combination': 'New Combination',
            'big_burr': 'Big Burr',
            'change_supplier': 'Change Supplier'
        }
        
        row_data = [
            cut_away.control_number,
            cut_away.date_prepared.strftime('%B %d, %Y'),
            cut_away.assy_line.line_name,
            cut_away.product_number,
            cut_away.schedule_of_cut_away.strftime('%B %d, %Y') if cut_away.schedule_of_cut_away else 'Not Set',
            cut_away.received_date_by_qa.strftime('%B %d, %Y') if cut_away.received_date_by_qa else 'Not Set',
            cut_away.machine_number or 'N/A',
            cut_away.applicator_number_code or 'N/A',
            cut_away.crimped_quantity or 'N/A',
            cut_away.terminal_part_number or 'N/A',
            purpose_labels.get(cut_away.purpose, 'N/A') if cut_away.purpose else 'N/A',
            cut_away.purpose_reason or 'N/A',
            cut_away.leadwire_partnumber or 'N/A',
            cut_away.rubber_seal_partnumber or 'N/A',
            cut_away.requested_by.name,
            status_text
        ]
        ws.append(row_data)
        
        # Get the current row
        current_row = ws.max_row
        
        # Apply borders to all cells in the row
        for col_num in range(1, 17):
            cell = ws.cell(row=current_row, column=col_num)
            cell.border = border
            
            # Apply status color
            if col_num == 16:  # Status column (now column P)
                if cut_away.status_of_cut_away == 'completed':
                    cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                elif cut_away.status_of_cut_away == 'cancelled':
                    cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                elif cut_away.status_of_cut_away == 'in_progress':
                    cell.fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')
                elif cut_away.status_of_cut_away == 'scheduled':
                    cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
                else:  # pending
                    cell.fill = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')
    
    # Adjust column widths
    column_widths = {
        'A': 18,  # Control Number
        'B': 18,  # Date Prepared
        'C': 15,  # Assembly Line
        'D': 18,  # Product Number
        'E': 18,  # Schedule Date
        'F': 18,  # Date Received
        'G': 18,  # Machine Number
        'H': 18,  # Applicator Code
        'I': 18,  # Crimped Quantity
        'J': 20,  # Terminal Part Number
        'K': 20,  # Purpose
        'L': 25,  # Reason for Changing Parts
        'M': 25,  # Leadwire Partnumber
        'N': 25,  # Rubber Seal Partnumber
        'O': 20,  # Requested By
        'P': 15   # Status
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'Cut_Away_Report_{date_from}_to_{date_to}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
