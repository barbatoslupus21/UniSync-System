from django.shortcuts import render, redirect, get_object_or_404
from django.core.paginator import Paginator
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta, date
from django.utils import timezone
import json

from .models import Meeting, MeetingRoom, MeetingType
from .forms import MeetingForm, MeetingRoomForm, MeetingTypeForm
from portalusers.models import Users


@login_required
def meeting_list(request):
    """Main view for the meeting scheduler page."""
    rooms = MeetingRoom.objects.filter(is_active=True)
    meeting_types = MeetingType.objects.filter(is_active=True)
    # Get all active users for the attendees list
    users = Users.objects.filter(is_active=True).exclude(name__isnull=True).exclude(name='').order_by('name')
    
    # Get today's meetings by default
    today = date.today()
    today_meetings = Meeting.objects.filter(
        date=today,
        status__in=['scheduled', 'in_progress']
    ).order_by('start_time')
    
    # Check if current user is a meeting admin
    is_meeting_admin = getattr(request.user, 'meetingadmin', False)
    
    context = {
        'rooms': rooms,
        'meeting_types': meeting_types,
        'users': users,
        'today_meetings': today_meetings,
        'today': today,
        'is_meeting_admin': is_meeting_admin,
    }
    return render(request, 'meetingscheduler/meeting.html', context)


@login_required
def get_meetings_by_date(request):
    """API endpoint to get meetings for a specific date."""
    date_str = request.GET.get('date')
    room_id = request.GET.get('room')
    
    if not date_str:
        return JsonResponse({'error': 'Date parameter is required'}, status=400)
    
    try:
        selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Invalid date format. Use YYYY-MM-DD'}, status=400)
    
    meetings = Meeting.objects.filter(
        date=selected_date,
        status__in=['scheduled', 'in_progress']
    ).select_related('room', 'meeting_type', 'organizer')
    
    if room_id:
        meetings = meetings.filter(room_id=room_id)
    
    meetings = meetings.order_by('start_time')
    
    meetings_data = []
    for meeting in meetings:
        attendees = list(meeting.attendees.values('id', 'name', 'avatar'))
        meetings_data.append({
            'id': meeting.id,
            'title': meeting.title,
            'description': meeting.description,
            'room': {
                'id': meeting.room.id,
                'name': meeting.room.name,
                'color': meeting.room.color
            },
            'meeting_type': {
                'id': meeting.meeting_type.id if meeting.meeting_type else None,
                'name': meeting.meeting_type.name if meeting.meeting_type else 'General',
                'icon': meeting.meeting_type.icon if meeting.meeting_type else 'fa-users',
                'color': meeting.meeting_type.color if meeting.meeting_type else '#3366ff'
            } if meeting.meeting_type else None,
            'date': meeting.date.strftime('%Y-%m-%d'),
            'start_time': meeting.start_time.strftime('%H:%M'),
            'end_time': meeting.end_time.strftime('%H:%M'),
            'location': meeting.location,
            'organizer': {
                'id': meeting.organizer.id,
                'name': meeting.organizer.name or meeting.organizer.username,
                'avatar': meeting.organizer.avatar.url if meeting.organizer.avatar else None
            },
            'attendees': attendees,
            'status': meeting.status,
            'is_past': meeting.is_past
        })
    
    return JsonResponse({
        'meetings': meetings_data,
        'date': selected_date.strftime('%Y-%m-%d'),
        'formatted_date': selected_date.strftime('%B %d, %Y')
    })


@login_required
def get_calendar_events(request):
    """API endpoint to get calendar events for a month with occupancy calculation."""
    year = request.GET.get('year', datetime.now().year)
    month = request.GET.get('month', datetime.now().month)
    room_id = request.GET.get('room')
    
    try:
        year = int(year)
        month = int(month)
    except ValueError:
        return JsonResponse({'error': 'Invalid year or month'}, status=400)
    
    # Get first and last day of month
    first_day = date(year, month, 1)
    if month == 12:
        last_day = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(year, month + 1, 1) - timedelta(days=1)
    
    # Get all active rooms for occupancy calculation
    active_rooms = MeetingRoom.objects.filter(is_active=True)
    total_rooms = active_rooms.count()
    
    meetings = Meeting.objects.filter(
        date__gte=first_day,
        date__lte=last_day,
        status__in=['scheduled', 'in_progress']
    ).select_related('room', 'meeting_type')
    
    if room_id:
        meetings = meetings.filter(room_id=room_id)
    
    # Group meetings by date and calculate occupancy
    events_by_date = {}
    occupancy_by_date = {}
    
    # Define working hours (8 AM to 6 PM = 10 hours = 600 minutes per room)
    working_minutes_per_room = 600  # 10 hours * 60 minutes
    
    for meeting in meetings:
        date_str = meeting.date.strftime('%Y-%m-%d')
        if date_str not in events_by_date:
            events_by_date[date_str] = []
            occupancy_by_date[date_str] = {'rooms_used': set(), 'total_minutes': 0}
        
        events_by_date[date_str].append({
            'id': meeting.id,
            'title': meeting.title,
            'start_time': meeting.start_time.strftime('%H:%M'),
            'end_time': meeting.end_time.strftime('%H:%M'),
            'room': meeting.room.name,
            'room_id': meeting.room.id,
            'color': meeting.meeting_type.color if meeting.meeting_type else meeting.room.color
        })
        
        # Calculate meeting duration in minutes
        start_dt = datetime.combine(meeting.date, meeting.start_time)
        end_dt = datetime.combine(meeting.date, meeting.end_time)
        duration = int((end_dt - start_dt).total_seconds() / 60)
        
        occupancy_by_date[date_str]['rooms_used'].add(meeting.room.id)
        occupancy_by_date[date_str]['total_minutes'] += duration
    
    # Calculate occupancy percentage for each date
    occupancy_percentages = {}
    if total_rooms > 0:
        total_available_minutes = total_rooms * working_minutes_per_room
        for date_str, data in occupancy_by_date.items():
            # Calculate percentage based on total booked minutes vs total available
            percentage = min(100, int((data['total_minutes'] / total_available_minutes) * 100))
            occupancy_percentages[date_str] = percentage
    
    return JsonResponse({
        'events': events_by_date,
        'occupancy': occupancy_percentages,
        'year': year,
        'month': month,
        'total_rooms': total_rooms
    })


@login_required
@require_http_methods(["POST"])
def create_meeting(request):
    """API endpoint to create a new meeting."""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    
    # Validate required fields
    required_fields = ['title', 'room', 'date', 'start_time', 'end_time']
    for field in required_fields:
        if field not in data or not data[field]:
            return JsonResponse({'error': f'{field} is required'}, status=400)
    
    try:
        room = MeetingRoom.objects.get(id=data['room'], is_active=True)
        meeting_date = datetime.strptime(data['date'], '%Y-%m-%d').date()
        start_time = datetime.strptime(data['start_time'], '%H:%M').time()
        end_time = datetime.strptime(data['end_time'], '%H:%M').time()
        
        # Check for overlapping meetings
        is_available, conflicting_meeting = Meeting.check_availability(
            room, meeting_date, start_time, end_time
        )
        
        if not is_available:
            return JsonResponse({
                'error': f'Time slot conflicts with "{conflicting_meeting.title}" '
                         f'({conflicting_meeting.start_time.strftime("%H:%M")} - '
                         f'{conflicting_meeting.end_time.strftime("%H:%M")})'
            }, status=400)
        
        meeting_type = None
        if data.get('meeting_type'):
            meeting_type = MeetingType.objects.filter(id=data['meeting_type']).first()
        
        meeting = Meeting.objects.create(
            title=data['title'],
            description=data.get('description', ''),
            room=room,
            meeting_type=meeting_type,
            date=meeting_date,
            start_time=start_time,
            end_time=end_time,
            location=data.get('location', ''),
            organizer=request.user
        )
        
        # Add attendees
        if data.get('attendees'):
            attendees = Users.objects.filter(id__in=data['attendees'])
            meeting.attendees.set(attendees)
        
        return JsonResponse({
            'success': True,
            'message': 'Meeting created successfully',
            'meeting': {
                'id': meeting.id,
                'title': meeting.title,
                'date': meeting.date.strftime('%Y-%m-%d'),
                'start_time': meeting.start_time.strftime('%H:%M'),
                'end_time': meeting.end_time.strftime('%H:%M')
            }
        })
        
    except MeetingRoom.DoesNotExist:
        return JsonResponse({'error': 'Meeting room not found'}, status=404)
    except ValueError as e:
        return JsonResponse({'error': str(e)}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["PUT"])
def update_meeting(request, meeting_id):
    """API endpoint to update an existing meeting."""
    try:
        meeting = Meeting.objects.get(id=meeting_id)
    except Meeting.DoesNotExist:
        return JsonResponse({'error': 'Meeting not found'}, status=404)
    
    # Check if user has permission to edit
    if meeting.organizer != request.user and not request.user.is_admin:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    
    try:
        if 'room' in data:
            meeting.room = MeetingRoom.objects.get(id=data['room'], is_active=True)
        if 'date' in data:
            meeting.date = datetime.strptime(data['date'], '%Y-%m-%d').date()
        if 'start_time' in data:
            meeting.start_time = datetime.strptime(data['start_time'], '%H:%M').time()
        if 'end_time' in data:
            meeting.end_time = datetime.strptime(data['end_time'], '%H:%M').time()
        if 'title' in data:
            meeting.title = data['title']
        if 'description' in data:
            meeting.description = data['description']
        if 'location' in data:
            meeting.location = data['location']
        if 'meeting_type' in data:
            meeting.meeting_type = MeetingType.objects.filter(id=data['meeting_type']).first()
        if 'status' in data:
            meeting.status = data['status']
        
        # Check for overlapping meetings (exclude current meeting)
        is_available, conflicting_meeting = Meeting.check_availability(
            meeting.room, meeting.date, meeting.start_time, meeting.end_time, meeting.id
        )
        
        if not is_available:
            return JsonResponse({
                'error': f'Time slot conflicts with "{conflicting_meeting.title}" '
                         f'({conflicting_meeting.start_time.strftime("%H:%M")} - '
                         f'{conflicting_meeting.end_time.strftime("%H:%M")})'
            }, status=400)
        
        meeting.save()
        
        # Update attendees if provided
        if 'attendees' in data:
            attendees = Users.objects.filter(id__in=data['attendees'])
            meeting.attendees.set(attendees)
        
        return JsonResponse({
            'success': True,
            'message': 'Meeting updated successfully'
        })
        
    except MeetingRoom.DoesNotExist:
        return JsonResponse({'error': 'Meeting room not found'}, status=404)
    except ValueError as e:
        return JsonResponse({'error': str(e)}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["DELETE"])
def delete_meeting(request, meeting_id):
    """API endpoint to delete a meeting."""
    try:
        meeting = Meeting.objects.get(id=meeting_id)
    except Meeting.DoesNotExist:
        return JsonResponse({'error': 'Meeting not found'}, status=404)
    
    # Check if user has permission to delete
    if meeting.organizer != request.user and not request.user.is_admin:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    meeting.delete()
    
    return JsonResponse({
        'success': True,
        'message': 'Meeting deleted successfully'
    })


@login_required
def get_meeting_detail(request, meeting_id):
    """API endpoint to get detailed meeting information."""
    try:
        meeting = Meeting.objects.select_related(
            'room', 'meeting_type', 'organizer'
        ).prefetch_related('attendees').get(id=meeting_id)
    except Meeting.DoesNotExist:
        return JsonResponse({'error': 'Meeting not found'}, status=404)
    
    attendees = [{
        'id': user.id,
        'name': user.name or user.username,
        'avatar': user.avatar.url if user.avatar else None
    } for user in meeting.attendees.all()]
    
    return JsonResponse({
        'id': meeting.id,
        'title': meeting.title,
        'description': meeting.description,
        'room': {
            'id': meeting.room.id,
            'name': meeting.room.name,
            'color': meeting.room.color
        },
        'meeting_type': {
            'id': meeting.meeting_type.id if meeting.meeting_type else None,
            'name': meeting.meeting_type.name if meeting.meeting_type else 'General',
            'icon': meeting.meeting_type.icon if meeting.meeting_type else 'fa-users',
            'color': meeting.meeting_type.color if meeting.meeting_type else '#3366ff'
        } if meeting.meeting_type else None,
        'date': meeting.date.strftime('%Y-%m-%d'),
        'start_time': meeting.start_time.strftime('%H:%M'),
        'end_time': meeting.end_time.strftime('%H:%M'),
        'location': meeting.location,
        'organizer': {
            'id': meeting.organizer.id,
            'name': meeting.organizer.name or meeting.organizer.username,
            'avatar': meeting.organizer.avatar.url if meeting.organizer.avatar else None
        },
        'attendees': attendees,
        'status': meeting.status,
        'is_past': meeting.is_past,
        'can_edit': meeting.organizer == request.user or request.user.is_admin
    })


@login_required
def check_availability(request):
    """API endpoint to check if a time slot is available."""
    room_id = request.GET.get('room')
    date_str = request.GET.get('date')
    start_time_str = request.GET.get('start_time')
    end_time_str = request.GET.get('end_time')
    exclude_id = request.GET.get('exclude')
    
    if not all([room_id, date_str, start_time_str, end_time_str]):
        return JsonResponse({'error': 'Missing required parameters'}, status=400)
    
    try:
        room = MeetingRoom.objects.get(id=room_id)
        meeting_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        start_time = datetime.strptime(start_time_str, '%H:%M').time()
        end_time = datetime.strptime(end_time_str, '%H:%M').time()
        
        is_available, conflicting_meeting = Meeting.check_availability(
            room, meeting_date, start_time, end_time, exclude_id
        )
        
        if is_available:
            return JsonResponse({'available': True})
        else:
            return JsonResponse({
                'available': False,
                'conflict': {
                    'id': conflicting_meeting.id,
                    'title': conflicting_meeting.title,
                    'start_time': conflicting_meeting.start_time.strftime('%H:%M'),
                    'end_time': conflicting_meeting.end_time.strftime('%H:%M')
                }
            })
            
    except MeetingRoom.DoesNotExist:
        return JsonResponse({'error': 'Room not found'}, status=404)
    except ValueError:
        return JsonResponse({'error': 'Invalid date or time format'}, status=400)


@login_required
def get_rooms(request):
    """API endpoint to get all active meeting rooms."""
    rooms = MeetingRoom.objects.filter(is_active=True).values(
        'id', 'name', 'description', 'capacity', 'location', 'color'
    )
    return JsonResponse({'rooms': list(rooms)})


@login_required
def get_meeting_types(request):
    """API endpoint to get all active meeting types."""
    types = MeetingType.objects.filter(is_active=True).values(
        'id', 'name', 'icon', 'color'
    )
    return JsonResponse({'types': list(types)})


# =============================================================================
# Admin API Endpoints for Meeting Settings
# =============================================================================

def check_meeting_admin(user):
    """Check if user is a meeting admin."""
    return getattr(user, 'meetingadmin', False)


@login_required
@require_http_methods(["GET"])
def get_all_rooms(request):
    """API endpoint to get all meeting rooms (including inactive) for admin."""
    if not check_meeting_admin(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    rooms = MeetingRoom.objects.all().values(
        'id', 'name', 'description', 'capacity', 'location', 'color', 'is_active'
    )
    return JsonResponse({'rooms': list(rooms)})


@login_required
@require_http_methods(["POST"])
def create_room(request):
    """API endpoint to create a new meeting room."""
    if not check_meeting_admin(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        data = json.loads(request.body)
        room = MeetingRoom.objects.create(
            name=data.get('name'),
            description=data.get('description', ''),
            capacity=int(data.get('capacity', 10)),
            location=data.get('location', ''),
            color=data.get('color', '#3366ff'),
            is_active=data.get('is_active', True)
        )
        return JsonResponse({
            'success': True,
            'room': {
                'id': room.id,
                'name': room.name,
                'description': room.description,
                'capacity': room.capacity,
                'location': room.location,
                'color': room.color,
                'is_active': room.is_active
            }
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["PUT"])
def update_room(request, room_id):
    """API endpoint to update a meeting room."""
    if not check_meeting_admin(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        room = get_object_or_404(MeetingRoom, id=room_id)
        data = json.loads(request.body)
        
        room.name = data.get('name', room.name)
        room.description = data.get('description', room.description)
        room.capacity = int(data.get('capacity', room.capacity))
        room.location = data.get('location', room.location)
        room.color = data.get('color', room.color)
        room.is_active = data.get('is_active', room.is_active)
        room.save()
        
        return JsonResponse({
            'success': True,
            'room': {
                'id': room.id,
                'name': room.name,
                'description': room.description,
                'capacity': room.capacity,
                'location': room.location,
                'color': room.color,
                'is_active': room.is_active
            }
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["DELETE"])
def delete_room(request, room_id):
    """API endpoint to delete a meeting room."""
    if not check_meeting_admin(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        room = get_object_or_404(MeetingRoom, id=room_id)
        # Check if room has any meetings
        if room.meetings.exists():
            # Soft delete - just deactivate
            room.is_active = False
            room.save()
            return JsonResponse({'success': True, 'soft_deleted': True, 'message': 'Room deactivated (has existing meetings)'})
        else:
            room.delete()
            return JsonResponse({'success': True, 'soft_deleted': False})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
def get_all_meeting_types(request):
    """API endpoint to get all meeting types (including inactive) for admin."""
    if not check_meeting_admin(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    types = MeetingType.objects.all().values(
        'id', 'name', 'icon', 'color', 'is_active'
    )
    return JsonResponse({'types': list(types)})


@login_required
@require_http_methods(["POST"])
def create_meeting_type(request):
    """API endpoint to create a new meeting type."""
    if not check_meeting_admin(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        data = json.loads(request.body)
        meeting_type = MeetingType.objects.create(
            name=data.get('name'),
            icon=data.get('icon', 'fa-users'),
            color=data.get('color', '#3366ff'),
            is_active=data.get('is_active', True)
        )
        return JsonResponse({
            'success': True,
            'type': {
                'id': meeting_type.id,
                'name': meeting_type.name,
                'icon': meeting_type.icon,
                'color': meeting_type.color,
                'is_active': meeting_type.is_active
            }
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["PUT"])
def update_meeting_type(request, type_id):
    """API endpoint to update a meeting type."""
    if not check_meeting_admin(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        meeting_type = get_object_or_404(MeetingType, id=type_id)
        data = json.loads(request.body)
        
        meeting_type.name = data.get('name', meeting_type.name)
        meeting_type.icon = data.get('icon', meeting_type.icon)
        meeting_type.color = data.get('color', meeting_type.color)
        meeting_type.is_active = data.get('is_active', meeting_type.is_active)
        meeting_type.save()
        
        return JsonResponse({
            'success': True,
            'type': {
                'id': meeting_type.id,
                'name': meeting_type.name,
                'icon': meeting_type.icon,
                'color': meeting_type.color,
                'is_active': meeting_type.is_active
            }
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["DELETE"])
def delete_meeting_type(request, type_id):
    """API endpoint to delete a meeting type."""
    if not check_meeting_admin(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        meeting_type = get_object_or_404(MeetingType, id=type_id)
        # Check if type has any meetings
        if meeting_type.meetings.exists():
            # Soft delete - just deactivate
            meeting_type.is_active = False
            meeting_type.save()
            return JsonResponse({'success': True, 'soft_deleted': True, 'message': 'Meeting type deactivated (has existing meetings)'})
        else:
            meeting_type.delete()
            return JsonResponse({'success': True, 'soft_deleted': False})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
def export_meetings(request):
    """API endpoint to export meetings to Excel for a specific month/year."""
    if not check_meeting_admin(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        month = int(request.GET.get('month', datetime.now().month))
        year = int(request.GET.get('year', datetime.now().year))
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Invalid month or year'}, status=400)
    
    # Validate month and year
    if month < 1 or month > 12:
        return JsonResponse({'error': 'Invalid month'}, status=400)
    if year < 2000 or year > 2100:
        return JsonResponse({'error': 'Invalid year'}, status=400)
    
    # Get first and last day of the month
    first_day = date(year, month, 1)
    if month == 12:
        last_day = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(year, month + 1, 1) - timedelta(days=1)
    
    # Get meetings for the selected month
    meetings = Meeting.objects.filter(
        date__gte=first_day,
        date__lte=last_day
    ).select_related('room', 'meeting_type', 'organizer').prefetch_related('attendees').order_by('date', 'start_time')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Meeting Schedule"
    
    # Define styles
    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=12)
    table_header_font = Font(bold=True)
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Get month name
    month_names = ['January', 'February', 'March', 'April', 'May', 'June',
                   'July', 'August', 'September', 'October', 'November', 'December']
    month_name = month_names[month - 1]
    
    # Header row - Company name
    ws.merge_cells('A1:G1')
    ws['A1'] = 'Ryonan Electric Philippines Corporation'
    ws['A1'].font = header_font
    ws['A1'].alignment = center_align
    
    # Subheader row - Report title with month/year
    ws.merge_cells('A2:G2')
    ws['A2'] = f'Meeting Schedule as of {month_name} {year}'
    ws['A2'].font = subheader_font
    ws['A2'].alignment = center_align
    
    # Empty row
    ws.append([])
    
    # Table headers
    table_headers = ['Date', 'Title', 'Room', 'Type', 'Time', 'Organizer', 'Attendees']
    ws.append(table_headers)
    
    # Style table headers (row 4)
    for col_num, header in enumerate(table_headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = table_header_font
        cell.fill = yellow_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # Add meeting data
    for meeting in meetings:
        # Get attendees names
        attendees_list = [a.name or a.username for a in meeting.attendees.all()]
        attendees_str = ', '.join(attendees_list) if attendees_list else 'None'
        
        # Format time
        time_str = f"{meeting.start_time.strftime('%I:%M %p')} - {meeting.end_time.strftime('%I:%M %p')}"
        
        row_data = [
            meeting.date.strftime('%Y-%m-%d'),
            meeting.title,
            meeting.room.name if meeting.room else 'N/A',
            meeting.meeting_type.name if meeting.meeting_type else 'General',
            time_str,
            meeting.organizer.name or meeting.organizer.username if meeting.organizer else 'N/A',
            attendees_str
        ]
        ws.append(row_data)
        
        # Apply borders to data row
        current_row = ws.max_row
        for col_num in range(1, len(table_headers) + 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.border = thin_border
            cell.alignment = left_align
    
    # Adjust column widths
    column_widths = [12, 30, 20, 15, 25, 20, 40]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col_num)].width = width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Meeting_Schedule_{month_name}_{year}.xlsx"'
    
    wb.save(response)
    return response
