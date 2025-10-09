from django.shortcuts import render, redirect, get_object_or_404
from django.core.paginator import Paginator
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from .models import StockDeclaration
from .forms import StockDeclarationForm
from settings.models import Line
from datetime import datetime, timedelta
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.db.models import Q, Case, When, IntegerField, Value

def stock_declaration_home(request):
    search_query = request.GET.get('search', '').strip()
    
    stock_declarations_list = StockDeclaration.objects.select_related('created_by').prefetch_related('lines').all()
    
    if search_query:
        stock_declarations_list = stock_declarations_list.filter(
            Q(control_number__icontains=search_query) |
            Q(product_number__icontains=search_query) |
            Q(product_name__icontains=search_query) |
            Q(stock_type__icontains=search_query) |
            Q(status__icontains=search_query) |
            Q(created_by__name__icontains=search_query) |
            Q(lines__line_name__icontains=search_query)
        ).distinct()
    
    # Ordering: if the current user is in purchasing, show 'filed' first then 'in_transit'
    if getattr(request.user, 'stock_declaration_purchasing', False):
        status_priority = Case(
            When(status='filed', then=Value(0)),
            When(status='in_transit', then=Value(1)),
            default=Value(2),
            output_field=IntegerField()
        )
        stock_declarations_list = stock_declarations_list.annotate(_status_order=status_priority).order_by('_status_order', '-created_at')
    else:
        # Default ordering: newest first
        stock_declarations_list = stock_declarations_list.order_by('-created_at')

    # Pagination
    paginator = Paginator(stock_declarations_list, 10)
    page_number = request.GET.get('page')
    stock_declarations = paginator.get_page(page_number)
    
    lines = Line.objects.all()
    
    if request.method == 'POST':
        form = StockDeclarationForm(request.POST, user=request.user)
        if form.is_valid():
            try:
                declaration = form.save()
                messages.success(request, f'Stock declaration {declaration.control_number} created successfully!')
                return redirect('stock_declaration_home')
            except Exception as e:
                messages.error(request, f'Error creating stock declaration: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = StockDeclarationForm()
    
    context = {
        'stock_declarations': stock_declarations,
        'lines': lines,
        'form': form,
        'search_query': search_query,
    }
    
    return render(request, "stockdeclaration/stockdeclaration.html", context)


@login_required
@require_POST
def edit_stock_declaration(request, declaration_id):
    """Edit an existing stock declaration"""
    declaration = get_object_or_404(StockDeclaration, id=declaration_id)
    
    # Check if user is the creator and status allows editing
    if declaration.created_by != request.user:
        return JsonResponse({
            'success': False,
            'message': 'You do not have permission to edit this declaration.'
        }, status=403)
    
    if declaration.status != 'in_transit':
        return JsonResponse({
            'success': False,
            'message': 'Only declarations with "In Transit" status can be edited.'
        }, status=400)
    
    # Update declaration fields
    declaration.product_number = request.POST.get('product_number')
    declaration.product_name = request.POST.get('product_name')
    declaration.quantity = request.POST.get('quantity')
    declaration.stock_type = request.POST.get('stock_type')
    
    # Update lines
    line_ids = request.POST.getlist('lines')
    if not line_ids:
        return JsonResponse({
            'success': False,
            'message': 'Please select at least one production line.'
        }, status=400)
    
    try:
        declaration.save()
        declaration.lines.set(line_ids)
        
        return JsonResponse({
            'success': True,
            'message': f'Stock declaration {declaration.control_number} updated successfully!'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating declaration: {str(e)}'
        }, status=500)


@login_required
@require_POST
def cancel_stock_declaration(request, declaration_id):
    """Cancel a stock declaration"""
    declaration = get_object_or_404(StockDeclaration, id=declaration_id)
    
    # Check if user is the creator
    if declaration.created_by != request.user:
        return JsonResponse({
            'success': False,
            'message': 'You do not have permission to cancel this declaration.'
        }, status=403)
    
    # Check if status allows cancellation
    if declaration.status not in ['filed', 'in_transit']:
        return JsonResponse({
            'success': False,
            'message': f'Cannot cancel a declaration with status "{declaration.get_status_display()}".'
        }, status=400)
    
    try:
        declaration.status = 'cancelled'
        declaration.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Stock declaration {declaration.control_number} has been cancelled.'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error cancelling declaration: {str(e)}'
        }, status=500)


@login_required
@require_POST
def mark_intransit_stock_declaration(request, declaration_id):
    """Mark a stock declaration as in transit"""
    declaration = get_object_or_404(StockDeclaration, id=declaration_id)
    
    # Check if user has purchasing permissions
    if not request.user.stock_declaration_purchasing:
        return JsonResponse({
            'success': False,
            'message': 'You do not have permission to mark declarations as in transit.'
        }, status=403)
    
    # Check if status allows marking as in transit
    if declaration.status != 'filed':
        return JsonResponse({
            'success': False,
            'message': f'Only declarations with "Filed" status can be marked as in transit.'
        }, status=400)
    
    # Get in transit notes
    intransit_notes = request.POST.get('intransit_notes', '').strip()
    if not intransit_notes:
        return JsonResponse({
            'success': False,
            'message': 'In transit notes are required.'
        }, status=400)
    
    try:
        declaration.status = 'in_transit'
        declaration.in_transit = intransit_notes
        declaration.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Stock declaration {declaration.control_number} has been marked as in transit.'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating declaration: {str(e)}'
        }, status=500)


@login_required
@require_POST
def update_intransit_notes(request, declaration_id):
    """Update in transit notes for a stock declaration"""
    declaration = get_object_or_404(StockDeclaration, id=declaration_id)

    # Check if declaration is in transit status
    if declaration.status != 'in_transit':
        return JsonResponse({
            'success': False,
            'message': f'Only declarations with "In Transit" status can have their notes updated.'
        }, status=400)

    # Get in transit notes
    intransit_notes = request.POST.get('intransit_notes', '').strip()
    if not intransit_notes:
        return JsonResponse({
            'success': False,
            'message': 'In transit notes are required.'
        }, status=400)

    try:
        declaration.in_transit = intransit_notes
        declaration.save()

        return JsonResponse({
            'success': True,
            'message': f'In transit notes for {declaration.control_number} have been updated successfully.'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating in transit notes: {str(e)}'
        }, status=500)


@login_required
@require_POST
def receive_stock_declaration(request, declaration_id):
    """Mark a stock declaration as received by production"""
    from django.utils import timezone
    
    declaration = get_object_or_404(StockDeclaration, id=declaration_id)
    
    # Check if user has a line assigned
    if not request.user.line:
        return JsonResponse({
            'success': False,
            'message': 'You do not have a production line assigned.'
        }, status=403)
    
    # Check if user's line is in the declaration's lines
    if not declaration.lines.filter(id=request.user.line.id).exists():
        return JsonResponse({
            'success': False,
            'message': 'This stock declaration is not assigned to your production line.'
        }, status=403)
    
    # Check if already received
    if declaration.received_by_production:
        return JsonResponse({
            'success': False,
            'message': 'This stock declaration has already been received by production.'
        }, status=400)
    
    # Check if status allows receiving (should be in_transit or arrived)
    if declaration.status not in ['in_transit', 'arrived', 'filed']:
        return JsonResponse({
            'success': False,
            'message': f'Stock declarations with "{declaration.get_status_display()}" status cannot be received.'
        }, status=400)
    
    try:
        declaration.received_by_production = True
        declaration.date_received_by_production = timezone.now()
        declaration.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Stock declaration {declaration.control_number} has been marked as received by production.',
            'date_received_by_production': declaration.date_received_by_production.isoformat(),
            'date_received_by_production_display': declaration.date_received_by_production.strftime('%b %d, %Y %I:%M %p')
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error updating declaration: {str(e)}'
        }, status=500)


@login_required
def export_stock_declaration_report(request):
    """Export stock declaration report to Excel"""
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    status = request.GET.get('status')
    stock_type = request.GET.get('stock_type', '')
    
    if not date_from or not date_to or not status:
        messages.error(request, 'Please provide all required fields.')
        return redirect('stock_declaration_home')
    
    # Convert string dates to date objects
    try:
        date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
        date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, 'Invalid date format.')
        return redirect('stock_declaration_home')
    
    # Filter stock declarations by date range
    stock_declarations = StockDeclaration.objects.filter(
        created_at__date__gte=date_from_obj,
        created_at__date__lte=date_to_obj
    ).select_related('created_by').prefetch_related('lines')
    
    # Filter by status if not 'all'
    if status != 'all':
        stock_declarations = stock_declarations.filter(status=status)
    
    # Filter by stock type if provided
    if stock_type:
        stock_declarations = stock_declarations.filter(stock_type=stock_type)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Declaration Report"
    
    # Format dates for display
    date_from_formatted = date_from_obj.strftime('%B %d, %Y')
    date_to_formatted = date_to_obj.strftime('%B %d, %Y')
    
    # Add header
    ws.merge_cells('A1:L1')
    header_cell = ws['A1']
    header_cell.value = "Ryonan Electric Philippines Corporation"
    header_cell.font = Font(bold=True, size=14)
    header_cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Add subtitle
    ws.merge_cells('A2:L2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = f"Stock Declaration Report for {date_from_formatted} to {date_to_formatted}"
    subtitle_cell.font = Font(italic=True, size=11)
    subtitle_cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Add empty row
    ws.append([])
    
    # Define table headers
    headers = [
        'Control Number',
        'Date Declared',
        'Product Number',
        'Product Name',
        'Quantity',
        'Stock Type',
        'Production Lines',
        'Status',
        'Received by Production',
        'Date Received',
        'In Transit Notes',
        'Declared By'
    ]
    
    # Add headers with formatting
    ws.append(headers)
    header_row = ws[4]
    
    # Style for headers
    header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    header_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in header_row:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data rows
    for declaration in stock_declarations:
        # Get production lines as comma-separated string
        lines_str = ", ".join([line.line_name for line in declaration.lines.all()])
        
        # Format received by production
        received_status = "Yes" if declaration.received_by_production else "No"
        received_date = declaration.date_received_by_production.strftime('%b %d, %Y %I:%M %p') if declaration.date_received_by_production else "N/A"
        
        row_data = [
            declaration.control_number,
            declaration.created_at.strftime('%b %d, %Y'),
            declaration.product_number,
            declaration.product_name,
            declaration.quantity,
            declaration.get_stock_type_display(),
            lines_str,
            declaration.get_status_display(),
            received_status,
            received_date,
            declaration.in_transit if declaration.in_transit else "N/A",
            declaration.created_by.name if declaration.created_by else "N/A"
        ]
        
        ws.append(row_data)
        
        # Apply border to all cells in the row
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = Alignment(vertical='center')
    
    # Adjust column widths
    column_widths = {
        'A': 18,  # Control Number
        'B': 15,  # Date Declared
        'C': 18,  # Product Number
        'D': 30,  # Product Name
        'E': 10,  # Quantity
        'F': 15,  # Stock Type
        'G': 25,  # Production Lines
        'H': 12,  # Status
        'I': 20,  # Received by Production
        'J': 20,  # Date Received
        'K': 35,  # In Transit Notes
        'L': 20   # Declared By
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Generate filename with date range
    filename = f'Stock_Declaration_Report_{date_from}_{date_to}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    # Save workbook to response
    wb.save(response)
    
    return response


@login_required
def stock_declaration_chart_data(request):

    time_range = request.GET.get('time_range', 'week')
    stock_type_filter = request.GET.get('stock_type', 'all')
    
    today = timezone.now()
    
    # Determine date range based on time_range parameter
    if time_range == 'today':
        start_date = today.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = today.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        # Generate hourly labels for today
        labels = []
        for hour in range(0, 24, 2):  # Every 2 hours
            labels.append(f'{hour:02d}:00')
        
        # Initialize data structure for hourly counts
        time_slots = 12  # 24 hours / 2
        
    elif time_range == 'week':
        # Get current week (Monday to Saturday)
        days_since_monday = today.weekday()  # Monday is 0, Sunday is 6
        start_date = (today - timedelta(days=days_since_monday)).replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = (start_date + timedelta(days=5)).replace(hour=23, minute=59, second=59, microsecond=999999)  # Saturday
        
        # Generate daily labels for the week (Mon-Sat)
        labels = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
        time_slots = 6
        
    elif time_range == 'month':
        # Get current month
        start_date = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        next_month = today.month + 1 if today.month < 12 else 1
        next_month_year = today.year if today.month < 12 else today.year + 1
        end_date = datetime(next_month_year, next_month, 1, tzinfo=timezone.get_current_timezone()) - timedelta(seconds=1)
        
        # Generate daily labels for the month
        days_in_month = (end_date - start_date).days + 1
        labels = [(start_date + timedelta(days=i)).strftime('%d') for i in range(days_in_month)]
        time_slots = days_in_month
        
    elif time_range == 'year':
        # Fiscal year: April of current year to May of next year
        if today.month >= 4:  # April or later
            fiscal_start = datetime(today.year, 4, 1, tzinfo=timezone.get_current_timezone())
            fiscal_end = datetime(today.year + 1, 5, 31, 23, 59, 59, 999999, tzinfo=timezone.get_current_timezone())
        else:  # Before April
            fiscal_start = datetime(today.year - 1, 4, 1, tzinfo=timezone.get_current_timezone())
            fiscal_end = datetime(today.year, 5, 31, 23, 59, 59, 999999, tzinfo=timezone.get_current_timezone())
        
        start_date = fiscal_start
        end_date = fiscal_end
        
        # Generate monthly labels (Apr to May - 14 months)
        labels = []
        current = start_date
        while current <= end_date:
            labels.append(current.strftime('%b %y'))
            # Move to next month
            if current.month == 12:
                current = current.replace(year=current.year + 1, month=1)
            else:
                current = current.replace(month=current.month + 1)
        
        time_slots = len(labels)
    
    else:
        # Default to week
        days_since_monday = today.weekday()
        start_date = (today - timedelta(days=days_since_monday)).replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = (start_date + timedelta(days=5)).replace(hour=23, minute=59, second=59, microsecond=999999)
        labels = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
        time_slots = 6
    
    # Get all stock declarations within the date range (always get all for consistent totals)
    all_declarations = StockDeclaration.objects.filter(
        created_at__gte=start_date,
        created_at__lte=end_date
    )
    
    if stock_type_filter == 'all':
        # Show separate lines for each stock type
        stock_types = {
            'critical': {'label': 'Critical Stock', 'data': [0] * time_slots, 'color': '#e74c3c'},
            'out_of_stock': {'label': 'Out Of Stock', 'data': [0] * time_slots, 'color': '#f39c12'},
            'overstock': {'label': 'Overstock', 'data': [0] * time_slots, 'color': '#3498db'}
        }
        
        # Count declarations for each time slot and stock type
        for declaration in all_declarations:
            # Determine which time slot this declaration belongs to
            if time_range == 'today':
                hour_index = declaration.created_at.hour // 2
                if 0 <= hour_index < time_slots:
                    if declaration.stock_type in stock_types:
                        stock_types[declaration.stock_type]['data'][hour_index] += 1
                        
            elif time_range == 'week':
                days_from_start = (declaration.created_at - start_date).days
                if 0 <= days_from_start < time_slots:
                    if declaration.stock_type in stock_types:
                        stock_types[declaration.stock_type]['data'][days_from_start] += 1
                        
            elif time_range == 'month':
                day_index = declaration.created_at.day - 1
                if 0 <= day_index < time_slots:
                    if declaration.stock_type in stock_types:
                        stock_types[declaration.stock_type]['data'][day_index] += 1
                        
            elif time_range == 'year':
                # Calculate month index from fiscal start
                months_from_start = (declaration.created_at.year - start_date.year) * 12 + (declaration.created_at.month - start_date.month)
                if 0 <= months_from_start < time_slots:
                    if declaration.stock_type in stock_types:
                        stock_types[declaration.stock_type]['data'][months_from_start] += 1
        
        # Build datasets array for all stock types
        datasets = []
        for stock_type, data in stock_types.items():
            datasets.append({
                'label': data['label'],
                'data': data['data'],
                'borderColor': data['color'],
                'backgroundColor': data['color'] + '33',  # Add transparency
                'borderWidth': 3,
                'tension': 0.4,  # Smooth curves
                'fill': True,
                'pointRadius': 4,
                'pointHoverRadius': 6,
                'pointBackgroundColor': data['color'],
                'pointBorderColor': '#fff',
                'pointBorderWidth': 2
            })
    else:
        # Show only the selected stock type
        declarations = all_declarations.filter(stock_type=stock_type_filter)
        
        # Initialize data array for the selected stock type
        type_data = [0] * time_slots
        
        # Count declarations for each time slot
        for declaration in declarations:
            # Determine which time slot this declaration belongs to
            if time_range == 'today':
                hour_index = declaration.created_at.hour // 2
                if 0 <= hour_index < time_slots:
                    type_data[hour_index] += 1
                    
            elif time_range == 'week':
                days_from_start = (declaration.created_at - start_date).days
                if 0 <= days_from_start < time_slots:
                    type_data[days_from_start] += 1
                    
            elif time_range == 'month':
                day_index = declaration.created_at.day - 1
                if 0 <= day_index < time_slots:
                    type_data[day_index] += 1
                    
            elif time_range == 'year':
                # Calculate month index from fiscal start
                months_from_start = (declaration.created_at.year - start_date.year) * 12 + (declaration.created_at.month - start_date.month)
                if 0 <= months_from_start < time_slots:
                    type_data[months_from_start] += 1
        
        # Get the display name for the stock type
        stock_type_display = dict(StockDeclaration.STOCK_TYPE_CHOICES).get(stock_type_filter, stock_type_filter.title())
        
        # Build single dataset for the selected type
        datasets = [{
            'label': f'{stock_type_display} Declarations',
            'data': type_data,
            'borderColor': '#3498db',
            'backgroundColor': '#3498db33',  # Add transparency
            'borderWidth': 3,
            'tension': 0.4,  # Smooth curves
            'fill': True,
            'pointRadius': 4,
            'pointHoverRadius': 6,
            'pointBackgroundColor': '#3498db',
            'pointBorderColor': '#fff',
            'pointBorderWidth': 2
        }]
    
    # Return chart configuration
    chart_config = {
        'type': 'line',
        'data': {
            'labels': labels,
            'datasets': datasets
        },
        'options': {
            'responsive': True,
            'maintainAspectRatio': False,
            'interaction': {
                'mode': 'index',
                'intersect': False
            },
            'plugins': {
                'legend': {
                    'display': True,
                    'position': 'top',
                    'labels': {
                        'usePointStyle': True,
                        'padding': 15,
                        'font': {
                            'size': 12,
                            'family': "'Segoe UI', Tahoma, Geneva, Verdana, sans-serif"
                        }
                    }
                },
                'tooltip': {
                    'backgroundColor': 'rgba(0, 0, 0, 0.8)',
                    'titleColor': '#fff',
                    'bodyColor': '#fff',
                    'borderColor': '#ddd',
                    'borderWidth': 1,
                    'padding': 12,
                    'displayColors': True,
                    'callbacks': {}
                }
            },
            'scales': {
                'y': {
                    'beginAtZero': True,
                    'ticks': {
                        'font': {
                            'size': 11
                        }
                    },
                    'grid': {
                        'color': 'rgba(0, 0, 0, 0.05)',
                        'drawBorder': False
                    },
                    'title': {
                        'display': True,
                        'text': 'Total Number of Declarations',
                        'font': {
                            'size': 12,
                            'weight': 'bold'
                        }
                    }
                },
                'x': {
                    'grid': {
                        'display': False,
                        'drawBorder': False
                    },
                    'ticks': {
                        'font': {
                            'size': 11
                        }
                    }
                }
            },
            'animation': {
                'duration': 750,
                'easing': 'easeInOutQuart'
            }
        }
    }
    
    return JsonResponse(chart_config)


@login_required
def stock_notifications_api(request):
    
    notifications = []
    
    if hasattr(request.user, 'line') and request.user.line:
        user_line = request.user.line
        today = timezone.now().date()
        
        stock_declarations = StockDeclaration.objects.filter(
            lines=user_line
        ).filter(created_at__date=today, received_by_production=False).distinct().select_related('created_by').prefetch_related('lines').exclude(status='cancelled')

        for declaration in stock_declarations:
            notification_data = {
                'id': declaration.id,
                'control_number': declaration.control_number,
                'stock_type': declaration.get_stock_type_display(),
                'stock_type_value': declaration.stock_type,
                'product_number': declaration.product_number,
                'product_name': declaration.product_name,
                'quantity': declaration.quantity,
                'status': declaration.get_status_display(),
                'created_at': declaration.created_at.strftime('%H:%M'),
            }
            notifications.append(notification_data)
        
        # Sort notifications by stock type priority
        order_map = {'out_of_stock': 0, 'critical': 1, 'overstock': 2}
        notifications.sort(key=lambda n: (order_map.get(n['stock_type_value'], 99), n['created_at']))
    
    return JsonResponse({
        'success': True,
        'notifications': notifications,
        'count': len(notifications)
    })
