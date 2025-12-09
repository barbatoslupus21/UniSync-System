import json
import requests
from datetime import datetime, timedelta, time
from io import BytesIO
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .models import (
    ShuttleVehicle,
    ShuttleProvider,
    Destination,
    DestinationGroup,
    UserShuttleAssignment,
    SubordinateGroup,
    OvertimeFiling,
    OvertimePasscode,
    Holiday
)
from .forms import (
    ShuttleVehicleForm,
    ShuttleProviderForm,
    DestinationForm,
    DestinationGroupForm,
    SubordinateGroupForm,
    OvertimeFilingForm,
    OvertimePasscodeForm,
    HolidayForm,
    BulkOvertimeFilingForm,
    ExportOvertimeForm
)


API_URL = "http://192.168.10.244:3001/api/users-with-line/"


def get_cutoff_status(filing_type, filing_date):
    """
    Check if filing is past the cutoff.
    Returns True if past cutoff (late filing), False otherwise.
    
    Shifting: Cutoff is Thursday 11:59 PM. Filing after Thursday is late.
    Daily: Cutoff is 11:00 AM on the same day.
    Saturday Off/Sunday: Cutoff is Friday 11:59 PM.
    Holiday: Cutoff is the day before the holiday.
    """
    now = timezone.localtime()
    today = now.date()
    current_time = now.time()
    current_weekday = today.weekday()  # Monday=0, Thursday=3, Friday=4

    if filing_type == 'shifting':
        # Shifting cutoff: Thursday 11:59 PM
        # If today is Friday (4), Saturday (5), Sunday (6), or any day after Thursday -> past cutoff
        # Thursday (3) itself is NOT past cutoff (can still file on Thursday)
        if current_weekday > 3:  # Friday, Saturday, Sunday
            return True
        return False

    elif filing_type == 'daily':
        # Daily cutoff: 11:00 AM on the filing date
        cutoff_time = time(11, 0)
        if filing_date < today:
            # Filing for a past date is always late
            return True
        elif filing_date == today:
            # Same day - check if past 11 AM
            return current_time > cutoff_time
        else:
            # Future date - not late
            return False

    elif filing_type in ['saturday_off', 'sunday']:
        # Saturday Off/Sunday cutoff: Friday 11:59 PM
        # If today is Saturday (5) or Sunday (6) -> past cutoff
        if current_weekday > 4:  # Saturday or Sunday
            return True
        return False

    elif filing_type == 'holiday':
        # Holiday cutoff: Day before the holiday
        if filing_date <= today:
            return True
        holiday = Holiday.objects.filter(date=filing_date, is_active=True).first()
        if holiday:
            cutoff_date = holiday.date - timedelta(days=1)
            return today > cutoff_date
        return False

    return False


@login_required
def user_filing(request):
    return render(request, 'overtime/userfiling.html')


@login_required
def admin_filing(request):
    return render(request, 'overtime/adminfiling.html')


@login_required
@require_http_methods(["GET"])
def get_subordinates(request):
    """Get subordinates from user's active subordinate groups"""
    # Get all active groups created by the current user
    user_groups = SubordinateGroup.objects.filter(created_by=request.user, is_active=True)
    
    # Collect all unique employee IDs from user's groups
    employee_ids = set()
    for group in user_groups:
        if group.employee_ids:
            employee_ids.update(group.employee_ids)
    
    # Get UserShuttleAssignment records for those employee IDs
    assignments = UserShuttleAssignment.objects.select_related('destination').filter(
        employee_id__in=employee_ids
    )
    
    data = [{
        'id': a.employee_id,
        'employee_id': a.employee_id,
        'full_name': a.employee_name,
        'department': a.department,
        'line_name': a.line_name,
        'destination': a.destination.name if a.destination else None
    } for a in assignments]
    
    return JsonResponse({'success': True, 'data': data})


@login_required
@require_http_methods(["GET"])
def get_all_shuttle_users_for_group(request):
    """Get all UserShuttleAssignment for creating/editing subordinate groups"""
    assignments = UserShuttleAssignment.objects.select_related('destination').all()
    
    data = [{
        'id': a.employee_id,
        'employee_id': a.employee_id,
        'full_name': a.employee_name,
        'department': a.department,
        'line_name': a.line_name,
        'destination': a.destination.name if a.destination else None
    } for a in assignments]
    
    return JsonResponse({'success': True, 'data': data})


# ============================================
# Shuttle User Views (UserShuttleAssignment CRUD)
# ============================================

@login_required
@require_http_methods(["GET"])
def get_shuttle_users(request):
    """Get all shuttle users with pagination and search"""
    search = request.GET.get('search', '')
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 10))
    
    users = UserShuttleAssignment.objects.all().select_related('destination')
    
    if search:
        users = users.filter(
            Q(employee_id__icontains=search) |
            Q(employee_name__icontains=search) |
            Q(line_name__icontains=search) |
            Q(department__icontains=search)
        )
    
    total = users.count()
    start = (page - 1) * per_page
    end = start + per_page
    users_page = users[start:end]
    
    # Get subordinate groups to determine which users are subordinates of whom
    subordinate_groups = SubordinateGroup.objects.filter(is_active=True).select_related('created_by')
    user_groups = {}
    for group in subordinate_groups:
        for emp_id in group.employee_ids:
            emp_id_str = str(emp_id)
            if emp_id_str not in user_groups:
                user_groups[emp_id_str] = []
            user_groups[emp_id_str].append({
                'group_name': group.name,
                'supervisor': group.created_by.name if group.created_by else 'Unknown'
            })
    
    data = [{
        'id': u.id,
        'employee_id': u.employee_id,
        'employee_name': u.employee_name,
        'line_name': u.line_name,
        'department': u.department,
        'destination': {
            'id': u.destination.id,
            'name': u.destination.name
        } if u.destination else None,
        'with_vehicle': u.with_vehicle,
        'groups': user_groups.get(str(u.employee_id), [])
    } for u in users_page]
    
    return JsonResponse({
        'success': True,
        'data': data,
        'total': total,
        'page': page,
        'per_page': per_page,
        'total_pages': (total + per_page - 1) // per_page if total > 0 else 1
    })


@login_required
@require_http_methods(["POST"])
def create_shuttle_user(request):
    """Create a new shuttle user assignment"""
    try:
        data = json.loads(request.body)
        
        employee_id = data.get('employee_id')
        employee_name = data.get('employee_name')
        line_name = data.get('line_name', '')
        department = data.get('department', '')
        destination_id = data.get('destination_id')
        with_vehicle = data.get('with_vehicle', False)
        
        if not employee_id or not employee_name:
            return JsonResponse({'success': False, 'error': 'Employee ID and Name are required'}, status=400)
        
        # Check if employee_id already exists
        if UserShuttleAssignment.objects.filter(employee_id=employee_id).exists():
            return JsonResponse({'success': False, 'error': 'Employee ID already exists'}, status=400)
        
        destination = None
        if destination_id:
            destination = get_object_or_404(Destination, id=destination_id)
        
        user = UserShuttleAssignment.objects.create(
            employee_id=employee_id,
            employee_name=employee_name,
            line_name=line_name or None,
            department=department or None,
            destination=destination,
            with_vehicle=with_vehicle
        )
        
        return JsonResponse({
            'success': True,
            'data': {
                'id': user.id,
                'employee_id': user.employee_id,
                'employee_name': user.employee_name,
                'line_name': user.line_name,
                'department': user.department,
                'destination_id': user.destination_id,
                'with_vehicle': user.with_vehicle
            }
        })
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["PUT"])
def update_shuttle_user(request, user_id):
    """Update an existing shuttle user assignment"""
    try:
        shuttle_user = get_object_or_404(UserShuttleAssignment, id=user_id)
        data = json.loads(request.body)
        
        employee_id = data.get('employee_id', shuttle_user.employee_id)
        employee_name = data.get('employee_name', shuttle_user.employee_name)
        line_name = data.get('line_name', shuttle_user.line_name)
        department = data.get('department', shuttle_user.department)
        destination_id = data.get('destination_id')
        with_vehicle = data.get('with_vehicle', shuttle_user.with_vehicle)
        
        # Check if new employee_id already exists (exclude current user)
        if UserShuttleAssignment.objects.filter(employee_id=employee_id).exclude(id=user_id).exists():
            return JsonResponse({'success': False, 'error': 'Employee ID already exists'}, status=400)
        
        shuttle_user.employee_id = employee_id
        shuttle_user.employee_name = employee_name
        shuttle_user.line_name = line_name or None
        shuttle_user.department = department or None
        shuttle_user.with_vehicle = with_vehicle
        
        if destination_id:
            shuttle_user.destination = get_object_or_404(Destination, id=destination_id)
        else:
            shuttle_user.destination = None
        
        shuttle_user.save()
        
        return JsonResponse({
            'success': True,
            'data': {
                'id': shuttle_user.id,
                'employee_id': shuttle_user.employee_id,
                'employee_name': shuttle_user.employee_name,
                'line_name': shuttle_user.line_name,
                'department': shuttle_user.department,
                'destination_id': shuttle_user.destination_id,
                'with_vehicle': shuttle_user.with_vehicle
            }
        })
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["DELETE"])
def delete_shuttle_user(request, user_id):
    """Delete a shuttle user assignment"""
    shuttle_user = get_object_or_404(UserShuttleAssignment, id=user_id)
    shuttle_user.delete()
    return JsonResponse({'success': True})


@login_required
@require_http_methods(["GET"])
def download_shuttle_template(request):
    """Download Excel template for shuttle user import with data validation"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Shuttle Users"
    
    # Get all active destinations for data validation
    destinations = list(Destination.objects.filter(is_active=True).values_list('name', flat=True))
    
    # Define headers
    headers = ['Employee ID', 'Employee Name', 'Line', 'Department', 'Shuttle Destination']
    
    # Style definitions
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1a3a5c', end_color='1a3a5c', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 30
    
    # Add data validation for Shuttle Destination column (column E)
    if destinations:
        # Create a hidden sheet for destination list
        dest_sheet = wb.create_sheet("DestinationList")
        for idx, dest in enumerate(destinations, 1):
            dest_sheet.cell(row=idx, column=1, value=dest)
        
        # Hide the destination sheet
        dest_sheet.sheet_state = 'hidden'
        
        # Create data validation
        dv = DataValidation(
            type="list",
            formula1=f"DestinationList!$A$1:$A${len(destinations)}",
            allow_blank=True
        )
        dv.error = "Please select a valid destination from the list"
        dv.errorTitle = "Invalid Destination"
        dv.prompt = "Select a destination"
        dv.promptTitle = "Shuttle Destination"
        
        # Apply validation to rows 2-1000 (column E)
        ws.add_data_validation(dv)
        dv.add('E2:E1000')
    
    # Add some sample rows with borders (empty data rows)
    for row in range(2, 12):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col, value='')
            cell.border = thin_border
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Create response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="shuttle_users_template.xlsx"'
    return response


@login_required
@require_http_methods(["POST"])
def import_shuttle_users(request):
    """Import shuttle users from Excel file"""
    if 'file' not in request.FILES:
        return JsonResponse({'success': False, 'error': 'No file provided'}, status=400)
    
    file = request.FILES['file']
    
    if not file.name.endswith(('.xlsx', '.xls')):
        return JsonResponse({'success': False, 'error': 'Invalid file format. Please upload an Excel file.'}, status=400)
    
    try:
        wb = load_workbook(file)
        ws = wb.active
        
        # Get destination mapping
        destinations = {dest.name.lower(): dest for dest in Destination.objects.filter(is_active=True)}
        
        imported = 0
        updated = 0
        errors = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Skip empty rows
                continue
            
            employee_id, employee_name, line_name, department, destination_name = row[:5] if len(row) >= 5 else (row + (None,) * (5 - len(row)))
            
            # Validate required fields
            row_errors = []
            if not employee_id:
                row_errors.append('Employee ID is required')
            if not employee_name:
                row_errors.append('Employee Name is required')
            
            # Validate destination
            destination = None
            if destination_name:
                destination = destinations.get(str(destination_name).lower().strip())
                if not destination:
                    row_errors.append(f'Invalid destination: {destination_name}')
            
            if row_errors:
                errors.append({
                    'row': row_num,
                    'employee_id': str(employee_id) if employee_id else '',
                    'employee_name': str(employee_name) if employee_name else '',
                    'line_name': str(line_name) if line_name else '',
                    'department': str(department) if department else '',
                    'destination': str(destination_name) if destination_name else '',
                    'errors': row_errors
                })
                continue
            
            # Create or update shuttle user
            try:
                shuttle_user, created = UserShuttleAssignment.objects.update_or_create(
                    employee_id=int(employee_id) if str(employee_id).isdigit() else employee_id,
                    defaults={
                        'employee_name': str(employee_name).strip(),
                        'line_name': str(line_name).strip() if line_name else None,
                        'department': str(department).strip() if department else None,
                        'destination': destination
                    }
                )
                if created:
                    imported += 1
                else:
                    updated += 1
            except Exception as e:
                errors.append({
                    'row': row_num,
                    'employee_id': str(employee_id),
                    'employee_name': str(employee_name),
                    'line_name': str(line_name) if line_name else '',
                    'department': str(department) if department else '',
                    'destination': str(destination_name) if destination_name else '',
                    'errors': [str(e)]
                })
        
        # If there are errors, create an error report Excel file
        error_file = None
        if errors:
            error_wb = Workbook()
            error_ws = error_wb.active
            error_ws.title = "Failed Imports"
            
            # Headers
            error_headers = ['Row', 'Employee ID', 'Employee Name', 'Line', 'Department', 'Shuttle Destination', 'Errors']
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='c0392b', end_color='c0392b', fill_type='solid')
            
            for col, header in enumerate(error_headers, 1):
                cell = error_ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Error data
            for idx, err in enumerate(errors, 2):
                error_ws.cell(row=idx, column=1, value=err['row'])
                error_ws.cell(row=idx, column=2, value=err['employee_id'])
                error_ws.cell(row=idx, column=3, value=err['employee_name'])
                error_ws.cell(row=idx, column=4, value=err['line_name'])
                error_ws.cell(row=idx, column=5, value=err['department'])
                error_ws.cell(row=idx, column=6, value=err['destination'])
                error_ws.cell(row=idx, column=7, value='; '.join(err['errors']))
            
            # Set column widths
            error_ws.column_dimensions['A'].width = 8
            error_ws.column_dimensions['B'].width = 15
            error_ws.column_dimensions['C'].width = 30
            error_ws.column_dimensions['D'].width = 20
            error_ws.column_dimensions['E'].width = 25
            error_ws.column_dimensions['F'].width = 25
            error_ws.column_dimensions['G'].width = 50
            
            # Save to base64
            import base64
            error_output = BytesIO()
            error_wb.save(error_output)
            error_output.seek(0)
            error_file = base64.b64encode(error_output.read()).decode('utf-8')
        
        return JsonResponse({
            'success': True,
            'imported': imported,
            'updated': updated,
            'error_count': len(errors),
            'error_file': error_file
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Failed to process file: {str(e)}'}, status=400)


@login_required
@require_http_methods(["GET"])
def get_subordinate_groups(request):
    groups = SubordinateGroup.objects.filter(
        created_by=request.user,
        is_active=True
    ).values('id', 'name', 'employee_ids', 'created_at')
    return JsonResponse({'success': True, 'data': list(groups)})


@login_required
@require_http_methods(["POST"])
def create_subordinate_group(request):
    try:
        data = json.loads(request.body)
        form = SubordinateGroupForm(data, user=request.user)
        if form.is_valid():
            group = form.save()
            return JsonResponse({
                'success': True,
                'data': {
                    'id': group.id,
                    'name': group.name,
                    'employee_ids': group.employee_ids
                }
            })
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)


@login_required
@require_http_methods(["PUT"])
def update_subordinate_group(request, group_id):
    try:
        group = get_object_or_404(SubordinateGroup, id=group_id, created_by=request.user)
        data = json.loads(request.body)
        form = SubordinateGroupForm(data, instance=group, user=request.user)
        if form.is_valid():
            group = form.save()
            return JsonResponse({
                'success': True,
                'data': {
                    'id': group.id,
                    'name': group.name,
                    'employee_ids': group.employee_ids
                }
            })
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)


@login_required
@require_http_methods(["DELETE"])
def delete_subordinate_group(request, group_id):
    group = get_object_or_404(SubordinateGroup, id=group_id, created_by=request.user)
    group.delete()
    return JsonResponse({'success': True})


@login_required
@require_http_methods(["POST"])
def check_cutoff(request):
    try:
        data = json.loads(request.body)
        filing_type = data.get('filing_type')
        date_str = data.get('date')
        filing_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        is_past_cutoff = get_cutoff_status(filing_type, filing_date)
        return JsonResponse({
            'success': True,
            'is_past_cutoff': is_past_cutoff
        })
    except (json.JSONDecodeError, ValueError) as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
@require_http_methods(["POST"])
def verify_passcode(request):
    try:
        data = json.loads(request.body)
        passcode = data.get('passcode')
        
        stored_passcode = OvertimePasscode.objects.filter(is_active=True).first()
        if stored_passcode and passcode == stored_passcode.passcode:
            return JsonResponse({'success': True, 'valid': True})
        return JsonResponse({'success': True, 'valid': False})
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)


@login_required
@require_http_methods(["POST"])
def check_duplicate_filings(request):
    """Check if any employees already have overtime filings for the given schedule"""
    try:
        data = json.loads(request.body)
        filing_type = data.get('filing_type')
        employee_ids = data.get('employee_ids', [])
        shift = data.get('shift')
        time_in = data.get('time_in')
        time_out = data.get('time_out')
        date_from = data.get('date_from')
        date_to = data.get('date_to')
        
        filing_date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
        filing_date_to = datetime.strptime(date_to, '%Y-%m-%d').date() if date_to else None
        filing_time_in = datetime.strptime(time_in, '%H:%M').time()
        filing_time_out = datetime.strptime(time_out, '%H:%M').time()
        
        duplicates = []
        
        for emp_id in employee_ids:
            # employee_id is a string field - keep as string
            if not emp_id:
                continue
            
            emp_id_str = str(emp_id)
            
            # Build query for existing filings
            existing_query = OvertimeFiling.objects.filter(
                employee_id=emp_id_str,
                filed_by=request.user,
                shift=shift,
                time_in=filing_time_in,
                time_out=filing_time_out
            )
            
            if filing_type == 'shifting':
                # For shifting: check if date ranges overlap
                # An overlap occurs when existing.date_from <= new.date_to AND existing.date_to >= new.date_from
                if filing_date_to:
                    existing_query = existing_query.filter(
                        filing_type='shifting',
                        date_from__lte=filing_date_to,
                        date_to__gte=filing_date_from
                    )
                else:
                    existing_query = existing_query.filter(
                        filing_type='shifting',
                        date_from=filing_date_from
                    )
            else:
                # For daily types: check exact date match
                existing_query = existing_query.filter(
                    filing_type=filing_type,
                    date_from=filing_date_from
                )
            
            existing = existing_query.first()
            if existing:
                duplicates.append({
                    'employee_id': emp_id,  # Keep original ID format for frontend
                    'employee_name': existing.employee_name,
                    'department': existing.department,
                    'line_name': existing.line_name,
                    'existing_filing_id': existing.id,
                    'existing_date_from': existing.date_from.strftime('%Y-%m-%d'),
                    'existing_date_to': existing.date_to.strftime('%Y-%m-%d') if existing.date_to else None,
                    'existing_status': existing.status
                })
        
        return JsonResponse({
            'success': True,
            'has_duplicates': len(duplicates) > 0,
            'duplicates': duplicates
        })
    except (json.JSONDecodeError, ValueError) as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def check_shifting_filings(request):
    """Check if employees have valid shifting filings for a given daily overtime date"""
    try:
        data = json.loads(request.body)
        employee_ids = data.get('employee_ids', [])
        daily_date = data.get('date')
        
        if not daily_date:
            return JsonResponse({'success': False, 'error': 'Date is required'}, status=400)
        
        check_date = datetime.strptime(daily_date, '%Y-%m-%d').date()
        
        employees_without_shifting = []
        
        for emp_id in employee_ids:
            if not emp_id:
                continue
            
            emp_id_str = str(emp_id)
            
            # Check if employee has a shifting filing where date_from <= daily_date <= date_to
            has_shifting = OvertimeFiling.objects.filter(
                employee_id=emp_id_str,
                filing_type='shifting',
                date_from__lte=check_date,
                date_to__gte=check_date
            ).exists()
            
            if not has_shifting:
                # Get employee info from UserShuttleAssignment
                assignment = UserShuttleAssignment.objects.filter(employee_id=emp_id_str).first()
                employees_without_shifting.append({
                    'employee_id': emp_id,
                    'employee_name': assignment.employee_name if assignment else f'Employee {emp_id}',
                    'department': assignment.department if assignment else '-',
                    'line_name': assignment.line_name if assignment else '-'
                })
        
        return JsonResponse({
            'success': True,
            'all_have_shifting': len(employees_without_shifting) == 0,
            'employees_without_shifting': employees_without_shifting
        })
    except (json.JSONDecodeError, ValueError) as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def submit_overtime(request):
    try:
        data = json.loads(request.body)
        filing_type = data.get('filing_type')
        employee_data = data.get('employee_data', [])
        shift = data.get('shift')
        time_in = data.get('time_in')
        time_out = data.get('time_out')
        date_from = data.get('date_from')
        date_to = data.get('date_to')
        passcode = data.get('passcode')
        update_duplicates = data.get('update_duplicates', False)  # Flag to update existing records

        filing_date = datetime.strptime(date_from, '%Y-%m-%d').date()
        filing_date_to = datetime.strptime(date_to, '%Y-%m-%d').date() if date_to else None
        filing_time_in = datetime.strptime(time_in, '%H:%M').time()
        filing_time_out = datetime.strptime(time_out, '%H:%M').time()
        is_late = get_cutoff_status(filing_type, filing_date)

        if is_late:
            stored_passcode = OvertimePasscode.objects.filter(is_active=True).first()
            if not stored_passcode or passcode != stored_passcode.passcode:
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid passcode for late filing'
                }, status=400)

        created_count = 0
        updated_count = 0
        
        for emp in employee_data:
            emp_id = emp.get('id')
            
            # employee_id is a string field - keep as string
            if not emp_id:
                continue
                
            emp_id_str = str(emp_id)
            
            # Check for existing filing if update_duplicates is True
            existing_filing = None
            if update_duplicates:
                existing_query = OvertimeFiling.objects.filter(
                    employee_id=emp_id_str,
                    filed_by=request.user,
                    shift=shift,
                    time_in=filing_time_in,
                    time_out=filing_time_out
                )
                
                if filing_type == 'shifting':
                    if filing_date_to:
                        existing_query = existing_query.filter(
                            filing_type='shifting',
                            date_from__lte=filing_date_to,
                            date_to__gte=filing_date
                        )
                    else:
                        existing_query = existing_query.filter(
                            filing_type='shifting',
                            date_from=filing_date
                        )
                else:
                    existing_query = existing_query.filter(
                        filing_type=filing_type,
                        date_from=filing_date
                    )
                
                existing_filing = existing_query.first()
            
            if existing_filing:
                # Update existing filing
                existing_filing.status = emp.get('status', 'ot')
                existing_filing.date_from = filing_date
                existing_filing.date_to = filing_date_to
                existing_filing.is_late_filing = is_late
                existing_filing.save()
                updated_count += 1
            else:
                # Create new filing
                filing = OvertimeFiling(
                    filing_type=filing_type,
                    employee_id=emp_id_str,
                    employee_name=emp.get('full_name'),
                    department=emp.get('department'),
                    line_name=emp.get('line_name'),
                    shift=shift,
                    time_in=filing_time_in,
                    time_out=filing_time_out,
                    date_from=filing_date,
                    date_to=filing_date_to,
                    status=emp.get('status', 'ot'),
                    filed_by=request.user,
                    is_late_filing=is_late
                )
                filing.save()
                created_count += 1

        # Build response message
        messages = []
        if created_count > 0:
            messages.append(f'Created {created_count} new filing(s)')
        if updated_count > 0:
            messages.append(f'Updated {updated_count} existing filing(s)')
        
        return JsonResponse({
            'success': True,
            'message': '. '.join(messages) if messages else 'No changes made',
            'created_count': created_count,
            'updated_count': updated_count
        })
    except (json.JSONDecodeError, ValueError) as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def get_overtime_filings(request):
    try:
        filing_type = request.GET.get('filing_type')
        date = request.GET.get('date')
        week = request.GET.get('week')
        year = request.GET.get('year')

        filings = OvertimeFiling.objects.filter(filed_by=request.user)

        if filing_type:
            filings = filings.filter(filing_type=filing_type)
        if date:
            filings = filings.filter(date_from=date)
        if week and year:
            filings = filings.filter(week_number=week, year=year)

        data = list(filings.values(
            'id', 'filing_type', 'employee_id', 'employee_name', 'department',
            'line_name', 'shift', 'time_in', 'time_out', 'date_from', 'date_to',
            'status', 'is_late_filing', 'created_at'
        ))
        return JsonResponse({'success': True, 'data': data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["DELETE"])
def delete_overtime_filing(request, filing_id):
    filing = get_object_or_404(OvertimeFiling, id=filing_id, filed_by=request.user)
    filing.delete()
    return JsonResponse({'success': True})


@login_required
@require_http_methods(["GET"])
def get_shuttle_vehicles(request):
    vehicles = ShuttleVehicle.objects.filter(is_active=True).values(
        'id', 'name', 'capacity', 'created_at'
    )
    return JsonResponse({'success': True, 'data': list(vehicles)})


@login_required
@require_http_methods(["POST"])
def create_shuttle_vehicle(request):
    try:
        data = json.loads(request.body)
        form = ShuttleVehicleForm(data)
        if form.is_valid():
            vehicle = form.save()
            return JsonResponse({
                'success': True,
                'data': {
                    'id': vehicle.id,
                    'name': vehicle.name,
                    'capacity': vehicle.capacity
                }
            })
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)


@login_required
@require_http_methods(["PUT"])
def update_shuttle_vehicle(request, vehicle_id):
    try:
        vehicle = get_object_or_404(ShuttleVehicle, id=vehicle_id)
        data = json.loads(request.body)
        form = ShuttleVehicleForm(data, instance=vehicle)
        if form.is_valid():
            vehicle = form.save()
            return JsonResponse({
                'success': True,
                'data': {
                    'id': vehicle.id,
                    'name': vehicle.name,
                    'capacity': vehicle.capacity
                }
            })
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)


@login_required
@require_http_methods(["DELETE"])
def delete_shuttle_vehicle(request, vehicle_id):
    vehicle = get_object_or_404(ShuttleVehicle, id=vehicle_id)
    vehicle.is_active = False
    vehicle.save()
    return JsonResponse({'success': True})


# ============================================
# Shuttle Provider Views
# ============================================

@login_required
@require_http_methods(["GET"])
def get_shuttle_providers(request):
    providers = ShuttleProvider.objects.filter(is_active=True).values(
        'id', 'name', 'contact_person', 'contact_number', 'created_at'
    )
    return JsonResponse({'success': True, 'data': list(providers)})


@login_required
@require_http_methods(["POST"])
def create_shuttle_provider(request):
    try:
        data = json.loads(request.body)
        form = ShuttleProviderForm(data)
        if form.is_valid():
            provider = form.save()
            return JsonResponse({
                'success': True,
                'data': {
                    'id': provider.id,
                    'name': provider.name,
                    'contact_person': provider.contact_person,
                    'contact_number': provider.contact_number
                }
            })
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)


@login_required
@require_http_methods(["PUT"])
def update_shuttle_provider(request, provider_id):
    try:
        provider = get_object_or_404(ShuttleProvider, id=provider_id)
        data = json.loads(request.body)
        form = ShuttleProviderForm(data, instance=provider)
        if form.is_valid():
            provider = form.save()
            return JsonResponse({
                'success': True,
                'data': {
                    'id': provider.id,
                    'name': provider.name,
                    'contact_person': provider.contact_person,
                    'contact_number': provider.contact_number
                }
            })
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)


@login_required
@require_http_methods(["DELETE"])
def delete_shuttle_provider(request, provider_id):
    provider = get_object_or_404(ShuttleProvider, id=provider_id)
    provider.is_active = False
    provider.save()
    return JsonResponse({'success': True})


# ============================================
# Destination Views (Individual Locations)
# ============================================

@login_required
@require_http_methods(["GET"])
def get_destinations(request):
    destinations = Destination.objects.filter(is_active=True).values(
        'id', 'name', 'created_at'
    )
    return JsonResponse({'success': True, 'data': list(destinations)})


@login_required
@require_http_methods(["GET"])
def get_available_destinations(request):
    """
    Get destinations available for assignment to a group.
    - If group_id is provided (editing), returns destinations belonging to that group + unassigned destinations
    - If no group_id (creating), returns only unassigned destinations
    """
    group_id = request.GET.get('group_id')
    
    # Get all destination IDs that are assigned to active groups
    assigned_destination_ids = set()
    active_groups = DestinationGroup.objects.filter(is_active=True).prefetch_related('destinations')
    
    for group in active_groups:
        # Skip the current group if we're editing
        if group_id and str(group.id) == str(group_id):
            continue
        assigned_destination_ids.update(group.destinations.filter(is_active=True).values_list('id', flat=True))
    
    # Get all active destinations that are NOT assigned to other groups
    available_destinations = Destination.objects.filter(
        is_active=True
    ).exclude(
        id__in=assigned_destination_ids
    ).values('id', 'name', 'created_at')
    
    return JsonResponse({'success': True, 'data': list(available_destinations)})


@login_required
@require_http_methods(["POST"])
def create_destination(request):
    try:
        data = json.loads(request.body)
        form = DestinationForm(data)
        if form.is_valid():
            destination = form.save()
            return JsonResponse({
                'success': True,
                'data': {
                    'id': destination.id,
                    'name': destination.name
                }
            })
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)


@login_required
@require_http_methods(["PUT"])
def update_destination(request, destination_id):
    try:
        destination = get_object_or_404(Destination, id=destination_id)
        data = json.loads(request.body)
        form = DestinationForm(data, instance=destination)
        if form.is_valid():
            destination = form.save()
            return JsonResponse({
                'success': True,
                'data': {
                    'id': destination.id,
                    'name': destination.name
                }
            })
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)


@login_required
@require_http_methods(["DELETE"])
def delete_destination(request, destination_id):
    destination = get_object_or_404(Destination, id=destination_id)
    destination.is_active = False
    destination.save()
    return JsonResponse({'success': True})


# ============================================
# Destination Group Views (Vehicle + Destinations)
# ============================================

@login_required
@require_http_methods(["GET"])
def get_destination_groups(request):
    groups = DestinationGroup.objects.filter(is_active=True).select_related('shuttle_vehicle', 'shuttle_provider').prefetch_related('destinations')
    data = [{
        'id': g.id,
        'name': g.name,
        'shuttle_provider': {
            'id': g.shuttle_provider.id,
            'name': g.shuttle_provider.name
        } if g.shuttle_provider else None,
        'shuttle_vehicle': {
            'id': g.shuttle_vehicle.id,
            'name': g.shuttle_vehicle.name,
            'capacity': g.shuttle_vehicle.capacity
        } if g.shuttle_vehicle else None,
        'destinations': [{'id': d.id, 'name': d.name} for d in g.destinations.filter(is_active=True)],
        'created_at': g.created_at
    } for g in groups]
    return JsonResponse({'success': True, 'data': data})


@login_required
@require_http_methods(["POST"])
def create_destination_group(request):
    try:
        data = json.loads(request.body)
        destination_ids = data.pop('destination_ids', [])
        
        form = DestinationGroupForm(data)
        if form.is_valid():
            group = form.save()
            # Add destinations to the group
            if destination_ids:
                destinations = Destination.objects.filter(id__in=destination_ids, is_active=True)
                group.destinations.set(destinations)
            
            return JsonResponse({
                'success': True,
                'data': {
                    'id': group.id,
                    'name': group.name,
                    'shuttle_vehicle_id': group.shuttle_vehicle_id,
                    'destination_ids': list(group.destinations.values_list('id', flat=True))
                }
            })
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)


@login_required
@require_http_methods(["PUT"])
def update_destination_group(request, group_id):
    try:
        group = get_object_or_404(DestinationGroup, id=group_id)
        data = json.loads(request.body)
        destination_ids = data.pop('destination_ids', [])
        
        form = DestinationGroupForm(data, instance=group)
        if form.is_valid():
            group = form.save()
            # Update destinations
            if destination_ids is not None:
                destinations = Destination.objects.filter(id__in=destination_ids, is_active=True)
                group.destinations.set(destinations)
            
            return JsonResponse({
                'success': True,
                'data': {
                    'id': group.id,
                    'name': group.name,
                    'shuttle_vehicle_id': group.shuttle_vehicle_id,
                    'destination_ids': list(group.destinations.values_list('id', flat=True))
                }
            })
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)


@login_required
@require_http_methods(["DELETE"])
def delete_destination_group(request, group_id):
    group = get_object_or_404(DestinationGroup, id=group_id)
    group.is_active = False
    group.save()
    return JsonResponse({'success': True})


@login_required
@require_http_methods(["GET"])
def get_user_shuttle_assignments(request):
    search = request.GET.get('search', '')
    assignments = UserShuttleAssignment.objects.select_related('destination_group')
    
    if search:
        assignments = assignments.filter(
            Q(employee_name__icontains=search) |
            Q(department__icontains=search) |
            Q(line_name__icontains=search)
        )

    data = [{
        'id': a.id,
        'employee_id': a.employee_id,
        'employee_name': a.employee_name,
        'department': a.department,
        'line_name': a.line_name,
        'destination_group': {
            'id': a.destination_group.id,
            'name': a.destination_group.name
        } if a.destination_group else None
    } for a in assignments]
    
    return JsonResponse({'success': True, 'data': data})


@login_required
@require_http_methods(["POST"])
def update_user_shuttle_assignment(request):
    try:
        data = json.loads(request.body)
        employee_id = data.get('employee_id')
        employee_name = data.get('employee_name')
        department = data.get('department')
        line_name = data.get('line_name')
        destination_group_id = data.get('destination_group_id')

        assignment, created = UserShuttleAssignment.objects.update_or_create(
            employee_id=employee_id,
            defaults={
                'employee_name': employee_name,
                'department': department,
                'line_name': line_name,
                'destination_group_id': destination_group_id if destination_group_id else None
            }
        )

        return JsonResponse({
            'success': True,
            'data': {
                'id': assignment.id,
                'employee_id': assignment.employee_id,
                'destination_group_id': assignment.destination_group_id
            }
        })
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)


@login_required
@require_http_methods(["GET"])
def get_holidays(request):
    holidays = Holiday.objects.filter(is_active=True).values('id', 'name', 'date')
    return JsonResponse({'success': True, 'data': list(holidays)})


@login_required
@require_http_methods(["POST"])
def create_holiday(request):
    try:
        data = json.loads(request.body)
        form = HolidayForm(data)
        if form.is_valid():
            holiday = form.save()
            return JsonResponse({
                'success': True,
                'data': {
                    'id': holiday.id,
                    'name': holiday.name,
                    'date': holiday.date.isoformat()
                }
            })
        return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)


@login_required
@require_http_methods(["DELETE"])
def delete_holiday(request, holiday_id):
    holiday = get_object_or_404(Holiday, id=holiday_id)
    holiday.is_active = False
    holiday.save()
    return JsonResponse({'success': True})


@login_required
@require_http_methods(["GET"])
def get_passcode(request):
    passcode = OvertimePasscode.objects.filter(is_active=True).first()
    if passcode:
        return JsonResponse({
            'success': True,
            'data': {'id': passcode.id, 'passcode': passcode.passcode}
        })
    return JsonResponse({'success': True, 'data': None})


@login_required
@require_http_methods(["POST"])
def update_passcode(request):
    try:
        data = json.loads(request.body)
        passcode_value = data.get('passcode')
        
        OvertimePasscode.objects.filter(is_active=True).update(is_active=False)
        passcode = OvertimePasscode.objects.create(
            passcode=passcode_value,
            is_active=True
        )
        
        return JsonResponse({
            'success': True,
            'data': {'id': passcode.id}
        })
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)


@login_required
@require_http_methods(["GET"])
def export_overtime(request):
    export_type = request.GET.get('export_type')
    date = request.GET.get('date')
    week = request.GET.get('week')
    year = request.GET.get('year')

    filings = OvertimeFiling.objects.all()

    if export_type == 'shifting':
        filings = filings.filter(
            filing_type='shifting',
            week_number=int(week),
            year=int(year)
        )
    elif export_type == 'daily':
        filings = filings.filter(
            filing_type='daily',
            date_from=date,
            status='not_ot'
        )
    elif export_type in ['sunday', 'saturday_off', 'holiday']:
        filings = filings.filter(
            filing_type=export_type,
            date_from=date,
            status='ot'
        )

    wb = Workbook()
    ws = wb.active
    ws.title = f"{export_type.title()} Overtime"

    header_fill = PatternFill(start_color="3366FF", end_color="3366FF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = ['No.', 'Employee ID', 'Employee Name', 'Department', 'Line', 
               'Shift', 'Time In', 'Time Out', 'Status', 'Shuttle Group']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    employee_assignments = {
        a.employee_id: a.destination_group.name if a.destination_group else 'Unassigned'
        for a in UserShuttleAssignment.objects.select_related('destination_group').all()
    }

    grouped_data = {}
    for filing in filings:
        group_name = employee_assignments.get(filing.employee_id, 'Unassigned')
        if group_name not in grouped_data:
            grouped_data[group_name] = []
        grouped_data[group_name].append(filing)

    row = 2
    for group_name, group_filings in sorted(grouped_data.items()):
        destination_group = DestinationGroup.objects.filter(name=group_name).first()
        capacity = destination_group.shuttle_vehicle.capacity if destination_group and destination_group.shuttle_vehicle else 13

        sub_groups = []
        current_group = []
        for filing in group_filings:
            current_group.append(filing)
            if len(current_group) >= capacity:
                sub_groups.append(current_group)
                current_group = []
        if current_group:
            sub_groups.append(current_group)

        for sub_idx, sub_group in enumerate(sub_groups):
            group_header = f"{group_name}"
            if len(sub_groups) > 1:
                group_header += f" - Group {sub_idx + 1}"
            
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
            cell = ws.cell(row=row, column=1, value=group_header)
            cell.fill = PatternFill(start_color="E0E8FF", end_color="E0E8FF", fill_type="solid")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            row += 1

            for idx, filing in enumerate(sub_group, 1):
                ws.cell(row=row, column=1, value=idx).border = thin_border
                ws.cell(row=row, column=2, value=filing.employee_id).border = thin_border
                ws.cell(row=row, column=3, value=filing.employee_name).border = thin_border
                ws.cell(row=row, column=4, value=filing.department or '').border = thin_border
                ws.cell(row=row, column=5, value=filing.line_name or '').border = thin_border
                ws.cell(row=row, column=6, value=filing.get_shift_display()).border = thin_border
                ws.cell(row=row, column=7, value=filing.time_in.strftime('%H:%M') if filing.time_in else '').border = thin_border
                ws.cell(row=row, column=8, value=filing.time_out.strftime('%H:%M') if filing.time_out else '').border = thin_border
                ws.cell(row=row, column=9, value=filing.get_status_display()).border = thin_border
                ws.cell(row=row, column=10, value=group_name).border = thin_border
                row += 1
            
            row += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    if export_type == 'shifting':
        filename = f"overtime_shifting_week{week}_{year}.xlsx"
    else:
        filename = f"overtime_{export_type}_{date}.xlsx"
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
@require_http_methods(["GET"])
def get_weeks_for_year(request):
    year = int(request.GET.get('year', timezone.now().year))
    weeks = []
    
    first_day = datetime(year, 1, 1)
    if first_day.weekday() > 3:
        first_day = first_day + timedelta(7 - first_day.weekday())
    else:
        first_day = first_day - timedelta(first_day.weekday())
    
    current_date = first_day
    week_num = 1
    
    while current_date.year <= year:
        week_end = current_date + timedelta(days=6)
        if current_date.year == year or week_end.year == year:
            weeks.append({
                'week_number': week_num,
                'start_date': current_date.strftime('%Y-%m-%d'),
                'end_date': week_end.strftime('%Y-%m-%d'),
                'display': f"Week {week_num}: {current_date.strftime('%b %d')} - {week_end.strftime('%b %d, %Y')}"
            })
        current_date += timedelta(days=7)
        week_num += 1
        if week_num > 53:
            break
    
    return JsonResponse({'success': True, 'data': weeks})


def calculate_shift_summary(filings, math):
    """Calculate vehicle requirements summary by shift (day/night) for daily filing type"""
    
    # Separate filings by shift
    dayshift_filings = [f for f in filings if f.get('shift') == 'day']
    nightshift_filings = [f for f in filings if f.get('shift') == 'night']
    
    def get_shift_vehicle_data(shift_filings):
        """Calculate vehicle requirements for a specific shift"""
        result = {
            'not_ot_vehicles': 0,
            'ot_vehicles': 0,
            'provider_breakdown': []
        }
        
        if not shift_filings:
            return result
        
        # Get destination groups with providers
        destination_groups = DestinationGroup.objects.filter(
            is_active=True
        ).select_related('shuttle_vehicle', 'shuttle_provider').prefetch_related('destinations')
        
        # Group data by provider
        provider_data = {}
        
        for status in ['not_ot', 'ot']:
            # Get employee IDs with this status, EXCLUDING employees with their own vehicle
            status_employee_ids = [str(f['employee_id']) for f in shift_filings if f.get('status') == status and not f.get('with_vehicle', False)]
            
            if not status_employee_ids:
                continue
                
            for group in destination_groups:
                destination_ids = list(group.destinations.values_list('id', flat=True))
                
                if not destination_ids:
                    continue
                
                # Count employees assigned to destinations in this group (excluding those with vehicles)
                employee_count = UserShuttleAssignment.objects.filter(
                    employee_id__in=status_employee_ids,
                    destination_id__in=destination_ids,
                    with_vehicle=False
                ).count()
                
                if employee_count > 0:
                    capacity = group.shuttle_vehicle.capacity if group.shuttle_vehicle else 13
                    vehicles_needed = math.ceil(employee_count / capacity)
                    
                    # Get provider name
                    provider_name = group.shuttle_provider.name if group.shuttle_provider else 'No Provider'
                    provider_id = group.shuttle_provider.id if group.shuttle_provider else 0
                    
                    # Initialize provider data if needed
                    if provider_id not in provider_data:
                        provider_data[provider_id] = {
                            'provider_name': provider_name,
                            'not_ot_vehicles': 0,
                            'ot_vehicles': 0,
                            'total_vehicles': 0,
                            'not_ot_employees': 0,
                            'ot_employees': 0
                        }
                    
                    # Add vehicles and employee counts to the right category
                    if status == 'not_ot':
                        provider_data[provider_id]['not_ot_vehicles'] += vehicles_needed
                        provider_data[provider_id]['not_ot_employees'] += employee_count
                        result['not_ot_vehicles'] += vehicles_needed
                    else:
                        provider_data[provider_id]['ot_vehicles'] += vehicles_needed
                        provider_data[provider_id]['ot_employees'] += employee_count
                        result['ot_vehicles'] += vehicles_needed
                    
                    provider_data[provider_id]['total_vehicles'] += vehicles_needed
        
        # Convert provider data to list
        result['provider_breakdown'] = list(provider_data.values())
        
        return result
    
    return {
        'dayshift': get_shift_vehicle_data(dayshift_filings),
        'nightshift': get_shift_vehicle_data(nightshift_filings)
    }


def get_ot_employees_from_shifting(daily_date, not_ot_employee_ids):
    """
    Get OT employees from shifting records where the daily date falls within the shifting date range.
    Excludes employees already filed as 'Not OT' in Daily filing.
    
    Logic:
    - Each shifting record has date_from and date_to
    - If daily_date falls within that range (date_from <= daily_date <= date_to), 
      the employee is classified as OT
    - Employees already in not_ot_employee_ids are excluded
    """
    from datetime import datetime
    
    # Convert daily_date to date object if it's a string
    if isinstance(daily_date, str):
        daily_date = datetime.strptime(daily_date, '%Y-%m-%d').date()
    
    # Find shifting records where the daily date falls within the date range
    # and status is 'ot' (confirmed in shifting)
    shifting_records = OvertimeFiling.objects.filter(
        filing_type='shifting',
        status='ot',
        date_from__lte=daily_date,
        date_to__gte=daily_date
    ).exclude(
        employee_id__in=not_ot_employee_ids
    ).values(
        'id', 'filing_type', 'employee_id', 'employee_name', 'department',
        'line_name', 'shift', 'time_in', 'time_out', 'date_from', 'date_to',
        'status', 'is_late_filing', 'created_at'
    )
    
    return list(shifting_records)


def get_not_ot_employees_from_daily(shifting_date_from, shifting_date_to, ot_employee_ids):
    """
    Get Not OT employees from daily records where the daily date falls within the shifting date range.
    Excludes employees already filed as 'OT' in Shifting filing.
    
    Logic:
    - For each daily filing, if the daily date (date_from) falls within shifting date range,
      the employee is classified as Not OT
    - Employees already in ot_employee_ids are excluded
    """
    from datetime import datetime
    
    # Convert dates to date objects if they're strings
    if isinstance(shifting_date_from, str):
        shifting_date_from = datetime.strptime(shifting_date_from, '%Y-%m-%d').date()
    if isinstance(shifting_date_to, str):
        shifting_date_to = datetime.strptime(shifting_date_to, '%Y-%m-%d').date()
    
    # Find daily records where the daily date falls within the shifting date range
    # and status is 'not_ot' (confirmed in daily)
    daily_records = OvertimeFiling.objects.filter(
        filing_type='daily',
        status='not_ot',
        date_from__gte=shifting_date_from,
        date_from__lte=shifting_date_to
    ).exclude(
        employee_id__in=ot_employee_ids
    ).values(
        'id', 'filing_type', 'employee_id', 'employee_name', 'department',
        'line_name', 'shift', 'time_in', 'time_out', 'date_from', 'date_to',
        'status', 'is_late_filing', 'created_at'
    )
    
    return list(daily_records)


@login_required
@require_http_methods(["GET"])
def get_overview_data(request):
    """Get overview data for admin dashboard with filings, counts, and vehicle requirements"""
    filing_type = request.GET.get('filing_type', 'daily')
    date = request.GET.get('date')
    week = request.GET.get('week')
    year = request.GET.get('year')
    
    import math
    
    # Build query based on filing type
    filings_query = OvertimeFiling.objects.all()
    
    # Determine the date for the query
    query_date = date
    if not query_date:
        query_date = timezone.localtime().date()
    
    if filing_type == 'shifting' and week and year:
        # Filter by week number and year
        filings_query = filings_query.filter(
            filing_type=filing_type,
            week_number=int(week),
            year=int(year)
        )
    elif filing_type == 'shifting' and date:
        # For Shifting: Find records where selected date falls within the shifting period
        # Condition: date_from <= selected_date <= date_to
        from datetime import datetime
        # Convert date string to date object if needed
        if isinstance(date, str):
            selected_date = datetime.strptime(date, '%Y-%m-%d').date()
        else:
            selected_date = date
        filings_query = filings_query.filter(
            filing_type=filing_type,
            date_from__lte=selected_date,
            date_to__gte=selected_date
        )
    elif date:
        # For other filing types: exact date match
        filings_query = filings_query.filter(
            filing_type=filing_type,
            date_from=date
        )
    else:
        # Default to today
        today = timezone.localtime().date()
        if filing_type == 'shifting':
            # For Shifting: Find records where today falls within the shifting period
            filings_query = filings_query.filter(
                filing_type=filing_type,
                date_from__lte=today,
                date_to__gte=today
            )
        else:
            filings_query = filings_query.filter(
                filing_type=filing_type,
                date_from=today
            )
    
    # Get all filings
    filings_raw = list(filings_query.values(
        'id', 'filing_type', 'employee_id', 'employee_name', 'department',
        'line_name', 'shift', 'time_in', 'time_out', 'date_from', 'date_to',
        'status', 'is_late_filing', 'created_at'
    ).order_by('-created_at'))
    
    # For Daily filing type: Auto-detect OT employees from Shifting records
    # Users only file "Not OT" employees; "OT" employees are derived from shifting
    if filing_type == 'daily':
        # Get employee IDs that are already filed as "Not OT" in this daily filing
        not_ot_employee_ids = [str(f['employee_id']) for f in filings_raw if f.get('status') == 'not_ot']
        
        # Get OT employees from shifting records where daily date falls within shifting date range
        ot_from_shifting = get_ot_employees_from_shifting(query_date, not_ot_employee_ids)
        
        # Mark these as auto-generated OT and add to filings
        for ot_filing in ot_from_shifting:
            ot_filing['auto_ot'] = True  # Flag to indicate this is auto-generated from shifting
            ot_filing['original_filing_type'] = 'shifting'  # Track the source
            filings_raw.append(ot_filing)
    
    # For Shifting filing type: Auto-detect Not OT employees from Daily records
    # Users only file "OT" employees; "Not OT" employees are derived from daily
    if filing_type == 'shifting' and filings_raw:
        # Get the date range from the first shifting filing (they all should have same week)
        first_filing = filings_raw[0] if filings_raw else None
        if first_filing:
            shifting_date_from = first_filing.get('date_from')
            shifting_date_to = first_filing.get('date_to')
            
            if shifting_date_from and shifting_date_to:
                # Get employee IDs that are already filed as "OT" in this shifting filing
                ot_employee_ids = [str(f['employee_id']) for f in filings_raw if f.get('status') == 'ot']
                
                # Get Not OT employees from daily records where daily date falls within shifting date range
                not_ot_from_daily = get_not_ot_employees_from_daily(shifting_date_from, shifting_date_to, ot_employee_ids)
                
                # Mark these as auto-generated Not OT and add to filings
                for not_ot_filing in not_ot_from_daily:
                    not_ot_filing['auto_not_ot'] = True  # Flag to indicate this is auto-generated from daily
                    not_ot_filing['original_filing_type'] = 'daily'  # Track the source
                    filings_raw.append(not_ot_filing)
    
    # Get employee IDs to fetch their destinations (convert to string for CharField lookup)
    employee_ids = [str(f['employee_id']) for f in filings_raw]
    
    # Get destination assignments for these employees (including with_vehicle status)
    destination_map = {}
    with_vehicle_map = {}
    if employee_ids:
        assignments = UserShuttleAssignment.objects.filter(
            employee_id__in=employee_ids
        ).select_related('destination').values('employee_id', 'destination__name', 'with_vehicle')
        
        for assignment in assignments:
            # Map both string and int versions of employee_id
            destination_map[assignment['employee_id']] = assignment['destination__name'] or '-'
            destination_map[int(assignment['employee_id']) if assignment['employee_id'].isdigit() else assignment['employee_id']] = assignment['destination__name'] or '-'
            # Track which employees have their own vehicle
            with_vehicle_map[assignment['employee_id']] = assignment['with_vehicle']
            with_vehicle_map[int(assignment['employee_id']) if assignment['employee_id'].isdigit() else assignment['employee_id']] = assignment['with_vehicle']
    
    # Add destination_name and with_vehicle to each filing
    filings = []
    for filing in filings_raw:
        filing['destination_name'] = destination_map.get(filing['employee_id'], destination_map.get(str(filing['employee_id']), '-'))
        filing['with_vehicle'] = with_vehicle_map.get(filing['employee_id'], with_vehicle_map.get(str(filing['employee_id']), False))
        filings.append(filing)
    
    
    # Calculate status counts
    status_counts = {
        'ot': 0,
        'not_ot': 0,
        'absent': 0,
        'leave': 0
    }
    
    for filing in filings:
        status = filing.get('status', 'not_ot')
        if status in status_counts:
            status_counts[status] += 1
    
    # Calculate vehicle requirements by destination group
    vehicle_requirements = []
    
    # Determine which status to count for vehicle requirements
    # Daily = Not OT, all others = OT
    if filing_type == 'daily':
        vehicle_status_filter = 'not_ot'
    else:
        vehicle_status_filter = 'ot'
    
    # Get employee IDs from filings with the appropriate status (convert to string for CharField lookup)
    # EXCLUDE employees who have their own vehicle (with_vehicle=True)
    filtered_employee_ids = [str(f['employee_id']) for f in filings if f.get('status') == vehicle_status_filter and not f.get('with_vehicle', False)]
    
    if filtered_employee_ids:
        # Get destination groups with their vehicles and destinations
        destination_groups = DestinationGroup.objects.filter(is_active=True).select_related('shuttle_vehicle', 'shuttle_provider').prefetch_related('destinations')
        
        for group in destination_groups:
            # Get destination IDs for this group
            destination_ids = list(group.destinations.values_list('id', flat=True))
            # Get destination names for display
            destination_names = list(group.destinations.filter(is_active=True).values_list('name', flat=True))
            
            if not destination_ids:
                continue
            
            # Count employees with the appropriate status assigned to destinations in this group
            # Employees with their own vehicle are already filtered out from filtered_employee_ids
            employee_count = UserShuttleAssignment.objects.filter(
                employee_id__in=filtered_employee_ids,
                destination_id__in=destination_ids,
                with_vehicle=False  # Extra safety check
            ).count()
            
            if employee_count > 0:
                capacity = group.shuttle_vehicle.capacity if group.shuttle_vehicle else 13
                vehicles_needed = math.ceil(employee_count / capacity)
                
                vehicle_requirements.append({
                    'group_id': group.id,
                    'group_name': group.name,
                    'vehicle_name': group.shuttle_vehicle.name if group.shuttle_vehicle else 'No Vehicle',
                    'vehicle_capacity': capacity,
                    'employee_count': employee_count,
                    'vehicles_needed': vehicles_needed,
                    'destinations': destination_names
                })
    
    # Calculate shift summary for daily and shifting filing types
    shift_summary = None
    if filing_type in ['daily', 'shifting']:
        shift_summary = calculate_shift_summary(filings, math)
    
    return JsonResponse({
        'success': True,
        'data': {
            'filings': filings,
            'status_counts': status_counts,
            'vehicle_requirements': vehicle_requirements,
            'total_employees': len(filings),
            'shift_summary': shift_summary
        }
    })


@login_required
@require_http_methods(["GET"])
def export_overview_data(request):
    """Export overview data to Excel with proper formatting and shuttle grouping"""
    import math
    
    filing_type = request.GET.get('filing_type', 'daily')
    date = request.GET.get('date')
    shift = request.GET.get('shift')  # 'day' or 'night' for daily filing type
    
    # Filing type display names
    filing_type_display = {
        'daily': 'Daily',
        'shifting': 'Shifting',
        'sunday': 'Sunday',
        'holiday': 'Holiday',
        'saturday_off': 'Saturday Off'
    }
    
    shift_display = {
        'day': 'Day Shift',
        'night': 'Night Shift'
    }
    
    # Build query based on filing type
    filings_query = OvertimeFiling.objects.all()
    
    if filing_type == 'shifting' and date:
        # For Shifting: Find records where selected date falls within the shifting period
        from datetime import datetime
        if isinstance(date, str):
            selected_date = datetime.strptime(date, '%Y-%m-%d').date()
        else:
            selected_date = date
        filings_query = filings_query.filter(
            filing_type=filing_type,
            date_from__lte=selected_date,
            date_to__gte=selected_date
        )
    elif date:
        filings_query = filings_query.filter(
            filing_type=filing_type,
            date_from=date
        )
    else:
        today = timezone.localtime().date()
        date = today.strftime('%Y-%m-%d')
        if filing_type == 'shifting':
            filings_query = filings_query.filter(
                filing_type=filing_type,
                date_from__lte=today,
                date_to__gte=today
            )
        else:
            filings_query = filings_query.filter(
                filing_type=filing_type,
                date_from=today
            )
    
    # For daily filing type with shift parameter, filter by shift and create 2-sheet export
    if filing_type == 'daily' and shift in ['day', 'night']:
        filings_query = filings_query.filter(shift=shift)
        return export_daily_with_shift(filings_query, date, shift, shift_display, math)
    
    # For shifting filing type with shift parameter, filter by shift and create 2-sheet export
    if filing_type == 'shifting' and shift in ['day', 'night']:
        filings_query = filings_query.filter(shift=shift)
        return export_shifting_with_not_ot(filings_query, date, shift, shift_display, math)
    
    # Standard export (original behavior)
    # Determine which status to filter by based on filing type
    if filing_type == 'daily':
        status_filter = 'not_ot'
    else:
        status_filter = 'ot'
    
    filings_query = filings_query.filter(status=status_filter)
    filings = list(filings_query)
    
    # Get employee IDs (convert to string for UserShuttleAssignment lookup)
    employee_ids = [str(f.employee_id) for f in filings]
    
    # Get employees with vehicles and exclude them from the export
    with_vehicle_employee_ids = set(
        UserShuttleAssignment.objects.filter(
            employee_id__in=employee_ids,
            with_vehicle=True
        ).values_list('employee_id', flat=True)
    )
    
    if with_vehicle_employee_ids:
        logger.info(f"Standard export: Excluding {len(with_vehicle_employee_ids)} employees with vehicles")
        filings = [f for f in filings if str(f.employee_id) not in with_vehicle_employee_ids]
        employee_ids = [str(f.employee_id) for f in filings]
    
    # Get destination assignments with destination group info
    employee_destination_map = {}
    employee_group_map = {}
    
    if employee_ids:
        assignments = UserShuttleAssignment.objects.filter(
            employee_id__in=employee_ids
        ).select_related('destination')
        
        for assignment in assignments:
            employee_destination_map[assignment.employee_id] = assignment.destination.name if assignment.destination else '-'
    
    # Get all destination groups with their destinations and vehicles
    destination_groups = DestinationGroup.objects.filter(is_active=True).select_related('shuttle_vehicle').prefetch_related('destinations')
    
    # Create a map of destination_id to group
    dest_to_group = {}
    for group in destination_groups:
        for dest in group.destinations.all():
            dest_to_group[dest.id] = group
    
    # Get which group each employee belongs to
    if employee_ids:
        assignments = UserShuttleAssignment.objects.filter(
            employee_id__in=employee_ids,
            destination__isnull=False
        ).select_related('destination')
        
        for assignment in assignments:
            if assignment.destination_id in dest_to_group:
                employee_group_map[assignment.employee_id] = dest_to_group[assignment.destination_id]
    
    # Group filings by destination group
    grouped_filings = {}
    unassigned_filings = []
    
    for filing in filings:
        # Use string employee_id for map lookups
        emp_id_str = str(filing.employee_id)
        group = employee_group_map.get(emp_id_str)
        if group:
            if group.id not in grouped_filings:
                grouped_filings[group.id] = {
                    'group': group,
                    'filings': [],
                    'destinations': set()
                }
            grouped_filings[group.id]['filings'].append(filing)
            dest_name = employee_destination_map.get(emp_id_str, '-')
            if dest_name != '-':
                grouped_filings[group.id]['destinations'].add(dest_name)
        else:
            unassigned_filings.append(filing)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Filed Overtime"
    
    # Write the standard export
    write_export_sheet(ws, grouped_filings, unassigned_filings, employee_destination_map, date, filing_type, filing_type_display, math)
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f"overtime_{filing_type}_{date}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


def export_daily_with_shift(filings_query, date, shift, shift_display, math):
    """Export daily overtime with specific shift into 2 sheets (Not OT and OT)"""
    
    filings = list(filings_query)
    
    # Separate filings by status - users only file "Not OT" for daily
    not_ot_filings = [f for f in filings if f.status == 'not_ot']
    
    # Get employee IDs that are "Not OT" to exclude from auto-OT
    not_ot_employee_ids = [str(f.employee_id) for f in not_ot_filings]
    
    # Auto-detect OT employees from Shifting records
    # Find shifting records where daily date falls within date range
    from datetime import datetime
    if isinstance(date, str):
        daily_date = datetime.strptime(date, '%Y-%m-%d').date()
    else:
        daily_date = date
    
    # Get OT employees from shifting with the same shift type
    ot_from_shifting = OvertimeFiling.objects.filter(
        filing_type='shifting',
        status='ot',
        shift=shift,
        date_from__lte=daily_date,
        date_to__gte=daily_date
    ).exclude(
        employee_id__in=not_ot_employee_ids
    )
    
    ot_filings = list(ot_from_shifting)
    
    # Combine all filings to get employee IDs
    all_filings = not_ot_filings + ot_filings
    
    # Get employee IDs (convert to string for UserShuttleAssignment lookup)
    employee_ids = [str(f.employee_id) for f in all_filings]
    
    # Get destination assignments with destination group info and with_vehicle status
    employee_destination_map = {}
    employee_group_map = {}
    employee_with_vehicle_map = {}  # Track employees with their own vehicle
    
    if employee_ids:
        assignments = UserShuttleAssignment.objects.filter(
            employee_id__in=employee_ids
        ).select_related('destination')
        
        for assignment in assignments:
            employee_destination_map[assignment.employee_id] = assignment.destination.name if assignment.destination else '-'
            employee_with_vehicle_map[assignment.employee_id] = assignment.with_vehicle
    
    # Filter out employees with their own vehicle from the export lists
    not_ot_filings = [f for f in not_ot_filings if not employee_with_vehicle_map.get(str(f.employee_id), False)]
    ot_filings = [f for f in ot_filings if not employee_with_vehicle_map.get(str(f.employee_id), False)]
    
    # Get all destination groups with their destinations and vehicles
    destination_groups = DestinationGroup.objects.filter(is_active=True).select_related('shuttle_vehicle').prefetch_related('destinations')
    
    # Create a map of destination_id to group
    dest_to_group = {}
    for group in destination_groups:
        for dest in group.destinations.all():
            dest_to_group[dest.id] = group
    
    # Get which group each employee belongs to (excluding those with vehicles)
    filtered_employee_ids = [str(f.employee_id) for f in not_ot_filings + ot_filings]
    if filtered_employee_ids:
        assignments = UserShuttleAssignment.objects.filter(
            employee_id__in=filtered_employee_ids,
            destination__isnull=False,
            with_vehicle=False
        ).select_related('destination')
        
        for assignment in assignments:
            if assignment.destination_id in dest_to_group:
                employee_group_map[assignment.employee_id] = dest_to_group[assignment.destination_id]
    
    def group_filings_by_destination(filings_list):
        """Group filings by destination group"""
        grouped = {}
        unassigned = []
        
        for filing in filings_list:
            emp_id_str = str(filing.employee_id)
            group = employee_group_map.get(emp_id_str)
            if group:
                if group.id not in grouped:
                    grouped[group.id] = {
                        'group': group,
                        'filings': [],
                        'destinations': set()
                    }
                grouped[group.id]['filings'].append(filing)
                dest_name = employee_destination_map.get(emp_id_str, '-')
                if dest_name != '-':
                    grouped[group.id]['destinations'].add(dest_name)
            else:
                unassigned.append(filing)
        
        return grouped, unassigned
    
    # Create workbook
    wb = Workbook()
    
    # Sheet 1: Not OT (manually filed by users)
    ws_not_ot = wb.active
    ws_not_ot.title = f"Not OT - {shift_display.get(shift, shift.title())}"
    not_ot_grouped, not_ot_unassigned = group_filings_by_destination(not_ot_filings)
    write_export_sheet(ws_not_ot, not_ot_grouped, not_ot_unassigned, employee_destination_map, date, 'daily', {'daily': 'Daily (Not OT)'}, math, shift_display.get(shift, ''))
    
    # Sheet 2: OT (auto-generated from shifting records)
    ws_ot = wb.create_sheet(f"OT - {shift_display.get(shift, shift.title())}")
    ot_grouped, ot_unassigned = group_filings_by_destination(ot_filings)
    write_export_sheet(ws_ot, ot_grouped, ot_unassigned, employee_destination_map, date, 'daily', {'daily': 'Daily (OT - from Shifting)'}, math, shift_display.get(shift, ''))
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f"overtime_daily_{shift}shift_{date}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


def export_shifting_with_not_ot(filings_query, date, shift, shift_display, math):
    """Export shifting overtime with specific shift into 2 sheets (OT and Not OT from Daily)"""
    
    filings = list(filings_query)
    
    # Separate filings by status - users only file "OT" for shifting
    ot_filings = [f for f in filings if f.status == 'ot']
    
    # Get date range from the first filing (all shifting filings should have same date range)
    if not ot_filings:
        # If no OT filings, create empty workbook
        wb = Workbook()
        ws_ot = wb.active
        ws_ot.title = f"OT - {shift_display.get(shift, shift.title())}"
        ws_not_ot = wb.create_sheet(f"Not OT - {shift_display.get(shift, shift.title())}")
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"overtime_shifting_{shift}shift_{date}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
    
    # Get the date range from the first shifting filing
    first_filing = ot_filings[0]
    shifting_date_from = first_filing.date_from
    shifting_date_to = first_filing.date_to
    
    # Get employee IDs that are "OT" to exclude from auto-Not OT
    ot_employee_ids = [str(f.employee_id) for f in ot_filings]
    
    # Auto-detect Not OT employees from Daily records
    # Find daily records where date falls within shifting date range
    not_ot_from_daily = OvertimeFiling.objects.filter(
        filing_type='daily',
        status='not_ot',
        shift=shift,
        date_from__gte=shifting_date_from,
        date_from__lte=shifting_date_to
    ).exclude(
        employee_id__in=ot_employee_ids
    )
    
    # Get unique employees (one employee may have multiple daily filings)
    seen_employees = set()
    not_ot_filings = []
    for filing in not_ot_from_daily:
        if filing.employee_id not in seen_employees:
            seen_employees.add(filing.employee_id)
            not_ot_filings.append(filing)
    
    # Combine all filings to get employee IDs
    all_filings = ot_filings + not_ot_filings
    
    # Get employee IDs (convert to string for UserShuttleAssignment lookup)
    employee_ids = [str(f.employee_id) for f in all_filings]
    
    # Get destination assignments with destination group info and with_vehicle status
    employee_destination_map = {}
    employee_group_map = {}
    employee_with_vehicle_map = {}  # Track employees with their own vehicle
    
    if employee_ids:
        assignments = UserShuttleAssignment.objects.filter(
            employee_id__in=employee_ids
        ).select_related('destination')
        
        for assignment in assignments:
            employee_destination_map[assignment.employee_id] = assignment.destination.name if assignment.destination else '-'
            employee_with_vehicle_map[assignment.employee_id] = assignment.with_vehicle
    
    # Filter out employees with their own vehicle from the export lists
    ot_filings = [f for f in ot_filings if not employee_with_vehicle_map.get(str(f.employee_id), False)]
    not_ot_filings = [f for f in not_ot_filings if not employee_with_vehicle_map.get(str(f.employee_id), False)]
    
    # Get all destination groups with their destinations and vehicles
    destination_groups = DestinationGroup.objects.filter(is_active=True).select_related('shuttle_vehicle').prefetch_related('destinations')
    
    # Create a map of destination_id to group
    dest_to_group = {}
    for group in destination_groups:
        for dest in group.destinations.all():
            dest_to_group[dest.id] = group
    
    # Get which group each employee belongs to (excluding those with vehicles)
    filtered_employee_ids = [str(f.employee_id) for f in ot_filings + not_ot_filings]
    if filtered_employee_ids:
        assignments = UserShuttleAssignment.objects.filter(
            employee_id__in=filtered_employee_ids,
            destination__isnull=False,
            with_vehicle=False
        ).select_related('destination')
        
        for assignment in assignments:
            if assignment.destination_id in dest_to_group:
                employee_group_map[assignment.employee_id] = dest_to_group[assignment.destination_id]
    
    def group_filings_by_destination(filings_list):
        """Group filings by destination group"""
        grouped = {}
        unassigned = []
        
        for filing in filings_list:
            emp_id_str = str(filing.employee_id)
            group = employee_group_map.get(emp_id_str)
            if group:
                if group.id not in grouped:
                    grouped[group.id] = {
                        'group': group,
                        'filings': [],
                        'destinations': set()
                    }
                grouped[group.id]['filings'].append(filing)
                dest_name = employee_destination_map.get(emp_id_str, '-')
                if dest_name != '-':
                    grouped[group.id]['destinations'].add(dest_name)
            else:
                unassigned.append(filing)
        
        return grouped, unassigned
    
    # Format date range for display
    from datetime import datetime
    try:
        if isinstance(shifting_date_from, str):
            date_from_obj = datetime.strptime(shifting_date_from, '%Y-%m-%d')
        else:
            date_from_obj = shifting_date_from
        if isinstance(shifting_date_to, str):
            date_to_obj = datetime.strptime(shifting_date_to, '%Y-%m-%d')
        else:
            date_to_obj = shifting_date_to
        formatted_date = f"{date_from_obj.strftime('%b %d')} - {date_to_obj.strftime('%b %d, %Y')}"
    except:
        formatted_date = f"{shifting_date_from} - {shifting_date_to}"
    
    # Create workbook
    wb = Workbook()
    
    # Sheet 1: OT (manually filed by users)
    ws_ot = wb.active
    ws_ot.title = f"OT - {shift_display.get(shift, shift.title())}"
    ot_grouped, ot_unassigned = group_filings_by_destination(ot_filings)
    write_export_sheet(ws_ot, ot_grouped, ot_unassigned, employee_destination_map, formatted_date, 'shifting', {'shifting': 'Shifting (OT)'}, math, shift_display.get(shift, ''))
    
    # Sheet 2: Not OT (auto-generated from daily records)
    ws_not_ot = wb.create_sheet(f"Not OT - {shift_display.get(shift, shift.title())}")
    not_ot_grouped, not_ot_unassigned = group_filings_by_destination(not_ot_filings)
    write_export_sheet(ws_not_ot, not_ot_grouped, not_ot_unassigned, employee_destination_map, formatted_date, 'shifting', {'shifting': 'Shifting (Not OT - from Daily)'}, math, shift_display.get(shift, ''))
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f"overtime_shifting_{shift}shift_{date}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


def write_export_sheet(ws, grouped_filings, unassigned_filings, employee_destination_map, date, filing_type, filing_type_display, math, shift_label=''):
    """Helper function to write data to an Excel sheet"""
    
    # Define styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    bold_font = Font(bold=True)
    header_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=12)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center')
    
    # Row 1: Main header
    ws.merge_cells('A1:E1')
    cell = ws.cell(row=1, column=1, value="RYONAN ELECTRIC PHILIPPINES CORPORATION")
    cell.font = title_font
    cell.alignment = center_align
    
    # Row 2: Sub header - Summary with date and type
    ws.merge_cells('A2:E2')
    # Format date
    try:
        date_obj = datetime.strptime(date, '%Y-%m-%d')
        formatted_date = date_obj.strftime('%B %d, %Y')
    except:
        formatted_date = date
    
    filing_type_name = filing_type_display.get(filing_type, filing_type.title())
    shift_suffix = f" – {shift_label}" if shift_label else ""
    cell = ws.cell(row=2, column=1, value=f"Summary for Filed Overtime for {formatted_date} – {filing_type_name}{shift_suffix}")
    cell.font = subtitle_font
    cell.alignment = center_align
    
    # Row 3: Empty row
    current_row = 4
    
    # Column headers
    headers = ['#', 'ID #', 'NAME OF EMPLOYEE', 'DEPT. / LINE', 'DESTINATION', 'SHUTTLE']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_align
    
    current_row += 1
    employee_number = 1
    
    # Process each destination group
    for group_id, group_data in sorted(grouped_filings.items(), key=lambda x: x[1]['group'].name):
        group = group_data['group']
        group_filings = group_data['filings']
        destinations = sorted(group_data['destinations'])
        
        if not group_filings:
            continue
        
        # Get vehicle capacity
        capacity = group.shuttle_vehicle.capacity if group.shuttle_vehicle else 13
        vehicle_name = group.shuttle_vehicle.name if group.shuttle_vehicle else 'VAN'
        
        # Count employees per destination within this group
        dest_count = {}
        for filing in group_filings:
            dest_name = employee_destination_map.get(str(filing.employee_id), '-')
            if dest_name not in dest_count:
                dest_count[dest_name] = 0
            dest_count[dest_name] += 1
        
        # Sort destinations by employee count (descending)
        sorted_destinations = sorted(dest_count.keys(), key=lambda d: dest_count[d], reverse=True)
        
        # Sort filings by destination (largest destination first)
        def get_dest_priority(filing):
            dest_name = employee_destination_map.get(str(filing.employee_id), '-')
            try:
                return sorted_destinations.index(dest_name)
            except ValueError:
                return len(sorted_destinations)
        
        group_filings_sorted = sorted(group_filings, key=get_dest_priority)
        
        # Calculate how many vehicles needed
        total_passengers = len(group_filings_sorted)
        vehicles_needed = math.ceil(total_passengers / capacity)
        
        # Create shuttle label with destinations (sorted by count)
        destinations_str = ' / '.join(sorted_destinations)
        if group.name not in destinations_str:
            shuttle_label = f"{destinations_str} / {group.name}"
        else:
            shuttle_label = destinations_str
        
        # Split filings into vehicle groups
        vehicle_groups = []
        for i in range(0, total_passengers, capacity):
            vehicle_groups.append(group_filings_sorted[i:i + capacity])
        
        # Process each vehicle group
        for vehicle_idx, vehicle_filings in enumerate(vehicle_groups):
            start_row = current_row
            
            for filing in vehicle_filings:
                # Format department/line as "Department – Line"
                dept = filing.department or ''
                line = filing.line_name or ''
                
                if dept and line:
                    if dept.strip().lower() == line.strip().lower():
                        dept_line = dept
                    else:
                        dept_line = f"{dept} – {line}"
                elif dept:
                    dept_line = dept
                elif line:
                    dept_line = line
                else:
                    dept_line = ''
                
                # Get destination (use string employee_id for map lookup)
                destination = employee_destination_map.get(str(filing.employee_id), '-')
                
                # Write data
                ws.cell(row=current_row, column=1, value=employee_number).border = thin_border
                ws.cell(row=current_row, column=1).alignment = center_align
                
                # Write employee_id as string to prevent Excel from treating it as a number
                emp_id_cell = ws.cell(row=current_row, column=2, value=str(filing.employee_id))
                emp_id_cell.border = thin_border
                emp_id_cell.number_format = '@'  # Text format
                
                ws.cell(row=current_row, column=3, value=filing.employee_name).border = thin_border
                
                ws.cell(row=current_row, column=4, value=dept_line).border = thin_border
                
                ws.cell(row=current_row, column=5, value=destination).border = thin_border
                
                # Shuttle column will be merged later
                ws.cell(row=current_row, column=6, value='').border = thin_border
                
                employee_number += 1
                current_row += 1
            
            end_row = current_row - 1
            
            # Merge and fill shuttle column for this vehicle group
            if start_row <= end_row:
                if start_row < end_row:
                    ws.merge_cells(start_row=start_row, start_column=6, end_row=end_row, end_column=6)
                
                shuttle_cell = ws.cell(row=start_row, column=6)
                
                # Show shuttle label with "1 VEHICLE_NAME" for each group
                shuttle_cell.value = f"{shuttle_label}\n1 {vehicle_name.upper()}"
                
                shuttle_cell.alignment = center_align
                shuttle_cell.border = thin_border
    
    # Process unassigned filings
    if unassigned_filings:
        for filing in unassigned_filings:
            # Format department/line as "Department – Line"
            dept = filing.department or ''
            line = filing.line_name or ''
            
            if dept and line:
                if dept.strip().lower() == line.strip().lower():
                    dept_line = dept
                else:
                    dept_line = f"{dept} – {line}"
            elif dept:
                dept_line = dept
            elif line:
                dept_line = line
            else:
                dept_line = ''
            
            # Get destination (use string employee_id for map lookup)
            destination = employee_destination_map.get(str(filing.employee_id), '-')
            
            ws.cell(row=current_row, column=1, value=employee_number).border = thin_border
            ws.cell(row=current_row, column=1).alignment = center_align
            
            # Write employee_id as string to prevent Excel from treating it as a number
            emp_id_cell = ws.cell(row=current_row, column=2, value=str(filing.employee_id))
            emp_id_cell.border = thin_border
            emp_id_cell.number_format = '@'  # Text format
            
            ws.cell(row=current_row, column=3, value=filing.employee_name).border = thin_border
            
            ws.cell(row=current_row, column=4, value=dept_line).border = thin_border
            
            ws.cell(row=current_row, column=5, value=destination).border = thin_border
            
            ws.cell(row=current_row, column=6, value='Unassigned').border = thin_border
            
            employee_number += 1
            current_row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 20


@login_required
@require_http_methods(["GET"])
def export_shuttle_users(request):
    """Export user shuttle assignments to Excel"""
    
    # Get all shuttle users
    shuttle_users = UserShuttleAssignment.objects.select_related('destination').all().order_by('employee_name')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Shuttle Assignments"
    
    # Define styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Column headers
    headers = ['ID Number', 'Employee Name', 'Department', 'Line', 'Shuttle Destination']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
    
    # Write data
    row = 2
    for user in shuttle_users:
        # Write employee_id as string to prevent Excel from treating it as a number
        emp_id_cell = ws.cell(row=row, column=1, value=str(user.employee_id))
        emp_id_cell.border = thin_border
        emp_id_cell.number_format = '@'  # Text format
        
        ws.cell(row=row, column=2, value=user.employee_name).border = thin_border
        ws.cell(row=row, column=3, value=user.department or '').border = thin_border
        ws.cell(row=row, column=4, value=user.line_name or '').border = thin_border
        ws.cell(row=row, column=5, value=user.destination.name if user.destination else '').border = thin_border
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    today = timezone.localtime().date().strftime('%Y-%m-%d')
    filename = f"shuttle_assignments_{today}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response