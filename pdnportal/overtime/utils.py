import openpyxl
import random
import string
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime, time, timedelta
from django.utils import timezone
from .models import OvertimePasscode

PASSCODE_REFRESH_HOUR = 7   # 7:00 AM
PASSCODE_REFRESH_MINUTE = 0


def generate_passcode():
    """Generate a secure 6-character passcode with uppercase, lowercase, digit, and special char."""
    special_chars = '!@#$%^&*'
    chars = [
        random.choice(string.ascii_uppercase),
        random.choice(string.ascii_lowercase),
        random.choice(string.digits),
        random.choice(special_chars),
    ]
    pool = string.ascii_uppercase + string.ascii_lowercase + string.digits + special_chars
    chars += random.choices(pool, k=2)
    random.shuffle(chars)
    return ''.join(chars)


def get_or_refresh_passcode():
    """
    Return the current active passcode, auto-generating a new one if stale.
    A passcode is stale when its last update date is before today AND
    the current local time is at or past PASSCODE_REFRESH_HOUR:PASSCODE_REFRESH_MINUTE.
    """
    now = timezone.localtime(timezone.now())
    refresh_due = now.hour * 60 + now.minute >= PASSCODE_REFRESH_HOUR * 60 + PASSCODE_REFRESH_MINUTE

    passcode_obj = OvertimePasscode.objects.filter(is_active=True).first()

    stale = (
        passcode_obj is None
        or (
            refresh_due
            and timezone.localtime(passcode_obj.updated_at).date() < now.date()
        )
    )

    if stale:
        new_passcode = generate_passcode()
        OvertimePasscode.objects.filter(is_active=True).update(is_active=False)
        passcode_obj = OvertimePasscode.objects.create(passcode=new_passcode, is_active=True)

    return passcode_obj

def is_late_filing(filing_type, filing_date, schedule_type=None):
    current_date = timezone.now().date()
    current_time = timezone.now().time()
    current_weekday = current_date.weekday()
    
    if filing_type == 'SHIFTING':
        return current_weekday > 3
    
    elif filing_type == 'DAILY':
        if schedule_type in ['SUNDAY', 'SATURDAY', 'HOLIDAY']:
            return current_weekday > 4 or current_weekday == 0
        else:
            cutoff_time = time(10, 0)
            return current_date == filing_date and current_time > cutoff_time
    
    return False

def proper_case(text):
    if not text:
        return text
    
    words = text.lower().split()
    lowercase_words = ['of', 'the', 'in', 'on', 'at', 'by', 'for', 'with', 'a', 'an', 'and', 'but', 'or', 'nor', 'to', 'from']
    
    lowercase_prefixes = ['mc', 'mac']
    
    result = []
    for i, word in enumerate(words):
        if i == 0 or word not in lowercase_words:
            prefix_found = False
            for prefix in lowercase_prefixes:
                if word.startswith(prefix) and len(word) > len(prefix):
                    result.append(prefix + word[len(prefix)].upper() + word[len(prefix)+1:])
                    prefix_found = True
                    break
            
            if not prefix_found:
                if '-' in word:
                    hyphenated = []
                    for part in word.split('-'):
                        if part:
                            hyphenated.append(part[0].upper() + part[1:])
                        else:
                            hyphenated.append('')
                    result.append('-'.join(hyphenated))
                else:
                    result.append(word[0].upper() + word[1:])
        else:
            result.append(word)
    
    return ' '.join(result)


def generate_excel_file(headers, data, filename):

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename.split('.')[0]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    for row_num, row_data in enumerate(data, 2):
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=row_num, column=col_num, value=row_data.get(header, ''))
    
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = max(15, len(header) + 5)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response