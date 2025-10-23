from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, HttpResponseForbidden, HttpResponse
from django.db.models import Count
from django.core.paginator import Paginator
from django.utils import timezone
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import calendar
from .models import ECIS
from .forms import ECISForm, FacilitatorReviewForm, CancelRequestForm
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Requestor Views
@login_required(login_url="user-login")
def ecis_list_requestor(request):
    current_month = timezone.now().month
    current_year = timezone.now().year

    all_ecis = ECIS.objects.filter(created_by=request.user).order_by('-last_updated')

    total_count = all_ecis.filter(date_prepared__month=current_month, date_prepared__year=current_year).count()
    approved_count = all_ecis.filter(status='Approved', date_prepared__month=current_month, date_prepared__year=current_year).count()
    onhold_count = all_ecis.filter(status='On Hold', date_prepared__month=current_month, date_prepared__year=current_year).count()
    review_count = all_ecis.filter(status='For Review', date_prepared__month=current_month, date_prepared__year=current_year).count()

    onhold_items = all_ecis.filter(status='On Hold').order_by('-last_updated')[:5]
    review_items = all_ecis.filter(status='For Review').order_by('-last_updated')[:5]
    revision_items = all_ecis.filter(status='Needs Revision').order_by('-last_updated')[:5]

    paginator = Paginator(all_ecis, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'ecis_list': page_obj,
        'total_count': total_count,
        'approved_count': approved_count,
        'onhold_count': onhold_count,
        'review_count': review_count,
        'onhold_items': onhold_items,
        'review_items': review_items,
        'revision_items': revision_items,
        'page_obj': page_obj,
        'paginator': paginator,
        'is_paginated': paginator.num_pages > 1,
    }

    return render(request, 'ecis/requestor.html', context)

@login_required(login_url="user-login")
def ecis_create(request):
    if request.method == 'POST':
        form = ECISForm(request.POST, user=request.user)
        if form.is_valid():
            ecis = form.save()
            return JsonResponse({
                'status': 'success',
                'message': 'ECIS created successfully',
                'ecis_number': ecis.number
            })
        else:
            return JsonResponse({
                'status': 'error',
                'errors': form.errors
            }, status=400)

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@login_required(login_url="user-login")
def ecis_detail_requestor(request, pk):
    ecis = get_object_or_404(ECIS, pk=pk, created_by=request.user)

    data = {
        'id': ecis.id,
        'number': ecis.number,
        'category': ecis.category,
        'category_display': ecis.get_category_display(),
        'department': ecis.department,
        'requested_by': ecis.requested_by,
        'customer': ecis.customer,
        'line_supervisor': ecis.line_supervisor,
        'affected_parts': ecis.affected_parts,
        'details_change': ecis.details_change,
        'implementation_date': ecis.implementation_date.strftime('%Y-%m-%d'),
        'status': ecis.status,
        'facilitator_remarks': ecis.facilitator_remarks,
        'date_prepared': ecis.date_prepared.strftime('%b %d, %Y'),
        'last_updated': ecis.last_updated.strftime('%b %d, %Y'),
        'can_edit': ecis.status in ['Approved', 'On Hold', 'Needs Revision'],
        'can_cancel': ecis.status in ['Approved', 'On Hold'],
    }
    return JsonResponse(data)

@login_required(login_url="user-login")
def ecis_edit(request, pk):
    ecis = get_object_or_404(ECIS, pk=pk, created_by=request.user)

    if ecis.status not in ['Approved', 'On Hold', 'Needs Revision']:
        return JsonResponse({
            'status': 'error',
            'message': 'This ECIS cannot be edited in its current status.'
        }, status=400)

    if request.method == 'POST':
        form = ECISForm(request.POST, instance=ecis, user=request.user)
        if form.is_valid():
            if ecis.status == 'On Hold' or ecis.status == 'Needs Revision':
                ecis.status = 'For Review'

            ecis = form.save()

            return JsonResponse({
                'status': 'success',
                'message': 'ECIS updated successfully',
                'ecis_number': ecis.number
            })
        else:
            return JsonResponse({
                'status': 'error',
                'errors': form.errors
            }, status=400)

    data = {
        'id': ecis.id,
        'number': ecis.number,
        'category': ecis.category,
        'department': ecis.department,
        'requested_by': ecis.requested_by,
        'customer': ecis.customer,
        'line_supervisor': ecis.line_supervisor,
        'affected_parts': ecis.affected_parts,
        'details_change': ecis.details_change,
        'implementation_date': ecis.implementation_date.strftime('%Y-%m-%d'),
    }
    return JsonResponse(data)

@login_required(login_url="user-login")
def ecis_cancel(request, pk):
    ecis = get_object_or_404(ECIS, pk=pk, created_by=request.user)

    if ecis.status not in ['Approved', 'On Hold']:
        return JsonResponse({
            'status': 'error',
            'message': 'This ECIS cannot be canceled in its current status.'
        }, status=400)

    if request.method == 'POST':
        form = CancelRequestForm(request.POST)
        if form.is_valid():
            ecis.status = 'Canceled'
            ecis.facilitator_remarks = form.cleaned_data.get('remarks', '')
            ecis.save()

            return JsonResponse({
                'status': 'success',
                'message': 'ECIS canceled successfully',
            })
        else:
            return JsonResponse({
                'status': 'error',
                'errors': form.errors
            }, status=400)

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


# Facilitator Views
@login_required(login_url="user-login")
def ecis_list_facilitator(request):
    current_month = timezone.now().month
    current_year = timezone.now().year

    all_ecis = ECIS.objects.all().order_by('-last_updated')

    pending_review = all_ecis.filter(status__in=['For Review', 'On Hold']).order_by('-last_updated')

    total_count = all_ecis.filter(date_prepared__month=current_month, date_prepared__year=current_year).count()
    approved_count = all_ecis.filter(status='Approved', date_prepared__month=current_month, date_prepared__year=current_year).count()
    onhold_count = all_ecis.filter(status='On Hold', date_prepared__month=current_month, date_prepared__year=current_year).count()
    review_count = all_ecis.filter(status='For Review', date_prepared__month=current_month, date_prepared__year=current_year).count()
    revision_count = all_ecis.filter(status='Needs Revision', date_prepared__month=current_month, date_prepared__year=current_year).count()

    paginator = Paginator(all_ecis, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'ecis_list': page_obj,
        'pending_review': pending_review,
        'total_count': total_count,
        'approved_count': approved_count,
        'onhold_count': onhold_count,
        'review_count': review_count,
        'revision_count': revision_count,
        'page_obj': page_obj,
        'paginator': paginator,
        'is_paginated': paginator.num_pages > 1,
    }

    return render(request, 'ecis/facilitator.html', context)

@login_required(login_url="user-login")
def ecis_detail_facilitator(request, pk):
    ecis = get_object_or_404(ECIS, pk=pk)

    data = {
        'id': ecis.id,
        'number': ecis.number,
        'category': ecis.category,
        'category_display': ecis.get_category_display(),
        'department': ecis.department,
        'requested_by': ecis.requested_by,
        'customer': ecis.customer,
        'line_supervisor': ecis.line_supervisor,
        'affected_parts': ecis.affected_parts,
        'details_change': ecis.details_change,
        'implementation_date': ecis.implementation_date.strftime('%Y-%m-%d'),
        'status': ecis.status,
        'facilitator_remarks': ecis.facilitator_remarks,
        'date_prepared': ecis.date_prepared.strftime('%b %d, %Y'),
        'last_updated': ecis.last_updated.strftime('%b %d, %Y'),
        'can_review': ecis.status in ['For Review', 'Approved', 'On Hold'],
        'created_by': ecis.created_by.get_full_name() or ecis.created_by.username,
    }
    return JsonResponse(data)

@login_required(login_url="user-login")
def ecis_review(request, pk):
    ecis = get_object_or_404(ECIS, pk=pk)

    if ecis.status != 'For Review' and ecis.status != 'Approved' and ecis.status != 'On Hold':
        return JsonResponse({
            'status': 'error',
            'message': f'This ECIS cannot be reviewed. Current status: {ecis.status}'
        }, status=400)

    if request.method == 'POST':
        form = FacilitatorReviewForm(request.POST)
        if form.is_valid():
            decision = form.cleaned_data['decision']
            remarks = form.cleaned_data['remarks']

            if decision == 'approve':
                ecis.status = 'Approved'
                if remarks:
                    ecis.facilitator_remarks = remarks
            elif decision == 'hold':
                ecis.status = 'On Hold'
                ecis.facilitator_remarks = remarks
            elif decision == 'revise':
                ecis.status = 'Needs Revision'
                ecis.facilitator_remarks = remarks

            ecis.save()

            return JsonResponse({
                'status': 'success',
                'message': 'Review submitted successfully',
            })
        else:
            return JsonResponse({
                'status': 'error',
                'errors': form.errors
            }, status=400)

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@login_required(login_url="user-login")
def ecis_chart_data(request):
    period = request.GET.get('period', '6month')

    now = timezone.now().date()

    if period == '3month':
        start_date = now - relativedelta(months=3)
    elif period == '1year':
        start_date = now - relativedelta(years=1)
    else:
        start_date = now - relativedelta(months=6)

    labels = []
    approved_values = []
    onhold_values = []
    review_values = []
    revision_values = []
    canceled_values = []

    current_date = start_date.replace(day=1)

    while current_date <= now:

        month_label = current_date.strftime('%b %Y')
        labels.append(month_label)

        base_query = ECIS.objects.filter(
            date_prepared__year=current_date.year,
            date_prepared__month=current_date.month
        )

        approved_count = base_query.filter(status='Approved').count()
        onhold_count = base_query.filter(status='On Hold').count()
        review_count = base_query.filter(status='For Review').count()
        revision_count = base_query.filter(status='Needs Revision').count()
        canceled_count = base_query.filter(status='Canceled').count()

        approved_values.append(approved_count)
        onhold_values.append(onhold_count)
        review_values.append(review_count)
        revision_values.append(revision_count)
        canceled_values.append(canceled_count)

        current_date = current_date + relativedelta(months=1)

    return JsonResponse({
        'labels': labels,
        'datasets': [
            {
                'label': 'Approved',
                'data': approved_values,
                'borderColor': '#4caf50',
                'backgroundColor': 'rgba(76, 175, 80, 0.5)',
                'fill': True
            },
            {
                'label': 'On Hold',
                'data': onhold_values,
                'borderColor': '#ff9800',
                'backgroundColor': 'rgba(255, 152, 0, 0.5)',
                'fill': True
            },
            {
                'label': 'For Review',
                'data': review_values,
                'borderColor': '#2196f3',
                'backgroundColor': 'rgba(33, 150, 243, 0.5)',
                'fill': True
            },
            {
                'label': 'Needs Revision',
                'data': revision_values,
                'borderColor': '#ff9800',
                'backgroundColor': 'rgba(255, 152, 0, 0.5)',
                'fill': True
            },
            {
                'label': 'Canceled',
                'data': canceled_values,
                'borderColor': '#f44336',
                'backgroundColor': 'rgba(244, 67, 54, 0.5)',
                'fill': True
            }
        ]
    })

@login_required(login_url="user-login")
def ecis_requestor_chart_data(request):
    period = request.GET.get('period', '6month')

    now = timezone.now().date()

    if period == '3month':
        start_date = now - relativedelta(months=3)
    elif period == '1year':
        start_date = now - relativedelta(years=1)
    else:
        start_date = now - relativedelta(months=6)

    labels = []
    approved_values = []
    onhold_values = []
    review_values = []
    revision_values = []
    canceled_values = []

    current_date = start_date.replace(day=1)

    while current_date <= now:

        month_label = current_date.strftime('%b %Y')
        labels.append(month_label)

        base_query = ECIS.objects.filter(
            date_prepared__year=current_date.year,
            date_prepared__month=current_date.month,
            created_by=request.user
        )

        approved_count = base_query.filter(status='Approved').count()
        onhold_count = base_query.filter(status='On Hold').count()
        review_count = base_query.filter(status='For Review').count()
        revision_count = base_query.filter(status='Needs Revision').count()
        canceled_count = base_query.filter(status='Canceled').count()

        approved_values.append(approved_count)
        onhold_values.append(onhold_count)
        review_values.append(review_count)
        revision_values.append(revision_count)
        canceled_values.append(canceled_count)

        current_date = current_date + relativedelta(months=1)

    return JsonResponse({
        'labels': labels,
        'datasets': [
            {
                'label': 'Approved',
                'data': approved_values,
                'borderColor': '#4caf50',
                'backgroundColor': 'rgba(76, 175, 80, 0.5)',
                'fill': True
            },
            {
                'label': 'On Hold',
                'data': onhold_values,
                'borderColor': '#ff9800',
                'backgroundColor': 'rgba(255, 152, 0, 0.5)',
                'fill': True
            },
            {
                'label': 'For Review',
                'data': review_values,
                'borderColor': '#2196f3',
                'backgroundColor': 'rgba(33, 150, 243, 0.5)',
                'fill': True
            },
            {
                'label': 'Needs Revision',
                'data': revision_values,
                'borderColor': '#ff9800',
                'backgroundColor': 'rgba(255, 152, 0, 0.5)',
                'fill': True
            },
            {
                'label': 'Canceled',
                'data': canceled_values,
                'borderColor': '#f44336',
                'backgroundColor': 'rgba(244, 67, 54, 0.5)',
                'fill': True
            }
        ]
    })

@login_required(login_url="user-login")
def ecis_export(request):
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    category = request.GET.get('category', 'all')
    
    # Parse dates
    try:
        date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
        date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return HttpResponse('Invalid date format', status=400)
    
    # Filter ECIS records
    ecis_queryset = ECIS.objects.filter(
        date_prepared__gte=date_from_obj,
        date_prepared__lte=date_to_obj
    )
    
    if category != 'all':
        ecis_queryset = ecis_queryset.filter(category=category)
    
    ecis_queryset = ecis_queryset.order_by('-date_prepared')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ECIS Summary Report'
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 30
    
    # Title
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = 'Ryonan Electric Philippines Corporation'
    title_cell.font = Font(name='Arial', size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Subtitle
    ws.merge_cells('A2:K2')
    subtitle_cell = ws['A2']
    date_from_formatted = date_from_obj.strftime('%B %d, %Y')
    date_to_formatted = date_to_obj.strftime('%B %d, %Y')
    subtitle_cell.value = f'ECIS Summary Report for {date_from_formatted} to {date_to_formatted}'
    subtitle_cell.font = Font(name='Arial', size=12)
    subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add space
    ws.append([])
    
    # Headers
    headers = [
        'ECIS Number',
        'Category',
        'Date Prepared',
        'Department',
        'Requested By',
        'Customer',
        'Line Supervisor',
        'Affected Parts',
        'Details of Change',
        'Implementation Date',
        'Status'
    ]
    
    ws.append(headers)
    
    # Style headers
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    bold_font = Font(name='Arial', size=10, bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col_num)
        cell.fill = yellow_fill
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Category display mapping
    category_display = {
        'OR': 'Orange (OR)',
        'YE': 'Yellow (YE)',
        'PN': 'Pink (PN)',
        'LG': 'Light Green (LG)',
        'GR': 'Green (GR)',
        'L': 'Blue (L)',
        'GY': 'Gray (GY)',
    }
    
    # Add data rows
    for ecis in ecis_queryset:
        row_data = [
            ecis.number,
            category_display.get(ecis.category, ecis.category),
            ecis.date_prepared.strftime('%B %d, %Y'),
            ecis.department,
            ecis.requested_by,
            ecis.customer or '',
            ecis.line_supervisor or '',
            ecis.affected_parts,
            ecis.details_change,
            ecis.implementation_date.strftime('%B %d, %Y'),
            ecis.status
        ]
        ws.append(row_data)
    
    # Apply borders and alignment to all data cells
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'ECIS_Summary_Report_{date_from_obj.strftime("%Y%m%d")}_to_{date_to_obj.strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

