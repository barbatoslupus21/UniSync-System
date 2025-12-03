import calendar
import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from datetime import datetime, timedelta
from .models import ManhoursLogsheet, Machine, Operators
from django.core.paginator import Paginator
from decimal import Decimal
from django.db.models import Sum, Q
from django.db.models.functions import TruncDay, TruncWeek, TruncHour
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from django.views.decorators.csrf import csrf_exempt
from portalusers.models import UserApprovers

@login_required(login_url="user-login")
def manhours(request):
    today = datetime.now()
    month_start = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    next_month = month_start.replace(month=month_start.month+1) if month_start.month < 12 else month_start.replace(year=month_start.year+1, month=1)

    if request.user.manhours_supervisor:
        user_requestors = UserApprovers.objects.filter(
            approver=request.user,
            module="Manhours", 
            approver_role="Approver"
        ).values_list('user', flat=True)

        # Include both the supervisor's own entries and entries from users they approve
        monthly_logs = ManhoursLogsheet.objects.filter(
            user__in=user_requestors,
            date_completed__gte=month_start,
            date_completed__lte=next_month
        ).order_by('-date_completed')

        all_records = ManhoursLogsheet.objects.filter(
            user__in=user_requestors,
        ).order_by('-date_completed')
    else:
        monthly_logs = ManhoursLogsheet.objects.filter(
            user=request.user,
            date_completed__gte=month_start,
            date_completed__lt=next_month
        )

        all_records = ManhoursLogsheet.objects.filter(
            user=request.user
        ).order_by('-date_completed')

    # Calculate statistics
    total_hours = monthly_logs.aggregate(total=Sum('manhours'))['total'] or 0
    total_output = monthly_logs.aggregate(total=Sum('output'))['total'] or 0
    total_setup = monthly_logs.aggregate(total=Sum('setup'))['total'] or 0

    active_operators = monthly_logs.values('operator').distinct().count()

    # Paginate the records
    paginator = Paginator(all_records, 10)
    page_number = request.GET.get('page')
    logsheet_entries = paginator.get_page(page_number)

    # Get all machines and operators
    machines = Machine.objects.all()
    operators = Operators.objects.all()

    # Calculate percentages
    total_all = total_hours + total_setup
    if total_all > 0:
        manhours_percentage = round((total_hours / total_all) * 100, 2)
        setup_percentage = round((total_setup / total_all) * 100, 2)
    else:
        manhours_percentage = 0
        setup_percentage = 0

    context = {
        'total_hours': round(total_hours, 2),
        'total_output': total_output,
        'total_setup': total_setup,
        'active_operators': active_operators,
        'manhours_percentage': manhours_percentage,
        'setup_percentage': setup_percentage,
        'logsheet_entries': logsheet_entries,
        'machines': machines,
        'operators': operators,
    }
    return render(request, 'manhours/manhours.html', context)

@login_required(login_url="user-login")
def create_manhours(request):
    if request.method == 'POST':
        try:
            operator = request.POST.get('operator')
            shift = request.POST.get('shift')
            line = request.POST.get('line')
            machine_id = request.POST.get('machine')
            setup = request.POST.get('setup')
            manhours = request.POST.get('manhours')
            output = request.POST.get('output')
            date_completed = request.POST.get('date_completed')

            if not all([operator, shift, line, machine_id, setup, manhours, output, date_completed]):
                messages.error(request, 'All fields are required')
                return redirect('manhours')

            machine = Machine.objects.get(id=machine_id)
            setup = int(setup)
            manhours = Decimal(manhours)
            output = Decimal(output)
            date_completed = datetime.fromisoformat(date_completed)

            entry = ManhoursLogsheet(
                user=request.user,
                operator=operator,
                shift=shift,
                line=line,
                machine=machine,
                setup=setup,
                manhours=manhours,
                output=output,
                date_completed=date_completed
            )
            entry.save()

            messages.success(request, 'Manhours entry added successfully')
            return redirect('manhours')

        except Exception as e:
            messages.error(request, f'Error creating entry: {str(e)}')
            return redirect('manhours')
    else:
        return redirect('manhours')

@login_required(login_url="user-login")
def update_manhours(request):
    if request.method == 'POST':
        entry_id = request.POST.get('entry_id')

        if not entry_id:
            messages.error(request, "Missing entry ID")
            return redirect('manhours')

        selected_manhours = get_object_or_404(ManhoursLogsheet, id=entry_id, user=request.user)

        try:
            operator = request.POST.get('operator')
            shift = request.POST.get('shift')
            line = request.POST.get('line')
            machine_id = request.POST.get('machine')

            try:
                setup = int(request.POST.get('setup', '0'))
            except ValueError:
                setup = 0

            try:
                manhours = Decimal(request.POST.get('manhours', '0'))
            except:
                manhours = Decimal('0')

            try:
                output = Decimal(request.POST.get('output', '0'))
            except:
                output = Decimal('0')

            date_completed = request.POST.get('date_completed')

            selected_manhours.operator = operator
            selected_manhours.shift = shift
            selected_manhours.line = line

            machine = Machine.objects.get(id=machine_id)
            selected_manhours.machine = machine

            selected_manhours.setup = setup
            selected_manhours.manhours = manhours
            selected_manhours.output = output

            # Parse date properly
            try:
                selected_manhours.date_completed = datetime.fromisoformat(date_completed)
            except ValueError:
                selected_manhours.date_completed = datetime.strptime(date_completed, '%Y-%m-%d')

            selected_manhours.save()

            messages.success(request, "Manhours entry updated successfully")

        except Machine.DoesNotExist:
            messages.error(request, "Selected machine does not exist")
        except ValueError as e:
            messages.error(request, f"Invalid input: {str(e)}")
        except Exception as e:
            messages.error(request, f"Error updating entry: {str(e)}")

    return redirect('manhours')

@login_required(login_url="user-login")
def search_manhours(request):
    """AJAX endpoint for searching manhours entries with pagination"""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    
    query = request.POST.get('search', '').strip()
    page = int(request.POST.get('page', 1))
    per_page = 10  # Number of results per page
    
    # Check if user is a manhours supervisor
    is_supervisor = request.user.manhours_supervisor
    
    if is_supervisor:
        # Get all users for which the current user is an approver in the Manhours module
        from portalusers.models import UserApprovers
        approved_users = UserApprovers.objects.filter(
            approver=request.user,
            module="Manhours"
        ).values_list('user', flat=True)
        
        # Include both the supervisor's own entries and entries from users they approve
        base_queryset = ManhoursLogsheet.objects.filter(
            user__in=list(approved_users) + [request.user.id]
        )
    else:
        # Regular user - only show their own entries
        base_queryset = ManhoursLogsheet.objects.filter(user=request.user)
    
    if query:
        # Search across multiple fields
        search_queryset = base_queryset.filter(
            Q(operator__icontains=query) |
            Q(line__icontains=query) |
            Q(machine__machine_name__icontains=query) |
            Q(shift__icontains=query) |
            Q(date_completed__icontains=query)
        ).order_by('-date_completed')
    else:
        # Return all entries when no search query
        search_queryset = base_queryset.order_by('-date_completed')
    
    # Apply pagination
    from django.core.paginator import Paginator
    paginator = Paginator(search_queryset, per_page)
    
    try:
        paginated_results = paginator.page(page)
    except:
        paginated_results = paginator.page(1)
    
    # Format results for JSON response matching the table structure
    results = []
    for entry in paginated_results:
        results.append({
            'id': entry.id,
            'date_completed': entry.date_completed.isoformat(),
            'operator': entry.operator,
            'operator_name': entry.operator,  # Alias for consistency
            'shift': entry.shift,
            'line': entry.line,
            'line_name': entry.line,  # Alias for consistency
            'machine_name': entry.machine.machine_name if entry.machine else 'N/A',
            'total_output': float(entry.total_output),
            'setup': entry.setup,
            'manhours': float(entry.manhours),
            'output': float(entry.output),
            'user_id': entry.user.id
        })
    
    return JsonResponse({
        'status': 'success',
        'results': results,
        'query': query,
        'pagination': {
            'current_page': paginated_results.number,
            'total_pages': paginator.num_pages,
            'total_results': paginator.count,
            'per_page': per_page,
            'has_previous': paginated_results.has_previous(),
            'has_next': paginated_results.has_next(),
            'previous_page': paginated_results.previous_page_number() if paginated_results.has_previous() else None,
            'next_page': paginated_results.next_page_number() if paginated_results.has_next() else None,
            'start_index': paginated_results.start_index(),
            'end_index': paginated_results.end_index(),
        }
    })
    try:
        entry = ManhoursLogsheet.objects.get(id=id, user=request.user)

        entry_data = {
            'id': id,
            'operator': entry.operator,
            'shift': entry.shift,
            'line': entry.line,
            'machine_id': entry.machine.id,
            'machine_name': entry.machine.machine_name,
            'setup': entry.setup,
            'manhours': float(entry.manhours),
            'output': float(entry.output),
            'total_output': float(entry.total_output),
            'date_completed': entry.date_completed.isoformat(),
            'date_submitted': entry.date_submitted.isoformat(),
            'user_name': entry.user.name
        }

        return JsonResponse({'status': 'success', 'entry': entry_data})

    except ManhoursLogsheet.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Entry not found'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@login_required(login_url="user-login")
def get_chart_data(request):
    from django.utils import timezone

    chart_type = request.GET.get('type', 'shiftOutput')
    period = request.GET.get('period', 'month')

    today = timezone.now()

    if period == 'today':
        start_date = today.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = today.replace(hour=23, minute=59, second=59, microsecond=59)

        date_trunc = TruncHour('date_completed')
        date_format = '%I %p'

    elif period == 'week':
        start_of_week = today - timedelta(days=today.weekday())
        start_date = start_of_week.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = start_date + timedelta(days=6, hours=23, minutes=59, seconds=59)

        date_trunc = TruncDay('date_completed')
        date_format = '%a'

    elif period == 'month':
        start_date = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        last_day = calendar.monthrange(today.year, today.month)[1]
        end_date = today.replace(day=last_day, hour=23, minute=59, second=59)

        date_trunc = TruncDay('date_completed')
        date_format = '%d'

    else:
        quarter_month = 3 * ((today.month - 1) // 3) + 1
        start_date = today.replace(month=quarter_month, day=1, hour=0, minute=0, second=0, microsecond=0)
        if quarter_month == 10:
            end_date = today.replace(month=12, day=31, hour=23, minute=59, second=59)
        else:
            if quarter_month + 2 > 12:
                year = today.year + 1
                month = (quarter_month + 2) % 12
            else:
                year = today.year
                month = quarter_month + 2
            last_day = calendar.monthrange(year, month)[1]
            end_date = today.replace(year=year, month=month, day=last_day,
                                   hour=23, minute=59, second=59)

        date_trunc = TruncWeek('date_completed')
        date_format = 'Week %W'

    # Check if user is a manhours supervisor
    is_supervisor = request.user.manhours_supervisor

    if is_supervisor:
        # Get all users for which the current user is an approver in the Manhours module
        from portalusers.models import UserApprovers
        approved_users = UserApprovers.objects.filter(
            approver=request.user,
            module="Manhours"
        ).values_list('user', flat=True)

        # Include both the supervisor's own entries and entries from users they approve
        entries = ManhoursLogsheet.objects.filter(
            date_completed__gte=start_date,
            date_completed__lte=end_date
        ).filter(
            user__in=list(approved_users) + [request.user.id]
        )
    else:
        # Regular user - only show their own entries
        entries = ManhoursLogsheet.objects.filter(
            user=request.user,
            date_completed__gte=start_date,
            date_completed__lte=end_date
        )

    if chart_type == 'shiftOutput':
        data = entries.annotate(
            date_group=date_trunc
        ).values('date_group', 'shift').annotate(
            total_output=Sum('output')
        ).order_by('date_group', 'shift')

        am_data = {}
        pm_data = {}
        all_dates = set()

        # Create date range information for the chart title/subtitle
        if period == 'today':
            today_str = start_date.strftime('%B %d, %Y')
            period_range = f"Today - {today_str}"
        elif period == 'week':
            week_start_str = start_date.strftime('%b %d')
            week_end_str = end_date.strftime('%b %d, %Y')
            period_range = f"Week of {week_start_str} - {week_end_str}"
        elif period == 'month':
            month_str = start_date.strftime('%B %Y')
            period_range = month_str
        else:
            quarter_num = ((start_date.month - 1) // 3) + 1
            quarter_str = f"Q{quarter_num} {start_date.year}"
            period_range = quarter_str

        # For month period, populate all_dates with all days of the month
        if period == 'month':
            last_day = calendar.monthrange(today.year, today.month)[1]
            for day in range(1, last_day + 1):
                date_str = f"{day:02d}"  # Format as 01, 02, 03, etc.
                all_dates.add(date_str)

        for item in data:
            # Format the date string based on the period
            if period == 'today':
                date_str = item['date_group'].strftime('%I %p')
            elif period == 'week':
                date_str = item['date_group'].strftime('%a')
            elif period == 'month':
                date_str = item['date_group'].strftime('%d')
            else:
                # For quarterly data, use week numbers with start date
                week_num = int(item['date_group'].strftime('%W'))
                date_str = f"W{week_num}"

            # For periods other than month, add dates to set as we process data
            if period != 'month':
                all_dates.add(date_str)

            if item['shift'] == 'AM':
                am_data[date_str] = float(item['total_output'])
            elif item['shift'] == 'PM':
                pm_data[date_str] = float(item['total_output'])

        # Create a mapping of date strings to full dates for display
        date_to_full_date = {}
        date_objects = {}  # Store actual date objects for sorting

        # For month period, pre-populate date_objects with all days
        if period == 'month':
            last_day = calendar.monthrange(today.year, today.month)[1]
            for day in range(1, last_day + 1):
                date_key = f"{day:02d}"
                date_obj = start_date.replace(day=day)
                date_to_full_date[date_key] = date_key
                date_objects[date_key] = date_obj

        for item in data:
            if period == 'today':
                # For today view, use hour format
                date_obj = item['date_group']
                date_key = date_obj.strftime('%I %p')  # 01 AM, 02 PM, etc.
                date_to_full_date[date_key] = date_key
                date_objects[date_key] = date_obj
            elif period == 'week':
                # For weekly view, use day of week with day number
                date_obj = item['date_group']
                date_key = date_obj.strftime('%a')  # Mon, Tue, etc.
                date_value = date_obj.strftime('%d')  # Day number: 01, 02, etc.
                date_to_full_date[date_key] = date_value
                date_objects[date_key] = date_obj
            elif period == 'month':
                # For monthly view, use day number (already pre-populated above)
                # But update if data exists to ensure correct date_obj
                date_obj = item['date_group']
                date_key = date_obj.strftime('%d')  # Day number: 01, 02, etc.
                date_objects[date_key] = date_obj  # Update with actual data date
            else:
                # For quarterly view, use week number
                date_obj = item['date_group']
                week_num = int(date_obj.strftime('%W'))
                date_key = f"W{week_num}"
                date_value = date_obj.strftime('%d')  # Day number of week start
                date_to_full_date[date_key] = date_value
                date_objects[date_key] = date_obj

        # Sort the dates appropriately
        if period == 'today':
            # Sort by hour (parse hour from format like "01 AM", "02 PM")
            def parse_hour(time_str):
                hour_str, period_str = time_str.split()
                hour = int(hour_str)
                if period_str == 'PM' and hour != 12:
                    hour += 12
                elif period_str == 'AM' and hour == 12:
                    hour = 0
                return hour
            sorted_dates = sorted(all_dates, key=parse_hour)
        elif period == 'week':
            day_order = {'Mon': 0, 'Tue': 1, 'Wed': 2, 'Thu': 3, 'Fri': 4, 'Sat': 5, 'Sun': 6}
            sorted_dates = sorted(all_dates, key=lambda x: day_order.get(x, 7))
        elif period == 'month':
            sorted_dates = sorted(all_dates, key=lambda x: int(x))
        else:
            sorted_dates = sorted(all_dates, key=lambda x: int(x[1:]) if x.startswith('W') else 0)

        # Create display labels with dates for tooltips
        tooltip_labels = []
        for date in sorted_dates:
            if period == 'today':
                # For today, just use the hour label
                tooltip_labels.append(date)
            elif period == 'week':
                date_obj = date_objects.get(date)
                if date_obj:
                    tooltip = f"{date} ({date_obj.strftime('%b %d')})"
                else:
                    tooltip = date
                tooltip_labels.append(tooltip)
            elif period == 'month':
                date_obj = date_objects.get(date)
                if date_obj:
                    tooltip = f"{date} ({date_obj.strftime('%b %d')})"
                else:
                    tooltip = date
                tooltip_labels.append(tooltip)
            else:
                date_obj = date_objects.get(date)
                if date_obj:
                    week_start = date_obj
                    week_end = week_start + timedelta(days=6)
                    tooltip = f"{date} ({week_start.strftime('%b %d')}-{week_end.strftime('%d')})"
                else:
                    tooltip = date
                tooltip_labels.append(tooltip)

        am_output_data = [am_data.get(date, 0) for date in sorted_dates]
        pm_output_data = [pm_data.get(date, 0) for date in sorted_dates]

        # Create period range for chart title
        if period == 'week':
            week_start_str = start_date.strftime('%b %d')
            week_end_str = end_date.strftime('%b %d, %Y')
            period_range = f"Week of {week_start_str} - {week_end_str}"
        elif period == 'month':
            month_str = start_date.strftime('%B %Y')
            period_range = month_str
        else:
            quarter_num = ((start_date.month - 1) // 3) + 1
            quarter_str = f"Q{quarter_num} {start_date.year}"
            period_range = quarter_str

        # For x-axis labels, we want to show the day numbers like in the image
        x_axis_labels = []
        for date in sorted_dates:
            day_number = date_to_full_date.get(date, '')
            x_axis_labels.append(day_number)

        result = {
            'labels': sorted_dates,      # Original labels (Mon, Tue, 01, 02, etc.)
            'tooltips': tooltip_labels,  # Detailed labels for tooltips
            'xAxisLabels': x_axis_labels, # Day numbers for x-axis display
            'title': period_range,       # Add period range as title
            'datasets': [
                {
                    'label': 'AM Shift',
                    'data': am_output_data,
                    'backgroundColor': 'rgba(51, 102, 255, 0.2)',
                    'borderColor': 'rgba(51, 102, 255, 1)',
                    'borderWidth': 2,
                    'tension': 0.3,
                    'fill': True,
                    'pointBackgroundColor': 'rgba(51, 102, 255, 1)',
                    'pointBorderColor': '#fff',
                    'pointRadius': 4
                },
                {
                    'label': 'PM Shift',
                    'data': pm_output_data,
                    'backgroundColor': 'rgba(241, 70, 104, 0.2)',
                    'borderColor': 'rgba(241, 70, 104, 1)',
                    'borderWidth': 2,
                    'tension': 0.3,
                    'fill': True,
                    'pointBackgroundColor': 'rgba(241, 70, 104, 1)',
                    'pointBorderColor': '#fff',
                    'pointRadius': 4
                }
            ]
        }

    elif chart_type == 'hoursByLine':
        data = entries.values('line').annotate(
            total_hours=Sum('manhours')
        ).order_by('line')

        result = {
            'labels': [item['line'] for item in data],
            'datasets': [{
                'label': 'Total Hours',
                'data': [float(item['total_hours']) for item in data],
                'backgroundColor': 'rgba(51, 102, 255, 0.5)',
                'borderColor': 'rgba(51, 102, 255, 1)',
                'borderWidth': 1
            }]
        }

    elif chart_type == 'outputByShift':
        data = entries.values('shift').annotate(
            total_output=Sum('output')
        ).order_by('shift')

        colors = {
            'AM': 'rgba(51, 102, 255, 0.7)',
            'PM': 'rgba(241, 70, 104, 0.7)'
        }

        border_colors = {
            'AM': 'rgba(51, 102, 255, 1)',
            'PM': 'rgba(241, 70, 104, 1)'
        }

        result = {
            'labels': [f"{item['shift']} Shift" for item in data],
            'datasets': [{
                'data': [float(item['total_output']) for item in data],
                'backgroundColor': [colors.get(item['shift'], 'rgba(72, 199, 116, 0.7)') for item in data],
                'borderColor': [border_colors.get(item['shift'], 'rgba(72, 199, 116, 1)') for item in data],
                'borderWidth': 1
            }]
        }

    else:
        data = entries.annotate(
            date_group=date_trunc
        ).values('date_group').annotate(
            total_hours=Sum('manhours'),
            total_output=Sum('output')
        ).order_by('date_group')

        result = {
            'labels': [item['date_group'].strftime(date_format) for item in data],
            'datasets': [
                {
                    'label': 'Manhours',
                    'data': [float(item['total_hours']) for item in data],
                    'backgroundColor': 'rgba(51, 102, 255, 0.5)',
                    'borderColor': 'rgba(51, 102, 255, 1)',
                    'borderWidth': 1
                },
                {
                    'label': 'Output',
                    'data': [float(item['total_output']) for item in data],
                    'backgroundColor': 'rgba(241, 70, 104, 0.5)',
                    'borderColor': 'rgba(241, 70, 104, 1)',
                    'borderWidth': 1
                }
            ]
        }

    return JsonResponse(result)

@login_required(login_url="user-login")
@csrf_exempt
def get_operators(request):
    operators = Operators.objects.all()
    search_term = request.GET.get('search', '').strip()
    
    if search_term:
        operators = operators.filter(
            Q(name__icontains=search_term) | Q(id_number__icontains=search_term)
        )
        
    return JsonResponse({
        'status': 'success',
        'search_query': search_term,
        'operators': [{
            'id': operator.id,
            'name': operator.name,
            'id_number': operator.id_number
        } for operator in operators]
    }, safe=False)

@login_required(login_url="user-login")
@csrf_exempt
def get_operator(request, operator_id):
    operator = get_object_or_404(Operators, id=operator_id)
    return JsonResponse({
        'status': 'success',
        'operator': {
            'id': operator.id,
            'name': operator.name,
            'id_number': operator.id_number
        }
    })

@login_required(login_url="user-login")
@csrf_exempt
def create_operator(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        id_number = request.POST.get('id_number', None)
        
        if name:
            operator = Operators(name=name, id_number=id_number)
            operator.save()
            return JsonResponse({
                'status': 'success',
                'message': 'Operator created successfully',
                'operator': {
                    'id': operator.id,
                    'name': operator.name,
                    'id_number': operator.id_number
                }
            })
        return JsonResponse({'status': 'error', 'message': 'Operator name is required'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

@login_required(login_url="user-login")
@csrf_exempt
def update_operator(request, operator_id):
    if request.method == 'POST':
        operator = get_object_or_404(Operators, id=operator_id)
        name = request.POST.get('name')
        id_number = request.POST.get('id_number', operator.id_number)
        
        if name:
            operator.name = name
            operator.id_number = id_number
            operator.save()
            return JsonResponse({
                'status': 'success',
                'message': 'Operator updated successfully',
                'operator': {
                    'id': operator.id,
                    'name': operator.name,
                    'id_number': operator.id_number
                }
            })
        return JsonResponse({'status': 'error', 'message': 'Operator name is required'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

@login_required(login_url="user-login")
@csrf_exempt
def delete_operator(request, operator_id):
    if request.method == 'POST' or request.method == 'DELETE':
        operator = get_object_or_404(Operators, id=operator_id)
        operator_name = operator.name
        operator.delete()
        return JsonResponse({
            'status': 'success',
            'message': f'Operator "{operator_name}" deleted successfully'
        })
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

@login_required(login_url="user-login")
def get_machine_performance(request):
    period = request.GET.get('period', 'month')
    today = datetime.now()

    if period == 'week':
        start_of_week = today - timedelta(days=today.weekday())
        start_date = start_of_week.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = start_date + timedelta(days=6, hours=23, minutes=59, seconds=59)
    else:
        start_date = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        last_day = calendar.monthrange(today.year, today.month)[1]
        end_date = today.replace(day=last_day, hour=23, minute=59, second=59, microsecond=999999)

    # Check if user is a manhours supervisor
    is_supervisor = request.user.manhours_supervisor

    if is_supervisor:
        # Get all users for which the current user is an approver in the Manhours module
        from portalusers.models import UserApprovers
        approved_users = UserApprovers.objects.filter(
            approver=request.user,
            module="Manhours"
        ).values_list('user', flat=True)

        # Include both the supervisor's own entries and entries from users they approve
        data = ManhoursLogsheet.objects.filter(
            date_completed__gte=start_date,
            date_completed__lte=end_date
        ).filter(
            user__in=list(approved_users) + [request.user.id]
        ).values(
            'machine__machine_name'
        ).annotate(
            total_output=Sum('output')
        ).order_by('-total_output')
    else:
        # Regular user - only show their own entries
        data = ManhoursLogsheet.objects.filter(
            user=request.user,
            date_completed__gte=start_date,
            date_completed__lte=end_date
        ).values(
            'machine__machine_name'
        ).annotate(
            total_output=Sum('output')
        ).order_by('-total_output')

    machine_names = []
    output_values = []
    bg_colors = []
    border_colors = []

    colors = [
        {'bg': 'rgba(51, 102, 255, 0.7)', 'border': 'rgba(51, 102, 255, 1)'},
        {'bg': 'rgba(72, 199, 116, 0.7)', 'border': 'rgba(72, 199, 116, 1)'},
        {'bg': 'rgba(255, 193, 7, 0.7)', 'border': 'rgba(255, 193, 7, 1)'},
        {'bg': 'rgba(241, 70, 104, 0.7)', 'border': 'rgba(241, 70, 104, 1)'},
        {'bg': 'rgba(153, 102, 255, 0.7)', 'border': 'rgba(153, 102, 255, 1)'},
        {'bg': 'rgba(75, 192, 192, 0.7)', 'border': 'rgba(75, 192, 192, 1)'},
        {'bg': 'rgba(255, 159, 64, 0.7)', 'border': 'rgba(255, 159, 64, 1)'},
        {'bg': 'rgba(201, 203, 207, 0.7)', 'border': 'rgba(201, 203, 207, 1)'}
    ]

    for i, item in enumerate(data):
        machine_names.append(item['machine__machine_name'])
        output_values.append(float(item['total_output']) if item['total_output'] else 0)
        color_index = i % len(colors)
        bg_colors.append(colors[color_index]['bg'])
        border_colors.append(colors[color_index]['border'])

    result = {
        'labels': machine_names,
        'datasets': [{
            'label': 'Total Output',
            'data': output_values,
            'backgroundColor': bg_colors,
            'borderColor': border_colors,
            'borderWidth': 1
        }]
    }

    return JsonResponse(result)

@login_required(login_url="user-login")
def export_reports(request):
    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        selected_reports = request.POST.getlist('export_reports[]')

        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')

        logs = ManhoursLogsheet.objects.filter(date_completed__range=(start_date, end_date)).order_by('date_completed')

        wb = Workbook()
        wb.remove(wb.active)

        bold_font = Font(bold=True)
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        def add_sheet(title, rows, headers):
            sheet = wb.create_sheet(title=title)
            sheet['A1'] = "RYONAN ELECTRIC PHILIPPINES"
            sheet['A1'].font = bold_font
            sheet['A2'] = title

            sheet.append([])
            sheet.append(headers)

            for col, cell in enumerate(sheet[4], start=1):
                cell.font = bold_font
                cell.fill = header_fill
                cell.border = thin_border

            for row_data in rows:
                sheet.append(row_data)
                for col in range(1, len(row_data) + 1):
                    cell = sheet.cell(row=sheet.max_row, column=col)
                    cell.border = thin_border

            for col in sheet.columns:
                max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
                sheet.column_dimensions[col[0].column_letter].width = max_length + 2

        if 'daily_summary' in selected_reports:
            headers = ['Date', 'Shift', 'Total Manhours']
            grouped = {}
            for log in logs:
                key = (log.date_completed.date(), log.shift)
                grouped.setdefault(key, 0)
                grouped[key] += float(log.manhours)
            rows = [[date.strftime('%Y-%m-%d'), shift, total] for (date, shift), total in grouped.items()]
            add_sheet("Daily Manhours Summary", rows, headers)

        if 'operator_performance' in selected_reports:
            headers = ['Operator', 'Date', 'Shift', 'Output', 'Manhours', 'Output per Hour']
            rows = [
                [log.operator, log.date_completed.date(), log.shift, float(log.output),
                 float(log.manhours), float(log.total_output)]
                for log in logs
            ]
            add_sheet("Operator Performance", rows, headers)

        if 'machine_utilization' in selected_reports:
            headers = ['Machine', 'Date', 'Shift', 'Manhours']
            rows = [
                [log.machine.machine_name if log.machine else 'N/A', log.date_completed.date(), log.shift, float(log.manhours)]
                for log in logs
            ]
            add_sheet("Machine Utilization", rows, headers)

        if 'daily_output' in selected_reports:
            headers = ['Date', 'Shift', 'Total Output']
            grouped = {}
            for log in logs:
                key = (log.date_completed.date(), log.shift)
                grouped.setdefault(key, 0)
                grouped[key] += float(log.output)
            rows = [[date.strftime('%Y-%m-%d'), shift, total] for (date, shift), total in grouped.items()]
            add_sheet("Daily Output Summary", rows, headers)

        if 'operator_machine_pairing' in selected_reports:
            headers = ['Date', 'Shift', 'Operator', 'Machine']
            rows = [
                [log.date_completed.date(), log.shift, log.operator, log.machine.machine_name if log.machine else 'N/A']
                for log in logs
            ]
            add_sheet("Operator-Machine Pairing", rows, headers)

        if 'shift_comparison' in selected_reports:
            headers = ['Date', 'Shift', 'Total Output', 'Total Manhours']
            grouped = {}
            for log in logs:
                key = (log.date_completed.date(), log.shift)
                if key not in grouped:
                    grouped[key] = {'output': 0, 'manhours': 0}
                grouped[key]['output'] += float(log.output)
                grouped[key]['manhours'] += float(log.manhours)
            rows = [[date.strftime('%Y-%m-%d'), shift, data['output'], data['manhours']] for (date, shift), data in grouped.items()]
            add_sheet("Shift Comparison", rows, headers)

        if 'setup_time_analysis' in selected_reports:
            headers = ['Date', 'Shift', 'Operator', 'Machine', 'Setup Time']
            rows = [
                [log.date_completed.date(), log.shift, log.operator, log.machine.machine_name if log.machine else 'N/A', log.setup]
                for log in logs if log.setup > 0
            ]
            add_sheet("Setup Time Analysis", rows, headers)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Manhours_Export_{start_date.date()}_to_{end_date.date()}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response

    return redirect(request, 'manhours')

@login_required(login_url="user-login")
def get_manhours_details(request, id):
    """Get details of a specific manhours entry for AJAX requests"""
    try:
        # Check if user is a manhours supervisor
        is_supervisor = request.user.manhours_supervisor

        if is_supervisor:
            # Get all users for which the current user is an approver in the Manhours module
            from portalusers.models import UserApprovers
            approved_users = UserApprovers.objects.filter(
                approver=request.user,
                module="Manhours"
            ).values_list('user', flat=True)

            # Include both the supervisor's own entries and entries from users they approve
            entry = get_object_or_404(
                ManhoursLogsheet,
                id=id,
                user__in=list(approved_users) + [request.user.id]
            )
        else:
            # Regular user - only show their own entries
            entry = get_object_or_404(ManhoursLogsheet, id=id, user=request.user)

        # Prepare entry data for JSON response
        entry_data = {
            'id': entry.id,
            'date_completed': entry.date_completed.isoformat(),
            'operator': entry.operator,
            'shift': entry.shift,
            'line': entry.line,
            'machine_id': entry.machine.id if entry.machine else None,
            'machine_name': entry.machine.machine_name if entry.machine else 'N/A',
            'setup': float(entry.setup),
            'manhours': float(entry.manhours),
            'output': entry.output,
            'total_output': entry.total_output,
            'date_submitted': entry.date_submitted.isoformat() if entry.date_submitted else None,
            'user_id': entry.user.id
        }

        return JsonResponse({
            'status': 'success',
            'entry': entry_data
        })

    except ManhoursLogsheet.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Entry not found or access denied'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)