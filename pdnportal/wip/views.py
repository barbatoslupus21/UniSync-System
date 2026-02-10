import json
import openpyxl
import pandas as pd
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse, Http404
from django.db.models import Count, Q
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.utils import timezone
from datetime import datetime
from django.core.paginator import Paginator
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from .models import InventorySession, InventoryEntry, ProductLists, Masterlist, Process, WipMasterlist
from .forms import InventorySessionForm, ProductSelectionForm, WIPProductForm
from settings.models import Line
from django.urls import reverse
from decimal import Decimal
from django.db import transaction

@login_required(login_url="user-login")
def dashboard(request):
    user = request.user
    
    if user.wip_counter:
        return counter_dashboard(request)
    elif user.wip_checker:
        return checker_dashboard(request)
    else:
        messages.error(request, "You don't have permission to access WIP module.")
        return redirect('/')

@login_required(login_url="user-login")
def counter_dashboard(request):
    if not request.user.wip_counter:
        messages.error(request, "Access denied. Counter permission required.")
        return redirect('/')
    
    user_sessions = InventorySession.objects.filter(created_by=request.user)
    
    # Get distinct dates from sessions
    from django.db.models.functions import TruncDate
    distinct_dates = user_sessions.annotate(
        date_only=TruncDate('created_at')
    ).values('date_only').distinct().order_by('-date_only')
    
    # Format dates for display
    formatted_dates = []
    for date_obj in distinct_dates:
        if date_obj['date_only']:
            formatted_date = date_obj['date_only'].strftime('%b %d, %Y')
            formatted_dates.append({
                'value': date_obj['date_only'].isoformat(),
                'display': formatted_date
            })
    
    # Filter by selected date if provided
    selected_date = request.GET.get('date', '')
    if selected_date:
        from datetime import datetime
        try:
            filter_date = datetime.fromisoformat(selected_date).date()
            user_sessions = user_sessions.filter(created_at__date=filter_date)
        except:
            pass
    
    # Annotate with entry count for each session
    user_sessions = user_sessions.annotate(entries_count=Count('entries')).order_by('-created_at')
    
    # Pagination
    paginator = Paginator(user_sessions, 10)  # 10 records per page
    page_number = request.GET.get('page', 1)
    sessions = paginator.get_page(page_number)
    
    total_sessions = InventorySession.objects.filter(created_by=request.user).count()
    # Get user's assigned lines
    user_lines = request.user.line.all()
    
    # Also include any lines from the user's existing sessions (for backward compatibility in edit modal)
    session_lines = InventorySession.objects.filter(created_by=request.user).values_list('line_id', flat=True).distinct()
    all_available_lines = Line.objects.filter(
        Q(id__in=user_lines.values_list('id', flat=True)) |
        Q(id__in=session_lines)
    ).distinct()
    
    total_items = InventoryEntry.objects.filter(session__created_by=request.user).count()
    checked_sessions = InventorySession.objects.filter(created_by=request.user, status='CHECKED').count()
    pending_sessions = InventorySession.objects.filter(created_by=request.user, status='FOR_CHECKING').count()
    
    context = {
        'sessions': sessions,
        'total_sessions': total_sessions,
        'total_items': total_items,
        'checked_sessions': checked_sessions,
        'pending_sessions': pending_sessions,
        'user_type': 'counter',
        'lines': user_lines,  # For create modal - only user's assigned lines
        'all_lines': all_available_lines,  # For edit modal - includes session history
        'formatted_dates': formatted_dates,
        'selected_date': selected_date,
        'paginator': paginator,
    }
    
    return render(request, 'wip/counter_dashboard.html', context)

@login_required(login_url="user-login")
def search_sessions(request):
    """AJAX endpoint for dynamic session search"""
    if not request.user.wip_counter and not request.user.wip_checker:
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    search_query = request.GET.get('q', '').strip()
    date_filter = request.GET.get('date', '')
    page_number = request.GET.get('page', 1)
    
    # Start with user's sessions only (created_by = current user)
    sessions = InventorySession.objects.filter(created_by=request.user)
    
    # Apply date filter if provided
    if date_filter:
        try:
            filter_date = datetime.fromisoformat(date_filter).date()
            sessions = sessions.filter(created_at__date=filter_date)
        except Exception as e:
            print(f"Date filter error: {e}")
    
    # Apply search query if provided
    if search_query:
        sessions = sessions.filter(
            Q(id__icontains=search_query) |
            Q(line__line_name__icontains=search_query) |
            Q(person_responsible__icontains=search_query)
        )
    
    # Annotate with entry count
    sessions = sessions.annotate(entries_count=Count('entries')).order_by('-created_at')
    
    # Pagination
    paginator = Paginator(sessions, 10)
    page_obj = paginator.get_page(page_number)
    
    # Render the partial template
    from django.template.loader import render_to_string
    html = render_to_string('wip/partials/session_table_rows.html', {
        'sessions': page_obj,
        'search_query': search_query,
    }, request=request)
    
    return JsonResponse({
        'html': html,
        'total_count': paginator.count,
        'page': int(page_number),
        'total_pages': paginator.num_pages,
        'has_next': page_obj.has_next() if page_obj else False,
        'has_previous': page_obj.has_previous() if page_obj else False,
    })

@login_required(login_url="user-login")
def checker_dashboard(request):
    if not request.user.wip_checker:
        messages.error(request, "Access denied. Checker permission required.")
        return redirect('/')

    sessions = InventorySession.objects.filter(checked_by__isnull=True)
    sessions_checking = InventorySession.objects.filter(checked_by=request.user)

    
    now = timezone.now()
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    session_summary = InventorySession.objects.filter(
        Q(status='FOR_CHECKING') | Q(status='CHECKED'),
        created_at__gte=month_start
    ).order_by('-created_at')
    processes = Process.objects.all()
    
    context = {
        'lines': sessions,
        'sessions': sessions_checking,
        'user_type': 'checker',
        'session_summary': session_summary,
        'processes': processes,
    }
    
    return render(request, 'wip/checker_dashboard.html', context)

@login_required(login_url="user-login")
def create_session(request):
    if not request.user.wip_counter:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'error': 'Access denied. Counter permission required.'}, status=403)
        messages.error(request, "Access denied. Counter permission required.")
        return redirect('wip_dashboard')
    
    if request.method == 'POST':
        form = InventorySessionForm(request.POST)
        if form.is_valid():
            line = form.cleaned_data.get('line')
            person_responsible = form.cleaned_data.get('person_responsible')
            
            # Check for existing session with same line, person_responsible, and created_at date
            today = timezone.now().date()
            existing_session = InventorySession.objects.filter(
                line=line,
                person_responsible=person_responsible,
                created_at__date=today
            ).first()
            
            if existing_session:
                error_msg = f"A session with the same Line, Person Responsible, and Date already exists."
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'error': error_msg}, status=400)
                messages.error(request, error_msg)
                return redirect('wip_dashboard')
            
            session = form.save(commit=False)
            session.created_by = request.user
            session.save()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'session_id': session.id, 'redirect_url': reverse('session_details', args=[session.id])})
            
            messages.success(request, f"Session {session.id} created successfully!")
            return redirect('session_details', session_id=session.id)
        else:
            errors = []
            for field, error_list in form.errors.items():
                for error in error_list:
                    errors.append(f"{field}: {error}")
            error_msg = "; ".join(errors)
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'error': error_msg}, status=400)
            
            for field, field_errors in form.errors.items():
                for error in field_errors:
                    messages.error(request, f"{field}: {error}")
        return redirect('wip_dashboard')
    return redirect('wip_dashboard')

@login_required(login_url="user-login")
def session_details(request, session_id):
    session = get_object_or_404(InventorySession, id=session_id)
    
    if request.user.wip_counter and session.created_by != request.user:
        messages.error(request, "You can only view your own sessions.")
        return redirect('wip_dashboard')
    
    if not (request.user.wip_counter or request.user.wip_checker):
        messages.error(request, "Access denied.")
        return redirect('/')
    
    raw_materials = session.entries.filter(inventory_type='RAW').select_related('product', 'material')
    finished_products = session.entries.filter(inventory_type='FINISHED').select_related('product')
    wip_products = session.entries.filter(inventory_type='WIP').select_related('product', 'process', 'material')
    
    # Attach WipMasterlist UOM to each WIP entry
    for entry in wip_products:
        wip_master = WipMasterlist.objects.filter(
            product_number=entry.product,
            material_name=entry.material
        ).first()
        entry.wip_uom = wip_master.uom if wip_master else ''
    
    # Attach ProductLists UOM to each raw material entry
    for entry in raw_materials:
        entry.product_uom = entry.product.uom if hasattr(entry.product, 'uom') else ''
    
    all_raw_materials_checked = raw_materials.exists() and not raw_materials.filter(status='FOR_CHECKING').exists()
    all_finished_products_checked = finished_products.exists() and not finished_products.filter(status='FOR_CHECKING').exists()
    all_wip_products_checked = wip_products.exists() and not wip_products.filter(status='FOR_CHECKING').exists()
    
    all_entries_checked = not session.entries.filter(status='FOR_CHECKING').exists()

    context = {
        'session': session,
        'raw_materials': raw_materials,
        'finished_products': finished_products,
        'wip_products': wip_products,
        'user_type': 'counter' if request.user.wip_counter else 'checker',
        'can_edit': session.can_edit and request.user.wip_counter,
        'can_check': session.can_check and request.user.wip_checker,
        'all_raw_materials_checked': all_raw_materials_checked,
        'all_finished_products_checked': all_finished_products_checked,
        'all_wip_products_checked': all_wip_products_checked,
        'all_entries_checked': all_entries_checked,
    }
    
    return render(request, 'wip/session_details.html', context)

@login_required(login_url="user-login")
def edit_session(request, session_id):
    if not request.user.wip_counter:
        messages.error(request, "Access denied. Counter permission required.")
        return redirect('wip_dashboard')
    
    session = get_object_or_404(InventorySession, id=session_id, created_by=request.user)
    
    if not session.can_edit:
        messages.error(request, "Cannot edit a checked session.")
        return redirect('session_details', session_id=session.id)
    
    if request.method == 'POST':
        form = InventorySessionForm(request.POST, instance=session)
        if form.is_valid():
            form.save()
            messages.success(request, "Session updated successfully!")
            return redirect('session_details', session_id=session.id)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    
    return redirect('session_details', session_id=session.id)

@login_required(login_url="user-login")
def delete_session(request, session_id):
    if not request.user.wip_counter:
        return JsonResponse({'success': False, 'message': 'Access denied.'})
    
    session = get_object_or_404(InventorySession, id=session_id, created_by=request.user)
    
    if not session.can_edit:
        return JsonResponse({'success': False, 'message': 'Cannot delete a checked session.'})
    
    if request.method == 'POST':
        session.delete()
        return JsonResponse({'success': True, 'message': 'Session deleted successfully!'})
    
    return JsonResponse({'success': False, 'message': 'Invalid request method.'})

@login_required(login_url="user-login")
@require_http_methods(["POST"])
def check_session(request, session_id):
    if not request.user.wip_checker:
        return JsonResponse({'success': False, 'message': 'Access denied.'})
    
    session = get_object_or_404(InventorySession, id=session_id)
    
    if not session.can_check:
        return JsonResponse({'success': False, 'message': 'Session cannot be checked.'})
    
    try:
        session.status = 'CHECKED'
        session.checked_by = request.user
        session.checked_at = timezone.now()
        session.save(update_fields=['status', 'checked_by', 'checked_at'])
        
        unchecked_entries = session.entries.filter(status='FOR_CHECKING')
        unchecked_entries.update(
            status='CHECKED',
            checked_by=request.user,
            checked_at=timezone.now()
        )
        
        return JsonResponse({
            'success': True, 
            'message': 'Session marked as checked successfully!'
        })
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'message': f'Error marking session as checked: {str(e)}'
        })

@login_required(login_url="user-login")
def line_sessions(request, line_id):
    if not request.user.wip_checker:
        messages.error(request, "Access denied. Checker permission required.")
        return redirect('wip_dashboard')
    
    line = get_object_or_404(Line, id=line_id)
    sessions = InventorySession.objects.filter(line=line).order_by('-created_at')
    
    context = {
        'line': line,
        'sessions': sessions,
        'user_type': 'checker'
    }
    
    return render(request, 'wip/line_sessions.html', context)

@login_required(login_url="user-login")
def get_products(request):
    products = ProductLists.objects.all().values('id', 'product_number', 'description')
    return JsonResponse({'products': list(products)})

@login_required(login_url="user-login")
def get_processes(request):
    processes = Process.objects.filter(is_active=True).values('id', 'process_name')
    return JsonResponse({'processes': list(processes)})

@login_required(login_url="user-login")
def get_materials(request):
    product_id = request.GET.get('product_id')
    material_type = request.GET.get('type', 'wip')  # Default to 'wip' for backward compatibility
    
    if product_id:
        try:
            if material_type == 'raw':
                # For raw materials, get materials from Masterlist (not WipMasterlist)
                # We need to get all materials that are used in WipMasterlist for this product
                wip_materials = WipMasterlist.objects.filter(
                    product_number_id=product_id,
                    is_active=True
                ).select_related('material_name')
                
                materials = []
                for wip_material in wip_materials:
                    masterlist_material = wip_material.material_name
                    materials.append({
                        'id': masterlist_material.id,  # Use Masterlist id for raw materials
                        'material_name': masterlist_material.material_name,
                        'description': masterlist_material.description,
                        'cutting_length': None,  # Not applicable for raw materials
                        'uom': masterlist_material.uom or '-',  # Use Masterlist UoM
                    })
            else:
                # For WIP materials, get from WipMasterlist (existing logic)
                wip_materials = WipMasterlist.objects.filter(
                    product_number_id=product_id,
                    is_active=True
                ).select_related('material_name').values(
                    'id',  # WipMasterlist id
                    'material_name__id',
                    'material_name__material_name',
                    'material_name__description',
                    'uom',  # WipMasterlist uom
                    'cutting_length'
                )
                
                # Format the response to match the expected structure
                materials = []
                for wip_material in wip_materials:
                    materials.append({
                        'id': wip_material['material_name__id'],  # Use Masterlist id for InventoryEntry.material
                        'material_name': wip_material['material_name__material_name'],
                        'description': wip_material['material_name__description'],
                        'cutting_length': wip_material['cutting_length'],
                        'uom': wip_material['uom'] or '-',
                    })
            
            return JsonResponse({'materials': materials})
        except Exception as e:
            print(f"Error in get_materials: {str(e)}")
            return JsonResponse({'materials': [], 'error': str(e)})
    return JsonResponse({'materials': []})

@login_required(login_url="user-login")
@require_http_methods(["POST"])
def add_raw_materials(request, session_id):
    if not request.user.wip_counter:
        return JsonResponse({'success': False, 'message': 'Access denied.'})
    
    session = get_object_or_404(InventorySession, id=session_id, created_by=request.user)
    
    if not session.can_edit:
        return JsonResponse({'success': False, 'message': 'Cannot edit a checked session.'})
    
    try:
        data = json.loads(request.body)
        entries = data.get('entries', [])
        valid_entries = []
        invalid_entries = []
        for entry_data in entries:
            product_id = entry_data.get('product_id')
            material_id = entry_data.get('material_id')
            count_tag = entry_data.get('count_tag')
            quantity = entry_data.get('quantity')
            if all([product_id, material_id, count_tag, quantity]):
                valid_entries.append(entry_data)
            else:
                invalid_entries.append(entry_data)
        if not valid_entries:
            return JsonResponse({'success': False, 'message': 'No valid entries to add.'})
        for entry_data in valid_entries:
            InventoryEntry.objects.create(
                session=session,
                inventory_type='RAW',
                product_id=entry_data['product_id'],
                material_id=entry_data['material_id'],
                count_tag=entry_data['count_tag'],
                quantity=Decimal(entry_data['quantity'])
            )
        msg = 'Raw materials added successfully!'
        if invalid_entries:
            msg += f' {len(invalid_entries)} incomplete row(s) were skipped.'
        return JsonResponse({'success': True, 'message': msg})
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required(login_url="user-login")
@require_http_methods(["POST"])
def add_finished_products(request, session_id):
    if not request.user.wip_counter:
        return JsonResponse({'success': False, 'message': 'Access denied.'})
    
    session = get_object_or_404(InventorySession, id=session_id, created_by=request.user)
    
    if not session.can_edit:
        return JsonResponse({'success': False, 'message': 'Cannot edit a checked session.'})
    
    try:
        data = json.loads(request.body)
        entries = data.get('entries', [])
        valid_entries = []
        invalid_entries = []
        for entry_data in entries:
            product_id = entry_data.get('product_id')
            count_tag = entry_data.get('count_tag')
            quantity = entry_data.get('quantity')
            if all([product_id, count_tag, quantity]):
                valid_entries.append(entry_data)
            else:
                invalid_entries.append(entry_data)
        if not valid_entries:
            return JsonResponse({'success': False, 'message': 'No valid entries to add.'})
        for entry_data in valid_entries:
            InventoryEntry.objects.create(
                session=session,
                inventory_type='FINISHED',
                product_id=entry_data['product_id'],
                count_tag=entry_data['count_tag'],
                quantity=Decimal(entry_data['quantity'])
            )
        msg = 'Finished products added successfully!'
        if invalid_entries:
            msg += f' {len(invalid_entries)} incomplete row(s) were skipped.'
        return JsonResponse({'success': True, 'message': msg})
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required(login_url="user-login")
@require_http_methods(["POST"])
def add_wip_products(request, session_id):
    if not request.user.wip_counter:
        return JsonResponse({'success': False, 'message': 'Access denied.'})
    
    session = get_object_or_404(InventorySession, id=session_id, created_by=request.user)
    
    if not session.can_edit:
        return JsonResponse({'success': False, 'message': 'Cannot edit a checked session.'})
    
    try:
        data = json.loads(request.body)
        product_id = data.get('product_id')
        process_id = data.get('process_id')
        material_ids = data.get('material_ids', [])
        count_tag = data.get('count_tag')
        quantity = data.get('quantity')
        
        if not all([product_id, process_id, material_ids, count_tag, quantity]):
            return JsonResponse({'success': False, 'message': 'All fields are required.'})
        
        # Only save entries for non-empty material_ids
        valid_material_ids = [mid for mid in material_ids if mid]
        if not valid_material_ids:
            return JsonResponse({'success': False, 'message': 'No valid materials selected.'})
        for material_id in valid_material_ids:
            InventoryEntry.objects.create(
                session=session,
                inventory_type='WIP',
                product_id=product_id,
                material_id=material_id,
                process_id=process_id,
                count_tag=count_tag,
                quantity=Decimal(quantity)
            )
        
        return JsonResponse({'success': True, 'message': 'WIP products added successfully!'})
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required(login_url="user-login")
@require_http_methods(["POST"])
def update_entry(request, entry_id):
    if not request.user.wip_counter:
        return JsonResponse({'success': False, 'message': 'Access denied.'})
    
    entry = get_object_or_404(InventoryEntry, id=entry_id, session__created_by=request.user)
    
    if not entry.session.can_edit:
        return JsonResponse({'success': False, 'message': 'Cannot edit entry from a checked session.'})
    
    try:
        data = json.loads(request.body)
        count_tag = data.get('count_tag')
        quantity = data.get('quantity')
        
        if not all([count_tag, quantity]):
            return JsonResponse({'success': False, 'message': 'Count tag and quantity are required.'})
        
        entry.count_tag = count_tag
        entry.quantity = Decimal(quantity)
        entry.save()
        
        return JsonResponse({'success': True, 'message': 'Entry updated successfully!'})
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required(login_url="user-login")
@require_http_methods(["POST"])
def delete_entry(request, entry_id):
    if not request.user.wip_counter:
        return JsonResponse({'success': False, 'message': 'Access denied.'})
    try:
        entry = get_object_or_404(InventoryEntry, id=entry_id, session__created_by=request.user)
        if not entry.session.can_edit:
            return JsonResponse({'success': False, 'message': 'Cannot delete entry from a checked session.'})
        entry.delete()
        return JsonResponse({'success': True, 'message': 'Entry deleted successfully!'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required(login_url="user-login")
@require_POST
def claim_session(request, session_id):
    if not request.user.wip_checker:
        return JsonResponse({'success': False, 'message': 'Access denied.'})
    session = get_object_or_404(InventorySession, id=session_id)
    if session.checked_by is not None:
        return JsonResponse({'success': False, 'message': 'Session already claimed.'})
    session.checked_by = request.user
    session.save(update_fields=['checked_by'])
    return JsonResponse({'success': True, 'message': 'Session successfully claimed for checking.'})

@login_required(login_url="user-login")
@require_http_methods(["POST"])
def mark_entry_checked(request, entry_id):
    try:
        print(f"DEBUG: mark_entry_checked called for entry {entry_id} by user {request.user.username}")
        
        # Parse the request body to get the section type
        data = json.loads(request.body) if request.body else {}
        section_type = data.get('section')
        
        # Check if entry exists first
        try:
            entry = InventoryEntry.objects.select_related('session').get(id=entry_id)
            print(f"DEBUG: Found entry {entry.id} with status {entry.status} and type {entry.inventory_type}")
            print(f"DEBUG: Entry belongs to session {entry.session.id} created by {entry.session.created_by.username}")
            print(f"DEBUG: Current user is {request.user.username}")
        except InventoryEntry.DoesNotExist:
            print(f"DEBUG: Entry {entry_id} not found in database")
            return JsonResponse({
                'success': False, 
                'message': f'Entry with ID {entry_id} not found in database'
            }, status=404)
        
        # Check if user has permission to check this entry
        if request.user.wip_checker:
            # Checkers can check any entry in sessions that are for checking
            if entry.session.status != 'FOR_CHECKING':
                return JsonResponse({
                    'success': False, 
                    'message': 'Cannot check entries in a session that is not for checking'
                }, status=403)
        elif request.user.wip_counter:
            # Counters can only check their own entries
            if entry.session.created_by != request.user:
                return JsonResponse({
                    'success': False, 
                    'message': 'You can only check entries in your own sessions'
                }, status=403)
        else:
            return JsonResponse({
                'success': False, 
                'message': 'You do not have permission to check entries'
            }, status=403)
        
        # Validate that the section type matches the entry type
        if section_type:
            section_to_type_map = {
                'raw': 'RAW',
                'finished': 'FINISHED', 
                'wip': 'WIP'
            }
            expected_type = section_to_type_map.get(section_type.lower())
            if expected_type and entry.inventory_type != expected_type:
                return JsonResponse({
                    'success': False, 
                    'message': f'Entry type mismatch. Expected {expected_type}, got {entry.inventory_type}'
                }, status=400)
        
        # Check if entry is already checked
        if entry.status == 'CHECKED':
            return JsonResponse({
                'success': False, 
                'message': 'Entry is already checked'
            }, status=400)
        
        entry.status = 'CHECKED'
        entry.checked_by = request.user
        entry.checked_at = timezone.now()
        entry.save(update_fields=['status', 'checked_by', 'checked_at'])
        
        print(f"DEBUG: Entry {entry.id} marked as checked successfully")
        return JsonResponse({'success': True, 'message': 'Entry marked as checked.'})
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        print(f"DEBUG: Error in mark_entry_checked: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

@login_required(login_url="user-login")
def admin_dashboard(request):    
    if not request.user.wip_facilitator:
        print(f"DEBUG: Access denied for user {request.user.username} - wip_facilitator=False")
        messages.error(request, "Access denied. Facilitator permission required.")
        return redirect('/')
    
    now = timezone.now()
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    current_month_sessions = InventorySession.objects.filter(
        Q(status='FOR_CHECKING') | Q(status='CHECKED'),
        created_at__gte=month_start
    ).select_related('line').order_by('-created_at')
    
    stats = {
        'active_products': ProductLists.objects.count(),
        'unique_materials': Masterlist.objects.count(),
        'sessions_for_checking': InventorySession.objects.filter(
            status='FOR_CHECKING',
            created_at__gte=month_start
        ).count(),
        'sessions_checked': InventorySession.objects.filter(
            status='CHECKED',
            created_at__gte=month_start
        ).count()
    }
    
    all_materials = Masterlist.objects.all().order_by('-created_at')

    materials = all_materials
    products = ProductLists.objects.all().order_by('-created_at')
    for product in products:
        product.wip_materials = WipMasterlist.objects.filter(product_number=product)
    processes = Process.objects.all()
    
    materials_paginator = Paginator(materials, 20)
    products_paginator = Paginator(products, 20)
    processes_paginator = Paginator(processes, 20)
    
    materials_page = request.GET.get('materials_page', 1)
    products_page = request.GET.get('products_page', 1)
    processes_page = request.GET.get('processes_page', 1)
    
    materials = materials_paginator.get_page(materials_page)
    products = products_paginator.get_page(products_page)
    processes = processes_paginator.get_page(processes_page)
    
    context = {
        'current_month_sessions': current_month_sessions,
        'materials': materials,
        'all_materials': all_materials,
        'products': products,
        'processes': processes,
        'stats': stats
    }
    
    return render(request, 'wip/admin_dashboard.html', context)

@login_required(login_url="user-login")
def export_masterlist_template(request):
    if not request.user.wip_facilitator:
        return JsonResponse({'success': False, 'message': 'Access denied.'})
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Masterlist Template"
    
    headers = ['Material Name', 'Description', 'Unit of Measure']
    
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="masterlist_template.xlsx"'
    
    wb.save(response)
    return response

@login_required(login_url="user-login")
def export_products_template(request):
    if not request.user.wip_facilitator:
        return JsonResponse({'success': False, 'message': 'Access denied.'})
    
    wb = openpyxl.Workbook()
    
    upload_ws = wb.active
    upload_ws.title = "Upload Sheet"
    
    upload_headers = ['Product Number', 'Description', 'Material Name', 'Cutting Length', 'Unit Of Measure']
    
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for col, header in enumerate(upload_headers, 1):
        cell = upload_ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for col in range(1, len(upload_headers) + 1):
        upload_ws.column_dimensions[chr(64 + col)].width = 20
    
    reference_ws = wb.create_sheet("Material Reference")
    ref_headers = ['Material', 'Description']
    
    for col, header in enumerate(ref_headers, 1):
        cell = reference_ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    materials = Masterlist.objects.all().order_by('material_name')
    for row, material in enumerate(materials, 2):
        reference_ws.cell(row=row, column=1, value=material.material_name)
        reference_ws.cell(row=row, column=2, value=material.description)
    
    for col in range(1, len(ref_headers) + 1):
        reference_ws.column_dimensions[chr(64 + col)].width = 25
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="products_template.xlsx"'
    
    wb.save(response)
    return response

@login_required(login_url="user-login")
@require_http_methods(["POST"])
def import_masterlist(request):
    if not request.user.wip_facilitator:
        return JsonResponse({'success': False, 'message': 'Access denied.'})
    if 'file' not in request.FILES:
        return JsonResponse({'success': False, 'message': 'No file uploaded.'})
    file = request.FILES['file']
    if not file.name.endswith(('.xlsx', '.xls')):
        return JsonResponse({'success': False, 'message': 'Please upload an Excel file.'})
    try:
        df = pd.read_excel(file)
        required_columns = ['Material Name', 'Description', 'Unit of Measure']
        if not all(col in df.columns for col in required_columns):
            return JsonResponse({
                'success': False, 
                'message': f'Missing required columns. Expected: {", ".join(required_columns)}'
            })
        duplicates = []
        imported_count = 0
        for _, row in df.iterrows():
            material_name = str(row['Material Name']).strip()
            description = str(row['Description']).strip() if pd.notna(row['Description']) else ''
            uom = str(row['Unit of Measure']).strip() if pd.notna(row['Unit of Measure']) else ''
            if not material_name:
                continue
            if Masterlist.objects.filter(material_name=material_name, description=description, uom=uom).exists():
                duplicates.append({
                    'material_name': material_name,
                    'description': description,
                    'uom': uom
                })
                continue
            Masterlist.objects.create(
                material_name=material_name,
                description=description,
                uom=uom
            )
            imported_count += 1
        if imported_count == 0 and duplicates:
            message = 'No new materials imported. All entries were duplicates.'
        else:
            message = f'Successfully imported {imported_count} materials.'
        return JsonResponse({
            'success': True,
            'imported_count': imported_count,
            'duplicates': duplicates,
            'message': message
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'An error occurred during file upload: {str(e)}', 'imported_count': 0, 'duplicates': []})

@login_required(login_url="user-login")
@require_http_methods(["POST"])
def import_products(request):
    if not request.user.wip_facilitator:
        return JsonResponse({'success': False, 'message': 'Access denied.'})
    if 'file' not in request.FILES:
        return JsonResponse({'success': False, 'message': 'No file uploaded.'})
    file = request.FILES['file']
    if not file.name.endswith(('.xlsx', '.xls')):
        return JsonResponse({'success': False, 'message': 'Please upload an Excel file.'})
    
    try:
        # Read only the "Upload Sheet" from the Excel file
        df = pd.read_excel(file, sheet_name='Upload Sheet')
        
        # Check for all required columns
        required_columns = ['Product Number', 'Description', 'Material Name', 'Cutting Length', 'Unit Of Measure']
        if not all(col in df.columns for col in required_columns):
            return JsonResponse({
                'success': False, 
                'message': f'Missing required columns. Expected: {", ".join(required_columns)}'
            })
        
        products_imported = 0
        wip_materials_imported = 0
        existing_products = []
        validation_errors = []
        processing_errors = []
        
        with transaction.atomic():
            # First pass: Process ProductLists
            for index, row in df.iterrows():
                row_number = index + 2  # Excel rows start at 1, and we have header
                
                product_number = str(row['Product Number']).strip() if pd.notna(row['Product Number']) else ''
                description = str(row['Description']).strip() if pd.notna(row['Description']) else ''
                
                if not product_number:
                    continue
                
                # Check if product already exists in database
                if ProductLists.objects.filter(product_number=product_number).exists():
                    existing_products.append(product_number)
                    continue
                
                # Create new product
                try:
                    ProductLists.objects.create(
                        product_number=product_number,
                        description=description
                    )
                    products_imported += 1
                except Exception as e:
                    processing_errors.append(f"Row {row_number}: Error creating product {product_number}: {str(e)}")
            
            # Second pass: Process WipMasterlist
            for index, row in df.iterrows():
                row_number = index + 2  # Excel rows start at 1, and we have header
                
                product_number = str(row['Product Number']).strip() if pd.notna(row['Product Number']) else ''
                material_name = str(row['Material Name']).strip() if pd.notna(row['Material Name']) else ''
                cutting_length_raw = row['Cutting Length']
                unit_of_measure = str(row['Unit Of Measure']).strip() if pd.notna(row['Unit Of Measure']) else ''
                
                if not product_number or not material_name:
                    continue
                
                # Validate cutting length
                cutting_length = None
                if pd.notna(cutting_length_raw):
                    try:
                        # Try to convert to float
                        cutting_length = float(cutting_length_raw)
                        if cutting_length < 0:
                            validation_errors.append(f"Row {row_number} (Product: {product_number}): Cutting Length must be positive, got {cutting_length_raw}")
                            continue
                    except (ValueError, TypeError):
                        validation_errors.append(f"Row {row_number} (Product: {product_number}): Invalid Cutting Length '{cutting_length_raw}', must be a valid number")
                        continue
                
                # Get or create material from Masterlist
                try:
                    material, created = Masterlist.objects.get_or_create(
                        material_name=material_name,
                        defaults={
                            'uom': unit_of_measure
                        }
                    )
                    
                    # Update UOM if material already existed and UOM is different
                    if not created and material.uom != unit_of_measure:
                        material.uom = unit_of_measure
                        material.save()
                        
                except Exception as e:
                    processing_errors.append(f"Row {row_number}: Error processing material {material_name}: {str(e)}")
                    continue
                
                # Get the product (should exist from first pass, but handle edge case)
                try:
                    product = ProductLists.objects.get(product_number=product_number)
                except ProductLists.DoesNotExist:
                    processing_errors.append(f"Row {row_number}: Product {product_number} not found. Make sure it was created in the first pass.")
                    continue
                
                # Create or update WipMasterlist entry
                try:
                    wip_master, created = WipMasterlist.objects.get_or_create(
                        product_number=product,
                        material_name=material,
                        defaults={
                            'cutting_length': cutting_length,
                            'uom': unit_of_measure
                        }
                    )
                    
                    if created:
                        wip_materials_imported += 1
                    else:
                        # Update existing WIP master entry
                        wip_master.cutting_length = cutting_length
                        wip_master.uom = unit_of_measure
                        wip_master.save()
                        
                except Exception as e:
                    processing_errors.append(f"Row {row_number}: Error creating WIP master entry for {product_number} - {material_name}: {str(e)}")
        
        # Prepare response message
        message_parts = []
        if products_imported > 0:
            message_parts.append(f"Successfully imported {products_imported} products")
        if wip_materials_imported > 0:
            message_parts.append(f"Successfully imported {wip_materials_imported} WIP materials")
        
        if existing_products:
            message_parts.append(f"Skipped {len(existing_products)} existing products: {', '.join(existing_products)}")
        
        if validation_errors:
            message_parts.append(f"Validation errors: {len(validation_errors)} rows skipped")
        
        if processing_errors:
            message_parts.append(f"Processing errors: {len(processing_errors)} rows failed")
        
        message = ". ".join(message_parts) if message_parts else "No data was imported"
        
        return JsonResponse({
            'success': True,
            'products_imported': products_imported,
            'wip_materials_imported': wip_materials_imported,
            'existing_products': existing_products,
            'validation_errors': validation_errors,
            'processing_errors': processing_errors,
            'message': message
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'message': f'An error occurred during file upload: {str(e)}',
            'products_imported': 0,
            'wip_materials_imported': 0,
            'existing_products': [],
            'validation_errors': [],
            'processing_errors': []
        })

@login_required(login_url="user-login")
def export_inventory_data(request):
    if not request.user.wip_facilitator:
        return JsonResponse({'success': False, 'message': 'Access denied.'})
    
    # Get date parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Validate date parameters
    if not start_date or not end_date:
        return JsonResponse({'success': False, 'message': 'Start and end date are required.'}, status=400)
    
    try:
        # Parse dates
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d').replace(hour=0, minute=0, second=0, microsecond=0)
        end_datetime = datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59, microsecond=999999)
    except ValueError:
        return JsonResponse({'success': False, 'message': 'Invalid date format.'}, status=400)
    
    wb = openpyxl.Workbook()
    
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    # Filter sessions by date range and status
    filtered_sessions = InventorySession.objects.filter(
        created_at__range=(start_datetime, end_datetime),
        status='CHECKED'
    )
    
    # Check if there are any sessions to export
    if not filtered_sessions.exists():
        return JsonResponse({
            'success': False, 
            'message': f'No checked sessions found between {start_date} and {end_date}.'
        }, status=404)
    
    # Raw Materials Sheet - 5 columns: Count Tag, Material Name, Description, WIP Qty, UoM
    raw_ws = wb.active
    raw_ws.title = "Raw Materials"
    
    raw_headers = [
        'Count Tag', 'Material Name', 'Description', 'WIP Qty', 'UoM'
    ]
    
    for col, header in enumerate(raw_headers, 1):
        cell = raw_ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    raw_entries = InventoryEntry.objects.filter(
        inventory_type='RAW',
        session__in=filtered_sessions
    ).select_related('product', 'material', 'session')
    
    for row, entry in enumerate(raw_entries, 2):
        raw_ws.cell(row=row, column=1, value=entry.count_tag)
        raw_ws.cell(row=row, column=2, value=entry.material.material_name)
        raw_ws.cell(row=row, column=3, value=entry.material.description)
        raw_ws.cell(row=row, column=4, value=entry.quantity)
        raw_ws.cell(row=row, column=5, value=entry.material.uom or '')
    
    # Finished Goods Sheet - 4 columns: Count Tag, Product Number, Description, WIP Quantity
    finished_ws = wb.create_sheet("Finished Goods")
    finished_headers = ['Count Tag', 'Product Number', 'Description', 'WIP Quantity']
    
    for col, header in enumerate(finished_headers, 1):
        cell = finished_ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    finished_entries = InventoryEntry.objects.filter(
        inventory_type='FINISHED',
        session__in=filtered_sessions
    ).select_related('product', 'session')
    
    for row, entry in enumerate(finished_entries, 2):
        finished_ws.cell(row=row, column=1, value=entry.count_tag)
        finished_ws.cell(row=row, column=2, value=entry.product.product_number)
        finished_ws.cell(row=row, column=3, value=entry.product.description)
        finished_ws.cell(row=row, column=4, value=entry.quantity)
    
    # WIP Product Sheet - 8 columns: Product Number, Count Tag, Material Name, Description, WIP Qty, Cutting Length, UoM, Total
    wip_ws = wb.create_sheet("In Process")
    wip_headers = [
        'Product Number', 'Count Tag', 'Material Name', 'Description', 
        'WIP Qty', 'Cutting Length', 'UoM', 'Total'
    ]
    
    for col, header in enumerate(wip_headers, 1):
        cell = wip_ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    wip_entries = InventoryEntry.objects.filter(
        inventory_type='WIP',
        session__in=filtered_sessions
    ).select_related('product', 'material', 'process', 'session')
    
    for row, entry in enumerate(wip_entries, 2):
        # Get WipMasterlist data for cutting length and UoM
        wip_master = WipMasterlist.objects.filter(
            product_number=entry.product,
            material_name=entry.material
        ).first()
        
        cutting_length = wip_master.cutting_length if wip_master else 0
        uom = wip_master.uom if wip_master else ''
        total = cutting_length * entry.quantity
        
        wip_ws.cell(row=row, column=1, value=entry.product.product_number)
        wip_ws.cell(row=row, column=2, value=entry.count_tag)
        wip_ws.cell(row=row, column=3, value=entry.material.material_name)
        wip_ws.cell(row=row, column=4, value=entry.material.description)
        wip_ws.cell(row=row, column=5, value=entry.quantity)
        wip_ws.cell(row=row, column=6, value=cutting_length)
        wip_ws.cell(row=row, column=7, value=uom)
        wip_ws.cell(row=row, column=8, value=total)
    
    # Set column widths
    for col in range(1, 6):  # Raw Materials - 5 columns
        raw_ws.column_dimensions[chr(64 + col)].width = 18
    
    for col in range(1, 5):  # Finished Goods - 4 columns
        finished_ws.column_dimensions[chr(64 + col)].width = 18
    
    for col in range(1, 9):  # WIP Products - 8 columns
        wip_ws.column_dimensions[chr(64 + col)].width = 18
    
    current_date = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"wip_inventory_data_{start_date}_to_{end_date}_{current_date}.xlsx"
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@login_required(login_url="user-login")
@require_http_methods(["POST"])
def create_material(request):
    if not request.user.wip_facilitator:
        return JsonResponse({'success': False, 'message': 'Access denied.'})
    
    try:
        data = json.loads(request.body)
        material_name = data.get('material_name', '').strip()
        description = data.get('description', '').strip()
        uom = data.get('uom')
        
        if not material_name:
            return JsonResponse({'success': False, 'message': 'Material name is required.'})
        
        if Masterlist.objects.filter(material_name=material_name).exists():
            return JsonResponse({'success': False, 'message': 'Material already exists.'})
        
        material = Masterlist.objects.create(
            material_name=material_name,
            description=description,
            uom=uom if uom else None
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Material created successfully!',
            'material': {
                'id': material.id,
                'material_name': material.material_name,
                'description': material.description,
                'uom': material.uom
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

@login_required(login_url="user-login")
@require_http_methods(["POST"])
def create_product(request):
    if not request.user.wip_facilitator:
        return JsonResponse({'success': False, 'message': 'Access denied.'})
    try:
        data = json.loads(request.body)
        product_number = data.get('product_number', '').strip()
        description = data.get('description', '').strip()
        materials = data.get('materials', [])
        if not product_number:
            return JsonResponse({'success': False, 'message': 'Product number is required.'})
        if ProductLists.objects.filter(product_number=product_number).exists():
            return JsonResponse({'success': False, 'message': 'Product already exists.'})
        with transaction.atomic():
            product = ProductLists.objects.create(
                product_number=product_number,
                description=description
            )
            for mat in materials:
                material_id = mat.get('material_id')
                cutting_length = mat.get('cutting_length')
                uom = mat.get('uom')
                if not material_id:
                    continue
                material_obj = Masterlist.objects.filter(id=material_id).first()
                if not material_obj:
                    continue
                WipMasterlist.objects.create(
                    product_number=product,
                    material_name=material_obj,
                    cutting_length=cutting_length,
                    uom=uom
                )
        return JsonResponse({
            'success': True,
            'message': 'Product created successfully!',
            'product': {
                'id': product.id,
                'product_number': product.product_number,
                'description': product.description
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

@login_required(login_url="user-login")
@require_http_methods(["POST"])
def create_process(request):
    if not request.user.wip_facilitator:
        return JsonResponse({'success': False, 'message': 'Access denied.'})
    
    try:
        data = json.loads(request.body)
        process_name = data.get('process_name', '').strip()
        
        if not process_name:
            return JsonResponse({'success': False, 'message': 'Process name is required.'})
        
        if Process.objects.filter(process_name=process_name).exists():
            return JsonResponse({'success': False, 'message': 'Process already exists.'})
        
        process = Process.objects.create(
            process_name=process_name
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Process created successfully!',
            'process': {
                'id': process.id,
                'process_name': process.process_name
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

@login_required(login_url="user-login")
def search_data(request):
    if not request.user.wip_facilitator:
        return JsonResponse({'success': False, 'message': 'Access denied.'})
    
    search_type = request.GET.get('type')
    query = request.GET.get('q', '').strip()
    page = int(request.GET.get('page', 1))
    
    if search_type == 'materials':
        queryset = Masterlist.objects.all()
        if query:
            queryset = queryset.filter(
                Q(material_name__icontains=query) | 
                Q(description__icontains=query)
            )
        queryset = queryset.order_by('material_name')
        
    elif search_type == 'products':
        queryset = ProductLists.objects.all()
        if query:
            queryset = queryset.filter(
                Q(product_number__icontains=query) | 
                Q(description__icontains=query)
            )
        queryset = queryset.order_by('product_number')
        
    elif search_type == 'processes':
        queryset = Process.objects.all().order_by('created_at')
        if query:
            queryset = queryset.filter(
                Q(process_name__icontains=query)
            )
        queryset = queryset.order_by('created_at')
        
    else:
        return JsonResponse({'success': False, 'message': 'Invalid search type.'})
    
    paginator = Paginator(queryset, 20)
    page_obj = paginator.get_page(page)
    
    results = []
    for item in page_obj:
        if search_type == 'materials':
            results.append({
                'id': item.id,
                'material_name': item.material_name,
                'description': item.description,
                'uom': item.uom
            })
        elif search_type == 'products':
            results.append({
                'id': item.id,
                'product_number': item.product_number,
                'description': item.description,
                'materials': [m.material_name for m in item.material.all()]
            })
        elif search_type == 'processes':
            results.append({
                'id': item.id,
                'process_name': item.process_name,
                'is_active': item.is_active
            })
    
    return JsonResponse({
        'success': True,
        'results': results,
        'has_next': page_obj.has_next(),
        'has_previous': page_obj.has_previous(),
        'current_page': page_obj.number,
        'total_pages': paginator.num_pages
    })

@login_required(login_url="user-login")
@require_http_methods(["POST"])
def delete_item(request, item_type, item_id):
    if not request.user.wip_facilitator:
        return JsonResponse({'success': False, 'message': 'Access denied.'})
    
    try:
        if item_type == 'material':
            item = get_object_or_404(Masterlist, id=item_id)
            item.delete()
        elif item_type == 'product':
            item = get_object_or_404(ProductLists, id=item_id)
            item.delete()
        elif item_type == 'process':
            item = get_object_or_404(Process, id=item_id)
            item.delete()
        else:
            return JsonResponse({'success': False, 'message': 'Invalid item type.'})
        
        return JsonResponse({'success': True, 'message': f'{item_type.title()} deleted successfully!'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error deleting {item_type}: {str(e)}'})

@login_required(login_url="user-login")
@require_http_methods(["POST"])
def update_item(request, item_type, item_id):
    if not request.user.wip_facilitator:
        return JsonResponse({'success': False, 'message': 'Access denied.'})
    
    try:
        data = json.loads(request.body)
        
        if item_type == 'material':
            item = get_object_or_404(Masterlist, id=item_id)
            item.material_name = data.get('material_name', item.material_name)
            item.description = data.get('description', item.description)
            item.uom = data.get('uom', item.uom)
            item.save()
            
        elif item_type == 'product':
            item = get_object_or_404(ProductLists, id=item_id)
            item.product_number = data.get('product_number', item.product_number)
            item.description = data.get('description', item.description)
            item.save()

            # Update WipMasterlist entries
            materials = data.get('materials', [])
            existing_wip = {wm.material_name_id: wm for wm in WipMasterlist.objects.filter(product_number=item)}
            submitted_ids = set()

            for m in materials:
                mat_id = int(m['material_id'])
                submitted_ids.add(mat_id)
                cutting_length = m.get('cutting_length') or None
                uom = m.get('uom') or None

                if mat_id in existing_wip:
                    # Update existing
                    wm = existing_wip[mat_id]
                    wm.cutting_length = cutting_length
                    wm.uom = uom
                    wm.save()
                else:
                    # Create new
                    WipMasterlist.objects.create(
                        product_number=item,
                        material_name_id=mat_id,
                        cutting_length=cutting_length,
                        uom=uom,
                        is_active=True
                    )

            # Remove materials that were unchecked
            for mat_id, wm in existing_wip.items():
                if mat_id not in submitted_ids:
                    wm.delete()
            
        elif item_type == 'process':
            item = get_object_or_404(Process, id=item_id)
            item.process_name = data.get('process_name', item.process_name)
            item.is_active = data.get('is_active', item.is_active)
            item.save()
            
        else:
            return JsonResponse({'success': False, 'message': 'Invalid item type.'})
        
        return JsonResponse({'success': True, 'message': f'{item_type.title()} updated successfully!'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error updating {item_type}: {str(e)}'})

@login_required(login_url="user-login")
def get_all_materials(request):
    print(f"DEBUG: API call - User {request.user.username} requesting materials")
    print(f"DEBUG: User wip_facilitator permission: {request.user.wip_facilitator}")
    
    if not request.user.wip_facilitator:
        print(f"DEBUG: API access denied for user {request.user.username} - wip_facilitator=False")
        return JsonResponse({'success': False, 'message': 'Access denied. Facilitator permission required.'})
    
    print(f"DEBUG: API access granted for user {request.user.username}")
    
    try:
        page = int(request.GET.get('page', 1))
        materials = Masterlist.objects.all().order_by('-created_at')
        paginator = Paginator(materials, 20)
        page_obj = paginator.get_page(page)
        
        materials_list = [{
            'id': material.id,
            'material_name': material.material_name,
            'description': material.description,
            'uom': material.uom,
            'created_at': material.created_at.strftime('%Y-%m-%d')
        } for material in page_obj]
        
        print(f"DEBUG: Returning {len(materials_list)} materials for user {request.user.username}")
        
        return JsonResponse({
            'success': True,
            'materials': materials_list,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous(),
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages
        })
    except Exception as e:
        print(f"DEBUG: Error in get_all_materials for user {request.user.username}: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

@login_required(login_url="user-login")
def get_all_products(request):
    
    if not request.user.wip_facilitator:
        return JsonResponse({'success': False, 'message': 'Access denied. Facilitator permission required.'})
    
    try:
        page = int(request.GET.get('page', 1))
        products = ProductLists.objects.all().order_by('product_number')
        paginator = Paginator(products, 20)
        page_obj = paginator.get_page(page)
        
        products_list = [{
            'id': product.id,
            'product_number': product.product_number,
            'description': product.description,
            'wip_materials': [wip.material_name.material_name for wip in WipMasterlist.objects.filter(product_number=product)],
            'created_at': product.created_at.strftime('%Y-%m-%d')
        } for product in page_obj]
        
        return JsonResponse({
            'success': True,
            'products': products_list,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous(),
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

@login_required(login_url="user-login")
def get_all_processes(request):
    
    if not request.user.wip_facilitator:
        return JsonResponse({'success': False, 'message': 'Access denied. Facilitator permission required.'})

    try:
        page = int(request.GET.get('page', 1))
        processes = Process.objects.all().order_by('created_at')
        paginator = Paginator(processes, 20)
        page_obj = paginator.get_page(page)
        
        processes_list = [{
            'id': process.id,
            'process_name': process.process_name,
            'is_active': process.is_active,
            'created_at': process.created_at.strftime('%Y-%m-%d')
        } for process in page_obj]
        
        return JsonResponse({
            'success': True,
            'processes': processes_list,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous(),
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

@login_required(login_url="user-login")
def debug_user_permissions(request):
    user = request.user
    return JsonResponse({
        'username': user.username,
        'wip_user': user.wip_user,
        'wip_counter': user.wip_counter,
        'wip_checker': user.wip_checker,
        'wip_facilitator': user.wip_facilitator,
        'is_authenticated': user.is_authenticated,
        'user_id': user.id
    })

@login_required(login_url="user-login")
def check_unverified_sessions(request):
    if not request.user.wip_facilitator:
        return JsonResponse({'success': False, 'message': 'Access denied. Facilitator permission required.'})

    # Get date parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Validate date parameters
    if not start_date or not end_date:
        return JsonResponse({'success': False, 'message': 'Start and end date are required.'}, status=400)
    
    try:
        # Parse dates
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d').replace(hour=0, minute=0, second=0, microsecond=0)
        end_datetime = datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59, microsecond=999999)
    except ValueError:
        return JsonResponse({'success': False, 'message': 'Invalid date format.'}, status=400)

    try:
        # Filter unverified sessions by date range
        unverified_sessions = InventorySession.objects.filter(
            status='FOR_CHECKING',
            created_at__range=(start_datetime, end_datetime)
        ).select_related('line', 'created_by').order_by('-created_at')
        
        sessions_list = []
        for session in unverified_sessions:
            try:
                session_data = {
                    'id': session.id,
                    'line_name': session.line.line_name,
                    'person_responsible': session.person_responsible,
                    'created_by': session.created_by.name,
                    'created_at': session.created_at.strftime('%Y-%m-%d %H:%M'),
                    'entry_count': session.entry_count
                }
                sessions_list.append(session_data)
            except Exception as session_error:
                continue
        
        return JsonResponse({
            'success': True,
            'has_unverified_sessions': len(sessions_list) > 0,
            'unverified_sessions': sessions_list,
            'count': len(sessions_list),
            'message': f'Found {len(sessions_list)} unverified session(s) between {start_date} and {end_date}.' if len(sessions_list) > 0 else f'No unverified sessions found between {start_date} and {end_date}.'
        })
    except Exception as e:
        import traceback
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

@login_required(login_url="user-login")
@require_http_methods(["POST"])
def checking_complete(request, session_id):
    try:
        print(f"DEBUG: checking_complete called for session {session_id} by user {request.user.username}")
        session = get_object_or_404(InventorySession, id=session_id)
        print(f"DEBUG: Found session {session.id} with status {session.status}")
        
        # Check if user has permission to check this session
        if not request.user.wip_checker:
            return JsonResponse({'success': False, 'message': 'Access denied. Checker permission required.'})
        
        # Check if all entries are checked
        unchecked_entries = InventoryEntry.objects.filter(
            session=session,
            status='FOR_CHECKING'
        )
        
        if unchecked_entries.exists():
            unchecked_list = list(unchecked_entries.values('id', 'count_tag', 'material__material_name', 'product__product_number'))
            return JsonResponse({
                'success': False, 
                'message': 'Cannot mark session as checked. Some entries are still pending.',
                'unchecked_entries': unchecked_list
            })
        
        # Mark session as checked
        session.status = 'CHECKED'
        session.checked_by = request.user
        session.checked_at = timezone.now()
        session.save(update_fields=['status', 'checked_by', 'checked_at'])
        
        print(f"DEBUG: Session {session.id} marked as checked successfully")
        
        # Redirect to dashboard
        redirect_url = reverse('wip_dashboard')
        
        return JsonResponse({
            'success': True, 
            'message': 'Session marked as checked successfully.',
            'redirect_url': redirect_url
        })
        
    except Exception as e:
        print(f"DEBUG: Error in checking_complete: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

@login_required(login_url="user-login")
def product_detail_api(request, product_id):
    """
    API endpoint to return product details and its materials for editing.
    """
    if request.method != 'GET':
        return JsonResponse({'error': 'GET only'}, status=405)
    try:
        product = ProductLists.objects.get(pk=product_id)
    except ProductLists.DoesNotExist:
        return JsonResponse({'error': 'Product not found'}, status=404)
    materials_qs = WipMasterlist.objects.filter(product_number=product)
    materials = [
        {
            'material_id': wm.material_name.id,
            'cutting_length': str(wm.cutting_length) if wm.cutting_length is not None else '',
            'uom': wm.uom or ''
        }
        for wm in materials_qs
    ]
    return JsonResponse({
        'product_number': product.product_number,
        'description': product.description,
        'materials': materials
    })
