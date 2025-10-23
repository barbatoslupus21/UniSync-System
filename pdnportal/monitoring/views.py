from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST, require_GET
from django.utils import timezone
from django.utils.timezone import localtime
from django.db.models import Sum, Count, F, Q
from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
import json
import pandas as pd
import openpyxl
from io import BytesIO
from datetime import datetime, timedelta, time
from .models import (
    Monitoring, Product, ProductionSchedulePlan, ProductionOutput, 
    LineToMonitor, SupervisorToMonitor, RecentActivity, OutputLog
)
from .forms import (
    MonitoringGroupForm, ProductForm, ScheduleForm, OutputForm, 
    FilterForm, ExportForm
)
from portalusers.models import Users
from settings.models import Line

from django.db.models import Sum, Count, F, Q, Avg, Max, Min
from django.core.cache import cache

@login_required(login_url="user-login")
def monitoring_dashboard(request):
    if not (request.user.monitoring_supervisor or request.user.monitoring_manager or request.user.monitoring_sales):
        messages.error(request, "You don't have permission to access this page.")
        return redirect('dashboard')

    today = timezone.now().date()

    if request.user.monitoring_sales:
        monitoring_groups = Monitoring.objects.all()
    else:
        assigned_monitoring_ids = SupervisorToMonitor.objects.filter(
            supervisor=request.user
        ).values_list('monitoring_id', flat=True)

        monitoring_groups = Monitoring.objects.filter(
            Q(id__in=assigned_monitoring_ids) | Q(created_by=request.user)
        )

    total_groups = monitoring_groups.count()
    total_lines = LineToMonitor.objects.filter(monitoring__in=monitoring_groups).values('line').distinct().count()
    todays_outputs = ProductionOutput.objects.filter(monitoring__in=monitoring_groups, recorded_at__date=today)
    todays_output = todays_outputs.aggregate(total=Sum('quantity_produced'))['total'] or 0
    backlog_issues = ProductionSchedulePlan.objects.filter(monitoring__in=monitoring_groups, status='Backlog').count()
    recent_activities = RecentActivity.objects.filter(monitoring__in=monitoring_groups)[:10]

    # Exclude lines already assigned to a monitoring group
    assigned_line_ids = LineToMonitor.objects.values_list('line_id', flat=True)
    available_lines = Line.objects.exclude(id__in=assigned_line_ids)

    supervisors = Users.objects.filter(
        Q(monitoring_user=True) & (Q(monitoring_supervisor=True) | Q(monitoring_manager=True))
    )

    form = MonitoringGroupForm()
    form.fields['lines'].queryset = available_lines
    filter_form = FilterForm(user=request.user)

    context = {
        'monitoring_groups': monitoring_groups,
        'total_groups': total_groups,
        'total_lines': total_lines,
        'todays_output': todays_output,
        'backlog_issues': backlog_issues,
        'recent_activities': recent_activities,
        'available_lines': available_lines,
        'supervisors': supervisors,
        'form': form,
        'filter_form': filter_form,
        'today_date': today.isoformat(),
    }

    return render(request, 'monitoring/monitoring-supervisor.html', context)

@login_required(login_url="user-login")
@require_POST
def create_monitoring_group(request):
    assigned_line_ids = LineToMonitor.objects.values_list('line_id', flat=True)
    available_lines = Line.objects.exclude(id__in=assigned_line_ids)
    form = MonitoringGroupForm(request.POST)
    form.fields['lines'].queryset = available_lines

    if form.is_valid():
        title = form.cleaned_data.get('title')
        existing = Monitoring.objects.filter(
            title=title,
            created_by=request.user
        ).first()

        if existing:
            return JsonResponse({'status': 'error', 'message': f"Monitoring group '{title}' already exists."}, status=400)
        else:
            # Use the form's custom save to handle everything
            monitoring = form.save(commit=True, created_by=request.user)
            monitoring.status = 'Running'
            monitoring.save()

            RecentActivity.objects.create(
                monitoring=monitoring,
                title="New Monitoring Group Created",
                description=f"{monitoring.title} has been created",
                activity_type='info',
                shift='AM' if timezone.now().hour < 12 else 'PM',
                created_by=request.user
            )

            return JsonResponse({'status': 'success', 'message': f"Monitoring group '{monitoring.title}' created successfully!"})

    errors = form.errors.get_json_data()
    return JsonResponse({'status': 'error', 'message': 'Please correct the errors in the form.', 'errors': errors}, status=400)

@login_required(login_url="user-login")
@require_POST
def edit_monitoring_group(request, group_id):
    monitoring = get_object_or_404(Monitoring, id=group_id)
    
    if not (request.user.monitoring_sales or monitoring.created_by == request.user or 
            SupervisorToMonitor.objects.filter(monitoring=monitoring, supervisor=request.user).exists()):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'error', 'message': "You don't have permission to edit this group."}, status=403)
        messages.error(request, "You don't have permission to edit this group.")
        return redirect('monitoring_dashboard')

    form = MonitoringGroupForm(request.POST, instance=monitoring)

    # Fix: Only show lines not assigned to other groups, plus lines already assigned to this group
    assigned_line_ids = LineToMonitor.objects.exclude(monitoring=monitoring).values_list('line_id', flat=True)
    available_lines = Line.objects.exclude(id__in=assigned_line_ids)

    if form.is_valid():
        monitoring = form.save(commit=False)
        # Set status manually from group_status
        group_status = request.POST.get('status')
        if group_status in ['Running', 'On Hold', 'Stopped']:
            monitoring.status = group_status
        monitoring.save()
        form.save_m2m()

        # Also update lines and supervisors
        lines = form.cleaned_data.get('lines', [])
        supervisors = form.cleaned_data.get('supervisors', [])
        LineToMonitor.objects.filter(monitoring=monitoring).delete()
        for line in lines:
            LineToMonitor.objects.get_or_create(monitoring=monitoring, line=line)
        SupervisorToMonitor.objects.filter(monitoring=monitoring).delete()
        for supervisor in supervisors:
            SupervisorToMonitor.objects.get_or_create(monitoring=monitoring, supervisor=supervisor)

        RecentActivity.objects.create(
            monitoring=monitoring,
            title="Monitoring Group Updated",
            description=f"{monitoring.title} has been updated",
            activity_type='info',
            shift='AM' if timezone.now().hour < 12 else 'PM',
            created_by=request.user
        )

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success', 'message': f"Monitoring group '{monitoring.title}' updated successfully!"})
        messages.success(request, f"Monitoring group '{monitoring.title}' updated successfully!")
        return redirect('monitoring_dashboard')

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        errors = form.errors.get_json_data()
        return JsonResponse({'status': 'error', 'message': 'Please correct the errors in the form.', 'errors': errors}, status=400)
    messages.error(request, "Please correct the errors in the form.")
    return redirect('monitoring_dashboard')

@login_required(login_url="user-login")
def get_monitoring_group(request, group_id):
    try:
        monitoring = get_object_or_404(Monitoring, id=group_id)

        if not (request.user.monitoring_sales or monitoring.created_by == request.user or 
                SupervisorToMonitor.objects.filter(monitoring=monitoring, supervisor=request.user).exists()):
            return JsonResponse({
                'status': 'error',
                'message': 'You do not have permission to view this group'
            }, status=403)

        data = {
            'id': monitoring.id,
            'title': monitoring.title,
            'group_status': monitoring.status,
            'description': monitoring.description or '',
            'line_ids': list(LineToMonitor.objects.filter(monitoring=monitoring).values_list('line_id', flat=True)),
            'supervisor_ids': list(SupervisorToMonitor.objects.filter(monitoring=monitoring).values_list('supervisor_id', flat=True)),
            'status': 'success'
        }

        return JsonResponse(data)

    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': 'Error retrieving group details'
        }, status=500)

@login_required(login_url="user-login")
@require_POST
def delete_monitoring_group(request, group_id):
    monitoring = get_object_or_404(Monitoring, id=group_id)
    
    if not (request.user.monitoring_sales or monitoring.created_by == request.user):
        return JsonResponse({
            'status': 'error',
            'message': 'You do not have permission to delete this group'
        }, status=403)

    try:
        monitoring_title = monitoring.title
        monitoring.delete()
        
        messages.success(request, f"Monitoring group '{monitoring_title}' deleted successfully!")
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': 'Failed to delete monitoring group'
        }, status=500)

@login_required(login_url="user-login")
def group_detail(request, group_id):
    monitoring = get_object_or_404(Monitoring, id=group_id)

    if not (request.user.monitoring_sales or monitoring.created_by == request.user or 
            SupervisorToMonitor.objects.filter(monitoring=monitoring, supervisor=request.user).exists()):
        messages.error(request, "You don't have permission to view this group.")
        return redirect('monitoring_dashboard')

    today = timezone.now().date()
    
    stats = {
        'total_products': monitoring.total_products,
        'total_schedules': monitoring.total_schedules,
        'active_schedules': monitoring.active_schedules,
        'completed_schedules': monitoring.completed_schedules,
        'efficiency_percentage': monitoring.efficiency_percentage,
        'total_output_today': monitoring.total_output_today,
    }

    products = Product.objects.filter(monitoring=monitoring).select_related('line')
    schedules = ProductionSchedulePlan.objects.filter(monitoring=monitoring).select_related('product_number__line')
    
    product_form = ProductForm(monitoring=monitoring)
    schedule_form = ScheduleForm(monitoring=monitoring)
    export_form = ExportForm(monitoring=monitoring)

    available_lines = LineToMonitor.objects.filter(monitoring=monitoring).all()

    context = {
        'monitoring': monitoring,
        'stats': stats,
        'products': products,
        'schedules': schedules,
        'product_form': product_form,
        'schedule_form': schedule_form,
        'export_form': export_form,
        'today_date': today.isoformat(),
        'available_lines': available_lines,
    }

    return render(request, 'monitoring/group-detail.html', context)

@login_required(login_url="user-login")
def group_dashboard_data(request, group_id):
    try:
        monitoring = get_object_or_404(Monitoring, id=group_id)
        if not _check_dashboard_permission(request.user, monitoring):
            return JsonResponse({'error': 'Permission denied'}, status=403)
        date_filter = request.GET.get('dateFilter', 'today')
        specific_date = request.GET.get('specificDate')
        shift_filter = request.GET.get('shiftFilter', 'all')
        date_range = _calculate_date_range(date_filter, specific_date)
        start_date, end_date = date_range
        base_filters = {
            'monitoring': monitoring,
            'date_range': (start_date, end_date),
            'shift_filter': shift_filter
        }
        dashboard_data = {
            **_get_production_metrics(base_filters),
            **_get_chart_data(base_filters),
            **_get_schedule_data(base_filters),
            'xLabels': _generate_chart_labels(start_date, end_date),
            'periodLabel': _get_period_label(date_filter, start_date, end_date),
            'lastUpdated': timezone.now().isoformat(),
            'refreshInterval': 60000
        }
        return JsonResponse(dashboard_data)
    except Exception as e:
        return JsonResponse({
            'error': 'Failed to load dashboard data',
            'message': str(e),
            'success': False
        }, status=500)

@login_required(login_url="user-login")
@require_POST
def add_product(request):
    monitoring_id = request.POST.get('monitoring_id')
    monitoring = get_object_or_404(Monitoring, id=monitoring_id)

    if not (request.user.monitoring_sales or monitoring.created_by == request.user or 
            SupervisorToMonitor.objects.filter(monitoring=monitoring, supervisor=request.user).exists()):
        return JsonResponse({
            'status': 'error',
            'message': 'You do not have permission to add products to this group'
        }, status=403)

    form = ProductForm(request.POST, monitoring=monitoring)

    if form.is_valid():
        product = form.save(commit=False)
        product.monitoring = monitoring
        product.save()

        RecentActivity.objects.create(
            monitoring=monitoring,
            title="New Product Added",
            description=f"Product '{product.product_name}' added to {product.line.line_name}",
            activity_type='info',
            shift='AM' if timezone.now().hour < 12 else 'PM',
            created_by=request.user
        )

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'status': 'success',
                'message': f"Product '{product.product_name}' added successfully!"
            })

        messages.success(request, f"Product '{product.product_name}' added successfully!")
        return redirect('group_detail', group_id=monitoring_id)

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'status': 'error',
            'message': "Please correct the errors in the product form.",
            'errors': form.errors
        })

    messages.error(request, "Please correct the errors in the product form.")
    return redirect('group_detail', group_id=monitoring_id)

@login_required(login_url="user-login")
def get_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    if not (request.user.monitoring_sales or product.monitoring.created_by == request.user or 
            SupervisorToMonitor.objects.filter(monitoring=product.monitoring, supervisor=request.user).exists()):
        return JsonResponse({
            'status': 'error',
            'message': 'You do not have permission to view this product'
        }, status=403)

    data = {
        'id': product.id,
        'product_name': product.product_name,
        'description': product.description,
        'line_id': product.line.id,
        'qty_per_box': product.qty_per_box,
        'qty_per_hour': product.qty_per_hour,
        'monitoring_id': product.monitoring.id
    }

    return JsonResponse(data)

@login_required(login_url="user-login")
@require_POST
def edit_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    if not (request.user.monitoring_sales or product.monitoring.created_by == request.user or 
            SupervisorToMonitor.objects.filter(monitoring=product.monitoring, supervisor=request.user).exists()):
        return JsonResponse({
            'status': 'error',
            'message': 'You do not have permission to edit this product'
        }, status=403)

    form = ProductForm(request.POST, instance=product, monitoring=product.monitoring)

    if form.is_valid():
        form.save()

        RecentActivity.objects.create(
            monitoring=product.monitoring,
            title="Product Updated",
            description=f"Product '{product.product_name}' has been updated",
            activity_type='info',
            shift='AM' if timezone.now().hour < 12 else 'PM',
            created_by=request.user
        )

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'status': 'success',
                'message': f"Product '{product.product_name}' updated successfully!"
            })

        messages.success(request, f"Product '{product.product_name}' updated successfully!")
        return redirect('group_detail', group_id=product.monitoring.id)

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'status': 'error',
            'message': "Please correct the errors in the form.",
            'errors': form.errors
        })

    messages.error(request, "Please correct the errors in the product form.")
    return redirect('group_detail', group_id=product.monitoring.id)

@login_required(login_url="user-login")
@require_POST
def delete_product(request, product_id):
    try:
        product = get_object_or_404(Product, id=product_id)

        if not (request.user.monitoring_sales or product.monitoring.created_by == request.user or 
                SupervisorToMonitor.objects.filter(monitoring=product.monitoring, supervisor=request.user).exists()):
            return JsonResponse({
                'status': 'error',
                'message': 'You do not have permission to delete this product'
            }, status=403)

        schedules = ProductionSchedulePlan.objects.filter(product_number=product)
        if schedules.exists():
            return JsonResponse({
                'status': 'error',
                'message': 'Cannot delete this product as it is used in production schedules'
            })

        product_name = product.product_name
        monitoring = product.monitoring
        monitoring_id = monitoring.id

        product.delete()

        RecentActivity.objects.create(
            monitoring=monitoring,
            title="Product Deleted",
            description=f"Product '{product_name}' has been deleted",
            activity_type='warning',
            shift='AM' if timezone.now().hour < 12 else 'PM',
            created_by=request.user
        )

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'status': 'success',
                'message': f"Product '{product_name}' deleted successfully!"
            })

        messages.success(request, f"Product '{product_name}' deleted successfully!")
        return redirect('group_detail', group_id=monitoring_id)

    except Exception as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'status': 'error',
                'message': f"Error deleting product: {str(e)}"
            })

        messages.error(request, f"Error deleting product: {str(e)}")
        return redirect('group_detail', group_id=product.monitoring.id)

@login_required(login_url="user-login")
@require_POST
def add_schedule(request):
    monitoring_id = request.POST.get('monitoring_id')
    monitoring = get_object_or_404(Monitoring, id=monitoring_id)

    if not (request.user.monitoring_sales or monitoring.created_by == request.user or 
            SupervisorToMonitor.objects.filter(monitoring=monitoring, supervisor=request.user).exists()):
        return JsonResponse({
            'status': 'error',
            'message': 'You do not have permission to add schedules to this group'
        }, status=403)

    form = ScheduleForm(request.POST, monitoring=monitoring)

    if form.is_valid():
        schedule = form.save(commit=False)
        schedule.monitoring = monitoring
        schedule.save()

        RecentActivity.objects.create(
            monitoring=monitoring,
            title=f"Schedule Added - {schedule.product_number.line.line_name}",
            description=f"Schedule for '{schedule.product_number.product_name}' added with target of {schedule.planned_qty} units",
            activity_type='info',
            shift=schedule.shift,
            created_by=request.user
        )

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'status': 'success',
                'message': 'Production schedule added successfully!'
            })

        messages.success(request, 'Production schedule added successfully!')
        return redirect('group_detail', group_id=monitoring_id)

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'status': 'error',
            'message': "Please correct the errors in the schedule form.",
            'errors': form.errors
        })

    messages.error(request, "Please correct the errors in the schedule form.")
    return redirect('group_detail', group_id=monitoring_id)

@login_required(login_url="user-login")
def get_schedule(request, schedule_id):
    schedule = get_object_or_404(ProductionSchedulePlan, id=schedule_id)

    if not (request.user.monitoring_sales or schedule.monitoring.created_by == request.user or 
            SupervisorToMonitor.objects.filter(monitoring=schedule.monitoring, supervisor=request.user).exists()):
        return JsonResponse({
            'status': 'error',
            'message': 'You do not have permission to view this schedule'
        }, status=403)

    data = {
        'id': schedule.id,
        'product_id': schedule.product_number.id,
        'product_name': schedule.product_number.product_name,
        'line_id': schedule.product_number.line.id,
        'line_name': schedule.product_number.line.line_name,
        'date_planned': schedule.date_planned.strftime('%Y-%m-%d'),
        'shift': schedule.shift,
        'planned_qty': schedule.planned_qty,
        'balance': schedule.balance,
        'status': schedule.status,
        'monitoring_id': schedule.monitoring.id
    }

    return JsonResponse(data)

@login_required(login_url="user-login")
@require_POST
def edit_schedule(request, schedule_id):
    schedule = get_object_or_404(ProductionSchedulePlan, id=schedule_id)

    if not (request.user.monitoring_sales or schedule.monitoring.created_by == request.user or 
            SupervisorToMonitor.objects.filter(monitoring=schedule.monitoring, supervisor=request.user).exists()):
        return JsonResponse({
            'status': 'error',
            'message': 'You do not have permission to edit this schedule'
        }, status=403)

    form = ScheduleForm(request.POST, instance=schedule, monitoring=schedule.monitoring)

    if form.is_valid():
        form.save()

        RecentActivity.objects.create(
            monitoring=schedule.monitoring,
            title="Schedule Updated",
            description=f"Schedule for '{schedule.product_number.product_name}' has been updated",
            activity_type='info',
            shift=schedule.shift,
            created_by=request.user
        )

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'status': 'success',
                'message': 'Schedule updated successfully!'
            })

        messages.success(request, 'Schedule updated successfully!')
        return redirect('group_detail', group_id=schedule.monitoring.id)

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'status': 'error',
            'message': "Please correct the errors in the form.",
            'errors': form.errors
        })

    messages.error(request, "Please correct the errors in the schedule form.")
    return redirect('group_detail', group_id=schedule.monitoring.id)

@login_required(login_url="user-login")
@require_POST
def delete_schedule(request, schedule_id):
    try:
        schedule = get_object_or_404(ProductionSchedulePlan, id=schedule_id)

        if not (request.user.monitoring_sales or schedule.monitoring.created_by == request.user or 
                SupervisorToMonitor.objects.filter(monitoring=schedule.monitoring, supervisor=request.user).exists()):
            return JsonResponse({
                'status': 'error',
                'message': 'You do not have permission to delete this schedule'
            }, status=403)

        outputs = ProductionOutput.objects.filter(schedule_plan=schedule)
        if outputs.exists():
            return JsonResponse({
                'status': 'error',
                'message': 'Cannot delete this schedule as it has recorded outputs'
            })

        schedule_name = f"{schedule.product_number.product_name} - {schedule.date_planned}"
        monitoring = schedule.monitoring
        monitoring_id = monitoring.id

        schedule.delete()

        RecentActivity.objects.create(
            monitoring=monitoring,
            title="Schedule Deleted",
            description=f"Schedule for '{schedule_name}' has been deleted",
            activity_type='warning',
            shift='AM' if timezone.now().hour < 12 else 'PM',
            created_by=request.user
        )

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'status': 'success',
                'message': f"Schedule '{schedule_name}' deleted successfully!"
            })

        messages.success(request, f"Schedule '{schedule_name}' deleted successfully!")
        return redirect('group_detail', group_id=monitoring_id)

    except Exception as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'status': 'error',
                'message': f"Error deleting schedule: {str(e)}"
            })

        messages.error(request, f"Error deleting schedule: {str(e)}")
        return redirect('group_detail', group_id=schedule.monitoring.id)

@login_required(login_url="user-login")
def get_chart_data(request):
    user = request.user
    
    if not (user.monitoring_supervisor or user.monitoring_manager or user.monitoring_sales):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    time_range = request.GET.get('time_range', 'today')
    monitoring_group_id = request.GET.get('monitoring_group')
    line_id = request.GET.get('line')
    shift = request.GET.get('shift', 'all')

    today = timezone.now().date()

    if time_range == 'today':
        start_date = today
        end_date = today
    elif time_range == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = start_date + timedelta(days=6)
    elif time_range == 'month':
        start_date = today.replace(day=1)
        next_month = (start_date.replace(day=28) + timedelta(days=4)).replace(day=1)
        end_date = next_month - timedelta(days=1)
    elif time_range == 'year':
        start_date = today.replace(month=1, day=1)
        end_date = today.replace(month=12, day=31)
    else:
        start_date = today
        end_date = today

    if user.monitoring_sales:
        monitoring_groups = Monitoring.objects.all()
    else:
        assigned_monitoring_ids = SupervisorToMonitor.objects.filter(
            supervisor=user
        ).values_list('monitoring_id', flat=True)
        
        monitoring_groups = Monitoring.objects.filter(
            Q(id__in=assigned_monitoring_ids) | Q(created_by=user)
        )

    if monitoring_group_id:
        monitoring_groups = monitoring_groups.filter(id=monitoring_group_id)

    schedule_filter = Q(monitoring__in=monitoring_groups, date_planned__range=[start_date, end_date])
    output_filter = Q(monitoring__in=monitoring_groups, recorded_at__date__range=[start_date, end_date])

    if line_id and line_id != 'all':
        schedule_filter &= Q(product_number__line_id=line_id)
        output_filter &= Q(line_id=line_id)

    if shift != 'all':
        schedule_filter &= Q(shift=shift)
        output_filter &= Q(shift=shift)

    schedules = ProductionSchedulePlan.objects.filter(schedule_filter)
    outputs = ProductionOutput.objects.filter(output_filter)

    labels = []
    target = []
    actual = []

    current_date = start_date
    while current_date <= end_date:
        labels.append(current_date.strftime('%Y-%m-%d'))
        
        daily_schedules = schedules.filter(date_planned=current_date)
        daily_outputs = outputs.filter(recorded_at__date=current_date)
        
        daily_planned = daily_schedules.aggregate(total=Sum('planned_qty'))['total'] or 0
        daily_produced = daily_outputs.aggregate(total=Sum('quantity_produced'))['total'] or 0
        
        target.append(daily_planned)
        actual.append(daily_produced)
        
        current_date += timedelta(days=1)

    return JsonResponse({
        'labels': labels,
        'target': target,
        'actual': actual,
    })

@login_required(login_url="user-login")
def get_performance_data(request, group_id):
    monitoring = get_object_or_404(Monitoring, id=group_id)
    
    if not (request.user.monitoring_sales or monitoring.created_by == request.user or 
            SupervisorToMonitor.objects.filter(monitoring=monitoring, supervisor=request.user).exists()):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    date = request.GET.get('date')
    line_id = request.GET.get('line', 'all')
    shift = request.GET.get('shift', 'all')

    if not date:
        date = timezone.now().date()
    else:
        try:
            date = datetime.strptime(date, '%Y-%m-%d').date()
        except ValueError:
            date = timezone.now().date()

    if shift == 'AM':
        hours_range = range(7, 19)
    elif shift == 'PM':
        hours_range = list(range(19, 24)) + list(range(0, 7))
    else:
        hours_range = list(range(7, 24)) + list(range(0, 7))

    labels = [f"{hour:02d}:00" for hour in hours_range]
    target_data = []
    actual_data = []

    for hour in hours_range:
        hour_start = timezone.make_aware(datetime.combine(date, timezone.datetime.min.time().replace(hour=hour)))
        hour_end = hour_start + timedelta(hours=1)
        
        if line_id != 'all':
            hour_outputs = ProductionOutput.objects.filter(
                monitoring=monitoring,
                line_id=line_id,
                recorded_at__range=[hour_start, hour_end]
            )
        else:
            hour_outputs = ProductionOutput.objects.filter(
                monitoring=monitoring,
                recorded_at__range=[hour_start, hour_end]
            )
        
        if shift != 'all':
            hour_outputs = hour_outputs.filter(shift=shift)
        
        hour_total = hour_outputs.aggregate(total=Sum('quantity_produced'))['total'] or 0
        actual_data.append(hour_total)
        
        if hour_outputs.exists():
            target_per_hour = hour_outputs.first().schedule_plan.product_number.qty_per_hour
            target_data.append(target_per_hour)
        else:
            target_data.append(0)

    return JsonResponse({
        'labels': labels,
        'target': target_data,
        'actual': actual_data
    })

@login_required(login_url="user-login")
def export_group_data(request, group_id):
    monitoring = get_object_or_404(Monitoring, id=group_id)
    
    if not (request.user.monitoring_sales or monitoring.created_by == request.user or 
            SupervisorToMonitor.objects.filter(monitoring=monitoring, supervisor=request.user).exists()):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    line_id = request.GET.get('line')
    shift = request.GET.get('shift', 'all')
    export_type = request.GET.get('export_type', 'hourly')

    try:
        from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
        to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        from_date = timezone.now().date()
        to_date = from_date

    if export_type == 'total':
        schedules = ProductionSchedulePlan.objects.filter(
            monitoring=monitoring,
            date_planned__range=[from_date, to_date]
        )
        if line_id != 'all':
            schedules = schedules.filter(product_number__line_id=line_id)
        if shift != 'all':
            schedules = schedules.filter(shift=shift.upper())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Production Total Output"
        
        headers = [
            'Planned Date', 'Line', 'Product', 'Planned Quantity', 'Total Output', 'Percentage', 'Status'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

        for row_num, schedule in enumerate(schedules, 2):
            total_output = schedule.outputs.aggregate(total=Sum('quantity_produced'))['total'] or 0
            planned_qty = schedule.planned_qty
            percent = (total_output / planned_qty * 100) if planned_qty else 0
            status = 'Met' if percent >= 100 else 'Not Met'

            ws.cell(row=row_num, column=1, value=schedule.date_planned)
            ws.cell(row=row_num, column=2, value=schedule.product_number.line.line_name)
            ws.cell(row=row_num, column=3, value=schedule.product_number.product_name)
            ws.cell(row=row_num, column=4, value=planned_qty)
            ws.cell(row=row_num, column=5, value=total_output)
            
            percent_cell = ws.cell(row=row_num, column=6, value=round(percent, 2))
            if percent >= 100:
                percent_cell.font = openpyxl.styles.Font(color="006100", bold=True)
            elif percent >= 90:
                percent_cell.font = openpyxl.styles.Font(color="007300")
            elif percent >= 80:
                percent_cell.font = openpyxl.styles.Font(color="B85450")
            elif percent >= 70:
                percent_cell.font = openpyxl.styles.Font(color="C55A11")
            else:
                percent_cell.font = openpyxl.styles.Font(color="9C0006", bold=True)
            
            status_cell = ws.cell(row=row_num, column=7, value=status)
            if status == 'Met':
                status_cell.fill = openpyxl.styles.PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                status_cell.font = openpyxl.styles.Font(color="006100", bold=True)
            else:
                status_cell.fill = openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                status_cell.font = openpyxl.styles.Font(color="9C0006", bold=True)

        for column in ws.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"{monitoring.title}_production_total_output_{from_date}_{to_date}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    else:
        output_logs_query = OutputLog.objects.filter(
            outputlog__monitoring=monitoring,
            time_recorded__date__range=[from_date, to_date]
        )

        if line_id and line_id != 'all':
            output_logs_query = output_logs_query.filter(outputlog__line_id=line_id)

        if shift != 'all':
            output_logs_query = output_logs_query.filter(outputlog__shift=shift.upper())

        output_logs = output_logs_query.select_related(
            'outputlog__line'
        ).order_by('time_recorded')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hourly Output Data"

        headers = ['Date', 'Line', 'Output', 'Status', 'Operator']
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

        for row_num, log in enumerate(output_logs, 2):
            ws.cell(row=row_num, column=1, value=log.time_recorded.date())
            ws.cell(row=row_num, column=2, value=log.outputlog.line.line_name)
            ws.cell(row=row_num, column=3, value=log.output)
            
            status_cell = ws.cell(row=row_num, column=4, value=log.status or 'N/A')
            if log.status == 'Met':
                status_cell.fill = openpyxl.styles.PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                status_cell.font = openpyxl.styles.Font(color="006100", bold=True)
            elif log.status == 'Not Met':
                status_cell.fill = openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                status_cell.font = openpyxl.styles.Font(color="9C0006", bold=True)
            
            ws.cell(row=row_num, column=5, value=log.operator or 'N/A')

        for column in ws.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"{monitoring.title}_hourly_output_{from_date}_{to_date}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

@login_required(login_url="user-login")
def export_product_template(request):
    monitoring_id = request.GET.get('monitoring_id')
    
    if not monitoring_id:
        messages.error(request, "Monitoring group ID is required")
        return redirect('monitoring_dashboard')
        
    try:
        monitoring = get_object_or_404(Monitoring, id=monitoring_id)
        
        if not (request.user.monitoring_sales or monitoring.created_by == request.user or 
                SupervisorToMonitor.objects.filter(monitoring=monitoring, supervisor=request.user).exists()):
            messages.error(request, "You don't have permission to access this template")
            return redirect('monitoring_dashboard')
    except:
        messages.error(request, "Invalid monitoring group ID")
        return redirect('monitoring_dashboard')

    wb = openpyxl.Workbook()
    
    ws_upload = wb.active
    ws_upload.title = "Upload Data"
    headers = ['Product Name', 'Description', 'Line', 'Qty per Box', 'Qty per Hour']
    for col_num, header in enumerate(headers, 1):
        ws_upload.cell(row=1, column=col_num, value=header)

    # Only include products for the selected monitoring group
    products = Product.objects.filter(monitoring=monitoring)
    for row_num, product in enumerate(products, 2):
        ws_upload.cell(row=row_num, column=1, value=product.product_name)
        ws_upload.cell(row=row_num, column=2, value=product.description)
        ws_upload.cell(row=row_num, column=3, value=product.line.line_name)
        ws_upload.cell(row=row_num, column=4, value=product.qty_per_box)
        ws_upload.cell(row=row_num, column=5, value=product.qty_per_hour)

    ws_lines = wb.create_sheet(title="Line Masterlist")
    ws_lines.cell(row=1, column=1, value="Line Name")
    lines = Line.objects.filter(assigned_lines__monitoring=monitoring).distinct()
    for row_num, line in enumerate(lines, 2):
        ws_lines.cell(row=row_num, column=1, value=line.line_name)

    for ws in [ws_upload, ws_lines]:
        for column in ws.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"product_template_{monitoring.title}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response

@login_required(login_url="user-login")
@require_POST
def import_products(request):
    monitoring_id = request.POST.get('monitoring_id')
    monitoring = get_object_or_404(Monitoring, id=monitoring_id)

    if not (request.user.monitoring_sales or monitoring.created_by == request.user or 
            SupervisorToMonitor.objects.filter(monitoring=monitoring, supervisor=request.user).exists()):
        messages.error(request, 'You do not have permission to import products to this group')
        return redirect(request.META.get('HTTP_REFERER', 'monitoring_dashboard'))

    if 'product_file' not in request.FILES:
        messages.error(request, 'No file uploaded')
        return redirect(request.META.get('HTTP_REFERER', 'monitoring_dashboard'))

    excel_file = request.FILES['product_file']

    if not excel_file.name.endswith(('.xlsx', '.xls')):
        messages.error(request, 'Uploaded file is not an Excel file')
        return redirect(request.META.get('HTTP_REFERER', 'monitoring_dashboard'))

    try:
        df = pd.read_excel(excel_file, sheet_name='Upload Data')

        required_columns = ['Product Name', 'Description', 'Line', 'Qty per Box', 'Qty per Hour']
        for column in required_columns:
            if column not in df.columns:
                messages.error(request, f'Missing required column: {column}')
                return redirect(request.META.get('HTTP_REFERER', 'monitoring_dashboard'))

        products_created = 0
        products_updated = 0
        errors = []

        valid_lines = set(Line.objects.filter(
            assigned_lines__monitoring=monitoring
        ).values_list('line_name', flat=True))

        with transaction.atomic():
            for index, row in df.iterrows():
                try:
                    product_name = str(row['Product Name']).strip()
                    description = str(row['Description']).strip()
                    line_name = str(row['Line']).strip()
                    qty_per_box = int(row['Qty per Box'])
                    qty_per_hour = int(row['Qty per Hour'])

                    if line_name not in valid_lines:
                        errors.append(f"Row {index+2}: Invalid line '{line_name}'")
                        continue

                    line = Line.objects.get(line_name=line_name)

                    existing_product = Product.objects.filter(
                        monitoring=monitoring,
                        product_name=product_name,
                        line=line
                    ).first()

                    if existing_product:
                        existing_product.description = description
                        existing_product.qty_per_box = qty_per_box
                        existing_product.qty_per_hour = qty_per_hour
                        existing_product.save()
                        products_updated += 1
                    else:
                        Product.objects.create(
                            monitoring=monitoring,
                            product_name=product_name,
                            description=description,
                            line=line,
                            qty_per_box=qty_per_box,
                            qty_per_hour=qty_per_hour
                        )
                        products_created += 1

                except Exception as e:
                    errors.append(f"Row {index+2}: {str(e)}")

        RecentActivity.objects.create(
            monitoring=monitoring,
            title="Products Imported",
            description=f"{products_created} products created, {products_updated} updated",
            activity_type='info',
            shift='AM' if timezone.now().hour < 12 else 'PM',
            created_by=request.user
        )

        if errors:
            messages.warning(request, f"Imported {products_created} products, updated {products_updated} with {len(errors)} errors. Example: {errors[0] if errors else ''}")
        else:
            messages.success(request, f"Successfully imported {products_created} products and updated {products_updated}")
        return redirect(request.META.get('HTTP_REFERER', 'monitoring_dashboard'))

    except Exception as e:
        messages.error(request, f"Error processing Excel file: {str(e)}")
        return redirect(request.META.get('HTTP_REFERER', 'monitoring_dashboard'))

@login_required(login_url="user-login")
def export_schedule_template(request):
    monitoring_id = request.GET.get('monitoring_id')
    
    if not monitoring_id:
        messages.error(request, "Monitoring group ID is required")
        return redirect('monitoring_dashboard')
        
    try:
        monitoring = get_object_or_404(Monitoring, id=monitoring_id)
        
        if not (request.user.monitoring_sales or monitoring.created_by == request.user or 
                SupervisorToMonitor.objects.filter(monitoring=monitoring, supervisor=request.user).exists()):
            messages.error(request, "You don't have permission to access this template")
            return redirect('monitoring_dashboard')
    except:
        messages.error(request, "Invalid monitoring group ID")
        return redirect('monitoring_dashboard')

    wb = openpyxl.Workbook()
    
    ws_upload = wb.active
    ws_upload.title = "Upload Schedule"
    headers = ['Date Planned', 'Product', 'Planned Qty', 'Shift', 'Status']
    for col_num, header in enumerate(headers, 1):
        ws_upload.cell(row=1, column=col_num, value=header)

    today = timezone.now().date().strftime('%Y-%m-%d')
    products = Product.objects.filter(monitoring=monitoring)
    
    sample_data = []
    for i, product in enumerate(products[:2]):
        sample_data.append([today, product.product_name, 1000, 'AM', 'Planned'])
        sample_data.append([today, product.product_name, 1500, 'PM', 'Planned'])

    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws_upload.cell(row=row_num, column=col_num, value=cell_value)

    ws_products = wb.create_sheet(title="Product Masterlist")
    ws_products.cell(row=1, column=1, value="Product Name")
    ws_products.cell(row=1, column=2, value="Line Name")
    for row_num, product in enumerate(products, 2):
        ws_products.cell(row=row_num, column=1, value=product.product_name)
        ws_products.cell(row=row_num, column=2, value=product.line.line_name)

    for ws in [ws_upload, ws_products]:
        for column in ws.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"schedule_template_{monitoring.title}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

@login_required(login_url="user-login")
@require_POST
def import_schedules(request):
    monitoring_id = request.POST.get('monitoring_id')
    monitoring = get_object_or_404(Monitoring, id=monitoring_id)

    if not (request.user.monitoring_sales or monitoring.created_by == request.user or 
            SupervisorToMonitor.objects.filter(monitoring=monitoring, supervisor=request.user).exists()):
        messages.error(request, 'You do not have permission to import schedules to this group')
        return redirect(request.META.get('HTTP_REFERER', 'monitoring_dashboard'))

    if 'schedule_file' not in request.FILES:
        messages.error(request, 'No file uploaded')
        return redirect(request.META.get('HTTP_REFERER', 'monitoring_dashboard'))

    excel_file = request.FILES['schedule_file']

    if not excel_file.name.endswith(('.xlsx', '.xls')):
        messages.error(request, 'Uploaded file is not an Excel file')
        return redirect(request.META.get('HTTP_REFERER', 'monitoring_dashboard'))

    try:
        import pandas as pd
        df = pd.read_excel(excel_file, sheet_name='Upload Schedule')

        required_columns = ['Date Planned', 'Product', 'Planned Qty', 'Shift', 'Status']
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            messages.error(request, f"Missing required columns: {', '.join(missing_columns)}")
            return redirect(request.META.get('HTTP_REFERER', 'monitoring_dashboard'))

        schedules_created = 0
        schedules_updated = 0
        errors = []

        valid_products = set(Product.objects.filter(
            monitoring=monitoring
        ).values_list('product_name', flat=True))

        from datetime import datetime
        with transaction.atomic():
            for index, row in df.iterrows():
                try:
                    date_planned_str = str(row['Date Planned']).strip()
                    product_name = str(row['Product']).strip()
                    planned_qty = int(row['Planned Qty'])
                    shift = str(row['Shift']).strip().upper()
                    status = str(row['Status']).strip()

                    if isinstance(row['Date Planned'], str):
                        date_planned = datetime.strptime(date_planned_str, '%Y-%m-%d').date()
                    else:
                        date_planned = row['Date Planned'].date()

                    if product_name not in valid_products:
                        errors.append(f"Row {index+2}: Invalid product '{product_name}'")
                        continue

                    if shift not in ['AM', 'PM']:
                        errors.append(f"Row {index+2}: Invalid shift '{shift}'. Must be 'AM' or 'PM'")
                        continue

                    valid_statuses = ['Planned', 'Change Load', 'Backlog']
                    if status not in valid_statuses:
                        errors.append(f"Row {index+2}: Invalid status '{status}'. Must be one of {', '.join(valid_statuses)}")
                        continue

                    product = Product.objects.get(
                        product_name=product_name,
                        monitoring=monitoring
                    )

                    existing_schedule = ProductionSchedulePlan.objects.filter(
                        monitoring=monitoring,
                        product_number=product,
                        date_planned=date_planned,
                        shift=shift
                    ).first()

                    if existing_schedule:
                        existing_schedule.planned_qty = planned_qty
                        existing_schedule.status = status
                        existing_schedule.balance = planned_qty
                        existing_schedule.save()
                        schedules_updated += 1
                    else:
                        ProductionSchedulePlan.objects.create(
                            monitoring=monitoring,
                            product_number=product,
                            date_planned=date_planned,
                            shift=shift,
                            planned_qty=planned_qty,
                            balance=planned_qty,
                            status=status
                        )
                        schedules_created += 1

                except Exception as e:
                    errors.append(f"Row {index+2}: {str(e)}")

        RecentActivity.objects.create(
            monitoring=monitoring,
            title="Schedules Imported",
            description=f"{schedules_created} schedules created, {schedules_updated} updated",
            activity_type='info',
            shift='AM' if timezone.now().hour < 12 else 'PM',
            created_by=request.user
        )

        if errors:
            messages.warning(request, f"Imported {schedules_created} schedules, updated {schedules_updated} with {len(errors)} errors. Example: {errors[0] if errors else ''}")
        else:
            messages.success(request, f"Successfully imported {schedules_created} schedules and updated {schedules_updated}")
        return redirect(request.META.get('HTTP_REFERER', 'monitoring_dashboard'))

    except Exception as e:
        messages.error(request, f"Error processing Excel file: {str(e)}")
        return redirect(request.META.get('HTTP_REFERER', 'monitoring_dashboard'))

# LINE DASHBOARD
@login_required(login_url="user-login")
def production_dashboard(request):
    now = localtime()
    today = now.date()
    current_hour = now.hour

    current_shift = 'AM' if 7 <= current_hour < 19 else 'PM'

    user_lines = request.user.line.all()

    schedule = ProductionSchedulePlan.objects.filter(
        product_number__line__in=user_lines,
        date_planned=today,
        shift=current_shift,
        status="Planned",
        balance__gt=0
    ).first()

    if not schedule:
        schedule = ProductionSchedulePlan.objects.filter(
            product_number__line__in=user_lines,
            date_planned=today,
            shift=current_shift,
            status="Change Load",
            balance__gt=0
        ).first()

    if not schedule:
        schedule = ProductionSchedulePlan.objects.filter(
            product_number__line__in=user_lines,
            date_planned=today,
            shift=current_shift,
            status="Backlog",
            balance__gt=0
        ).first()

    if not schedule:
        context = {
            'schedule_exists': False,
            'page_title': 'Production Monitoring',
            'subtitle': 'No scheduled production for this shift',
            'active_nav': 'monitoring',
            'form': OutputForm()
        }
        return render(request, 'monitoring/line-dashboard.html', context)

    if request.method == 'POST':
        form = OutputForm(request.POST)
        
        if form.is_valid():
            operator = form.cleaned_data['operator']
            quantity = form.cleaned_data['quantity']
            
            production_output, created = ProductionOutput.objects.get_or_create(
                monitoring=schedule.monitoring,
                schedule_plan=schedule,
                line=schedule.product_number.line,
                shift=current_shift,
                defaults={
                    'inspector': operator,
                    'quantity_produced': 0
                }
            )

            if operator and production_output.inspector != operator:
                production_output.inspector = operator

            production_output.quantity_produced += quantity
            production_output.save()

            output_log = OutputLog.objects.create(
                outputlog=production_output,
                output=quantity,
                time_recorded=timezone.now()
            )

            request.session['last_operator'] = operator
            
            target_per_hour = schedule.product_number.qty_per_hour
            status = "Met" if quantity >= target_per_hour else "Not Met"
            current_time = timezone.now()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'success',
                    'quantity': quantity,
                    'target': target_per_hour,
                    'evaluation': status,
                    'operator': operator,
                    'line_name': schedule.product_number.line.line_name,
                    'time_recorded': localtime(current_time).strftime('%H:%M'),
                    'message': 'Production output added successfully!'
                })
            
            messages.success(request, 'Production output added successfully!')
            return redirect('line_dashboard')
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'error',
                    'errors': form.errors,
                    'message': 'Please correct the form errors.'
                })
    
    form = OutputForm()
    if 'last_operator' in request.session:
        form.initial['operator'] = request.session['last_operator']

    outputs = ProductionOutput.objects.filter(schedule_plan=schedule)
    total_produced = outputs.aggregate(total=Sum('quantity_produced'))['total'] or 0

    if schedule.planned_qty > 0:
        completion_percentage = (total_produced / schedule.planned_qty) * 100
    else:
        completion_percentage = 0

    target_per_hour = schedule.product_number.qty_per_hour

    output_logs = OutputLog.objects.filter(
        outputlog__schedule_plan=schedule
    ).order_by('time_recorded')

    display_logs = []
    for log in output_logs:
        local_time = localtime(log.time_recorded)
        variance = log.output - target_per_hour

        display_logs.append({
            'time': local_time,
            'operator': log.outputlog.inspector,
            'line': log.outputlog.line.line_name,
            'output': log.output,
            'target': target_per_hour,
            'variance': variance,
            'status': log.status if log.status else ("Met" if variance >= 0 else "Not Met")
        })

    chart_labels = []
    chart_outputs = []
    chart_targets = []

    for log in output_logs:
        local_time = localtime(log.time_recorded)
        time_label = local_time.strftime('%H:%M')
        chart_labels.append(time_label)
        chart_outputs.append(log.output)
        chart_targets.append(target_per_hour)

    if not chart_labels:
        current_time = localtime()
        time_label = current_time.strftime('%H:%M')
        chart_labels = [time_label]
        chart_outputs = [0]
        chart_targets = [target_per_hour]

    chart_data = {
        'labels': chart_labels,
        'datasets': [
            {
                'label': 'Output',
                'data': chart_outputs
            },
            {
                'label': 'Target',
                'data': chart_targets
            }
        ]
    }

    context = {
        'schedule_exists': True,
        'page_title': 'Production Monitoring',
        'subtitle': f'Line: {schedule.product_number.line.line_name} - {current_shift} Shift',
        'active_nav': 'monitoring',
        'schedule': schedule,
        'product_name': schedule.product_number.product_name,
        'line_name': schedule.product_number.line.line_name,
        'planned_qty': schedule.planned_qty,
        'total_produced': total_produced,
        'completion_percentage': completion_percentage,
        'balance': schedule.balance,
        'target_per_hour': target_per_hour,
        'logs': display_logs,
        'chart_data': json.dumps(chart_data),
        'form': form
    }

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'status': 'success',
            'total_produced': total_produced,
            'completion_percentage': completion_percentage,
            'balance': schedule.balance,
            'chart_data': chart_data
        })

    return render(request, 'monitoring/line-dashboard.html', context)

@login_required(login_url="user-login")
@require_POST
def delete_monitoring_group(request, group_id):
    monitoring = get_object_or_404(Monitoring, id=group_id)
    # Only allow creator or assigned supervisor/manager to delete
    if not (monitoring.created_by == request.user or SupervisorToMonitor.objects.filter(monitoring=monitoring, supervisor=request.user).exists()):
        return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)
    try:
        monitoring.delete()
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@login_required(login_url="user-login")
def facilitator_dashboard(request):
    if not (request.user.monitoring_supervisor or request.user.monitoring_manager or request.user.monitoring_sales):
        messages.error(request, "You do not have permission to access this page")
        return redirect('monitoring_dashboard')

    if request.user.monitoring_sales:
        monitoring_groups = Monitoring.objects.all()
    else:
        assigned_monitoring_ids = SupervisorToMonitor.objects.filter(
            supervisor=request.user
        ).values_list('monitoring_id', flat=True)

        monitoring_groups = Monitoring.objects.filter(
            Q(id__in=assigned_monitoring_ids) | Q(created_by=request.user)
        )

    today = timezone.now().date()
    
    total_groups = monitoring_groups.count()
    total_lines = LineToMonitor.objects.filter(monitoring__in=monitoring_groups).values('line').distinct().count()
    todays_outputs = ProductionOutput.objects.filter(monitoring__in=monitoring_groups, recorded_at__date=today)
    todays_output = todays_outputs.aggregate(total=Sum('quantity_produced'))['total'] or 0
    backlog_issues = ProductionSchedulePlan.objects.filter(monitoring__in=monitoring_groups, status='Backlog').count()

    context = {
        'monitoring_groups': monitoring_groups,
        'total_groups': total_groups,
        'total_lines': total_lines,
        'todays_output': todays_output,
        'backlog_issues': backlog_issues,
        'available_lines': Line.objects.all(),
        'supervisors': Users.objects.filter(Q(monitoring_user=True) & (Q(monitoring_supervisor=True) | Q(monitoring_manager=True))),
        'form': MonitoringGroupForm(),
        'today_date': today.isoformat()
    }

    return render(request, 'monitoring/monitoring-supervisor.html', context)

@login_required(login_url="user-login")
def facilitator_chart_data(request):
    if not (request.user.monitoring_supervisor or request.user.monitoring_manager or request.user.monitoring_sales):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    time_range = request.GET.get('timeRange', 'today')
    group_id = request.GET.get('groupId', 'all')

    today = timezone.now().date()

    if time_range == 'today':
        start_date = today
        end_date = today
    elif time_range == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = start_date + timedelta(days=6)
    elif time_range == 'month':
        start_date = today.replace(day=1)
        next_month = (start_date.replace(day=28) + timedelta(days=4)).replace(day=1)
        end_date = next_month - timedelta(days=1)
    elif time_range == 'year':
        start_date = today.replace(month=1, day=1)
        end_date = today.replace(month=12, day=31)
    else:
        start_date = today
        end_date = today

    if request.user.monitoring_sales:
        monitoring_groups = Monitoring.objects.all()
    else:
        monitoring_groups = Monitoring.objects.filter(created_by=request.user)

    if group_id != 'all':
        monitoring_groups = monitoring_groups.filter(id=group_id)

    schedules = ProductionSchedulePlan.objects.filter(
        monitoring__in=monitoring_groups,
        date_planned__range=[start_date, end_date]
    )
    
    outputs = ProductionOutput.objects.filter(
        monitoring__in=monitoring_groups,
        recorded_at__date__range=[start_date, end_date]
    )

    labels = []
    target_data = []
    actual_data = []

    current_date = start_date
    while current_date <= end_date:
        labels.append(current_date.strftime('%Y-%m-%d'))
        
        daily_schedules = schedules.filter(date_planned=current_date)
        daily_outputs = outputs.filter(recorded_at__date=current_date)
        
        daily_target = daily_schedules.aggregate(
            total=Sum('planned_qty')
        )['total'] or 0
        
        daily_actual = daily_outputs.aggregate(
            total=Sum('quantity_produced')
        )['total'] or 0
        
        target_data.append(daily_target)
        actual_data.append(daily_actual)
        
        current_date += timedelta(days=1)

    return JsonResponse({
        'labels': labels,
        'target': target_data,
        'actual': actual_data
    })

@login_required(login_url="user-login")
def facilitator_group_detail(request, group_id):
    if not (request.user.monitoring_supervisor or request.user.monitoring_manager or request.user.monitoring_sales):
        messages.error(request, "You do not have permission to access this page")
        return redirect('overview')

    monitoring = get_object_or_404(Monitoring, id=group_id)

    if not request.user.monitoring_sales:
        if not (
            SupervisorToMonitor.objects.filter(monitoring=monitoring, supervisor=request.user).exists() or
            monitoring.created_by == request.user
        ):
            messages.error(request, "You do not have permission to view this group")
            return redirect('facilitator_dashboard')

    today = timezone.now().date()
    
    # Only show lines assigned to this monitoring group
    line_ids = LineToMonitor.objects.filter(monitoring=monitoring).values_list('line_id', flat=True)
    available_lines = Line.objects.filter(id__in=line_ids)
    
    today_schedules = ProductionSchedulePlan.objects.filter(monitoring=monitoring, date_planned=today)
    today_outputs = ProductionOutput.objects.filter(monitoring=monitoring, recorded_at__date=today)
    
    total_planned = today_schedules.aggregate(total=Sum('planned_qty'))['total'] or 0
    total_produced = today_outputs.aggregate(total=Sum('quantity_produced'))['total'] or 0
    
    efficiency_percentage = round((total_produced / total_planned) * 100) if total_planned > 0 else 0
    
    products = monitoring.monitoring_product.all().order_by('-created_at')
    schedules = ProductionSchedulePlan.objects.filter(monitoring=monitoring).order_by('-created_at')
    
    product_form = ProductForm(monitoring=monitoring)
    schedule_form = ScheduleForm(monitoring=monitoring)
    export_form = ExportForm(monitoring=monitoring)

    context = {
        'monitoring': monitoring,
        'available_lines': available_lines,
        'total_planned': total_planned,
        'total_produced': total_produced,
        'efficiency_percentage': efficiency_percentage,
        'products': products,
        'schedules': schedules,
        'product_form': product_form,
        'schedule_form': schedule_form,
        'export_form': export_form,
        'today_date': today.isoformat()
    }

    return render(request, 'monitoring/group-detail.html', context)

@login_required(login_url="user-login")
def facilitator_group_chart_data(request, group_id):
    if not (request.user.monitoring_supervisor or request.user.monitoring_manager or request.user.monitoring_sales):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    monitoring = get_object_or_404(Monitoring, id=group_id)
    
    date_filter = request.GET.get('date', timezone.now().date().isoformat())
    line_filter = request.GET.get('line', 'all')
    shift_filter = request.GET.get('shift', 'all')

    try:
        filter_date = datetime.strptime(date_filter, '%Y-%m-%d').date()
    except ValueError:
        filter_date = timezone.now().date()

    schedules_query = ProductionSchedulePlan.objects.filter(
        monitoring=monitoring,
        date_planned=filter_date
    )
    
    outputs_query = ProductionOutput.objects.filter(
        monitoring=monitoring,
        recorded_at__date=filter_date
    )

    if line_filter != 'all':
        schedules_query = schedules_query.filter(product_number__line_id=line_filter)
        outputs_query = outputs_query.filter(line_id=line_filter)

    if shift_filter != 'all':
        schedules_query = schedules_query.filter(shift=shift_filter.upper())
        outputs_query = outputs_query.filter(shift=shift_filter.upper())

    if shift_filter == 'AM':
        hours_range = range(7, 19)
    elif shift_filter == 'PM':
        hours_range = list(range(19, 24)) + list(range(0, 7))
    else:
        hours_range = list(range(7, 24)) + list(range(0, 7))

    labels = [f"{hour:02d}:00" for hour in hours_range]
    
    target_per_hour = schedules_query.aggregate(
        total=Sum('product_number__qty_per_hour')
    )['total'] or 0
    
    target_data = [target_per_hour] * len(hours_range)
    actual_data = [0] * len(hours_range)

    output_logs = OutputLog.objects.filter(
        outputlog__in=outputs_query
    ).order_by('time_recorded')

    for log in output_logs:
        log_time = timezone.localtime(log.time_recorded)
        hour = log_time.hour
        
        if (shift_filter == 'PM' or shift_filter == 'all') and hour < 7:
            hour += 24
            
        if hour in hours_range:
            hour_index = list(hours_range).index(hour)
            actual_data[hour_index] = log.output

    return JsonResponse({
        'labels': labels,
        'target': target_data,
        'actual': actual_data
    })

@login_required(login_url="user-login")
def facilitator_export_output_data(request, group_id):
    if not (request.user.monitoring_supervisor or request.user.monitoring_manager or request.user.monitoring_sales):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    monitoring = get_object_or_404(Monitoring, id=group_id)
    
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    line_id = request.GET.get('line', 'all')
    shift = request.GET.get('shift', 'all')

    try:
        from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
        to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        from_date = timezone.now().date()
        to_date = from_date

    outputs_query = ProductionOutput.objects.filter(
        monitoring=monitoring,
        recorded_at__date__range=[from_date, to_date]
    ).select_related('line', 'schedule_plan__product_number')

    if line_id != 'all':
        outputs_query = outputs_query.filter(line_id=line_id)

    if shift != 'all':
        outputs_query = outputs_query.filter(shift=shift.upper())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Output Data"

    headers = ['Date', 'Time', 'Line', 'Product', 'Shift', 'Quantity', 'Inspector']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    for row_num, output in enumerate(outputs_query, 2):
        local_time = timezone.localtime(output.recorded_at)
        ws.cell(row=row_num, column=1, value=local_time.date())
        ws.cell(row=row_num, column=2, value=local_time.time())
        ws.cell(row=row_num, column=3, value=output.line.line_name if output.line else 'N/A')
        ws.cell(row=row_num, column=4, value=output.schedule_plan.product_number.product_name if output.schedule_plan and output.schedule_plan.product_number else 'N/A')
        ws.cell(row=row_num, column=5, value=output.shift)
        ws.cell(row=row_num, column=6, value=output.quantity_produced)
        ws.cell(row=row_num, column=7, value=output.inspector or 'N/A')

    for column in ws.columns:
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{monitoring.title}_output_data_{from_date}_{to_date}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response

# GROUP DASHBOARD
def group_dashboard(request, group_id):
    monitoring = get_object_or_404(Monitoring, id=group_id)
    
    if not _check_dashboard_permission(request.user, monitoring):
        messages.error(request, "You do not have permission to view this dashboard")
        return redirect('monitoring_dashboard')
    
    today = timezone.now().date()
    
    dashboard_stats = _get_dashboard_overview(monitoring, today)
    
    context = {
        'monitoring': monitoring,
        'today_date': today.isoformat(),
        'dashboard_stats': dashboard_stats,
        'user_permissions': {
            'can_edit': monitoring.created_by == request.user,
            'can_view_all': request.user.monitoring_sales,
            'is_supervisor': request.user.monitoring_supervisor or request.user.monitoring_manager
        }
    }
    
    return render(request, 'monitoring/group-dashboard.html', context)

@login_required(login_url="user-login")
def group_dashboard_data(request, group_id):
    try:
        monitoring = get_object_or_404(Monitoring, id=group_id)
        
        if not _check_dashboard_permission(request.user, monitoring):
            return JsonResponse({'error': 'Permission denied'}, status=403)
        
        date_filter = request.GET.get('dateFilter', 'today')
        specific_date = request.GET.get('specificDate')
        shift_filter = request.GET.get('shiftFilter', 'all')
        
        date_range = _calculate_date_range(date_filter, specific_date)
        start_date, end_date = date_range
        
        base_filters = {
            'monitoring': monitoring,
            'date_range': (start_date, end_date),
            'shift_filter': shift_filter
        }
        
        dashboard_data = {
            **_get_production_metrics(base_filters),
            **_get_chart_data(base_filters),
            **_get_schedule_data(base_filters),
            'xLabels': _generate_chart_labels(start_date, end_date),
            'periodLabel': _get_period_label(date_filter, start_date, end_date),
            'lastUpdated': timezone.now().isoformat(),
            'refreshInterval': 60000
        }
        return JsonResponse(dashboard_data)
        
    except Exception as e:
        return JsonResponse({
            'error': 'Failed to load dashboard data',
            'message': str(e),
            'success': False
        }, status=500)

def _check_dashboard_permission(user, monitoring):
    """Check if user has permission to view dashboard"""
    if user.monitoring_sales:
        return True
    
    if monitoring.created_by == user:
        return True
        
    if SupervisorToMonitor.objects.filter(monitoring=monitoring, supervisor=user).exists():
        return True
        
    return False

def _calculate_date_range(date_filter, specific_date):
    """Calculate start and end dates based on filter"""
    today = timezone.localdate()
    
    if date_filter == 'today':
        return today, today
    elif date_filter == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = start_date + timedelta(days=6)
        return start_date, end_date
    elif date_filter == 'month':
        start_date = today.replace(day=1)
        next_month = (start_date.replace(day=28) + timedelta(days=4)).replace(day=1)
        end_date = next_month - timedelta(days=1)
        return start_date, end_date
    elif date_filter == 'year':
        start_date = today.replace(month=1, day=1)
        end_date = today.replace(month=12, day=31)
        return start_date, end_date
    elif date_filter == 'customDate' and specific_date:
        try:
            custom_date = datetime.strptime(specific_date, '%Y-%m-%d').date()
            return custom_date, custom_date
        except ValueError:
            return today, today

    else:
        return today, today

def _get_dashboard_overview(monitoring, today):
    """Get basic dashboard overview data"""
    schedules_today = ProductionSchedulePlan.objects.filter(
        monitoring=monitoring,
        date_planned=today
    )
    
    outputs_today = ProductionOutput.objects.filter(
        monitoring=monitoring,
        recorded_at__date=today
    )
    
    total_planned_today = schedules_today.aggregate(total=Sum('planned_qty'))['total'] or 0
    total_produced_today = outputs_today.aggregate(total=Sum('quantity_produced'))['total'] or 0
    
    efficiency = round((total_produced_today / total_planned_today) * 100) if total_planned_today > 0 else 0
    
    return {
        'total_schedules_today': schedules_today.count(),
        'total_planned_today': total_planned_today,
        'total_produced_today': total_produced_today,
        'efficiency_today': efficiency,
        'active_lines': schedules_today.values('product_number__line').distinct().count()
    }

def _get_production_metrics(base_filters):
    """Calculate production metrics with optimized queries"""
    monitoring = base_filters['monitoring']
    start_date, end_date = base_filters['date_range']
    shift_filter = base_filters['shift_filter']
    schedule_query = ProductionSchedulePlan.objects.filter(
        monitoring=monitoring,
        date_planned__range=[start_date, end_date]
    ).select_related('product_number__line')
    output_query = ProductionOutput.objects.filter(
        monitoring=monitoring,
        recorded_at__date__range=[start_date, end_date]
    ).select_related('line', 'schedule_plan')
    if shift_filter != 'all':
        schedule_query = schedule_query.filter(shift=shift_filter.upper())
        output_query = output_query.filter(shift=shift_filter.upper())
    schedules = schedule_query
    outputs = output_query
    metrics = schedules.aggregate(
        total_schedules=Count('id'),
        total_planned=Sum('planned_qty'),
        avg_planned=Avg('planned_qty')
    )
    output_metrics = outputs.aggregate(
        total_produced=Sum('quantity_produced'),
        avg_produced=Avg('quantity_produced')
    )
    total_schedules = metrics['total_schedules'] or 0
    total_planned = metrics['total_planned'] or 0
    total_produced = output_metrics['total_produced'] or 0
    production_progress = round((total_produced / total_planned) * 100) if total_planned > 0 else 0
    not_met_target = 100 - production_progress if production_progress <= 100 else 0
    lines = LineToMonitor.objects.filter(monitoring=monitoring).values_list('line', flat=True)
    active_lines = schedules.values('product_number__line').distinct().count()
    total_lines = lines.count()
    total_schedules_target = ProductionSchedulePlan.objects.filter(monitoring=monitoring).count()
    return {
        'totalSchedules': total_schedules,
        'totalSchedulesTarget': total_schedules_target,
        'totalPlanned': total_planned,
        'totalProduced': total_produced,
        'productionProgress': production_progress,
        'productionProgressTarget': 100,
        'activeLines': active_lines,
        'totalLines': total_lines,
        'notMetTarget': not_met_target,
        'averagePlanned': round(metrics['avg_planned'] or 0, 2),
        'averageProduced': round(output_metrics['avg_produced'] or 0, 2)
    }

def _get_chart_data(base_filters):
    """Generate optimized chart data"""
    monitoring = base_filters['monitoring']
    start_date, end_date = base_filters['date_range']
    shift_filter = base_filters['shift_filter']
    
    output_per_day = []
    efficiency_data = []
    
    current_date = start_date
    while current_date <= end_date:
        daily_schedules = ProductionSchedulePlan.objects.filter(
            monitoring=monitoring,
            date_planned=current_date
        )
        
        daily_outputs = ProductionOutput.objects.filter(
            monitoring=monitoring,
            recorded_at__date=current_date
        )
        
        if shift_filter != 'all':
            daily_schedules = daily_schedules.filter(shift=shift_filter.upper())
            daily_outputs = daily_outputs.filter(shift=shift_filter.upper())
        
        daily_metrics = daily_schedules.aggregate(
            planned=Sum('planned_qty')
        )
        
        daily_output_metrics = daily_outputs.aggregate(
            produced=Sum('quantity_produced')
        )
        
        daily_planned = daily_metrics['planned'] or 0
        daily_produced = daily_output_metrics['produced'] or 0
        
        output_per_day.append({
            'date': current_date.strftime('%Y-%m-%d'),
            'quantity': daily_produced,
            'label': current_date.strftime('%b %d')
        })
        
        efficiency = round((daily_produced / daily_planned) * 100) if daily_planned > 0 else 0
        efficiency_data.append({
            'date': current_date.strftime('%Y-%m-%d'),
            'efficiency': efficiency,
            'label': current_date.strftime('%b %d')
        })
        
        current_date += timedelta(days=1)
    
    output_by_line = _get_line_performance_data(monitoring, start_date, end_date, shift_filter)
    shift_output = _get_shift_distribution_data(monitoring, start_date, end_date)
    status_distribution = _get_status_distribution_data(monitoring, start_date, end_date, shift_filter)
    
    return {
        'outputPerDay': output_per_day,
        'efficiencyData': efficiency_data,
        'outputByLine': output_by_line,
        'shiftOutput': shift_output,
        'statusDistribution': status_distribution
    }

def _get_line_performance_data(monitoring, start_date, end_date, shift_filter):
    """Get performance data by production line"""
    lines = LineToMonitor.objects.filter(monitoring=monitoring).select_related('line')
    
    output_by_line = []
    
    for line_monitor in lines:
        line = line_monitor.line
        
        outputs = ProductionOutput.objects.filter(
            monitoring=monitoring,
            line=line,
            recorded_at__date__range=[start_date, end_date]
        )
        
        if shift_filter != 'all':
            outputs = outputs.filter(shift=shift_filter.upper())
        
        quantity = outputs.aggregate(total=Sum('quantity_produced'))['total'] or 0
        
        output_by_line.append({
            'line': line.line_name,
            'quantity': quantity,
            'lineId': line.id
        })
    
    return sorted(output_by_line, key=lambda x: x['quantity'], reverse=True)

def _get_shift_distribution_data(monitoring, start_date, end_date):
    """Get shift distribution data"""
    outputs = ProductionOutput.objects.filter(
        monitoring=monitoring,
        recorded_at__date__range=[start_date, end_date]
    )
    
    am_outputs = outputs.filter(shift='AM')
    pm_outputs = outputs.filter(shift='PM')
    
    am_total = am_outputs.aggregate(total=Sum('quantity_produced'))['total'] or 0
    pm_total = pm_outputs.aggregate(total=Sum('quantity_produced'))['total'] or 0
    
    return [
        {'shift': 'AM Shift', 'quantity': am_total},
        {'shift': 'PM Shift', 'quantity': pm_total}
    ]

def _get_status_distribution_data(monitoring, start_date, end_date, shift_filter):
    """Get schedule status distribution"""
    schedules = ProductionSchedulePlan.objects.filter(
        monitoring=monitoring,
        date_planned__range=[start_date, end_date]
    )
    
    if shift_filter != 'all':
        schedules = schedules.filter(shift=shift_filter.upper())
    
    status_counts = schedules.values('status').annotate(count=Count('id')).order_by('-count')
    
    return [
        {
            'status': item['status'], 
            'count': item['count'],
            'percentage': round((item['count'] / schedules.count()) * 100, 1) if schedules.count() > 0 else 0
        } 
        for item in status_counts
    ]

def _get_schedule_data(base_filters):
    """Get formatted schedule data for table"""
    monitoring = base_filters['monitoring']
    start_date, end_date = base_filters['date_range']
    shift_filter = base_filters['shift_filter']
    
    schedules = ProductionSchedulePlan.objects.filter(
        monitoring=monitoring,
        date_planned__range=[start_date, end_date]
    ).select_related('product_number__line').prefetch_related('outputs')
    
    if shift_filter != 'all':
        schedules = schedules.filter(shift=shift_filter.upper())
    
    schedule_list = []
    
    for schedule in schedules.order_by('-date_planned', '-created_at')[:100]:
        outputs = schedule.outputs.all()
        produced_qty = sum(output.quantity_produced for output in outputs)
        progress = round((produced_qty / schedule.planned_qty) * 100, 1) if schedule.planned_qty > 0 else 0
        
        schedule_list.append({
            'id': schedule.id,
            'date': schedule.date_planned.strftime('%Y-%m-%d'),
            'product': schedule.product_number.product_name,
            'line': schedule.product_number.line.line_name,
            'shift': schedule.shift,
            'plannedQty': schedule.planned_qty,
            'producedQty': produced_qty,
            'progress': progress,
            'status': schedule.status,
            'balance': schedule.balance,
            'efficiency': min(progress, 100)
        })
    
    return {'schedules': schedule_list}

def _generate_chart_labels(start_date, end_date):
    """Generate labels for charts"""
    labels = []
    current_date = start_date
    current_year = timezone.localdate().year
    
    if start_date == end_date:
        if start_date.year == current_year:
            return [start_date.strftime('%b %d')]
        else:
            return [start_date.strftime('%b %d, %Y')]
    
    while current_date <= end_date:
        if current_date.year == current_year:
            labels.append(current_date.strftime('%b %d'))
        else:
            labels.append(current_date.strftime('%b %d, %Y'))
        current_date += timedelta(days=1)
    
    return labels

def _get_period_label(date_filter, start_date, end_date):
    """Generate period label for dashboard"""
    current_year = timezone.localdate().year
    
    if start_date == end_date:
        if start_date.year == current_year:
            return f"{date_filter.title()} - {start_date.strftime('%b %d')}"
        else:
            return f"{date_filter.title()} - {start_date.strftime('%b %d, %Y')}"
    else:
        if start_date.year != end_date.year:
            return f"{date_filter.title()} - {start_date.strftime('%b %d, %Y')} to {end_date.strftime('%b %d, %Y')}"
        elif start_date.year == current_year:
            return f"{date_filter.title()} - {start_date.strftime('%b %d')} to {end_date.strftime('%b %d')}"
        else:
            return f"{date_filter.title()} - {start_date.strftime('%b %d')} to {end_date.strftime('%b %d, %Y')}"

@login_required(login_url="user-login")
def get_real_time_metrics(request, group_id):
    """API endpoint for real-time metric updates"""
    monitoring = get_object_or_404(Monitoring, id=group_id)
    
    if not _check_dashboard_permission(request.user, monitoring):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    today = timezone.localdate()
    
    schedules_today = ProductionSchedulePlan.objects.filter(
        monitoring=monitoring,
        date_planned=today
    )
    
    outputs_today = ProductionOutput.objects.filter(
        monitoring=monitoring,
        recorded_at__date=today
    )
    
    metrics = {
        'totalPlannedToday': schedules_today.aggregate(total=Sum('planned_qty'))['total'] or 0,
        'totalProducedToday': outputs_today.aggregate(total=Sum('quantity_produced'))['total'] or 0,
        'activeSchedules': schedules_today.count(),
        'completedSchedules': schedules_today.filter(balance=0).count(),
        'lastUpdate': timezone.now().isoformat()
    }
    
    metrics['efficiencyToday'] = round(
        (metrics['totalProducedToday'] / metrics['totalPlannedToday']) * 100
    ) if metrics['totalPlannedToday'] > 0 else 0
    
    return JsonResponse(metrics)

@login_required(login_url="user-login")
def export_dashboard_data(request, group_id):
    """Export dashboard data as CSV or Excel"""
    monitoring = get_object_or_404(Monitoring, id=group_id)
    
    if not _check_dashboard_permission(request.user, monitoring):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    export_format = request.GET.get('format', 'csv')
    date_filter = request.GET.get('dateFilter', 'today')
    specific_date = request.GET.get('specificDate')
    shift_filter = request.GET.get('shiftFilter', 'all')
    
    date_range = _calculate_date_range(date_filter, specific_date)
    start_date, end_date = date_range
    
    # Get all dashboard data
    base_filters = {
        'monitoring': monitoring,
        'date_range': (start_date, end_date),
        'shift_filter': shift_filter
    }
    
    schedule_data = _get_schedule_data(base_filters)
    schedules = schedule_data['schedules']
    
    if export_format == 'csv':
        return _export_csv(schedules, monitoring.title, start_date, end_date)
    elif export_format == 'excel':
        return _export_excel(schedules, monitoring.title, start_date, end_date)
    else:
        return JsonResponse({'error': 'Invalid export format'}, status=400)

def _export_csv(schedules, title, start_date, end_date):
    """Export schedules as CSV"""
    import csv
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/csv')
    filename = f"{title}_dashboard_{start_date}_{end_date}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Date', 'Product', 'Line', 'Shift', 'Planned Qty', 
        'Produced Qty', 'Progress %', 'Status', 'Balance', 'Efficiency %'
    ])
    
    for schedule in schedules:
        writer.writerow([
            schedule['date'],
            schedule['product'],
            schedule['line'],
            schedule['shift'],
            schedule['plannedQty'],
            schedule['producedQty'],
            f"{schedule['progress']:.1f}%",
            schedule['status'],
            schedule['balance'],
            f"{schedule['efficiency']:.1f}%"
        ])
    
    return response

def _export_excel(schedules, title, start_date, end_date):
    """Export schedules as Excel"""
    import openpyxl
    from django.http import HttpResponse
    from io import BytesIO
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dashboard Data"
    
    # Headers
    headers = [
        'Date', 'Product', 'Line', 'Shift', 'Planned Qty', 
        'Produced Qty', 'Progress %', 'Status', 'Balance', 'Efficiency %'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(
            start_color="0ea5e9", end_color="0ea5e9", fill_type="solid"
        )
        cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
    
    # Data rows
    for row_num, schedule in enumerate(schedules, 2):
        ws.cell(row=row_num, column=1, value=schedule['date'])
        ws.cell(row=row_num, column=2, value=schedule['product'])
        ws.cell(row=row_num, column=3, value=schedule['line'])
        ws.cell(row=row_num, column=4, value=schedule['shift'])
        ws.cell(row=row_num, column=5, value=schedule['plannedQty'])
        ws.cell(row=row_num, column=6, value=schedule['producedQty'])
        ws.cell(row=row_num, column=7, value=f"{schedule['progress']:.1f}%")
        ws.cell(row=row_num, column=8, value=schedule['status'])
        ws.cell(row=row_num, column=9, value=schedule['balance'])
        ws.cell(row=row_num, column=10, value=f"{schedule['efficiency']:.1f}%")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{title}_dashboard_{start_date}_{end_date}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

@login_required(login_url="user-login")
def get_dashboard_summary(request, group_id):
    """Get a quick summary of dashboard metrics"""
    monitoring = get_object_or_404(Monitoring, id=group_id)
    
    if not _check_dashboard_permission(request.user, monitoring):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    today = timezone.localdate()
    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)
    
    # Today's metrics
    today_metrics = _get_dashboard_overview(monitoring, today)
    
    # Week metrics
    week_schedules = ProductionSchedulePlan.objects.filter(
        monitoring=monitoring,
        date_planned__range=[week_start, today]
    )
    week_outputs = ProductionOutput.objects.filter(
        monitoring=monitoring,
        recorded_at__date__range=[week_start, today]
    )
    
    week_planned = week_schedules.aggregate(total=Sum('planned_qty'))['total'] or 0
    week_produced = week_outputs.aggregate(total=Sum('quantity_produced'))['total'] or 0
    week_efficiency = round((week_produced / week_planned) * 100) if week_planned > 0 else 0
    
    # Month metrics
    month_schedules = ProductionSchedulePlan.objects.filter(
        monitoring=monitoring,
        date_planned__range=[month_start, today]
    )
    month_outputs = ProductionOutput.objects.filter(
        monitoring=monitoring,
        recorded_at__date__range=[month_start, today]
    )
    
    month_planned = month_schedules.aggregate(total=Sum('planned_qty'))['total'] or 0
    month_produced = month_outputs.aggregate(total=Sum('quantity_produced'))['total'] or 0
    month_efficiency = round((month_produced / month_planned) * 100) if month_planned > 0 else 0
    
    summary = {
        'today': {
            'efficiency': today_metrics['efficiency_today'],
            'planned': today_metrics['total_planned_today'],
            'produced': today_metrics['total_produced_today'],
            'schedules': today_metrics['total_schedules_today']
        },
        'week': {
            'efficiency': week_efficiency,
            'planned': week_planned,
            'produced': week_produced,
            'schedules': week_schedules.count()
        },
        'month': {
            'efficiency': month_efficiency,
            'planned': month_planned,
            'produced': month_produced,
            'schedules': month_schedules.count()
        },
        'trends': {
            'efficiency_trend': 'up' if week_efficiency > today_metrics['efficiency_today'] else 'down',
            'production_trend': 'up' if week_produced > today_metrics['total_produced_today'] * 7 else 'down'
        },
        'lastUpdated': timezone.now().isoformat()
    }
    
    return JsonResponse(summary)