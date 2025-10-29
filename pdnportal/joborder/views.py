from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone
from datetime import timedelta, date
from django.http import JsonResponse
from django.utils.timezone import now
from rest_framework import serializers
from rest_framework.response import Response
from rest_framework.decorators import api_view
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from .serializers import JORequestSerializer, JORoutingSerializer, JORequestApproverSerializer, JORequestAdminSerializer
from django.db.models.functions import TruncMonth
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from .models import GreenControlNumber, YellowControlNumber, WhiteControlNumber, OrangeControlNumber, JOLogsheet, JORouting
from portalusers.models import UserApprovers, Users
from notification.models import Notification
import calendar
import datetime
import openpyxl
import json
import io
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse, JsonResponse
from datetime import datetime
from django.db.models import Count, Q
from django.views.decorators.http import require_GET, require_POST
from django.db.models.functions import ExtractMonth, ExtractYear
from django.core.paginator import Paginator
from settings.models import Line
from .forms import JobOrderRequestForm
from django.db.models import Case, When, Value, IntegerField, Q, BooleanField, Exists, OuterRef, Subquery, Max

# REQUESTORS VIEW
@login_required(login_url="user-login")
def requestor_page(request):
    current_month = now().month
    current_year = now().year

    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', 'all')

    if request.user.job_order_approver or request.user.job_order_pmd:
        pendingJO = JOLogsheet.objects.filter(
            joRouting__approver=request.user,
            joRouting__status='pending'
        ).distinct().order_by('-joRouting__request_at')
    elif request.user.job_order_spectator:
        pendingJO = JOLogsheet.objects.all().order_by('-date_created')[:10]
    elif request.user.job_order_maintenance:
        pendingJO = JOLogsheet.objects.filter(in_charge=request.user).filter(Q(target_date__isnull=True) | Q(action_taken__isnull=True)).order_by("-date_created")
    elif request.user.job_order_checker:
        pendingJO = JOLogsheet.objects.filter(Q(joRouting__approver=request.user, joRouting__status='pending') | Q(prepared_by=request.user, status='Completed') | Q(prepared_by=request.user, status='Checked')).order_by('-date_created')
    else:
        pendingJO = JOLogsheet.objects.filter(prepared_by=request.user
        ).annotate(
            status_order=Case(
                When(status="Completed", then=Value(1)),
                When(status="Routing", then=Value(2)),
                output_field=IntegerField(),
            )
        ).order_by("status_order", "-date_created")

    #Table Data
    if request.user.job_order_approver:
        pending_routing_qs = JORouting.objects.filter(
            jo_request=OuterRef('pk'),
            approver=request.user,
            status='pending'
        )

        latest_request_at = JORouting.objects.filter(jo_request=OuterRef('pk')).order_by('-request_at').values('request_at')[:1]

        joRequests = JOLogsheet.objects.all().prefetch_related('joRouting').annotate(
            is_pending=Case(When(Exists(pending_routing_qs), then=Value(1)), default=Value(0), output_field=IntegerField()),
            latest_request_at=Subquery(latest_request_at)
        ).order_by('-is_pending', '-latest_request_at').distinct()

    elif request.user.job_order_maintenance:
        joRequests = JOLogsheet.objects.filter(in_charge=request.user).annotate(
            priority_order=Case(
                When(priority_level='Urgent', then=Value(1)),
                When(priority_level='High', then=Value(2)),
                When(priority_level='Medium', then=Value(3)),
                When(priority_level='Low', then=Value(4)),
                default=Value(5),
                output_field=IntegerField(),
            ),
            step_order=Case(
                When(Q(target_date__isnull=True) & Q(status='Assigned'), then=Value(1)),
                When(Q(target_date__isnull=False) & Q(action_taken__isnull=True) & Q(status='Assigned'), then=Value(2)),
                When(
                    Q(target_date__isnull=False) & 
                    Q(action_taken__isnull=False) & 
                    Q(status__in=['Completed', 'Closed', 'Checked']),
                    then=Value(3)
                ),
                default=Value(4),
                output_field=IntegerField(),
            )
        ).order_by('step_order', 'priority_order', 'date_of_completion', '-date_created')
        
    elif request.user.job_order_checker:
        joRequests = JOLogsheet.objects.filter(
            Q(joRouting__approver=request.user, joRouting__status='pending') |
            Q(prepared_by=request.user) | Q(joRouting__approver=request.user, joRouting__status='approved')
        ).annotate(
            is_pending_approval=Case(
                When(joRouting__approver=request.user, joRouting__status='pending', then=Value(1)),
                default=Value(0),
                output_field=IntegerField(),
            )
        ).distinct().order_by('-is_pending_approval', '-date_created')

    elif request.user.job_order_pmd:
        pending_routing_qs = JORouting.objects.filter(
            jo_request=OuterRef('pk'),
            approver=request.user,
            status='pending'
        )

        latest_request_at = JORouting.objects.filter(jo_request=OuterRef('pk')).order_by('-request_at').values('request_at')[:1]

        joRequests = JOLogsheet.objects.all().prefetch_related('joRouting').annotate(
            is_pending=Case(When(Exists(pending_routing_qs), then=Value(1)), default=Value(0), output_field=IntegerField()),
            latest_request_at=Subquery(latest_request_at)
        ).order_by('-is_pending', '-latest_request_at').distinct()

    elif request.user.job_order_spectator:
        joRequests = JOLogsheet.objects.all().order_by("-date_created")
        
    else:
       joRequests = (
            JOLogsheet.objects.filter(prepared_by=request.user).annotate(
                is_pending_approval=Max(Case(
                    When(joRouting__approver=request.user, joRouting__status='pending', then=Value(1)),
                    default=Value(0),
                    output_field=IntegerField(),
                ))
            )).order_by('-is_pending_approval', '-date_created')
    
    # Apply search filter
    if search_query:
        joRequests = joRequests.filter(
            Q(jo_number__icontains=search_query) |
            Q(jo_color__icontains=search_query) |
            Q(jo_tools__icontains=search_query) |
            Q(jo_type__icontains=search_query) |
            Q(requestor__icontains=search_query) |
            Q(line__line_name__icontains=search_query) |
            Q(status__icontains=search_query) |
            Q(details__icontains=search_query)
        ).distinct()
    
    # Apply status filter
    if status_filter and status_filter != 'all':
        status_map = {
            'closed': 'Closed',
            'routing': 'Routing',
            'assigned': 'Assigned',
            'cancelled': 'Cancelled',
            'rejected': 'Rejected',
            'completed': 'Completed'
        }
        if status_filter in status_map:
            joRequests = joRequests.filter(status=status_map[status_filter])
    
    # Pagination
    paginator = Paginator(joRequests, 10)
    page_number = request.GET.get('page')
    all_requests = paginator.get_page(page_number)
    
    if request.user.job_order_approver or request.user.job_order_pmd or request.user.job_order_maintenance or request.user.job_order_checker or request.user.job_order_requestor:
        for req in all_requests:
            routing_entry = req.joRouting.filter(approver=request.user).first()
            req.has_pending_routing = routing_entry and routing_entry.status.lower() == 'pending'

            date_entry = req.joRouting.filter(approver=request.user).first()
            req.has_date_entry = date_entry and date_entry.status.lower() == 'pending' and date_entry.jo_request.target_date is None
            req.for_completion = date_entry and date_entry.status.lower() == 'pending' and date_entry.jo_request.target_date is not None and date_entry.jo_request.date_complete is None

            closing_entry = req.joRouting.filter(approver=request.user).last()
            req.has_closing_entry = (
                closing_entry
                and closing_entry.status.lower() == 'pending'
                and req.status in ["Completed", "Checked", "On Hold"]
            )

    else:
        for req in all_requests:
            req.has_pending_routing = False
            req.has_date_entry = False
            req.for_completion = False
            req.has_closing_entry = False

    joRequestsCount = JOLogsheet.objects.filter(prepared_by=request.user, date_created__year=current_year, date_created__month=current_month).count()
    pendingJOCount = JOLogsheet.objects.filter(status = "Routing", prepared_by=request.user, date_created__year=current_year, date_created__month=current_month).count()
    approvedJOCount = JOLogsheet.objects.filter(status = "Closed", prepared_by=request.user, date_created__year=current_year, date_created__month=current_month).count()
    lines = Line.objects.all()

    context={
        'pendingJO':pendingJO,
        'all_requests':all_requests,
        'joRequestsCount':joRequestsCount,
        'pendingJOCount':pendingJOCount,
        'approvedJOCount':approvedJOCount,
        'lines': lines,
        'form': JobOrderRequestForm(user=request.user),
        'search_query': search_query,
        'status_filter': status_filter,
    }
    return render(request, 'joborder/jo-requestor.html', context)

@login_required
@require_GET
def job_order_chart_data(request, period):
    today = timezone.now().date()

    if period == '3month':
        months_back = 3
    elif period == '1year':
        months_back = 12
    else:
        months_back = 6

    start_date = today.replace(day=1)
    for _ in range(months_back - 1):
        prev_month = start_date.month - 1
        year = start_date.year
        if prev_month == 0:
            prev_month = 12
            year -= 1
        start_date = start_date.replace(year=year, month=prev_month, day=1)

    months = []
    month_year_map = {}

    current = start_date
    index = 0
    while current <= today.replace(day=28):
        month_name = calendar.month_name[current.month][:3]
        year = current.year
        month_year = f"{month_name} {year}"

        months.append(month_year)
        month_year_map[f"{current.year}-{current.month:02d}"] = index

        if current.month == 12:
            current = current.replace(year=current.year + 1, month=1)
        else:
            current = current.replace(month=current.month + 1)

        index += 1

    green_data = [0] * len(months)
    yellow_data = [0] * len(months)
    white_data = [0] * len(months)
    orange_data = [0] * len(months)

    if request.user.job_order_approver:
        job_orders = JOLogsheet.objects.filter(
            joRouting__approver=request.user,
            date_created__gte=timezone.make_aware(timezone.datetime.combine(start_date, timezone.datetime.min.time())),
            date_created__lte=timezone.now()
        )
    elif request.user.job_order_maintenance:
        job_orders = JOLogsheet.objects.filter(
            in_charge=request.user,
            date_created__gte=timezone.make_aware(timezone.datetime.combine(start_date, timezone.datetime.min.time())),
            date_created__lte=timezone.now()
        )
    elif request.user.job_order_pmd or request.user.job_order_spectator:
        job_orders = JOLogsheet.objects.filter(
            date_created__gte=timezone.make_aware(timezone.datetime.combine(start_date, timezone.datetime.min.time())),
            date_created__lte=timezone.now()
        )
    else:
        job_orders = JOLogsheet.objects.filter(
            prepared_by=request.user,
            date_created__gte=timezone.make_aware(timezone.datetime.combine(start_date, timezone.datetime.min.time())),
            date_created__lte=timezone.now()
        )

    for order in job_orders:
        date = order.date_created.date()
        month_year_key = f"{date.year}-{date.month:02d}"

        if month_year_key in month_year_map:
            idx = month_year_map[month_year_key]
            category = order.jo_color.lower() if order.jo_color else "unknown"

            if category == 'green':
                green_data[idx] += 1
            elif category == 'yellow':
                yellow_data[idx] += 1
            elif category == 'white':
                white_data[idx] += 1
            elif category == 'orange':
                orange_data[idx] += 1

    # After calculating green_data, yellow_data, white_data, orange_data
    total_data = [
        green_data[i] + yellow_data[i] + white_data[i] + orange_data[i]
        for i in range(len(green_data))
    ]

    return JsonResponse({
        'labels': months,
        'green': green_data,
        'yellow': yellow_data,
        'white': white_data,
        'orange': orange_data,
        'total': total_data
    })

@login_required(login_url="user-login")
def create_jo_request(request):
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    if request.method == 'POST':
        form = JobOrderRequestForm(request.POST, user=request.user)

        if form.is_valid():
            category = request.POST.get("jo-category")
            tooling = request.POST.get("tooling")
            nature = request.POST.get("nature")
            details = form.cleaned_data['details']
            line = form.cleaned_data['line']
            complaint = request.POST.get("complaint")
            requestor = request.POST.get("requestor")

            # Validate additional required fields
            if not all([category, tooling, nature, requestor]):
                missing = []
                if not category: missing.append('category')
                if not tooling: missing.append('tooling')
                if not nature: missing.append('nature')
                if not requestor: missing.append('requestor')

                error_message = f"Missing required fields: {', '.join(missing)}"

                if is_ajax:
                    return JsonResponse({'status': 'error', 'message': error_message})

                messages.error(request, error_message)
                return redirect("requestor-homepage")

            if (category == "orange" and
                nature in ["countermeasure-cri", "countermeasure-ecc", "safety"] and
                not complaint):
                error_message = "Please specify your complaint for orange category requests."

                if is_ajax:
                    return JsonResponse({'status': 'error', 'message': error_message})

                messages.error(request, error_message)
                return redirect("requestor-homepage")

            changes_request = f"{nature}: {complaint}" if complaint else nature

            try:
                jo_approver = UserApprovers.objects.get(user=request.user, module="Job Order", approver_role="Approver")

                if not jo_approver.approver:
                    error_message = "Invalid approver configuration. Please contact administrator."

                    if is_ajax:
                        return JsonResponse({'status': 'error', 'message': error_message})

                    messages.error(request, error_message)
                    return redirect("requestor-homepage")

            except UserApprovers.DoesNotExist:
                error_message = "No approver assigned for JO Request. Please contact administrator."

                if is_ajax:
                    return JsonResponse({'status': 'error', 'message': error_message})

                messages.error(request, error_message)
                return redirect("requestor-homepage")
            except Exception as e:
                error_message = f"Error finding approver: {str(e)}"

                if is_ajax:
                    return JsonResponse({'status': 'error', 'message': error_message})

                messages.error(request, "Failed to find approver. Please contact administrator.")
                return redirect("requestor-homepage")

            control_number = None
            try:
                if category == "green":
                    instance = GreenControlNumber.objects.create(prepared_by=request.user)
                    control_number = f'G-{instance.id:04}'
                elif category == "yellow":
                    instance = YellowControlNumber.objects.create(prepared_by=request.user)
                    control_number = f'Y-{instance.id:04}'
                elif category == "white":
                    instance = WhiteControlNumber.objects.create(prepared_by=request.user)
                    control_number = f'W-{instance.id:04}'
                elif category == "orange":
                    instance = OrangeControlNumber.objects.create(prepared_by=request.user)
                    control_number = f'O-{instance.id:04}'
                else:
                    error_message = "Invalid category selected."

                    if is_ajax:
                        return JsonResponse({'status': 'error', 'message': error_message})

                    messages.error(request, error_message)
                    return redirect("requestor-homepage")
            except Exception as e:
                error_message = f"Failed to generate control number: {str(e)}"

                if is_ajax:
                    return JsonResponse({'status': 'error', 'message': "Failed to generate control number"})

                messages.error(request, "Failed to generate control number")
                return redirect("requestor-homepage")

            try:
                new_request = JOLogsheet.objects.create(
                    jo_number=control_number,
                    prepared_by=request.user,
                    requestor=requestor,
                    line=line,  # Now passing the Line instance rather than the ID
                    jo_type=changes_request,
                    jo_tools=tooling,
                    jo_color=category.capitalize(),
                    details=details,
                )
            except Exception as e:
                error_message = f"JOLogsheet creation failed: {str(e)}"

                if is_ajax:
                    return JsonResponse({'status': 'error', 'message': "Failed to create job order"})

                messages.error(request, "Failed to create job order")
                return redirect("requestor-homepage")

            try:
                JORouting.objects.create(
                    jo_request=new_request,
                    approver=request.user,
                    status="submitted"
                )

                JORouting.objects.create(
                    jo_request=new_request,
                    approver=jo_approver.approver,
                    status="pending"
                )
            except Exception as e:
                error_message = f"Routing creation failed: {str(e)}"

                if is_ajax:
                    return JsonResponse({'status': 'error', 'message': "Failed to create approval routing"})

                messages.error(request, "Failed to create approval routing")
                return redirect("requestor-homepage")

            try:
                Notification.objects.create(
                    sender=request.user,
                    recipient=jo_approver.approver,
                    title="Approval",
                    message=f'You have a JO request from {request.user.name} awaiting your approval.'
                )
            except Exception as e:
                pass

            success_message = f"JO request {control_number} submitted successfully!"

            if is_ajax:
                return JsonResponse({
                    'status': 'success',
                    'message': success_message,
                    'control_number': control_number
                })

            messages.success(request, success_message)
            return redirect("requestor-homepage")

        # Form is not valid
        if not form.is_valid():
            error_message = "Please correct the errors in the form."

            if is_ajax:
                return JsonResponse({'status': 'error', 'message': error_message})

            messages.error(request, error_message)
            return redirect("requestor-homepage")

    # GET request - show form
    form = JobOrderRequestForm(user=request.user)
    context = {
        'form': form,
    }
    return render(request, 'joborder/create_jo_request.html', context)

@login_required(login_url="user-login")
@require_GET
def check_user_approver(request):
    """
    Check if the current user has a valid approver assigned for Job Order module
    """
    try:
        user_approver = UserApprovers.objects.filter(
            user=request.user,
            module="Job Order",
            approver_role="Approver"
        ).first()

        # Check if user has approver
        if not user_approver or not user_approver.approver:
            return JsonResponse({
                'status': 'error',
                'has_approver': False,
                'message': 'No approver assigned. Please contact administrator to assign an approver first.'
            })
        
        # Approver is assigned
        return JsonResponse({
            'status': 'success',
            'has_approver': True,
            'approver_name': user_approver.approver.name
        })

    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'has_approver': False,
            'message': f'Error checking approver: {str(e)}'
        })

@login_required(login_url="user-login")
def get_job_order_details(request, jo_id):
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    try:
        print(f"Looking for job order with ID: {jo_id}")

        job_order = JOLogsheet.objects.get(id=jo_id)
        print(f"Found job order: {job_order.jo_number}")

        routing_entries = JORouting.objects.filter(jo_request=job_order).order_by('request_at')
        print(f"Found {routing_entries.count()} routing entries")

        routing_data = []
        for entry in routing_entries:
            # Determine approver role
            approver_role = "Unknown"
            if hasattr(entry.approver, 'job_order_section_head') and entry.approver.job_order_section_head:
                approver_role = "Section Head"
            elif hasattr(entry.approver, 'job_order_department_head') and entry.approver.job_order_department_head:
                approver_role = "Department Head"
            elif hasattr(entry.approver, 'job_order_checker') and entry.approver.job_order_checker:
                approver_role = "Quality Control"
            elif hasattr(entry.approver, 'job_order_maintenance') and entry.approver.job_order_maintenance:
                approver_role = "Maintenance"
            
            routing_data.append({
                'approver': str(entry.approver),
                'approver_name': entry.approver.name if hasattr(entry.approver, 'name') else str(entry.approver),
                'approver_role': approver_role,
                'status': entry.status.capitalize(),
                'request_at': timezone.localtime(entry.request_at).strftime('%b %d, %Y, %I:%M %p') if entry.request_at else None,
                'approved_at': timezone.localtime(entry.approved_at).strftime('%b %d, %Y, %I:%M %p') if entry.approved_at else None,
                'remarks': entry.remarks if entry.remarks else ""
            })

        data = {
            'status': 'success',
            'id': job_order.id,
            'jo_number': job_order.jo_number,
            'category': job_order.jo_color,
            'line': job_order.line.line_name if job_order.line else None,
            'jo_status': job_order.status,
            'submitted_date': timezone.localtime(job_order.date_created).strftime('%b %d, %Y') if job_order.date_created else None,
            'tool': job_order.jo_tools,
            'nature': job_order.jo_type,
            'prepared_by': job_order.prepared_by.name if job_order.prepared_by else None,
            'requestor': job_order.requestor,
            'details': job_order.details,
            'in_charge': job_order.in_charge.name if job_order.in_charge else None,
            'date_received': timezone.localtime(job_order.date_received).strftime('%b %d, %Y') if job_order.date_received else None,
            'target_date': timezone.localtime(job_order.target_date).strftime('%b %d, %Y') if job_order.target_date else None,
            'date_complete': timezone.localtime(job_order.date_complete).strftime('%b %d, %Y') if job_order.date_complete else None,
            'action_taken': job_order.action_taken if job_order.action_taken else None,
            'target_date_reason': job_order.target_date_reason if job_order.target_date_reason else None,
            'priority_level': job_order.priority_level,
            'quality_matter': job_order.quality_matter,
            'date_of_completion': job_order.date_of_completion.strftime('%b %d, %Y') if job_order.date_of_completion else None,
            'is_creator': request.user == job_order.prepared_by,
            'routing': routing_data
        }


        return JsonResponse(data)

    except JOLogsheet.DoesNotExist:
        print(f"Job order with ID {jo_id} not found")
        return JsonResponse({'status': 'error', 'message': f'Job order with ID {jo_id} not found'})

    except Exception as e:
        print(f"Error retrieving job order details: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'status': 'error', 'message': f"Error retrieving job order details: {str(e)}"})

@login_required(login_url="user-login")
@require_GET
def search_job_orders(request):
    search_query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', 'all')
    
    if request.user.job_order_approver:
        job_orders = JOLogsheet.objects.filter(joRouting__status="pending", joRouting__approver=request.user).order_by("-joRouting__request_at")
    elif request.user.job_order_spectator:
        job_orders = JOLogsheet.objects.all().order_by("-date_created")
    elif request.user.job_order_maintenance:
        job_orders = JOLogsheet.objects.filter(in_charge=request.user).order_by("-date_created")
    elif request.user.job_order_checker:
        joRequests = JOLogsheet.objects.filter(
            Q(joRouting__approver=request.user, joRouting__status='pending') |
            Q(prepared_by=request.user) | Q(joRouting__approver=request.user, joRouting__status='approved')).order_by('-date_created')
    else:
        job_orders = JOLogsheet.objects.filter(prepared_by=request.user).order_by("-date_created")
    
    # Apply status filter
    if status_filter != 'all':
        status_map = {
            'closed': 'Closed',
            'routing': 'Routing',
            'cancelled': 'Cancelled',
            'rejected': 'Rejected',
            'completed': 'Completed'
        }
        if status_filter in status_map:
            job_orders = job_orders.filter(status=status_map[status_filter])
    
    # Apply search query if provided
    if search_query:
        job_orders = job_orders.filter(
            Q(jo_number__icontains=search_query) |
            Q(jo_color__icontains=search_query) |
            Q(jo_tools__icontains=search_query) |
            Q(jo_type__icontains=search_query) |
            Q(requestor__icontains=search_query) |
            Q(line__line_name__icontains=search_query) |
            Q(status__icontains=search_query) |
            Q(details__icontains=search_query)
        )
    
    # Order by date created (most recent first)
    job_orders = job_orders.order_by('-date_created')
    
    # Build response data
    results = []
    for jo in job_orders:
        results.append({
            'id': jo.id,
            'jo_number': jo.jo_number,
            'jo_color': jo.jo_color,
            'jo_tools': jo.jo_tools,
            'jo_type': jo.jo_type,
            'requestor': jo.requestor,
            'line': jo.line.line_name if jo.line else '',
            'status': jo.status,
            'date_created': jo.date_created.strftime('%b %d, %Y') if jo.date_created else '',
            'can_close': jo.status == 'Completed',
        })
    
    return JsonResponse({
        'status': 'success',
        'count': len(results),
        'results': results
    })

@login_required(login_url="user-login")
@require_POST
def review_job_order(request, jo_id):

    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    try:
        job_order = JOLogsheet.objects.get(id=jo_id)
        action = request.POST.get('action')
        remarks = request.POST.get('remarks', '').strip()
        
        if action not in ['approve', 'disapprove']:
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid action specified'
            })
        
        if action == 'disapprove' and not remarks:
            return JsonResponse({
                'status': 'error',
                'message': 'Remarks are required when disapproving'
            })
        
        current_routing = JORouting.objects.filter(
            jo_request=job_order,
            approver=request.user,
            status='pending'
        ).first()
        
        if not current_routing:
            return JsonResponse({
                'status': 'error',
                'message': 'You are not authorized to review this job order or it has already been processed'
            })
        
        if request.user.job_order_approver:
            modified_fields_json = request.POST.get('modified_fields', '{}')
            try:
                modified_fields = json.loads(modified_fields_json)
                
                if 'details' in modified_fields:
                    job_order.details = modified_fields['details']
                    job_order.save()
                
            except json.JSONDecodeError:
                pass 
        
        if action == 'approve':

            current_routing.status = 'approved'
            current_routing.remarks = remarks
            current_routing.approved_at = timezone.now()
            current_routing.save()
            
            if request.user.job_order_approver:
                quality_matter = request.POST.get('quality_matter', 'false').lower() == 'true'
                priority_level = request.POST.get('priority_level', 'Low')
                date_of_completion_str = request.POST.get('date_of_completion', '')
                
                job_order.quality_matter = quality_matter
                job_order.priority_level = priority_level
                
                if date_of_completion_str:
                    try:
                        from datetime import datetime
                        date_of_completion = datetime.strptime(date_of_completion_str, '%Y-%m-%d').date()
                        job_order.date_of_completion = date_of_completion
                    except (ValueError, TypeError) as e:
                        pass 
                
                job_order.save()
            
            elif request.user.job_order_checker:

                job_order.status = 'Checked'
                job_order.save()
                
                JORouting.objects.create(
                    jo_request=job_order,
                    approver=job_order.prepared_by,
                    status='pending'
                )
                
                try:
                    Notification.objects.create(
                        sender=request.user,
                        recipient=job_order.prepared_by,
                        title="Job Order Checked",
                        message=f'Your job order {job_order.jo_number} has been checked and is ready for closure.'
                    )
                except Exception as e:
                    pass
                
                return JsonResponse({
                    'status': 'success',
                    'message': f'Job Order {job_order.jo_number} has been checked successfully'
                })
            
            if job_order.jo_color and job_order.jo_color.lower() == 'orange':
                if request.user.job_order_pmd:
                    next_approvers = Users.objects.filter(
                        job_order_user=True,
                        job_order_facilitator=True
                    )
                else:
                    next_approvers = Users.objects.filter(
                        job_order_user=True,
                        job_order_pmd=True
                    )
            else:
                next_approvers = Users.objects.filter(
                    job_order_user=True,
                    job_order_facilitator=True
                )
            
            for approver in next_approvers:
                JORouting.objects.create(
                    jo_request=job_order,
                    approver=approver,
                    status='pending'
                )
                
                try:
                    Notification.objects.create(
                        sender=request.user,
                        recipient=approver,
                        title="Approval",
                        message=f'You have a JO request {job_order.jo_number} awaiting your approval.'
                    )
                except Exception as e:
                    pass
            
            return JsonResponse({
                'status': 'success',
                'message': f'Job Order {job_order.jo_number} has been approved successfully'
            })
            
        else:
            
            current_routing.status = 'disapproved'
            current_routing.remarks = remarks
            current_routing.approved_at = timezone.now()
            current_routing.save()
            
            job_order.status = 'Rejected'
            job_order.save()
            
            try:
                Notification.objects.create(
                    sender=request.user,
                    recipient=job_order.prepared_by,
                    title="Approval",
                    message=f'Your job order request {job_order.jo_number} has been disapproved by {request.user.name}. Reason: {remarks}'
                )
            except Exception as e:
                pass
            
            return JsonResponse({
                'status': 'success',
                'message': f'Job Order {job_order.jo_number} has been disapproved'
            })
            
    except JOLogsheet.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Job order not found'
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': f'An error occurred: {str(e)}'
        })

@login_required(login_url="user-login")
@require_GET
def check_preparer_has_checker(request, jo_id):
    """
    API endpoint to check if the preparer of a job order has an assigned checker.
    Only checks if quality_matter is True.
    Returns: JSON with status and has_checker boolean
    """
    try:
        # Get the job order
        job_order = JOLogsheet.objects.get(id=jo_id)
        
        # If quality_matter is False, no checker is required
        if not job_order.quality_matter:
            return JsonResponse({
                'status': 'success',
                'has_checker': True,  # Return True since checker is not required
                'quality_matter': False,
                'preparer_name': job_order.prepared_by.name if job_order.prepared_by else 'Unknown',
                'jo_number': job_order.jo_number
            })
        
        # If quality_matter is True, check if the preparer has a checker assigned
        has_checker = UserApprovers.objects.filter(
            user=job_order.prepared_by,
            module="Job Order",
            approver_role="Checker"
        ).exists()
        
        return JsonResponse({
            'status': 'success',
            'has_checker': has_checker,
            'quality_matter': True,
            'preparer_name': job_order.prepared_by.name if job_order.prepared_by else 'Unknown',
            'jo_number': job_order.jo_number
        })
    except JOLogsheet.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Job order not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': f'An error occurred: {str(e)}'
        }, status=500)

@login_required(login_url="user-login")
@require_POST
def complete_job_order(request, jo_id):

    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    # Check if user has maintenance role
    if not request.user.job_order_maintenance:
        return JsonResponse({
            'status': 'error',
            'message': 'You do not have permission to complete job orders'
        })
    
    try:
        job_order = JOLogsheet.objects.get(id=jo_id)
        
        # CRITICAL: Only check if preparer has a checker assigned when quality_matter is True
        if job_order.quality_matter:
            has_checker = UserApprovers.objects.filter(
                user=job_order.prepared_by,
                module="Job Order",
                approver_role="Checker"
            ).exists()
            
            if not has_checker:
                preparer_name = job_order.prepared_by.name if job_order.prepared_by else 'Unknown'
                return JsonResponse({
                    'status': 'error',
                    'message': f'Cannot complete this request. Please inform the requestor ({preparer_name}) to assign a checker to Admin before proceeding with completion.'
                })
        
        action_taken = request.POST.get('action_taken', '').strip()
        remarks = request.POST.get('remarks', '').strip()
        
        # Validate action_taken
        if not action_taken:
            return JsonResponse({
                'status': 'error',
                'message': 'Action taken is required'
            })
        
        # Update the current user's routing to approved
        current_routing = JORouting.objects.filter(
            jo_request=job_order,
            approver=request.user,
            status='pending'
        ).first()
        
        if not current_routing:
            return JsonResponse({
                'status': 'error',
                'message': 'No pending routing found for this job order'
            })
        
        # Update routing status
        current_routing.status = 'approved'
        current_routing.remarks = remarks if remarks else f'Completed: {action_taken}'
        current_routing.approved_at = timezone.now()
        current_routing.save()
        
        # Update job order
        job_order.action_taken = action_taken
        job_order.date_complete = timezone.now().date()
        job_order.status = 'Completed'
        job_order.save()
        
        # Check if quality_matter is True and create routing accordingly
        if job_order.quality_matter:
            try:
                # Get the checker for the job order's prepared_by user
                checker_approver = UserApprovers.objects.filter(
                    user=job_order.prepared_by,
                    module="Job Order",
                    approver_role="Checker"
                ).first()
                
                if checker_approver and checker_approver.approver:
                    # Create routing for checker
                    JORouting.objects.create(
                        jo_request=job_order,
                        approver=checker_approver.approver,
                        status='pending'
                    )
                    
                    # Notify the checker
                    try:
                        Notification.objects.create(
                            sender=request.user,
                            recipient=checker_approver.approver,
                            title="Quality Check Required",
                            message=f'Job order {job_order.jo_number} requires quality checking.'
                        )
                    except Exception as e:
                        pass
                else:
                    # If no checker found, log a warning but don't fail
                    print(f"Warning: No checker found for user {job_order.prepared_by.name}")
            except Exception as e:
                print(f"Error creating checker routing: {str(e)}")
                # Don't fail the completion if checker routing fails
        else:
            # If not quality matter, create routing for the prepared_by user (requestor)
            try:
                JORouting.objects.create(
                    jo_request=job_order,
                    approver=job_order.prepared_by,
                    status='pending'
                )
                
                # The notification to prepared_by is already sent below, so no duplicate needed
            except Exception as e:
                print(f"Error creating requestor routing: {str(e)}")
                # Don't fail the completion if routing creation fails
        
        # Notify the requestor
        try:
            Notification.objects.create(
                sender=request.user,
                recipient=job_order.prepared_by,
                title="Job Order Completed",
                message=f'Your job order {job_order.jo_number} has been completed by {request.user.name}.'
            )
        except Exception as e:
            pass  # Don't fail if notification fails
        
        return JsonResponse({
            'status': 'success',
            'message': f'Job Order {job_order.jo_number} has been completed successfully'
        })
        
    except JOLogsheet.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Job order not found'
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': f'An error occurred: {str(e)}'
        })

@login_required(login_url="user-login")
def cancel_jo_request(request, jo_id):
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    if request.method != 'POST':
        if is_ajax:
            return JsonResponse({'status': 'error', 'message': 'Invalid request method'})
        messages.error(request, "Invalid request method")
        return redirect("requestor-homepage")

    try:

        job_order = JOLogsheet.objects.get(id=jo_id)
        last_routing = JORouting.objects.filter(jo_request=job_order)
        last_routing.delete()

        if job_order.prepared_by != request.user:
            if is_ajax:
                return JsonResponse({'status': 'error', 'message': 'You can only cancel your own job order requests'})
            messages.error(request, "You can only cancel your own job order requests")
            return redirect("requestor-homepage")

        if job_order.status != 'Routing':
            if is_ajax:
                return JsonResponse({'status': 'error', 'message': 'This job order cannot be cancelled in its current state'})
            messages.error(request, "This job order cannot be cancelled in its current state")
            return redirect("requestor-homepage")

        job_order.status = 'Cancelled'
        job_order.save()


        JORouting.objects.create(
            jo_request=job_order,
            approver=request.user,
            status='Cancelled',
            remarks=request.POST.get('remarks', '')
        )

        if is_ajax:
            return JsonResponse({
                'status': 'success',
                'message': f'Job order {job_order.jo_number} has been cancelled'
            })

        messages.success(request, f'Job order {job_order.jo_number} has been cancelled')
        return redirect("requestor-homepage")

    except JOLogsheet.DoesNotExist:
        if is_ajax:
            return JsonResponse({'status': 'error', 'message': 'Job order not found'})
        messages.error(request, "Job order not found")
        return redirect("requestor-homepage")

    except Exception as e:
        if is_ajax:
            return JsonResponse({'status': 'error', 'message': f'An error occurred: {str(e)}'})
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect("requestor-homepage")

@login_required(login_url="user-login")
def close_transaction(request, jo_id):
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    if request.method != 'POST':
        if is_ajax:
            return JsonResponse({'status': 'error', 'message': 'Invalid request method'})
        messages.error(request, "Invalid request method")
        return redirect("requestor-homepage")

    try:
        job_order = get_object_or_404(JOLogsheet, id=jo_id)

        if job_order.prepared_by != request.user:
            if is_ajax:
                return JsonResponse({'status': 'error', 'message': 'You can only close your own job order transactions'})
            messages.error(request, "You can only close your own job order transactions")
            return redirect("requestor-homepage")

        if job_order.status not in ['Checked', 'Completed', 'On Hold']:
            if is_ajax:
                return JsonResponse({'status': 'error', 'message': 'This job order cannot be closed in its current state'})
            messages.error(request, "This job order cannot be closed in its current state")
            return redirect("requestor-homepage")

        # Check if remarks are required for On Hold status
        remarks = request.POST.get('remarks', '').strip()
        if job_order.status == 'On Hold' and not remarks:
            if is_ajax:
                return JsonResponse({'status': 'error', 'message': 'Feedback is required when closing a job order that is on hold'})
            messages.error(request, "Feedback is required when closing a job order that is on hold")
            return redirect("requestor-homepage")

        try:
            routing_entry = JORouting.objects.filter(
                jo_request=job_order,
                approver=request.user,
            ).last()

            # Update routing entry fields
            routing_entry.status = "Approved"
            if remarks:
                routing_entry.remarks = remarks
            routing_entry.approved_at = timezone.now()
            routing_entry.save()

            # Update job order status to Closed
            job_order.status = "Closed"
            job_order.save()

            if is_ajax:
                return JsonResponse({
                    'status': 'success',
                    'message': f'Job order {job_order.jo_number} has been closed successfully'
                })

            messages.success(request, f'Job order {job_order.jo_number} has been closed successfully')
            return redirect("requestor-homepage")

        except JORouting.DoesNotExist:
            if is_ajax:
                return JsonResponse({'status': 'error', 'message': 'Routing entry does not exist'})
            messages.error(request, "Routing entry does not exist")
            return redirect("requestor-homepage")

    except Exception as e:
        if is_ajax:
            return JsonResponse({'status': 'error', 'message': f'An error occurred: {str(e)}'})
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect("requestor-homepage")

@login_required(login_url="user-login")
def hold_request(request, jo_id):
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    if request.method != 'POST':
        if is_ajax:
            return JsonResponse({'status': 'error', 'message': 'Invalid request method'})
        messages.error(request, "Invalid request method")
        return redirect("requestor-homepage")

    try:
        job_order = get_object_or_404(JOLogsheet, id=jo_id)

        # Check if user is the creator
        if job_order.prepared_by != request.user:
            if is_ajax:
                return JsonResponse({'status': 'error', 'message': 'You can only hold your own job order requests'})
            messages.error(request, "You can only hold your own job order requests")
            return redirect("requestor-homepage")

        # Check if status is Completed or Checked (not already On Hold)
        if job_order.status not in ['Checked', 'Completed']:
            if is_ajax:
                return JsonResponse({'status': 'error', 'message': 'This job order cannot be put on hold in its current state'})
            messages.error(request, "This job order cannot be put on hold in its current state")
            return redirect("requestor-homepage")

        # Validate remarks
        remarks = request.POST.get('remarks', '').strip()
        if not remarks:
            if is_ajax:
                return JsonResponse({'status': 'error', 'message': 'Remarks are required when putting a request on hold'})
            messages.error(request, "Remarks are required when putting a request on hold")
            return redirect("requestor-homepage")

        try:
            # Get the latest routing entry for this user
            routing_entry = JORouting.objects.filter(
                jo_request=job_order,
                approver=request.user,
            ).last()

            if routing_entry:
                # Update the routing entry with remarks
                routing_entry.remarks = remarks
                routing_entry.save()

            # Update job order status to On Hold
            job_order.status = "On Hold"
            job_order.save()

            if is_ajax:
                return JsonResponse({
                    'status': 'success',
                    'message': f'Job order {job_order.jo_number} has been put on hold successfully'
                })

            messages.success(request, f'Job order {job_order.jo_number} has been put on hold successfully')
            return redirect("requestor-homepage")

        except JORouting.DoesNotExist:
            if is_ajax:
                return JsonResponse({'status': 'error', 'message': 'Routing entry does not exist'})
            messages.error(request, "Routing entry does not exist")
            return redirect("requestor-homepage")

    except Exception as e:
        if is_ajax:
            return JsonResponse({'status': 'error', 'message': f'An error occurred: {str(e)}'})
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect("requestor-homepage")

# FACILITATOR VIEW
@login_required(login_url="user-login")
def job_order_facilitator(request):
    current_month = now().month
    current_year = now().year
    today = timezone.now().date()
    maintenance_personnel = Users.objects.filter(job_order_maintenance=True)

    pending_assignments_list = JORouting.objects.filter(jo_request__in_charge__isnull=True, approver=request.user).order_by('-request_at')[:10]
    for jo in pending_assignments_list:
        if jo.jo_request and jo.jo_request.date_created:
            jo.is_overdue_warning = (timezone.now() - jo.jo_request.date_created) >= timedelta(days=1)
            jo.is_overdue_critical = (timezone.now() - jo.jo_request.date_created) >= timedelta(days=2)
        else:
            jo.is_overdue_warning = False
            jo.is_overdue_critical = False

    # Get search and filter parameters
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', 'all')

    # Base queryset with ordering
    job_orders_queryset = (
        JOLogsheet.objects.annotate(
            in_charge_null=Case(
                When(in_charge__isnull=True, then=Value(True)),
                default=Value(False),
                output_field=BooleanField(),
            ),
            priority_order=Case(
                When(priority_level="Urgent", then=Value(1)),
                When(priority_level="High", then=Value(2)),
                When(priority_level="Medium", then=Value(3)),
                When(priority_level="Low", then=Value(4)),
                default=Value(5),
                output_field=IntegerField(),
            ),
        )
        .order_by(
            "-in_charge_null",
            "priority_order", 
            "date_of_completion", 
            "-date_created",  
        )
    )

    # Apply search filter - searches across all records in the database
    if search_query:
        job_orders_queryset = job_orders_queryset.filter(
            Q(jo_number__icontains=search_query) |
            Q(line__line_name__icontains=search_query) |
            Q(jo_type__icontains=search_query) |
            Q(jo_tools__icontains=search_query) |
            Q(requestor__icontains=search_query) |
            Q(jo_color__icontains=search_query) |
            Q(status__icontains=search_query)
        )

    # Apply status filter
    if status_filter and status_filter != 'all':
        job_orders_queryset = job_orders_queryset.filter(status__iexact=status_filter)

    # Pagination - 10 items per page
    page_number = request.GET.get('page', 1)
    paginator = Paginator(job_orders_queryset, 10)
    all_requests = paginator.get_page(page_number)

    # Add has_pending_routing attribute to each request
    for request_obj in all_requests:
        request_obj.has_pending_routing = request_obj.joRouting.filter(approver=request.user, status='pending').exists()
        request_obj.is_reviewable = not request_obj.in_charge and request_obj.status not in ["Rejected", "Cancelled"]
        
    # Calculate statistics after all other processing
    try:
        joRequestsCount = JOLogsheet.objects.filter(status__in=["Routing", "Completed", "Assigned"]).count()
        overdueCount = JOLogsheet.objects.filter(in_charge__isnull=False, status__in=["Closed", "Completed", "Checked"]).count()
        assignJOCount = JOLogsheet.objects.filter(status="Routing").count()
        pendingJOCount = JOLogsheet.objects.filter(in_charge__isnull=False, status="Closed").count()
    except Exception as e:
        joRequestsCount = 0
        overdueCount = 0
        assignJOCount = 0
        pendingJOCount = 0

    context = {
        'joRequestsCount': joRequestsCount,
        'overdueCount': overdueCount,
        'assignJOCount': assignJOCount,
        'pendingJOCount': pendingJOCount,
        'pending_assignments_list': pending_assignments_list,
        'all_requests': all_requests,
        'maintenance_personnel': maintenance_personnel,
        'search_query': search_query,
        'status_filter': status_filter,
    }

    return render(request, 'joborder/jo-facilitator.html', context)

@login_required(login_url="user-login")
def job_order_analytics(request):
    try:
        from datetime import timedelta
        import calendar
        
        period = request.GET.get('period', 'thismonth')
        today = timezone.now()
        default_labels = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun']
        default_data = [0, 0, 0, 0, 0, 0]

        labels = []
        total_by_month = []
        completed_by_month = []
        assigned_by_month = []

        # For "This Month" or "Last Month", show daily data
        if period == 'thismonth' or period == '1month':
            if period == 'thismonth':
                # For "This Month", use the current month
                target_month = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            else:
                # For "Last Month", go back one month
                first_of_this_month = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                prev_month = first_of_this_month.month - 1 or 12
                prev_year = first_of_this_month.year if first_of_this_month.month > 1 else first_of_this_month.year - 1
                target_month = first_of_this_month.replace(year=prev_year, month=prev_month, day=1)
            
            # Get the number of days in the target month
            days_in_month = calendar.monthrange(target_month.year, target_month.month)[1]
            
            # Calculate the first day of the next month for range checking
            if target_month.month == 12:
                next_month = target_month.replace(year=target_month.year+1, month=1, day=1)
            else:
                next_month = target_month.replace(month=target_month.month+1, day=1)
            
            # Loop through each day of the month
            for day in range(1, days_in_month + 1):
                day_start = target_month.replace(day=day, hour=0, minute=0, second=0, microsecond=0)
                day_end = day_start + timedelta(days=1)
                
                # Label is just the day number
                labels.append(str(day))
                
                try:
                    # Count total job orders created on this day
                    total_count = JOLogsheet.objects.filter(
                        date_created__gte=day_start,
                        date_created__lt=day_end
                    ).count()
                except Exception as db_error:
                    print(f"Database error when fetching total count for day {day}: {str(db_error)}")
                    total_count = 0
                
                try:
                    # Count completed job orders (excluding Assigned status)
                    completed_count = JOLogsheet.objects.filter(
                        date_created__gte=day_start,
                        date_created__lt=day_end,
                        status__in=['Completed', 'Checked', 'Closed']
                    ).count()
                except Exception as db_error:
                    print(f"Database error when fetching completed count for day {day}: {str(db_error)}")
                    completed_count = 0
                
                try:
                    # Count assigned job orders separately
                    assigned_count = JOLogsheet.objects.filter(
                        date_created__gte=day_start,
                        date_created__lt=day_end,
                        status='Assigned'
                    ).count()
                except Exception as db_error:
                    print(f"Database error when fetching assigned count for day {day}: {str(db_error)}")
                    assigned_count = 0
                
                total_by_month.append(total_count)
                completed_by_month.append(completed_count)
                assigned_by_month.append(assigned_count)
        
        # For "Last 3 Months" or "Last 6 Months", show monthly data
        else:
            if period == '3month':
                months = 3
            else:
                months = 6
            
            # Calculate the first day of the current month
            first_of_this_month = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            # Calculate the first month to include
            start_month = first_of_this_month
            for _ in range(months - 1):
                prev_month = start_month.month - 1 or 12
                prev_year = start_month.year if start_month.month > 1 else start_month.year - 1
                start_month = start_month.replace(year=prev_year, month=prev_month, day=1)
            
            # Now, for each month, calculate the start and end
            for i in range(months):
                # Move forward i months from start_month
                month = (start_month.month - 1 + i) % 12 + 1
                year = start_month.year + ((start_month.month - 1 + i) // 12)
                month_start = start_month.replace(year=year, month=month, day=1)
                # Calculate the first day of the next month
                if month == 12:
                    next_month = month_start.replace(year=year+1, month=1, day=1)
                else:
                    next_month = month_start.replace(month=month+1, day=1)
                month_end = next_month
                month_label = month_start.strftime('%b')
                try:
                    # Include "Assigned" status in total count
                    total_count = JOLogsheet.objects.filter(
                        date_created__gte=month_start,
                        date_created__lt=month_end
                    ).count()
                except Exception as db_error:
                    print(f"Database error when fetching total count: {str(db_error)}")
                    total_count = 0
                try:
                    # Count completed job orders (excluding Assigned status)
                    completed_count = JOLogsheet.objects.filter(
                        date_created__gte=month_start,
                        date_created__lt=month_end,
                        status__in=['Completed', 'Checked', 'Closed']
                    ).count()
                except Exception as db_error:
                    print(f"Database error when fetching completed count: {str(db_error)}")
                    completed_count = 0
                
                try:
                    # Count assigned job orders separately
                    assigned_count = JOLogsheet.objects.filter(
                        date_created__gte=month_start,
                        date_created__lt=month_end,
                        status='Assigned'
                    ).count()
                except Exception as db_error:
                    print(f"Database error when fetching assigned count: {str(db_error)}")
                    assigned_count = 0
                
                labels.append(month_label)
                total_by_month.append(total_count)
                completed_by_month.append(completed_count)
                assigned_by_month.append(assigned_count)

        # Fallback if no data
        if not labels:
            labels = default_labels
            total_by_month = default_data
            completed_by_month = default_data
            assigned_by_month = default_data

        data = {
            'status': 'success',
            'period': period,
            'labels': labels,
            'total_by_month': total_by_month,
            'completed_by_month': completed_by_month,
            'assigned_by_month': assigned_by_month
        }
        return JsonResponse(data)
    except Exception as e:
        import traceback
        print(f"Error in job_order_analytics: {str(e)}")
        traceback.print_exc()
        return JsonResponse({
            'status': 'success',
            'period': request.GET.get('period', 'thismonth'),
            'labels': default_labels,
            'total_by_month': default_data,
            'completed_by_month': default_data,
            'assigned_by_month': default_data,
            'error_logged': True
        })

@login_required(login_url="user-login")
def maintenance_workload(request):
    try:
        maintenance_staff = Users.objects.filter(job_order_maintenance=True)

        workload_data = []
        for staff in maintenance_staff:
            try:

                active_tasks = JOLogsheet.objects.filter(
                    in_charge=staff,
                    status='Assigned'
                ).count()

                max_capacity = 20

                if active_tasks > max_capacity:
                    max_capacity = active_tasks

                if max_capacity > 0:
                    workload_percentage = min(int((active_tasks / max_capacity) * 100), 100)
                else:
                    workload_percentage = 0

                workload_data.append({
                    'name': f"{staff.name}" if hasattr(staff, 'name') and staff.name else f"Staff #{staff.id}",
                    'active_tasks': active_tasks,
                    'workload_percentage': workload_percentage
                })

            except Exception as staff_error:
                import traceback
                print(f"Error processing workload for staff {staff.id}: {str(staff_error)}")
                traceback.print_exc()

                workload_data.append({
                    'name': f"Staff #{staff.id}" if hasattr(staff, 'id') else "Unknown Staff",
                    'active_tasks': 0,
                    'workload_percentage': 0,
                    'error': True
                })

        return JsonResponse({
            'status': 'success',
            'workload_data': workload_data
        })

    except Exception as e:
        # Log the error
        import traceback
        print(f"Error in maintenance_workload: {str(e)}")
        traceback.print_exc()

        # Return error response with empty data
        return JsonResponse({
            'status': 'success',  # Still return success to prevent UI breaking
            'workload_data': [],
            'error_logged': True
        })

@login_required(login_url="user-login")
@require_POST
def assign_person_in_charge(request):
    try:
        jo_id = request.POST.get('jo_id')
        assignee_id = request.POST.get('assignee_id')

        if not jo_id or not assignee_id:
            messages.error(request, 'Missing required information for assignment.')
            return redirect('facilitator')

        job_order = JOLogsheet.objects.get(id=jo_id)
        assignee = Users.objects.get(id=assignee_id)

        job_order.in_charge = assignee
        job_order.status = "Assigned"
        job_order.date_received=timezone.now()
        job_order.save()

        routing_entry = JORouting.objects.filter(jo_request=job_order, approver=request.user).first()
        routing_entry.status='approved'
        routing_entry.remarks=f'Request assigned to {assignee.name}.'
        routing_entry.approved_at=timezone.now()
        routing_entry.save()

        new_routing = JORouting.objects.create(
            jo_request=job_order,
            approver=assignee,
            status='pending'
        )

        messages.success(request, f'Job Order {job_order.jo_number} successfully assigned to {assignee.name}.')
        return redirect('facilitator')

    except JOLogsheet.DoesNotExist:
        messages.error(request, f'Job Order with ID {jo_id} not found.')

    except Users.DoesNotExist:
        messages.error(request, f'User with ID {assignee_id} not found.')

    except Exception as e:
        messages.error(request, f'An error occurred: {str(e)}')
    return redirect('facilitator')

@login_required(login_url="user-login")
def export_job_orders(request):
    if request.method != 'POST':
        print("Method not allowed, expected POST")
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        period = request.POST.get('period', 'custom')
        date_from = request.POST.get('date_from')
        date_to = request.POST.get('date_to')
        status_filters = request.POST.getlist('status')
        category_filters = request.POST.getlist('category')

        # Helper function to get custom month range (24th to 24th)
        def get_custom_month_range(target_date):
            """
            Calculate custom month range from 24th to 24th (ending on 24th).
            For October: Sept 24 to Oct 24
            """
            target_day = target_date.day
            
            if target_day < 24:
                # We're before the 24th, so the period started last month on the 24th
                if target_date.month == 1:
                    start_date = target_date.replace(year=target_date.year - 1, month=12, day=24, hour=0, minute=0, second=0, microsecond=0)
                else:
                    start_date = target_date.replace(month=target_date.month - 1, day=24, hour=0, minute=0, second=0, microsecond=0)
                end_date = target_date.replace(day=24, hour=23, minute=59, second=59, microsecond=999999)
            else:
                # We're on or after the 24th, so the period starts this month on the 24th
                start_date = target_date.replace(day=24, hour=0, minute=0, second=0, microsecond=0)
                if target_date.month == 12:
                    end_date = target_date.replace(year=target_date.year + 1, month=1, day=24, hour=23, minute=59, second=59, microsecond=999999)
                else:
                    end_date = target_date.replace(month=target_date.month + 1, day=24, hour=23, minute=59, second=59, microsecond=999999)
            
            return timezone.make_aware(start_date), timezone.make_aware(end_date)

        # Determine date range based on period selection
        if period == 'current_month':
            # Use current month with 24th-to-24th logic
            today = datetime.now()
            start_date, end_date = get_custom_month_range(today)
            # Cap end_date at current time if it extends into the future
            now = timezone.now()
            if end_date > now:
                end_date = now
        elif period == 'last_month':
            # Use last month with 24th-to-24th logic
            today = datetime.now()
            if today.month == 1:
                prev_month = today.replace(year=today.year - 1, month=12, day=15)
            else:
                prev_month = today.replace(month=today.month - 1, day=15)
            start_date, end_date = get_custom_month_range(prev_month)
        else:
            # Custom date range
            if date_from:
                try:
                    start_date = datetime.strptime(date_from, '%Y-%m-%d')
                    start_date = timezone.make_aware(start_date.replace(hour=0, minute=0, second=0))
                except ValueError:
                    return JsonResponse({'error': 'Invalid start date format'}, status=400)
            else:
                start_date = timezone.now().replace(hour=0, minute=0, second=0) - timezone.timedelta(days=30)

            if date_to:
                try:
                    end_date = datetime.strptime(date_to, '%Y-%m-%d')
                    end_date = timezone.make_aware(end_date.replace(hour=23, minute=59, second=59))
                except ValueError:
                    return JsonResponse({'error': 'Invalid end date format'}, status=400)
            else:
                end_date = timezone.now().replace(hour=23, minute=59, second=59)

        if 'all' in status_filters:
            status_filters = [status[0] for status in JOLogsheet.STATUS_CHOICES]

        if 'all' in category_filters:
            category_filters = list(JOLogsheet.objects.values_list('jo_color', flat=True).distinct())

        query_filters = {
            'date_created__gte': start_date,
            'date_created__lte': end_date
        }

        if status_filters and 'all' not in request.POST.getlist('status'):
            query_filters['status__in'] = status_filters

        if category_filters and 'all' not in request.POST.getlist('category'):
            query_filters['jo_color__in'] = category_filters

        job_orders = JOLogsheet.objects.filter(**query_filters).order_by('-date_created')

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Job Orders'

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3366FF", end_color="3366FF", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        cell_alignment = Alignment(vertical='top', wrap_text=True)

        status_styles = {
            'Routing': {
                'fill': PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),
                'font': Font(color="FFC107")
            },
            'Completed': {
                'fill': PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
                'font': Font(color="4CAF50")
            },
            'Checked': {
                'fill': PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
                'font': Font(color="2196F3")
            },
            'Cancelled': {
                'fill': PatternFill(start_color="FBE9E7", end_color="FBE9E7", fill_type="solid"),
                'font': Font(color="FF5722")
            },
            'Closed': {
                'fill': PatternFill(start_color="E0F2F1", end_color="E0F2F1", fill_type="solid"),
                'font': Font(color="009688")
            },
            'Rejected': {
                'fill': PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
                'font': Font(color="F44336")
            }
        }

        category_styles = {
            'green': {
                'fill': PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
                'font': Font(color="4CAF50")
            },
            'yellow': {
                'fill': PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),
                'font': Font(color="FFC107")
            },
            'white': {
                'fill': PatternFill(start_color="ECEFF1", end_color="ECEFF1", fill_type="solid"),
                'font': Font(color="607D8B")
            },
            'orange': {
                'fill': PatternFill(start_color="FBE9E7", end_color="FBE9E7", fill_type="solid"),
                'font': Font(color="FF5722")
            }
        }

        column_widths = {
            'A': 15,  # Date Submitted
            'B': 15,  # JO Number
            'C': 20,  # Requested By
            'D': 15,  # Line
            'E': 12,  # Category
            'F': 15,  # Tool
            'G': 15,  # Nature of Changes
            'H': 30,  # Details
            'I': 12,  # Status
            'J': 20,  # Person In Charge
            'K': 15,  # Date Received
            'L': 15,  # Expected Date
            'M': 15,  # Date Completed
            'N': 30,  # Action Taken
            'O': 30,  # Remarks
        }

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        headers = [
            'Date Submitted',
            'JO Number',
            'Requested By',
            'Line',
            'Category',
            'Tool',
            'Nature of Changes',
            'Details',
            'Status',
            'Person In Charge',
            'Date Received',
            'Expected Date',
            'Date Completed',
            'Action Taken',
            'Remarks',
        ]

        for col, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_idx, jo in enumerate(job_orders, start=2):
            # Date Created
            formatted_date = jo.date_created.strftime('%Y-%m-%d') if jo.date_created else ''
            cell = worksheet.cell(row=row_idx, column=1, value=formatted_date)
            cell.border = thin_border
            cell.alignment = cell_alignment

            # JO Number
            cell = worksheet.cell(row=row_idx, column=2, value=jo.jo_number)
            cell.border = thin_border
            cell.alignment = cell_alignment

            # Requested By
            cell = worksheet.cell(row=row_idx, column=3, value=jo.requestor)
            cell.border = thin_border
            cell.alignment = cell_alignment

            # Line
            cell = worksheet.cell(row=row_idx, column=4, value=jo.line.line_name if jo.line else '')
            cell.border = thin_border
            cell.alignment = cell_alignment

            # Category with color formatting
            cell = worksheet.cell(row=row_idx, column=5, value=jo.jo_color)
            if jo.jo_color and jo.jo_color.lower() in category_styles:
                style = category_styles[jo.jo_color.lower()]
                cell.fill = style['fill']
                cell.font = style['font']
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Tool
            cell = worksheet.cell(row=row_idx, column=6, value=jo.jo_tools)
            cell.border = thin_border
            cell.alignment = cell_alignment

            # Nature
            cell = worksheet.cell(row=row_idx, column=7, value=jo.jo_type)
            cell.border = thin_border
            cell.alignment = cell_alignment

            # Details
            cell = worksheet.cell(row=row_idx, column=8, value=jo.details)
            cell.border = thin_border
            cell.alignment = cell_alignment

            # Status with color formatting
            cell = worksheet.cell(row=row_idx, column=9, value=jo.status)
            if jo.status in status_styles:
                style = status_styles[jo.status]
                cell.fill = style['fill']
                cell.font = style['font']
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Person In Charge
            in_charge_name = ''
            if jo.in_charge:
                try:
                    if hasattr(jo.in_charge, 'name') and jo.in_charge.name:
                        in_charge_name = jo.in_charge.name
                    else:
                        in_charge_name = jo.in_charge.username
                except:
                    in_charge_name = str(jo.in_charge.id) if jo.in_charge.id else 'Unknown'

            cell = worksheet.cell(row=row_idx, column=10, value=in_charge_name)
            cell.border = thin_border
            cell.alignment = cell_alignment

            # Date Received
            if jo.date_received:
                formatted_date = jo.date_received.strftime('%Y-%m-%d')
                cell = worksheet.cell(row=row_idx, column=11, value=formatted_date)
            else:
                cell = worksheet.cell(row=row_idx, column=11, value='')
            cell.border = thin_border
            cell.alignment = cell_alignment

            # Expected Date
            if jo.target_date:
                formatted_date = jo.target_date.strftime('%Y-%m-%d')
                cell = worksheet.cell(row=row_idx, column=12, value=formatted_date)
            else:
                cell = worksheet.cell(row=row_idx, column=12, value='')
            cell.border = thin_border
            cell.alignment = cell_alignment

            # Date Completed
            if jo.date_complete:
                formatted_date = jo.date_complete.strftime('%Y-%m-%d')
                cell = worksheet.cell(row=row_idx, column=13, value=formatted_date)
            else:
                cell = worksheet.cell(row=row_idx, column=13, value='')
            cell.border = thin_border
            cell.alignment = cell_alignment

            # Action Taken
            cell = worksheet.cell(row=row_idx, column=14, value=jo.action_taken if jo.action_taken else '')
            cell.border = thin_border
            cell.alignment = cell_alignment

            # Remarks
            cell = worksheet.cell(row=row_idx, column=15, value=jo.target_date_reason if jo.target_date_reason else '')
            cell.border = thin_border
            cell.alignment = cell_alignment

        summary_sheet = workbook.create_sheet(title="Summary")

        summary_sheet.column_dimensions['A'].width = 40
        summary_sheet.column_dimensions['B'].width = 15

        # Title
        title_cell = summary_sheet.cell(row=1, column=1, value='Job Orders Export Summary')
        title_cell.font = Font(bold=True, size=14)
        summary_sheet.merge_cells('A1:B1')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Date range
        summary_sheet.cell(row=3, column=1, value='Date Range:').font = Font(bold=True)
        summary_sheet.cell(row=3, column=2, value=f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")

        # Total count
        summary_sheet.cell(row=4, column=1, value='Total Job Orders:').font = Font(bold=True)
        summary_sheet.cell(row=4, column=2, value=len(job_orders))

        # Calculate metrics based on filtered job_orders (filtered by date range)
        active_jo_count = job_orders.filter(status__in=["Routing", "Assigned"]).count()
        routing_requests_count = job_orders.filter(status="Routing").count()
        in_progress_count = job_orders.filter(in_charge__isnull=False, status="Assigned").count()
        completed_requests_count = job_orders.filter(in_charge__isnull=False, status__in=["Closed", "Completed", "Checked"]).count()
        compliance_rate = (completed_requests_count / active_jo_count * 100) if active_jo_count > 0 else 0

        current_row = 6

        # ===== TABLE 1: STATISTIC SUMMARY =====
        summary_sheet.cell(row=current_row, column=1, value='Statistic Summary').font = Font(bold=True, size=12)
        current_row += 1
        
        # Table headers with blue background and white font
        header_cell_1 = summary_sheet.cell(row=current_row, column=1, value='Metric')
        header_cell_1.font = Font(bold=True, color="FFFFFF")
        header_cell_1.fill = PatternFill(start_color="3366FF", end_color="3366FF", fill_type="solid")
        header_cell_1.alignment = Alignment(horizontal='center', vertical='center')
        header_cell_1.border = thin_border
        
        header_cell_2 = summary_sheet.cell(row=current_row, column=2, value='Count')
        header_cell_2.font = Font(bold=True, color="FFFFFF")
        header_cell_2.fill = PatternFill(start_color="3366FF", end_color="3366FF", fill_type="solid")
        header_cell_2.alignment = Alignment(horizontal='center', vertical='center')
        header_cell_2.border = thin_border

        current_row += 1

        # Metrics data with borders
        metrics_data = [
            ('Active JO', active_jo_count),
            ('Routing Requests', routing_requests_count),
            ('In Progress Requests', in_progress_count),
            ('Completed Requests', completed_requests_count),
            ('Compliance Rate', f"{compliance_rate:.2f}%")
        ]

        for metric_name, metric_value in metrics_data:
            cell_name = summary_sheet.cell(row=current_row, column=1, value=metric_name)
            cell_name.border = thin_border
            cell_name.alignment = Alignment(vertical='center')
            
            cell_value = summary_sheet.cell(row=current_row, column=2, value=metric_value)
            cell_value.border = thin_border
            cell_value.alignment = Alignment(horizontal='center', vertical='center')
            
            current_row += 1

        current_row += 2  # Add spacing

        # ===== TABLE 2: STATUS BREAKDOWN =====
        summary_sheet.cell(row=current_row, column=1, value='Status Breakdown').font = Font(bold=True, size=12)
        current_row += 1

        # Status table headers
        status_header_1 = summary_sheet.cell(row=current_row, column=1, value='Status')
        status_header_1.font = Font(bold=True, color="FFFFFF")
        status_header_1.fill = PatternFill(start_color="3366FF", end_color="3366FF", fill_type="solid")
        status_header_1.alignment = Alignment(horizontal='center', vertical='center')
        status_header_1.border = thin_border
        
        status_header_2 = summary_sheet.cell(row=current_row, column=2, value='Count')
        status_header_2.font = Font(bold=True, color="FFFFFF")
        status_header_2.fill = PatternFill(start_color="3366FF", end_color="3366FF", fill_type="solid")
        status_header_2.alignment = Alignment(horizontal='center', vertical='center')
        status_header_2.border = thin_border

        current_row += 1

        # Status breakdown data
        status_counts = {}
        for jo in job_orders:
            status_counts[jo.status] = status_counts.get(jo.status, 0) + 1

        for status, count in status_counts.items():
            cell_status = summary_sheet.cell(row=current_row, column=1, value=status)
            cell_status.border = thin_border
            cell_status.alignment = Alignment(vertical='center')
            if status in status_styles:
                cell_status.font = status_styles[status]['font']
            
            cell_count = summary_sheet.cell(row=current_row, column=2, value=count)
            cell_count.border = thin_border
            cell_count.alignment = Alignment(horizontal='center', vertical='center')
            
            current_row += 1

        current_row += 2  # Add spacing

        # ===== TABLE 3: PRIORITY LEVEL BREAKDOWN =====
        summary_sheet.cell(row=current_row, column=1, value='Priority Level Breakdown').font = Font(bold=True, size=12)
        current_row += 1

        # Priority table headers
        priority_header_1 = summary_sheet.cell(row=current_row, column=1, value='Priority Level')
        priority_header_1.font = Font(bold=True, color="FFFFFF")
        priority_header_1.fill = PatternFill(start_color="3366FF", end_color="3366FF", fill_type="solid")
        priority_header_1.alignment = Alignment(horizontal='center', vertical='center')
        priority_header_1.border = thin_border
        
        priority_header_2 = summary_sheet.cell(row=current_row, column=2, value='Count')
        priority_header_2.font = Font(bold=True, color="FFFFFF")
        priority_header_2.fill = PatternFill(start_color="3366FF", end_color="3366FF", fill_type="solid")
        priority_header_2.alignment = Alignment(horizontal='center', vertical='center')
        priority_header_2.border = thin_border

        current_row += 1

        # Priority breakdown data
        priority_counts = {}
        for jo in job_orders:
            priority = jo.priority_level if jo.priority_level else 'Not Set'
            priority_counts[priority] = priority_counts.get(priority, 0) + 1

        # Sort priorities in order: Urgent, High, Medium, Low, Not Set
        priority_order = ['Urgent', 'High', 'Medium', 'Low', 'Not Set']
        for priority in priority_order:
            if priority in priority_counts:
                cell_priority = summary_sheet.cell(row=current_row, column=1, value=priority)
                cell_priority.border = thin_border
                cell_priority.alignment = Alignment(vertical='center')
                
                cell_count = summary_sheet.cell(row=current_row, column=2, value=priority_counts[priority])
                cell_count.border = thin_border
                cell_count.alignment = Alignment(horizontal='center', vertical='center')
                
                current_row += 1

        current_row += 2  # Add spacing

        # ===== TABLE 4: CATEGORY BREAKDOWN =====
        summary_sheet.cell(row=current_row, column=1, value='Category Breakdown').font = Font(bold=True, size=12)
        current_row += 1

        # Category table headers
        category_header_1 = summary_sheet.cell(row=current_row, column=1, value='Category')
        category_header_1.font = Font(bold=True, color="FFFFFF")
        category_header_1.fill = PatternFill(start_color="3366FF", end_color="3366FF", fill_type="solid")
        category_header_1.alignment = Alignment(horizontal='center', vertical='center')
        category_header_1.border = thin_border
        
        category_header_2 = summary_sheet.cell(row=current_row, column=2, value='Count')
        category_header_2.font = Font(bold=True, color="FFFFFF")
        category_header_2.fill = PatternFill(start_color="3366FF", end_color="3366FF", fill_type="solid")
        category_header_2.alignment = Alignment(horizontal='center', vertical='center')
        category_header_2.border = thin_border

        current_row += 1

        # Category breakdown data
        category_counts = {}
        for jo in job_orders:
            if jo.jo_color:
                category_counts[jo.jo_color] = category_counts.get(jo.jo_color, 0) + 1

        for category, count in category_counts.items():
            cell_category = summary_sheet.cell(row=current_row, column=1, value=category)
            cell_category.border = thin_border
            cell_category.alignment = Alignment(vertical='center')
            if category.lower() in category_styles:
                cell_category.font = category_styles[category.lower()]['font']
            
            cell_count = summary_sheet.cell(row=current_row, column=2, value=count)
            cell_count.border = thin_border
            cell_count.alignment = Alignment(horizontal='center', vertical='center')
            
            current_row += 1

        # Save workbook to BytesIO object
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        # Generate a unique filename with timestamp
        filename = f"job_orders_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        # Create response with proper headers
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # Set headers for file download
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Content-Length'] = str(len(response.content))
        response['Access-Control-Expose-Headers'] = 'Content-Disposition, Content-Length, X-Export-Success'
        response['X-Export-Success'] = 'true'

        print(f"Export successful, returning file: {filename}, size: {len(response.content)} bytes")
        return response

    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        print(f"Export error: {str(e)}")
        print(error_traceback)

        # Return a more detailed error response
        error_message = str(e)
        if "CSRF" in error_message:
            error_message = "CSRF verification failed. Please refresh the page and try again."

        return JsonResponse({
            'error': error_message,
            'details': 'An error occurred while generating the Excel file. Please try again or contact support.'
        }, status=500)

# Queue
@login_required(login_url="user-login")
def queue_overview(request):
    current_month = now().month
    current_year = now().year
    today = timezone.now().date()

    pendingJO = JORouting.objects.filter(status = "pending", approver=request.user).order_by("-request_at")

    joRequestsCount = JORouting.objects.filter(approver=request.user, request_at__year=current_year, request_at__month=current_month).count()
    formatted_joRequestsCount = f"{joRequestsCount:04}"
    noTargetJOCount = JORouting.objects.filter(status = "pending", approver=request.user, jo_request__target_date__isnull=True,request_at__year=current_year, request_at__month=current_month).count()
    formatted_noTargetJOCount = f"{noTargetJOCount:04}"
    noInchargeJOCount = JORouting.objects.filter(status = "pending", approver=request.user, jo_request__target_date__isnull=True, request_at__year=current_year, request_at__month=current_month).count()
    formatted_noInchargeJOCount = f"{noInchargeJOCount:04}"
    pendingJOCount = JORouting.objects.filter(status = "pending", approver=request.user, request_at__year=current_year, request_at__month=current_month).count()
    formatted_pendingJOCount = f"{pendingJOCount:04}"

    # Calculate statistics for dashboard cards
    try:
        active_jo_count = JOLogsheet.objects.filter(status__in=["Routing", "Completed", "Assigned"]).count()
    except Exception:
        active_jo_count = 0

    try:
        routing_requests_count = JOLogsheet.objects.filter(status="Routing").count()
    except Exception:
        routing_requests_count = 0

    try:
        in_progress_requests_count = JOLogsheet.objects.filter(in_charge__isnull=False, status="Assigned").count()
    except Exception:
        in_progress_requests_count = 0

    try:
        completed_requests_count = JOLogsheet.objects.filter(in_charge__isnull=False, status__in=["Closed", "Completed", "Checked"]).count()
    except Exception:
        completed_requests_count = 0

    # Set placeholder values for comparison metrics (these will be calculated later if needed)
    active_jo_percentage = 0
    routing_requests_change = 0
    in_progress_requests_change = 0
    in_progress_requests_improved = False
    completed_requests_change = 0
    completed_requests_improved = False

    # Get today's date information for timeline
    today_name = today.strftime('%A')
    today_date = today.strftime('%b %d')

    context={
        'queueList':pendingJO,
        'joRequestsCount':formatted_joRequestsCount,
        'noTargetJOCount':formatted_noTargetJOCount,
        'noInchargeJOCount':formatted_noInchargeJOCount,
        'pendingJOCount':formatted_pendingJOCount,
        'active_jo_count': active_jo_count,
        'active_jo_percentage': active_jo_percentage,
        'routing_requests_count': routing_requests_count,
        'routing_requests_change': routing_requests_change,
        'in_progress_requests_count': in_progress_requests_count,
        'in_progress_requests_change': in_progress_requests_change,
        'in_progress_requests_improved': in_progress_requests_improved,
        'completed_requests_count': completed_requests_count,
        'completed_requests_change': completed_requests_change,
        'completed_requests_improved': completed_requests_improved,
        'today_name': today_name,
        'today_date': today_date,
    }
    return render(request, 'joborder/overall-dashboard.html', context)

@require_GET
def job_order_stats_api(request):
    try:
        now = timezone.now()
        today = now.date()
        thirty_days_ago = today - timedelta(days=30)
        sixty_days_ago = today - timedelta(days=60)

        maintenance_users = Users.objects.filter(job_order_maintenance=True)

        try:
            active_jo_count = JOLogsheet.objects.filter(status__in=["Routing", "Completed", "Assigned"]).count()
        except Exception as e:
            active_jo_count = 0

        try:
            routing_requests_count = JOLogsheet.objects.filter(status="Routing").count()
        except Exception as e:
            routing_requests_count = 0

        try:
            in_progress_requests_count = JOLogsheet.objects.filter(in_charge__isnull=False, status="Assigned").count()
        except Exception as e:
            in_progress_requests_count = 0

        try:
            completed_requests_count = JOLogsheet.objects.filter(in_charge__isnull=False, status__in=["Closed", "Completed", "Checked"]).count()
        except Exception as e:
            completed_requests_count = 0

        # Helper function to get custom month range (24th to 23rd)
        def get_custom_month_range(target_date):
            """
            Calculate custom month range from 24th to 23rd.
            If target_date.day < 24: use prev_month 24th to current_month 23rd
            If target_date.day >= 24: use current_month 24th to next_month 23rd
            """
            if target_date.day < 24:
                # Period is prev month 24th to current month 23rd
                if target_date.month == 1:
                    start_date = target_date.replace(year=target_date.year - 1, month=12, day=24)
                else:
                    start_date = target_date.replace(month=target_date.month - 1, day=24)
                end_date = target_date.replace(day=23)
            else:
                # Period is current month 24th to next month 23rd
                start_date = target_date.replace(day=24)
                if target_date.month == 12:
                    end_date = target_date.replace(year=target_date.year + 1, month=1, day=23)
                else:
                    end_date = target_date.replace(month=target_date.month + 1, day=23)
            
            return start_date, end_date

        # Calculate active job orders for current custom month and last custom month
        current_month_start, current_month_end = get_custom_month_range(today)
        # Cap current month end at today
        if current_month_end > today:
            current_month_end = today
        
        # Get last custom month range
        if today.month == 1:
            prev_month = today.replace(year=today.year - 1, month=12, day=15)
        else:
            prev_month = today.replace(month=today.month - 1, day=15)
        last_month_start, last_month_end = get_custom_month_range(prev_month)
        last_month_start, last_month_end = get_custom_month_range(prev_month)

        try:
            current_month_active = JOLogsheet.objects.filter(
                status__in=["Routing", "Completed", "Assigned"],
                date_created__date__gte=current_month_start,
                date_created__date__lte=current_month_end
            ).count()
        except Exception as e:
            current_month_active = 0

        try:
            last_month_active = JOLogsheet.objects.filter(
                status__in=["Routing", "Completed", "Assigned"],
                date_created__date__gte=last_month_start,
                date_created__date__lte=last_month_end
            ).count()
        except Exception as e:
            last_month_active = 0

        # Calculate percentage change from last month
        if last_month_active > 0:
            active_jo_percentage = round(((current_month_active - last_month_active) / last_month_active) * 100, 1)
        else:
            active_jo_percentage = 0 if current_month_active == 0 else 100

        # Calculate routing requests for current and last month
        try:
            current_month_routing = JOLogsheet.objects.filter(
                status="Routing",
                date_created__date__gte=current_month_start,
                date_created__date__lte=current_month_end
            ).count()
        except Exception as e:
            current_month_routing = 0

        try:
            last_month_routing = JOLogsheet.objects.filter(
                status="Routing",
                date_created__date__gte=last_month_start,
                date_created__date__lte=last_month_end
            ).count()
        except Exception as e:
            last_month_routing = 0

        # Calculate routing requests percentage change
        if last_month_routing > 0:
            routing_requests_change = round(((current_month_routing - last_month_routing) / last_month_routing) * 100, 1)
        else:
            routing_requests_change = 0 if current_month_routing == 0 else 100

        # Calculate in-progress requests for current and last month
        try:
            current_month_in_progress = JOLogsheet.objects.filter(
                in_charge__isnull=False,
                status="Assigned",
                date_created__date__gte=current_month_start,
                date_created__date__lte=current_month_end
            ).count()
        except Exception as e:
            current_month_in_progress = 0

        try:
            last_month_in_progress = JOLogsheet.objects.filter(
                in_charge__isnull=False,
                status="Assigned",
                date_created__date__gte=last_month_start,
                date_created__date__lte=last_month_end
            ).count()
        except Exception as e:
            last_month_in_progress = 0

        # Calculate in-progress requests percentage change
        if last_month_in_progress > 0:
            in_progress_requests_change = round(((current_month_in_progress - last_month_in_progress) / last_month_in_progress) * 100, 1)
            in_progress_requests_improved = current_month_in_progress < last_month_in_progress  # Lower is better for in-progress
        else:
            in_progress_requests_change = 0 if current_month_in_progress == 0 else 100
            in_progress_requests_improved = current_month_in_progress == 0

        # Calculate completed requests for current and last month
        try:
            current_month_completed = JOLogsheet.objects.filter(
                in_charge__isnull=False,
                status__in=["Closed", "Completed", "Checked"],
                date_created__date__gte=current_month_start,
                date_created__date__lte=current_month_end
            ).count()
        except Exception as e:
            current_month_completed = 0

        try:
            last_month_completed = JOLogsheet.objects.filter(
                in_charge__isnull=False,
                status__in=["Closed", "Completed", "Checked"],
                date_created__date__gte=last_month_start,
                date_created__date__lte=last_month_end
            ).count()
        except Exception as e:
            last_month_completed = 0

        # Calculate completed requests percentage change
        if last_month_completed > 0:
            completed_requests_change = round(((current_month_completed - last_month_completed) / last_month_completed) * 100, 1)
            completed_requests_improved = current_month_completed > last_month_completed  # Higher is better for completed
        else:
            completed_requests_change = 0 if current_month_completed == 0 else 100
            completed_requests_improved = current_month_completed > 0

        return JsonResponse({
            'status': 'success',
            'active_jo_count': active_jo_count,
            'active_jo_percentage': active_jo_percentage,
            'routing_requests_count': routing_requests_count,
            'routing_requests_change': routing_requests_change,
            'in_progress_requests_count': in_progress_requests_count,
            'in_progress_requests_change': in_progress_requests_change,
            'in_progress_requests_improved': in_progress_requests_improved,
            'completed_requests_count': completed_requests_count,
            'completed_requests_change': completed_requests_change,
            'completed_requests_improved': completed_requests_improved,
            'last_updated': timezone.now().strftime("%Y-%m-%d %H:%M:%S")
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'status': 'error', 'message': str(e)})

@require_GET
def job_order_queue_api(request):
    """API endpoint for Job Order Queue data"""
    try:
        now = timezone.now()
        today = now.date()
        tomorrow = today + timedelta(days=1)
        two_days_from_now = today + timedelta(days=2)
        three_days_from_now = today + timedelta(days=3)

        # Get all job orders with status "Assigned"
        all_jobs = JOLogsheet.objects.filter(
            status='Assigned'
        )

        overdue_jobs = []
        urgent_jobs = []
        critical_jobs = []
        upcoming_jobs = []
        normal_jobs = []

        for job in all_jobs:
            # Get requestor information
            requestor = job.requestor if job.requestor else "Unknown"

            # Get department/line information
            department_line = "N/A"
            try:
                if job.line:
                    department_line = job.line.line_name
            except Exception as e:
                print(f"Error getting line information: {str(e)}")

            # Calculate waiting time
            days_waiting = (today - job.date_created.date()).days
            if days_waiting == 0:
                waiting_time = "Created today"
            elif days_waiting == 1:
                waiting_time = "1 day ago"
            else:
                waiting_time = f"{days_waiting} days ago"

            job_data = {
                'id': job.id,
                'jo_number': job.jo_number,
                'category': job.jo_color,
                'description': f"{job.jo_tools} - {job.jo_type}",
                'requestor': requestor,
                'department': department_line,
                'priority_level': job.priority_level if job.priority_level else "Low",
                'waiting_time': waiting_time,
                'days_waiting': days_waiting,
                'date_created': job.date_created.strftime("%b %d"),
                'target_date': job.target_date.strftime("%b %d") if job.target_date else None,
                'target_date_raw': job.target_date,
                'date_of_completion': job.date_of_completion
            }

            # Categorize jobs based on Assignment Reminders logic
            # 1. OVERDUE: Jobs past their target date (highest priority)
            is_overdue = False
            if job.target_date:
                target_date_local = timezone.localtime(job.target_date).date()
                if target_date_local < today:
                    is_overdue = True
                    job_data['urgency_type'] = 'overdue'
                    job_data['sort_date'] = target_date_local
                    overdue_jobs.append(job_data)
            
            # Check date_of_completion for Urgent jobs
            elif job.priority_level == 'Urgent' and job.date_of_completion and job.date_of_completion < today:
                is_overdue = True
                job_data['urgency_type'] = 'overdue'
                job_data['sort_date'] = job.date_of_completion
                overdue_jobs.append(job_data)

            if is_overdue:
                continue  # Skip other categorizations

            # 2. URGENT: priority_level = "Urgent"
            if job.priority_level and job.priority_level == "Urgent":
                job_data['urgency_type'] = 'urgent'
                if job.date_of_completion:
                    job_data['sort_date'] = job.date_of_completion
                else:
                    job_data['sort_date'] = today + timedelta(days=999)
                urgent_jobs.append(job_data)
            
            # 3. CRITICAL: target_date is today or tomorrow
            elif job.target_date:
                target_date_local = timezone.localtime(job.target_date).date()
                if target_date_local in [today, tomorrow]:
                    job_data['urgency_type'] = 'critical'
                    job_data['sort_date'] = target_date_local
                    critical_jobs.append(job_data)
                
                # 4. UPCOMING: target_date is 2-3 days from now
                elif target_date_local in [two_days_from_now, three_days_from_now]:
                    job_data['urgency_type'] = 'upcoming'
                    job_data['sort_date'] = target_date_local
                    upcoming_jobs.append(job_data)
                
                # 5. NORMAL: target_date is more than 3 days away
                else:
                    job_data['urgency_type'] = 'normal'
                    job_data['sort_date'] = target_date_local
                    normal_jobs.append(job_data)
            
            # Default: no target date
            else:
                job_data['urgency_type'] = 'normal'
                job_data['sort_date'] = today + timedelta(days=999)
                normal_jobs.append(job_data)

        # Sort each category by sort_date (earliest first)
        overdue_jobs.sort(key=lambda x: x['sort_date'])
        urgent_jobs.sort(key=lambda x: x['sort_date'])
        critical_jobs.sort(key=lambda x: x['sort_date'])
        upcoming_jobs.sort(key=lambda x: x['sort_date'])
        normal_jobs.sort(key=lambda x: x['sort_date'])

        # Combine all categories in order: overdue, urgent, critical, upcoming, normal
        queue_data = overdue_jobs + urgent_jobs + critical_jobs + upcoming_jobs + normal_jobs

        return JsonResponse({
            'status': 'success',
            'queue': queue_data,
            'total_count': len(queue_data),
            'overdue_count': len(overdue_jobs),
            'urgent_count': len(urgent_jobs),
            'critical_count': len(critical_jobs),
            'upcoming_count': len(upcoming_jobs),
            'normal_count': len(normal_jobs),
            'last_updated': now.strftime("%Y-%m-%d %H:%M:%S")
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })

@require_GET
def job_order_timeline_api(request, view_type):
    try:
        now = timezone.now()
        today = now.date()
        maintenance = Users.objects.filter(job_order_maintenance=True)

        response_data = {
            'status': 'success',
            'view_type': view_type,
            'events': []
        }

        # Get all pending job orders assigned to maintenance users regardless of date
        all_pending_jobs = JORouting.objects.filter(
            approver__in=maintenance,
            status='Processing'
        ).order_by('-request_at')

        # Helper function to create event data with common fields
        def create_event_data(job, additional_fields=None):
            # Get the JO number
            jo_number = job.jo_number.jo_number if hasattr(job.jo_number, 'jo_number') else str(job.jo_number)

            # Get the requestor name
            requestor = job.jo_number.requestor if hasattr(job.jo_number, 'requestor') else "Unknown"

            # Get target date for color coding
            target_date = job.jo_number.target_date if hasattr(job.jo_number, 'target_date') else None

            # Calculate days until target date for color coding
            days_until_target = None
            has_target_date = target_date is not None
            is_overdue = False
            is_approaching = False

            if has_target_date:
                # Ensure we're comparing dates in the same timezone
                # Convert timezone-aware datetime to date in local timezone
                if hasattr(target_date, 'date'):
                    # If target_date is timezone-aware, convert to local timezone first
                    if hasattr(target_date, 'tzinfo') and target_date.tzinfo is not None:
                        target_date_local = timezone.localtime(target_date)
                        target_date_only = target_date_local.date()
                    else:
                        target_date_only = target_date.date()
                else:
                    target_date_only = target_date
                
                days_until_target = (target_date_only - today).days
                is_overdue = days_until_target < 0
                is_approaching = 0 <= days_until_target <= 3

            # Format the time with AM/PM
            time_str = job.request_at.strftime("%I:%M %p").strip()

            # Format the date for month view
            date_str = job.request_at.strftime("%b %d")

            # Get department/line information
            department_line = "N/A"
            try:
                if job.jo_number.line:
                    department_line = job.jo_number.line.line_name
            except Exception as e:
                print(f"Error getting line information: {str(e)}")

            # Create the base event data
            event_data = {
                'jo_number': jo_number,
                'time': time_str,
                'date': date_str,
                'title': f"Assigned to {job.approver.name}",
                'requestor': requestor,
                'requestor_dept': department_line,
                'status': job.status,
                'has_target_date': has_target_date,
                'is_overdue': is_overdue,
                'is_approaching': is_approaching,
                'days_until_target': days_until_target,
                'event_date': job.request_at.strftime("%Y-%m-%d")  # For date filtering
            }

            # Add any additional fields
            if additional_fields:
                event_data.update(additional_fields)

            return event_data

        if view_type == 'timeline-today':
            # For today view, still filter by today but include all pending jobs
            today_start = timezone.make_aware(datetime.combine(today, datetime.min.time()))
            today_end = timezone.make_aware(datetime.combine(today, datetime.max.time()))

            # Get jobs created today
            today_jobs = all_pending_jobs.filter(
                request_at__range=[today_start, today_end]
            )

            # If no jobs today, show some of the most recent pending jobs
            if not today_jobs.exists():
                today_jobs = all_pending_jobs[:5]  # Show up to 5 recent pending jobs

            for job in today_jobs:
                event_data = create_event_data(job)
                response_data['events'].append(event_data)

        elif view_type == 'timeline-week':
            start_of_week = today - timedelta(days=today.weekday())
            end_of_week = start_of_week + timedelta(days=6)

            # Create day columns for the current week
            day_columns = {(start_of_week + timedelta(days=i)).day: i+1 for i in range(7)}
            day_rows = {i+1: 1 for i in range(7)}

            # Get jobs from this week
            week_jobs = all_pending_jobs.filter(
                request_at__date__gte=start_of_week,
                request_at__date__lte=end_of_week
            )

            # If no jobs this week, show some of the most recent pending jobs
            if not week_jobs.exists():
                week_jobs = all_pending_jobs[:10]  # Show up to 10 recent pending jobs

                # For jobs outside the current week, distribute them evenly across the week
                for i, job in enumerate(week_jobs):
                    day_column = (i % 7) + 1  # Distribute across the 7 days of the week
                    additional_fields = {
                        'day_column': day_column,
                        'row': day_rows[day_column]
                    }
                    event_data = create_event_data(job, additional_fields)
                    response_data['events'].append(event_data)
                    day_rows[day_column] += 1
            else:
                # For jobs within the current week, place them on the correct day
                for job in week_jobs:
                    job_day = job.request_at.day
                    # Find the closest day column if the exact day isn't in our columns
                    closest_day = min(day_columns.keys(), key=lambda x: abs(x - job_day))
                    day_column = day_columns.get(closest_day, 1)

                    additional_fields = {
                        'day_column': day_column,
                        'row': day_rows[day_column]
                    }
                    event_data = create_event_data(job, additional_fields)
                    response_data['events'].append(event_data)
                    day_rows[day_column] += 1

        elif view_type == 'timeline-month':
            start_of_month = today.replace(day=1)

            # Create week columns for the current month
            weeks_in_month = (today.day - 1) // 7 + 1
            week_columns = {i+1: i+1 for i in range(weeks_in_month)}
            week_rows = {i+1: 1 for i in range(weeks_in_month)}

            # Get jobs from this month
            month_jobs = all_pending_jobs.filter(
                request_at__year=today.year,
                request_at__month=today.month
            )

            # If no jobs this month, show some of the most recent pending jobs
            if not month_jobs.exists():
                month_jobs = all_pending_jobs[:15]  # Show up to 15 recent pending jobs

                # For jobs outside the current month, distribute them evenly across the weeks
                for i, job in enumerate(month_jobs):
                    week_column = (i % weeks_in_month) + 1  # Distribute across the weeks of the month
                    additional_fields = {
                        'week_column': week_column,
                        'row': week_rows[week_column]
                    }
                    event_data = create_event_data(job, additional_fields)
                    response_data['events'].append(event_data)
                    week_rows[week_column] += 1
            else:
                # For jobs within the current month, place them in the correct week
                for job in month_jobs:
                    week_of_month = (job.request_at.day - 1) // 7 + 1
                    if week_of_month not in week_columns:
                        week_columns[week_of_month] = len(week_columns) + 1
                        week_rows[week_columns[week_of_month]] = 1

                    week_column = week_columns[week_of_month]
                    additional_fields = {
                        'week_column': week_column,
                        'row': week_rows[week_column]
                    }
                    event_data = create_event_data(job, additional_fields)
                    response_data['events'].append(event_data)
                    week_rows[week_column] += 1

        response_data['events'].sort(key=lambda x: x.get('row', 0))

        return JsonResponse(response_data)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'status': 'error', 'message': str(e)})

@require_GET
def job_order_deadlines_api(request):
    try:
        now = timezone.now()
        today = now.date()
        tomorrow = today + timedelta(days=1)
        two_days_from_now = today + timedelta(days=2)
        three_days_from_now = today + timedelta(days=3)

        maintenance_personnel = Users.objects.filter(job_order_maintenance=True)

        # Get all active job orders that need attention (exclude Cancelled and Rejected)
        active_jobs = JOLogsheet.objects.filter(in_charge__isnull=False,
            status__in=['Routing', 'Assigned', 'Completed', 'Checked']
        ).exclude(
            status__in=['Cancelled', 'Rejected']
        ).distinct()

        urgent_jobs = []
        critical_jobs = []
        upcoming_jobs = []
        overdue_assigned_jobs = []

        # Categorize job orders
        for job in active_jobs:
            assigned_person = job.in_charge.name if job.in_charge else None
            assigned_to = f"Assigned to {assigned_person}" if assigned_person else "Not yet assigned"
            
            # Get requestor information
            requestor = job.requestor if job.requestor else "Unknown"

            # Get department/line information
            department_line = "N/A"
            try:
                if job.line:
                    department_line = job.line.line_name
            except Exception as e:
                print(f"Error getting line information: {str(e)}")

            # Build base job data
            job_data = {
                'jo_number': job.jo_number,
                'category': job.jo_color,
                'description': f"{job.jo_tools} - {job.jo_type} - {assigned_to}",
                'requestor': requestor,
                'department': department_line,
                'priority_level': job.priority_level if job.priority_level else "Low"
            }

            # Check for OVERDUE ASSIGNED jobs first (highest priority)
            is_overdue_assigned = False
            if job.status == 'Assigned':
                # Check if past target date
                if job.target_date:
                    # Convert timezone-aware datetime to local date
                    target_date_local = timezone.localtime(job.target_date).date()
                    if target_date_local < today:
                        is_overdue_assigned = True
                        days_overdue = (today - target_date_local).days
                        job_data['day'] = target_date_local.day
                        job_data['month'] = timezone.localtime(job.target_date).strftime('%b')
                        job_data['countdown'] = f"{days_overdue} {'day' if days_overdue == 1 else 'days'} overdue"
                        job_data['sort_date'] = target_date_local
                # Check if past date of completion for Urgent jobs
                elif job.priority_level == 'Urgent' and job.date_of_completion and job.date_of_completion < today:
                    is_overdue_assigned = True
                    days_overdue = (today - job.date_of_completion).days
                    job_data['day'] = job.date_of_completion.day
                    job_data['month'] = job.date_of_completion.strftime('%b')
                    job_data['countdown'] = f"{days_overdue} {'day' if days_overdue == 1 else 'days'} overdue"
                    job_data['sort_date'] = job.date_of_completion

            if is_overdue_assigned:
                job_data['urgency_type'] = 'overdue_assigned'
                job_data['is_critical'] = True
                overdue_assigned_jobs.append(job_data)
                continue  # Skip other categorizations

            # Categorize as URGENT (priority level = "Urgent")
            if job.priority_level and job.priority_level == "Urgent":
                if job.date_of_completion:
                    job_data['day'] = job.date_of_completion.day
                    job_data['month'] = job.date_of_completion.strftime('%b')
                    days_until = (job.date_of_completion - today).days
                    if days_until == 0:
                        job_data['countdown'] = "Today"
                    elif days_until == 1:
                        job_data['countdown'] = "Tomorrow"
                    elif days_until < 0:
                        job_data['countdown'] = f"{abs(days_until)} {'day' if abs(days_until) == 1 else 'days'} overdue"
                    else:
                        job_data['countdown'] = f"{days_until} {'day' if days_until == 1 else 'days'}"
                    job_data['sort_date'] = job.date_of_completion
                else:
                    job_data['day'] = '-'
                    job_data['month'] = '-'
                    job_data['countdown'] = "No date set"
                    job_data['sort_date'] = today + timedelta(days=999)  # Push to end
                
                job_data['urgency_type'] = 'urgent'
                job_data['is_critical'] = True
                urgent_jobs.append(job_data)
            
            # Categorize as CRITICAL (target_date is today or tomorrow)
            elif job.target_date:
                target_date_local = timezone.localtime(job.target_date).date()
                if target_date_local in [today, tomorrow]:
                    job_data['day'] = target_date_local.day
                    job_data['month'] = timezone.localtime(job.target_date).strftime('%b')
                    days_until = (target_date_local - today).days
                    if days_until == 0:
                        job_data['countdown'] = "Today"
                    else:
                        job_data['countdown'] = "Tomorrow"
                    job_data['urgency_type'] = 'critical'
                    job_data['is_critical'] = True
                    job_data['sort_date'] = target_date_local
                    critical_jobs.append(job_data)
            
                # Categorize as UPCOMING (target_date is 2-3 days from now)
                elif target_date_local in [two_days_from_now, three_days_from_now]:
                    job_data['day'] = target_date_local.day
                    job_data['month'] = timezone.localtime(job.target_date).strftime('%b')
                    days_until = (target_date_local - today).days
                    job_data['countdown'] = f"{days_until} {'day' if days_until == 1 else 'days'}"
                    job_data['urgency_type'] = 'upcoming'
                    job_data['is_critical'] = False
                    job_data['sort_date'] = target_date_local
                    upcoming_jobs.append(job_data)

        # Sort each category
        overdue_assigned_jobs.sort(key=lambda x: x['sort_date'])  # Earliest overdue first
        urgent_jobs.sort(key=lambda x: x['sort_date'])
        critical_jobs.sort(key=lambda x: x['sort_date'])
        upcoming_jobs.sort(key=lambda x: x['sort_date'])

        # Combine in order: Overdue Assigned -> Urgent -> Critical -> Upcoming
        deadlines_data = overdue_assigned_jobs + urgent_jobs + critical_jobs + upcoming_jobs

        return JsonResponse({
            'status': 'success',
            'deadlines': deadlines_data,
            'urgent_count': len(urgent_jobs),
            'critical_count': len(critical_jobs),
            'upcoming_count': len(upcoming_jobs),
            'overdue_assigned_count': len(overdue_assigned_jobs),
            'last_updated': now.strftime("%Y-%m-%d %H:%M:%S")
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })

@require_GET
def job_order_alerts_api(request):
    try:
        now = timezone.now()
        today = now.date()
        maintenance_personnel = Users.objects.filter(job_order_maintenance=True)
        alerts_data = []

        overdue_jobs = JOLogsheet.objects.filter(
            target_date__lt=now,
            status='Routing',
            joRouting__status='Processing',
            joRouting__approver_sequence=6,
            joRouting__approver__in=maintenance_personnel
        ).distinct().order_by('target_date')[:3]

        for job in overdue_jobs:
            days_overdue = (today - job.target_date.date()).days
            assigned_to = job.in_charge.name if job.in_charge else "No assignee yet"
            requestor = job.requestor if job.requestor else "Unknown"
            department_line = "N/A"
            try:
                if job.line:
                    department_line = job.line.line_name
            except Exception as e:
                print(f"Error getting line information: {str(e)}")

            message = f"{job.jo_tools} is overdue by {days_overdue} days. Assigned to {assigned_to}."

            alert = {
                'type': 'critical',
                'icon': 'exclamation-triangle',
                'title': f"Overdue Job Order: {job.jo_number}",
                'time': f"{days_overdue} days ago",
                'message': message,
                'requestor': requestor,
                'department': department_line,
                'actions': [
                    {
                        'text': 'View Details',
                        'icon': 'eye',
                        'data_jo': job.jo_number
                    },
                    {
                        'text': 'Send Reminder',
                        'icon': 'bell',
                        'data_jo': None
                    }
                ]
            }

            alerts_data.append(alert)

        overloaded_staff = JOLogsheet.objects.filter(
            status='Routing',
            joRouting__status='Pending',
            joRouting__approver_sequence=6,
            joRouting__approver__in=maintenance_personnel
        ).values('in_charge').annotate(
            task_count=Count('id')
        ).filter(task_count__gte=3).order_by('-task_count')[:2]

        for staff in overloaded_staff:
            if staff['in_charge']:
                try:
                    staff_member = Users.objects.get(id=staff['in_charge'])
                    workload_percent = min(round((staff['task_count'] / 6) * 100), 100)

                    alert = {
                        'type': 'resource',
                        'icon': 'cogs',
                        'title': f"Resource Bottleneck: {staff_member.name}",
                        'time': 'Active',
                        'message': f"{staff_member.name} has {workload_percent}% workload capacity with {staff['task_count']} pending tasks.",
                        'actions': [
                            {
                                'text': 'Reassign Tasks',
                                'icon': 'exchange-alt',
                                'data_jo': None
                            },
                            {
                                'text': 'Resource Planning',
                                'icon': 'users',
                                'data_jo': None
                            }
                        ]
                    }

                    alerts_data.append(alert)
                except Users.DoesNotExist:
                    continue

        return JsonResponse({
            'status': 'success',
            'alerts': alerts_data,
            'last_updated': now.strftime("%Y-%m-%d %H:%M:%S")
        })

    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })

@require_GET
def job_order_compliance_chart_api(request):
    """API endpoint for Job Order Compliance chart data"""
    try:
        period = request.GET.get('period', '6month')
        today = timezone.now().date()
        
        # Helper function to get custom month range (24th to 24th)
        def get_custom_month_range(target_date):
            """
            Returns the start and end date for a custom month period (24th to 24th).
            For example, October 2025 covers Sep 24 to Oct 24.
            """
            # If today is before the 24th, we're still in the previous month's period
            if target_date.day < 24:
                # Use previous month's 24th to current month's 23rd
                month_end = target_date.replace(day=23)
                # Get previous month's 24th
                if target_date.month == 1:
                    month_start = target_date.replace(year=target_date.year - 1, month=12, day=24)
                else:
                    month_start = target_date.replace(month=target_date.month - 1, day=24)
            else:
                # From this month's 24th to next month's 23rd
                month_start = target_date.replace(day=24)
                # Get next month's 23rd
                if target_date.month == 12:
                    month_end = target_date.replace(year=target_date.year + 1, month=1, day=23)
                else:
                    month_end = target_date.replace(month=target_date.month + 1, day=23)
            
            return month_start, month_end
        
        # Determine the start date based on the period
        if period == 'thismonth':
            # Current custom month period (24th to 24th)
            start_date, end_date = get_custom_month_range(today)
            # Don't go beyond today
            if end_date > today:
                end_date = today
        elif period == '1month':
            # Previous custom month period
            # Go back one month from today
            if today.month == 1:
                prev_month_date = today.replace(year=today.year - 1, month=12, day=15)
            else:
                prev_month_date = today.replace(month=today.month - 1, day=15)
            start_date, end_date = get_custom_month_range(prev_month_date)
        elif period == '3month':
            # Last 3 custom months
            start_date = today - timedelta(days=90)
            end_date = today
        elif period == '6month':
            # Last 6 custom months (default)
            start_date = today - timedelta(days=180)
            end_date = today
        else:
            # Default to 6 months
            start_date = today - timedelta(days=180)
            end_date = today
        
        # For thismonth and 1month, we'll show daily data
        # For 3month and 6month, we'll group by month
        
        if period in ['thismonth', '1month']:
            # Daily data
            labels = []
            filed_data = []
            completed_data = []
            routing_data = []
            inprogress_data = []
            
            current_date = start_date
            while current_date <= end_date:
                labels.append(current_date.strftime('%b %d'))
                
                # Count filed JOs on this date
                filed_count = JOLogsheet.objects.filter(
                    date_created__date=current_date
                ).count()
                filed_data.append(filed_count)
                
                # Count completed JOs on this date
                completed_count = JOLogsheet.objects.filter(
                    date_complete__date=current_date,
                    status__in=['Completed', 'Checked', 'Closed']
                ).count()
                completed_data.append(completed_count)
                
                # Count routing JOs on this date
                routing_count = JOLogsheet.objects.filter(
                    date_created__date=current_date,
                    status='Routing'
                ).count()
                routing_data.append(routing_count)
                
                # Count in-progress JOs on this date
                inprogress_count = JOLogsheet.objects.filter(
                    date_created__date=current_date,
                    status='Assigned'
                ).count()
                inprogress_data.append(inprogress_count)
                
                current_date += timedelta(days=1)
        else:
            # Monthly data for 3month and 6month using custom 24th-to-24th ranges
            labels = []
            filed_data = []
            completed_data = []
            routing_data = []
            inprogress_data = []
            
            # Generate list of custom months (24th to 24th) in the range
            months = []
            current_date = start_date
            
            # Start from the 24th of the month containing start_date
            if start_date.day < 24:
                # Start from previous month's 24th
                if start_date.month == 1:
                    month_start = start_date.replace(year=start_date.year - 1, month=12, day=24)
                else:
                    month_start = start_date.replace(month=start_date.month - 1, day=24)
            else:
                month_start = start_date.replace(day=24)
            
            while month_start <= end_date:
                # Calculate month_end as 23rd of next month
                if month_start.month == 12:
                    month_end = month_start.replace(year=month_start.year + 1, month=1, day=23)
                else:
                    month_end = month_start.replace(month=month_start.month + 1, day=23)
                
                # Don't go beyond end_date
                if month_end > end_date:
                    month_end = end_date
                
                months.append((month_start, month_end))
                
                # Move to next custom month (next month's 24th)
                if month_start.month == 12:
                    month_start = month_start.replace(year=month_start.year + 1, month=1, day=24)
                else:
                    month_start = month_start.replace(month=month_start.month + 1, day=24)
            
            for month_start, month_end in months:
                # Label as "Month Year" based on the end date of the period
                # (e.g., Sep 24 - Oct 23 is labeled "Oct 2025")
                label = month_end.strftime('%b %Y')
                labels.append(label)
                
                # Count filed JOs in this custom month
                filed_count = JOLogsheet.objects.filter(
                    date_created__date__gte=month_start,
                    date_created__date__lte=month_end
                ).count()
                filed_data.append(filed_count)
                
                # Count completed JOs in this custom month
                completed_count = JOLogsheet.objects.filter(
                    date_complete__date__gte=month_start,
                    date_complete__date__lte=month_end,
                    status__in=['Completed', 'Checked', 'Closed']
                ).count()
                completed_data.append(completed_count)
                
                # Count routing JOs in this custom month
                routing_count = JOLogsheet.objects.filter(
                    date_created__date__gte=month_start,
                    date_created__date__lte=month_end,
                    status='Routing'
                ).count()
                routing_data.append(routing_count)
                
                # Count in-progress JOs in this custom month
                inprogress_count = JOLogsheet.objects.filter(
                    date_created__date__gte=month_start,
                    date_created__date__lte=month_end,
                    status='Assigned'
                ).count()
                inprogress_data.append(inprogress_count)
        
        return JsonResponse({
            'status': 'success',
            'labels': labels,
            'filed': filed_data,
            'completed': completed_data,
            'routing': routing_data,
            'inprogress': inprogress_data
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })

def compliance_rate_api(request):
    """API endpoint for Compliance Rate pie chart data"""
    try:
        period = request.GET.get('period', '6month')
        today = timezone.now().date()
        
        # Helper function to get custom month range (24th to 23rd)
        def get_custom_month_range(target_date):
            """
            Calculate custom month range from 24th to 23rd.
            If target_date.day < 24: use prev_month 24th to current_month 23rd
            If target_date.day >= 24: use current_month 24th to next_month 23rd
            """
            if target_date.day < 24:
                # Period is prev month 24th to current month 23rd
                if target_date.month == 1:
                    start_date = target_date.replace(year=target_date.year - 1, month=12, day=24)
                else:
                    start_date = target_date.replace(month=target_date.month - 1, day=24)
                end_date = target_date.replace(day=23)
            else:
                # Period is current month 24th to next month 23rd
                start_date = target_date.replace(day=24)
                if target_date.month == 12:
                    end_date = target_date.replace(year=target_date.year + 1, month=1, day=23)
                else:
                    end_date = target_date.replace(month=target_date.month + 1, day=23)
            
            return start_date, end_date
        
        # Determine the start date based on the period
        if period == 'thismonth':
            # Custom month range (24th to 23rd) for current period, ending today
            start_date, month_end = get_custom_month_range(today)
            end_date = today  # Cap at today
        elif period == '1month':
            # Previous custom month range (24th to 23rd)
            # Go back to previous month first
            if today.month == 1:
                prev_month = today.replace(year=today.year - 1, month=12, day=15)
            else:
                prev_month = today.replace(month=today.month - 1, day=15)
            start_date, end_date = get_custom_month_range(prev_month)
        elif period == '6month':
            # Last 6 custom months (approximately 180 days from 24th to 23rd)
            # Calculate 6 months back from current custom month start
            current_start, _ = get_custom_month_range(today)
            if current_start.month > 6:
                start_date = current_start.replace(month=current_start.month - 6)
            else:
                start_date = current_start.replace(year=current_start.year - 1, month=current_start.month + 6)
            end_date = today
        elif period == 'fiscalyear':
            # Fiscal year: May 24 to April 23
            current_year = today.year
            if today.month > 4 or (today.month == 4 and today.day >= 24):
                # May 24 of current year to April 23 of next year
                start_date = date(current_year, 5, 24)
                end_date = date(current_year + 1, 4, 23)
            else:
                # May 24 of previous year to April 23 of current year
                start_date = date(current_year - 1, 5, 24)
                end_date = date(current_year, 4, 23)
            # If end_date is in the future, use today
            if end_date > today:
                end_date = today
        else:
            # Default to 6 months
            current_start, _ = get_custom_month_range(today)
            if current_start.month > 6:
                start_date = current_start.replace(month=current_start.month - 6)
            else:
                start_date = current_start.replace(year=current_start.year - 1, month=current_start.month + 6)
            end_date = today
        
        # Count total JOs filed in the period
        total_jos = JOLogsheet.objects.filter(
            date_created__date__gte=start_date,
            date_created__date__lte=end_date
        ).count()
        
        # Count completed JOs in the period
        completed_jos = JOLogsheet.objects.filter(
            date_created__date__gte=start_date,
            date_created__date__lte=end_date,
            status__in=['Completed', 'Checked', 'Closed']
        ).count()
        
        # Count routing JOs in the period
        routing_jos = JOLogsheet.objects.filter(
            date_created__date__gte=start_date,
            date_created__date__lte=end_date,
            status='Routing'
        ).count()
        
        # Count in-progress JOs in the period (Processing status)
        inprogress_jos = JOLogsheet.objects.filter(
            date_created__date__gte=start_date,
            date_created__date__lte=end_date,
            status='Assigned'
        ).count()
        
        # Calculate percentages
        if total_jos > 0:
            completed_percentage = round((completed_jos / total_jos) * 100, 1)
            pending_percentage = round(((total_jos - completed_jos) / total_jos) * 100, 1)
        else:
            completed_percentage = 0
            pending_percentage = 0
        
        pending_jos = total_jos - completed_jos
        
        return JsonResponse({
            'status': 'success',
            'total_jos': total_jos,
            'completed_jos': completed_jos,
            'pending_jos': pending_jos,
            'routing_jos': routing_jos,
            'inprogress_jos': inprogress_jos,
            'completed_percentage': completed_percentage,
            'pending_percentage': pending_percentage
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })

# Maintenance
@login_required
def get_maintenance_personnel(request):
    """API endpoint to get list of maintenance personnel for assignment"""
    try:
        # Get all users with job_order_maintenance = True
        maintenance_users = Users.objects.filter(job_order_maintenance=True).order_by('name')
        
        # Prepare data for response
        personnel_data = []
        for user in maintenance_users:
            personnel_data.append({
                'id': user.id,
                'name': user.name,
                'username': user.username
            })
        
        return JsonResponse({
            'status': 'success',
            'personnel': personnel_data
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })

@login_required
def maintenance_job_orders_api(request):
    """API endpoint to get job orders for maintenance personnel"""
    try:
        # Get job orders assigned to the current user
        job_orders = JOLogsheet.objects.filter(
            joRouting__approver=request.user,
            joRouting__approver_sequence=6
        ).order_by("-joRouting__request_at", "joRouting__status")

        # Prepare data for response
        job_orders_data = []
        for job in job_orders:
            # Check if job has pending routing
            has_pending_routing = job.joRouting.filter(approver_sequence=6, status='Processing').exists()

            # Check if job is overdue
            is_overdue = False
            if job.target_date and job.target_date < now().date() and job.status != 'Completed':
                is_overdue = True

            job_orders_data.append({
                'id': job.id,
                'jo_number': job.jo_number,
                'jo_color': job.jo_color,
                'jo_tools': job.jo_tools,
                'requestor': job.requestor,
                'target_date': job.target_date.strftime('%b %d, %Y') if job.target_date else None,
                'status': job.status,
                'is_overdue': is_overdue,
                'has_pending_routing': has_pending_routing
            })

        return JsonResponse({
            'status': 'success',
            'job_orders': job_orders_data
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)

@login_required
def get_job_order_trends(request):
    period = request.GET.get('period', 'month')
    current_user = request.user
    today = timezone.now()

    # Helper function to get custom month range (24th to 24th)
    def get_custom_month_range(target_date):
        """
        Calculate custom month range from 24th to 24th (ending on 24th).
        For October: Sept 24 to Oct 24
        """
        target_day = target_date.day
        
        if target_day < 24:
            # We're before the 24th, so the period started last month on the 24th
            if target_date.month == 1:
                start_date = target_date.replace(year=target_date.year - 1, month=12, day=24)
            else:
                start_date = target_date.replace(month=target_date.month - 1, day=24)
            end_date = target_date.replace(day=24)
        else:
            # We're on or after the 24th, so the period starts this month on the 24th
            start_date = target_date.replace(day=24)
            if target_date.month == 12:
                end_date = target_date.replace(year=target_date.year + 1, month=1, day=24)
            else:
                end_date = target_date.replace(month=target_date.month + 1, day=24)
        
        return start_date, end_date

    if (period == 'month'):
        # Use custom 24th-to-24th month range
        start_date, end_date = get_custom_month_range(today)
        
        # If end_date is in the future, cap it at today
        if end_date > today:
            end_date = today
        
        # Calculate the number of days in this custom month period
        days_in_period = (end_date.date() - start_date.date()).days + 1
        
        labels = [(start_date + timedelta(days=i)).strftime('%d') for i in range(days_in_period)]
        new_orders_data = [0] * days_in_period
        completed_data = [0] * days_in_period

        # Get new routings within the custom month range
        new_routings = JORouting.objects.filter(
            approver=current_user,
            approver_sequence=6,
            request_at__gte=start_date,
            request_at__lte=end_date
        )

        # Get completed routings within the custom month range
        completed_routings = JORouting.objects.filter(
            approver=current_user,
            approver_sequence=6,
            approved_at__gte=start_date,
            approved_at__lte=end_date,
            status='Approved'
        )

        for routing in new_routings:
            day_diff = (routing.request_at.date() - start_date.date()).days
            if 0 <= day_diff < len(new_orders_data):
                new_orders_data[day_diff] += 1

        for routing in completed_routings:
            if routing.approved_at:
                day_diff = (routing.approved_at.date() - start_date.date()).days
                if 0 <= day_diff < len(completed_data):
                    completed_data[day_diff] += 1

    elif period == 'quarter':
        current_month = today.month
        quarter_start_month = 3 * ((current_month - 1) // 3) + 1

        months = []
        for i in range(3):
            month = (quarter_start_month + i - 1) % 12 + 1
            year = today.year if month >= quarter_start_month or quarter_start_month > today.month else today.year - 1
            months.append((year, month))

        labels = [datetime(year, month, 1).strftime('%b') for year, month in months]
        new_orders_data = [0] * 3
        completed_data = [0] * 3

        for i, (year, month) in enumerate(months):
            month_new_routings = JORouting.objects.filter(
                approver=current_user,
                approver_sequence=6,
                request_at__year=year,
                request_at__month=month
            ).count()
            new_orders_data[i] = month_new_routings

            month_completed_routings = JORouting.objects.filter(
                approver=current_user,
                approver_sequence=6,
                approved_at__year=year,
                approved_at__month=month,
                status='Approved'
            ).count()
            completed_data[i] = month_completed_routings

    elif period == 'year':
        months_data = []
        for i in range(11, -1, -1):
            target_date = today - timedelta(days=i*30)
            year = target_date.year
            month = target_date.month
            months_data.append((year, month, target_date.strftime('%b %y')))

        months_data.sort()

        labels = [month_label for _, _, month_label in months_data]
        new_orders_data = [0] * 12
        completed_data = [0] * 12

        for i, (year, month, _) in enumerate(months_data):
            month_new_routings = JORouting.objects.filter(
                approver=current_user,
                approver_sequence=6,
                request_at__year=year,
                request_at__month=month
            ).count()
            new_orders_data[i] = month_new_routings

            month_completed_routings = JORouting.objects.filter(
                approver=current_user,
                approver_sequence=6,
                approved_at__year=year,
                approved_at__month=month,
                status='Approved'
            ).count()
            completed_data[i] = month_completed_routings

    data = {
        'labels': labels,
        'datasets': [
            {
                'label': 'New Orders',
                'data': new_orders_data,
                'borderColor': '#3366ff',
                'backgroundColor': 'rgba(51, 102, 255, 0.8)',
                'borderWidth': 1
            },
            {
                'label': 'Completed',
                'data': completed_data,
                'borderColor': '#4caf50',
                'backgroundColor': 'rgba(76, 175, 80, 0.8)',
                'borderWidth': 1
            }
        ]
    }

    return JsonResponse(data)

@login_required
def get_upcoming_deadlines(request):
    today = timezone.now().date()
    tomorrow = today + timedelta(days=1)
    next_week = today + timedelta(days=7)
    two_weeks = today + timedelta(days=14)

    # Base query for JORouting objects assigned to the current user
    base_query = JORouting.objects.filter(
        approver=request.user,
        approver_sequence=6,
        status="Processing"  # Use Processing instead of Pending to match other queries
    )

    # Count job orders with target date today
    today_count = base_query.filter(
        jo_number__target_date__date=today
    ).count()

    # Count job orders with target date tomorrow
    tomorrow_count = base_query.filter(
        jo_number__target_date__date=tomorrow
    ).count()

    # Count job orders with target date in this week (after tomorrow, up to 7 days from today)
    this_week_count = base_query.filter(
        jo_number__target_date__date__gt=tomorrow,
        jo_number__target_date__date__lte=next_week
    ).count()

    # Count job orders with target date in next week (8-14 days from today)
    next_week_count = base_query.filter(
        jo_number__target_date__date__gt=next_week,
        jo_number__target_date__date__lte=two_weeks
    ).count()

    # Count job orders with target date beyond two weeks
    later_count = base_query.filter(
        jo_number__target_date__date__gt=two_weeks
    ).count()

    # Count overdue job orders (target date before today)
    overdue_count = base_query.filter(
        jo_number__target_date__lt=today
    ).count()


    data = {
        'labels': ['Overdue', 'Today', 'Tomorrow', 'This Week', 'Next Week', 'Later'],
        'datasets': [{
            'label': 'Number of Job Orders',
            'data': [overdue_count, today_count, tomorrow_count, this_week_count, next_week_count, later_count],
            'backgroundColor': [
                'rgba(220, 53, 69, 0.7)',   # Overdue (red)
                'rgba(241, 70, 104, 0.7)',  # Today (urgent)
                'rgba(255, 159, 64, 0.7)',  # Tomorrow
                'rgba(255, 193, 7, 0.7)',   # This Week
                'rgba(76, 175, 80, 0.7)',   # Next Week
                'rgba(51, 102, 255, 0.7)'   # Later
            ],
            'borderColor': [
                'rgba(220, 53, 69, 1)',
                'rgba(241, 70, 104, 1)',
                'rgba(255, 159, 64, 1)',
                'rgba(255, 193, 7, 1)',
                'rgba(76, 175, 80, 1)',
                'rgba(51, 102, 255, 1)'
            ],
            'borderWidth': 1,
            'barThickness': 30
        }]
    }

    return JsonResponse(data)

@login_required
@require_GET
def get_job_order_for_date_setting(request, jo_id):
    try:
        job_order = JOLogsheet.objects.get(id=jo_id)
        
        # Get the last approved routing record
        last_approved_routing = JORouting.objects.filter(
            jo_request=job_order,
            status='approved'
        ).order_by('-approved_at').first()
        
        last_approved_date = None
        if last_approved_routing and last_approved_routing.approved_at:
            last_approved_date = last_approved_routing.approved_at.strftime('%Y-%m-%d')
        
        # Get date of completion if priority is Urgent
        date_of_completion = None
        if job_order.date_of_completion:
            date_of_completion = job_order.date_of_completion.strftime('%Y-%m-%d')
        
        return JsonResponse({
            'status': 'success',
            'job_order': {
                'id': job_order.id,
                'jo_number': job_order.jo_number,
                'jo_color': job_order.jo_color,
                'jo_tools': job_order.jo_tools,
                'jo_type': job_order.jo_type,
                'line': job_order.line.line_name if job_order.line else None,
                'requestor': job_order.requestor,
                'status': job_order.status,
                'prepared_by': job_order.prepared_by.name if job_order.prepared_by else None,
                'priority_level': job_order.priority_level,
                'details': job_order.details,
                'date_of_completion': date_of_completion,
                'last_approved_date': last_approved_date,
                'current_target_date': job_order.target_date.strftime('%Y-%m-%d') if job_order.target_date else None,
                'current_target_date_reason': job_order.target_date_reason or ''
            }
        })
        
    except JOLogsheet.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Job order not found'
        }, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)

@login_required
def set_target_date(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)

    try:
        job_id = request.POST.get('job_id')
        target_date_str = request.POST.get('target_date')
        target_date_reason = request.POST.get('target_date_reason', '')

        if not job_id or not target_date_str:
            return JsonResponse({
                'status': 'error',
                'message': 'Missing required fields'
            }, status=400)

        target_date = datetime.strptime(target_date_str, '%Y-%m-%d')
        target_date = timezone.make_aware(target_date)

        job_order = JOLogsheet.objects.get(id=job_id)

        if job_order.in_charge != request.user:
            return JsonResponse({
                'status': 'error',
                'message': 'You are not authorized to update this job order'
            }, status=403)
        
        # Validate if reason is required based on business rules
        last_approved_routing = JORouting.objects.filter(
            jo_request=job_order,
            status='approved'
        ).order_by('-approved_at').first()
        
        reason_required = False
        reason_message = ''
        
        if last_approved_routing and last_approved_routing.approved_at:
            last_approved_date = last_approved_routing.approved_at.date()
            target_date_only = target_date.date()
            
            # Calculate the difference in days
            days_difference = (target_date_only - last_approved_date).days
            
            # Check based on JO color
            if job_order.jo_color in ['Green', 'Yellow', 'White']:
                # More than 1 month (30 days) after last approved date
                if days_difference > 30:
                    reason_required = True
                    reason_message = 'Target date is more than 1 month after the last approved routing date'
            elif job_order.jo_color == 'Orange':
                # More than 1 week (7 days) after last approved date
                if days_difference > 7:
                    reason_required = True
                    reason_message = 'Target date is more than 1 week after the last approved routing date'
        
        # Check for Urgent priority and date_of_completion
        if job_order.priority_level == 'Urgent' and job_order.date_of_completion:
            if target_date.date() > job_order.date_of_completion:
                reason_required = True
                reason_message = 'Target date is later than the date of completion for an Urgent job order'
        
        # If reason is required but not provided, return error
        if reason_required and not target_date_reason.strip():
            return JsonResponse({
                'status': 'error',
                'message': f'Reason is required: {reason_message}',
                'reason_required': True
            }, status=400)

        job_order.target_date = target_date
        job_order.target_date_reason = target_date_reason
        job_order.save()

        return JsonResponse({
            'status': 'success',
            'message': 'Target date set successfully',
            'target_date': target_date.strftime('%b %d, %Y')
        })

    except JOLogsheet.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Job order not found'
        }, status=404)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)

@login_required
def complete_job_order_old(request):
    """
    OLD DEPRECATED VERSION - DO NOT USE
    This function has been replaced by the new complete_job_order at line 755
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)

    try:
        job_id = request.POST.get('job_id')
        action_taken = request.POST.get('action_taken')
        completion_remarks = request.POST.get('completion_remarks', '')

        if not job_id or not action_taken:
            return JsonResponse({
                'status': 'error',
                'message': 'Missing required fields'
            }, status=400)

        job_order = JOLogsheet.objects.get(id=job_id)

        if job_order.in_charge != request.user:
            return JsonResponse({
                'status': 'error',
                'message': 'You are not authorized to update this job order'
            }, status=403)

        current_routing = JORouting.objects.filter(
            jo_number=job_order,
            approver=request.user,
            status='Processing'
        ).first()

        if not current_routing:
            return JsonResponse({
                'status': 'error',
                'message': 'No pending routing found'
            }, status=404)

        checker_approver = UserApprovers.objects.filter(
            user=job_order.prepared_by,
            module="Job Order",
            approver_role="Checker"
        ).first()

        if not checker_approver or not checker_approver.approver:
            return JsonResponse({
                'status': 'error',
                'message': 'Checker not found'
            }, status=404)

        # Proceed with updating only if checker is available
        job_order.action_taken = action_taken
        job_order.date_complete = timezone.now()
        job_order.status = 'Completed'
        job_order.save()

        current_routing.status = 'Approved'
        current_routing.approved_at = timezone.now()
        current_routing.remarks = completion_remarks
        current_routing.save()

        checker_routing = JORouting.objects.filter(jo_number=job_order,approver_sequence=7,status="Rejected").first()
        if checker_routing:
            checker_routing.status="Processing"
            checker_routing.save()
        else:
            JORouting.objects.create(
                jo_number=job_order,
                jo_request=job_order.prepared_by,
                approver=checker_approver.approver,
                approver_sequence=current_routing.approver_sequence + 1,
                status='Processing'
            )

        return JsonResponse({
            'status': 'success',
            'message': 'Job order marked as complete',
            'completion_date': job_order.date_complete.strftime('%b %d, %Y')
        })

    except JOLogsheet.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Job order not found'
        }, status=404)

    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)